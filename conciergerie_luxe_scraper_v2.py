"""
conciergerie_luxe_scraper_v2.py
==============================
Version ULTRA-RAPIDE avec parallélisation massive
Architecture: Phase 1 (SerpAPI séquentiel avec cache) → Phase 2 (Email scraping parallèle)
"""

import os
import time
import re
import hashlib
import json
import random
import logging
import argparse
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional, List, Dict, Tuple, Set
from dataclasses import dataclass
from urllib.parse import urljoin, urlparse
from concurrent.futures import ThreadPoolExecutor, as_completed

import requests
import pandas as pd
from dotenv import load_dotenv

# ── CHARGEMENT CONFIGURATION ───────────────────────────────────────────────────
load_dotenv()

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s | %(levelname)-8s | %(message)s',
    datefmt='%H:%M:%S'
)
logger = logging.getLogger(__name__)

# ── CONFIGURATION ──────────────────────────────────────────────────────────────
@dataclass
class Config:
    SERP_API_KEY: str = os.getenv("SERP_API_KEY", "")
    PAUSE_ENTRE_REQUETES: float = 2.0          # Réduit (était 3.5)
    PAUSE_ENTRE_PAGES: float = 1.0              # Réduit (était 2.0)
    MAX_PAGES_PAR_REQUETE: int = 3              # Réduit (était 5)
    REQUEST_TIMEOUT: int = 20
    MAX_RETRIES: int = 3
    BASE_BACKOFF: float = 2.0
    MAX_BACKOFF: float = 60.0                   # Réduit (était 120)
    JITTER_FACTOR: float = 0.2
    CACHE_DIR: str = "./cache_serpapi"
    CACHE_TTL_HOURS: int = 48

    # === PARAMÈTRES EMAIL SCRAPING PARALLÈLE ===
    EMAIL_SCRAPING_ENABLED: bool = True
    EMAIL_MAX_WORKERS: int = 15                 # Workers parallèles pour scraping email
    EMAIL_TIMEOUT: int = 8                      # Réduit drastiquement (était 15)
    EMAIL_MAX_PAGES_PER_SITE: int = 2           # Seulement homepage + contact
    EMAIL_PAUSE_BETWEEN_BATCHES: float = 0.3    # Pause entre batchs de workers

    def __post_init__(self):
        if not self.SERP_API_KEY:
            raise ValueError("SERP_API_KEY manquant. Définissez-la dans .env")

CONFIG = Config()

# ── USER-AGENTS ROTATION ───────────────────────────────────────────────────────
USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:121.0) Gecko/20100101 Firefox/121.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.1 Safari/605.1.15",
]

def get_headers() -> dict:
    return {
        'User-Agent': random.choice(USER_AGENTS),
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
        'Accept-Language': 'fr-FR,fr;q=0.9,en-US;q=0.8,en;q=0.7',
        'Accept-Encoding': 'gzip, deflate',
        'Connection': 'keep-alive',
    }

# ── CACHE DISQUE ───────────────────────────────────────────────────────────────
class DiskCache:
    def __init__(self, cache_dir: str, ttl_hours: int):
        self.cache_dir = Path(cache_dir)
        self.cache_dir.mkdir(parents=True, exist_ok=True)
        self.ttl = timedelta(hours=ttl_hours)

    def _get_path(self, key: str) -> Path:
        return self.cache_dir / f"{hashlib.md5(key.encode()).hexdigest()}.json"

    def get(self, key: str) -> Optional[dict]:
        path = self._get_path(key)
        if not path.exists():
            return None
        try:
            with open(path, 'r', encoding='utf-8') as f:
                entry = json.load(f)
            if datetime.now() - datetime.fromisoformat(entry['timestamp']) > self.ttl:
                path.unlink()
                return None
            return entry['data']
        except:
            return None

    def set(self, key: str, data: dict) -> None:
        with open(self._get_path(key), 'w', encoding='utf-8') as f:
            json.dump({'timestamp': datetime.now().isoformat(), 'data': data}, f, ensure_ascii=False)

CACHE = DiskCache(CONFIG.CACHE_DIR, CONFIG.CACHE_TTL_HOURS)

# ── UTILITAIRES ────────────────────────────────────────────────────────────────

def normaliser_nom(nom: str) -> str:
    if not nom:
        return ""
    return re.sub(r'[^\w\s]', '', re.sub(r'\s+', ' ', nom.lower().strip()))

def normaliser_adresse(adresse: str) -> str:
    if not adresse:
        return ""
    return re.sub(r'[^\w\s]', '', re.sub(r'\s+', ' ', adresse.lower().strip()))[:40]

def generer_cle_dedup(nom: str, adresse: str) -> str:
    return hashlib.md5(f"{normaliser_nom(nom)}|{normaliser_adresse(adresse)}".encode()).hexdigest()

def normaliser_telephone(tel: str) -> str:
    if not tel:
        return ""
    tel = re.sub(r'[^\d+]', '', tel)
    if tel.startswith('0') and len(tel) == 10:
        tel = '+33' + tel[1:]
    return tel

def extraire_email(texte: str) -> Optional[str]:
    if not texte:
        return None
    pattern = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
    matches = re.findall(pattern, texte)
    emails_filtrés = [e for e in matches if not any(x in e.lower() for x in ['noreply', 'no-reply', 'donotreply', 'example', 'test@', 'user@', 'utilisateur@'])]
    return emails_filtrés[0] if emails_filtrés else (matches[0] if matches else None)

def calculer_backoff(attempt: int) -> float:
    exponential = CONFIG.BASE_BACKOFF * (2 ** attempt)
    capped = min(exponential, CONFIG.MAX_BACKOFF)
    jitter = 1 + (random.random() * CONFIG.JITTER_FACTOR)
    return capped * jitter

# ── SERPAPI AVEC RETRY ─────────────────────────────────────────────────────────

def scrape_serpapi(query: str, start: int = 0) -> Tuple[List[dict], bool, Optional[str]]:
    cache_key = f"serpapi:{query}:{start}"
    cached = CACHE.get(cache_key)
    if cached is not None:
        return cached.get("results", []), cached.get("has_more", False), None

    params = {
        "engine": "google_maps",
        "q": query,
        "api_key": CONFIG.SERP_API_KEY,
        "hl": "fr",
        "gl": "fr",
        "type": "search",
        "start": start,
    }

    for attempt in range(CONFIG.MAX_RETRIES):
        try:
            response = requests.get("https://serpapi.com/search", params=params, timeout=CONFIG.REQUEST_TIMEOUT)

            if response.status_code == 429:
                retry_after = response.headers.get('Retry-After')
                wait_time = int(retry_after) if retry_after and retry_after.isdigit() else int(calculer_backoff(attempt))
                logger.warning(f"   ⏳ 429 → attente {wait_time}s (tentative {attempt+1}/{CONFIG.MAX_RETRIES})")
                time.sleep(wait_time)
                continue

            if response.status_code >= 400:
                logger.error(f"   ❌ HTTP {response.status_code}")
                time.sleep(calculer_backoff(attempt))
                continue

            response.raise_for_status()
            data = response.json()

            if "error" in data:
                if "rate limit" in str(data["error"]).lower():
                    time.sleep(calculer_backoff(attempt))
                    continue
                return [], False, str(data["error"])

            resultats = []
            for place in data.get("local_results", []):
                gps = place.get("gps_coordinates", {})
                pid = place.get("place_id", "")
                resultats.append({
                    "nom": place.get("title", ""),
                    "adresse": place.get("address", ""),
                    "telephone": normaliser_telephone(place.get("phone", "")),
                    "site_web": place.get("website", ""),
                    "email": "",
                    "note": place.get("rating", ""),
                    "nb_avis": place.get("reviews", 0),
                    "categorie": place.get("type", ""),
                    "latitude": gps.get("latitude"),
                    "longitude": gps.get("longitude"),
                    "place_id": pid,
                    "lien_google_maps": f"https://www.google.com/maps/place/?q=place_id:{pid}" if pid else "",
                    "requete_source": query,
                })

            has_more = "next" in data.get("serpapi_pagination", {}) or "next_page_token" in data.get("serpapi_pagination", {})
            CACHE.set(cache_key, {"results": resultats, "has_more": has_more})
            return resultats, has_more, None

        except requests.exceptions.Timeout:
            time.sleep(calculer_backoff(attempt))
        except requests.exceptions.RequestException as e:
            logger.error(f"   🌐 Erreur: {e}")
            time.sleep(calculer_backoff(attempt))

    return [], False, f"Échec après {CONFIG.MAX_RETRIES} tentatives"

# ── REQUÊTES OPTIMISÉES (RÉDUITES POUR RAPIDITÉ) ──────────────────────────────

REQUETES_CONCIERGERIE = [
    # Requêtes principales (français)
    "conciergerie de luxe Paris",
    "conciergerie privée Paris",
    "concierge service Paris luxe",
    "concierge personnel Paris",
    "conciergerie exclusive Paris",
    "conciergerie premium Paris",
    "conciergerie de prestige Paris",
    "conciergerie sur mesure Paris",

    # Arrondissements luxe
    "conciergerie luxe Paris 1er",
    "conciergerie luxe Paris 8e",
    "conciergerie luxe Paris 16e",
    "conciergerie luxe Paris 7e",
    "conciergerie luxe Paris 6e",
    "conciergerie luxe Paris 4e",

    # # Services spécifiques
    # "gestion location luxe Paris conciergerie",
    # "conciergerie Airbnb luxe Paris",
    # "property management luxury Paris",

    # # Anglais
    # "luxury concierge service Paris",
    # "private concierge Paris France",
    # "VIP concierge service Paris",
    # "high end concierge Paris",
    # "elite concierge Paris",
    # "luxury lifestyle management Paris",

    # # Services associés
    # "personal assistant Paris luxe",
    # "lifestyle management Paris",
    # "luxury travel concierge Paris",
    # "VIP travel service Paris",
    # "chauffeur privé Paris luxe",
    # "shopping personal shopper Paris luxe",
]

# ── PHASE 1: SCRAPING SERPAPI RAPIDE ───────────────────────────────────────────

def phase1_scraper_serpapi(max_pages: int = None) -> pd.DataFrame:
    """Phase 1: Récupération rapide des données SerpAPI (séquentiel, respect rate limit)."""
    max_pages = max_pages or CONFIG.MAX_PAGES_PAR_REQUETE

    print("=" * 75)
    print("  PHASE 1: SCRAPING SERPAPI (Google Maps)")
    print("=" * 75)
    print(f"   • Requêtes: {len(REQUETES_CONCIERGERIE)}")
    print(f"   • Pages max/requête: {max_pages}")
    print(f"   • Pause entre requêtes: {CONFIG.PAUSE_ENTRE_REQUETES}s")
    print()

    tous_resultats: List[dict] = []
    vus: Set[str] = set()
    stats = {"requetes_ok": 0, "requetes_ko": 0, "total_trouves": 0, "nouveaux": 0}

    debut = time.time()

    for idx, requete in enumerate(REQUETES_CONCIERGERIE, 1):
        logger.info(f"[{idx:02d}/{len(REQUETES_CONCIERGERIE)}] '{requete}'")

        page = 0
        requete_ok = False

        while page < max_pages:
            start = page * 20
            resultats, a_encore, erreur = scrape_serpapi(requete, start=start)

            if erreur and not resultats:
                logger.error(f"      💥 {erreur}")
                break

            if resultats:
                requete_ok = True

            nouveaux = 0
            for place in resultats:
                cle = generer_cle_dedup(place["nom"], place["adresse"])
                if cle not in vus and place["nom"]:
                    vus.add(cle)
                    stats["total_trouves"] += 1
                    stats["nouveaux"] += 1
                    tous_resultats.append(place)
                    nouveaux += 1

            logger.info(f"      📊 {len(resultats)} résultats, {nouveaux} nouveaux → Total unique: {len(tous_resultats)}")

            if not a_encore or len(resultats) == 0:
                break

            page += 1
            time.sleep(CONFIG.PAUSE_ENTRE_PAGES)

        if requete_ok:
            stats["requetes_ok"] += 1
        else:
            stats["requetes_ko"] += 1

        if idx < len(REQUETES_CONCIERGERIE):
            time.sleep(CONFIG.PAUSE_ENTRE_REQUETES + random.uniform(0, 0.5))

    duree = time.time() - debut

    print(f"✅ Phase 1 terminée en {duree:.1f}s")
    print(f"   • Requêtes OK: {stats['requetes_ok']}/{len(REQUETES_CONCIERGERIE)}")
    print(f"   • Résultats bruts: {stats['total_trouves']}")
    print(f"   • Entrées uniques: {stats['nouveaux']}")

    return pd.DataFrame(tous_resultats)

# ── PHASE 2: SCRAPING EMAIL PARALLÈLE ──────────────────────────────────────────

def scraper_email_unique(place: dict) -> dict:
    """
    Scrape l'email d'un seul site web. Appelé par ThreadPoolExecutor.
    Ultra-rapide: timeout court, seulement 2 pages testées.
    """
    url = place.get("site_web", "")
    if not url or place.get("email"):
        return place

    parsed = urlparse(url)
    base_url = f"{parsed.scheme}://{parsed.netloc}"

    pages = [url, urljoin(base_url, "/contact")]
    headers = get_headers()

    for page_url in pages[:CONFIG.EMAIL_MAX_PAGES_PER_SITE]:
        try:
            response = requests.get(page_url, headers=headers, timeout=CONFIG.EMAIL_TIMEOUT, allow_redirects=True)
            if response.status_code == 200:
                email = extraire_email(response.text)
                if email:
                    place["email"] = email
                    return place
        except:
            pass

    return place

def phase2_scraper_email_parallele(df: pd.DataFrame) -> pd.DataFrame:
    """Phase 2: Scraping d'email en parallèle massif avec ThreadPoolExecutor."""
    if df.empty or not CONFIG.EMAIL_SCRAPING_ENABLED:
        return df

    # Filtrer seulement les entrées avec un site web mais sans email
    a_scraper = df[df["site_web"] != ""].copy()

    if a_scraper.empty:
        logger.info("   ℹ️ Aucun site web à scraper pour email")
        return df

    print("" + "=" * 75)
    print("  PHASE 2: SCRAPING EMAIL EN PARALLÈLE")
    print("=" * 75)
    print(f"   • Entrées à traiter: {len(a_scraper)}")
    print(f"   • Workers parallèles: {CONFIG.EMAIL_MAX_WORKERS}")
    print(f"   • Timeout par requête: {CONFIG.EMAIL_TIMEOUT}s")
    print(f"   • Pages testées/site: {CONFIG.EMAIL_MAX_PAGES_PER_SITE}")
    print()

    debut = time.time()
    emails_trouves = 0

    # Convertir en liste de dicts pour le threading
    places = a_scraper.to_dict('records')

    with ThreadPoolExecutor(max_workers=CONFIG.EMAIL_MAX_WORKERS) as executor:
        # Soumettre toutes les tâches
        future_to_place = {
            executor.submit(scraper_email_unique, place): place 
            for place in places
        }

        total = len(future_to_place)
        completed = 0

        for future in as_completed(future_to_place):
            completed += 1
            place = future_to_place[future]

            try:
                result = future.result()
                if result.get("email") and not place.get("email"):
                    emails_trouves += 1
                    logger.info(f"   📧 [{completed}/{total}] {place.get('nom', '')[:35]}... → {result['email']}")
                elif completed % 10 == 0:
                    logger.info(f"   ⏳ Progression: {completed}/{total} ({completed/total*100:.0f}%)")
            except Exception as e:
                logger.debug(f"   ⚠️ Erreur scraping email pour {place.get('nom', '')[:35]}...: {e}")

    duree = time.time() - debut

    # Reconstruire le DataFrame avec les emails trouvés
    result_map = {generer_cle_dedup(p["nom"], p["adresse"]): p for p in places}

    for idx, row in df.iterrows():
        cle = generer_cle_dedup(row["nom"], row["adresse"])
        if cle in result_map and result_map[cle].get("email"):
            df.at[idx, "email"] = result_map[cle]["email"]

    print(f"✅ Phase 2 terminée en {duree:.1f}s")
    print(f"   • Emails trouvés: {emails_trouves}/{len(a_scraper)}")
    print(f"   • Taux de succès: {emails_trouves/len(a_scraper)*100:.1f}%")

    return df

# ── ANALYSE COMPLÉTUDE ─────────────────────────────────────────────────────────

def analyser_completude(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df

    df = df.copy()
    statuts, manquants_list, priorites, scores = [], [], [], []

    for _, row in df.iterrows():
        manquants = []
        score = 0
        champs = {"telephone": row.get("telephone"), "email": row.get("email"), 
                  "site_web": row.get("site_web"), "adresse": row.get("adresse")}

        for champ, val in champs.items():
            if val and str(val).strip():
                score += 25
            else:
                manquants.append(champ)

        if score == 100:
            statut, priorite = "✅ COMPLET", "A"
        elif score >= 75:
            statut, priorite = "⚠️ QUASI COMPLET", "B"
        elif score >= 50:
            statut, priorite = "⚡ MOYENNEMENT COMPLET", "B"
        else:
            statut, priorite = "❌ INCOMPLET", "C"

        statuts.append(statut)
        manquants_list.append(", ".join(manquants) if manquants else "Aucun")
        priorites.append(priorite)
        scores.append(score)

    df["statut_completude"] = statuts
    df["champs_manquants"] = manquants_list
    df["priorite_completion"] = priorites
    df["score_completude"] = scores
    df["email_manuel"] = ""
    df["telephone_manuel"] = ""
    df["notes"] = ""

    ordre = {"C": 0, "B": 1, "A": 2}
    df["_ordre"] = df["priorite_completion"].map(ordre)
    df = df.sort_values(["_ordre", "score_completude", "nom"], ascending=[True, False, True])
    df = df.drop(columns=["_ordre"])

    return df.reset_index(drop=True)

# ── EXPORT EXCEL ───────────────────────────────────────────────────────────────

def export_excel(df: pd.DataFrame, fichier: str):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        df.to_csv(fichier.replace('.xlsx', '.csv'), index=False, encoding='utf-8-sig')
        return

    df_incomplet = df[df["priorite_completion"] == "C"].copy()
    df_a_completer = df[df["priorite_completion"] == "B"].copy()
    df_complet = df[df["priorite_completion"] == "A"].copy()

    cols = ["priorite_completion", "statut_completude", "score_completude", "champs_manquants",
            "nom", "adresse", "telephone", "telephone_manuel", "email", "email_manuel",
            "site_web", "note", "nb_avis", "categorie", "lien_google_maps", "notes"]

    fills = {
        "header": PatternFill("solid", fgColor="1F4E78"),
        "incomplet": PatternFill("solid", fgColor="FCE4D6"),
        "a_completer": PatternFill("solid", fgColor="FFF2CC"),
        "complet": PatternFill("solid", fgColor="E2EFDA"),
    }

    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    bold_font = Font(name="Calibri", bold=True, size=10)
    normal_font = Font(name="Calibri", size=10)
    link_font = Font(name="Calibri", size=10, color="0563C1", underline="single")
    border = Border(bottom=Side(style="thin", color="D0D0D0"), right=Side(style="thin", color="D0D0D0"))

    wb = Workbook()
    wb.remove(wb.active)

    def creer_onglet(wb, df_onglet, titre, fill_data):
        ws = wb.create_sheet(titre)
        if df_onglet.empty:
            ws.cell(1, 1, "Aucune entrée").font = normal_font
            return

        df_out = df_onglet[cols].copy()
        headers = {
            "priorite_completion": "Priorité", "statut_completude": "Statut",
            "score_completude": "Score %", "champs_manquants": "Manquants",
            "nom": "Nom", "adresse": "Adresse", "telephone": "Tél (auto)",
            "telephone_manuel": "Tél (manuel)", "email": "Email (auto)",
            "email_manuel": "Email (manuel)", "site_web": "Site Web",
            "note": "Note", "nb_avis": "Avis", "categorie": "Catégorie",
            "lien_google_maps": "🔗 Maps", "notes": "Notes",
        }

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
        c = ws.cell(1, 1, f"CONCIERGERIES LUXE PARIS — {titre}")
        c.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
        c.fill = fills["header"]
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        for col_idx, col_name in enumerate(cols, 1):
            cell = ws.cell(2, col_idx, headers.get(col_name, col_name))
            cell.font = header_font
            cell.fill = fills["header"]
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[2].height = 40

        for row_idx, (_, data_row) in enumerate(df_out.iterrows(), 3):
            for col_idx, col_name in enumerate(cols, 1):
                val = data_row[col_name]
                cell = ws.cell(row_idx, col_idx)
                if col_name in ("lien_google_maps", "site_web") and val:
                    cell.value = str(val)
                    cell.font = link_font
                    cell.hyperlink = str(val)
                else:
                    cell.value = val if pd.notna(val) and val != "" else ""
                    cell.font = bold_font if col_name == "nom" else normal_font
                cell.fill = fill_data
                cell.alignment = Alignment(vertical="top", wrap_text=(col_name in ("adresse", "champs_manquants", "notes")))
                cell.border = border
            ws.row_dimensions[row_idx].height = 30

        largeurs = {"priorite_completion": 8, "statut_completude": 22, "score_completude": 8,
                    "champs_manquants": 20, "nom": 35, "adresse": 40, "telephone": 16,
                    "telephone_manuel": 16, "email": 30, "email_manuel": 30,
                    "site_web": 32, "note": 8, "nb_avis": 8, "categorie": 20,
                    "lien_google_maps": 24, "notes": 30}
        for col_idx, col_name in enumerate(cols, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = largeurs.get(col_name, 15)

        ws.freeze_panes = "A3"

    creer_onglet(wb, df_incomplet, "1_INCOMPLETS", fills["incomplet"])
    creer_onglet(wb, df_a_completer, "2_A_COMPLETER", fills["a_completer"])
    creer_onglet(wb, df_complet, "3_COMPLETS", fills["complet"])

    # Stats
    ws_stats = wb.create_sheet("STATISTIQUES", 0)
    ws_stats.column_dimensions["A"].width = 40
    ws_stats.column_dimensions["B"].width = 15
    c = ws_stats.cell(1, 1, "STATISTIQUES")
    c.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    c.fill = fills["header"]
    ws_stats.merge_cells("A1:B1")
    ws_stats.row_dimensions[1].height = 28

    total = len(df)
    complet = len(df_complet)
    taux = (complet / total * 100) if total > 0 else 0

    stats = [
        ("Total conciergeries", total), ("", ""),
        ("✅ Complets (100%)", len(df_complet)),
        ("⚠️ Quasi complets (75-99%)", len(df[df["score_completude"] >= 75]) - len(df_complet)),
        ("⚡ Moyennement complets (50-74%)", len(df[(df["score_completude"] >= 50) & (df["score_completude"] < 75)])),
        ("❌ Incomplets (<50%)", len(df_incomplet)), ("", ""),
        ("Taux de complétude", f"{taux:.1f}%"),
        ("Score moyen", f"{df['score_completude'].mean():.1f}%"), ("", ""),
        ("Sans téléphone", len(df[df["telephone"] == ""])),
        ("Sans email", len(df[df["email"] == ""])),
        ("Sans site web", len(df[df["site_web"] == ""])),
        ("Sans adresse", len(df[df["adresse"] == ""])),
    ]

    for i, (label, valeur) in enumerate(stats, 2):
        c1 = ws_stats.cell(i, 1, label)
        c1.font = bold_font if label else normal_font
        c2 = ws_stats.cell(i, 2, valeur)
        c2.font = bold_font if label else normal_font
        c2.alignment = Alignment(horizontal="right")
        if label:
            c1.fill = PatternFill("solid", fgColor="F2F2F2")
            c2.fill = PatternFill("solid", fgColor="F2F2F2")
        ws_stats.row_dimensions[i].height = 22

    wb.save(fichier)

    print(f"✅ Fichier Excel: {fichier}")
    print(f"   ✅ Complets: {len(df_complet)} | ⚠️ À compléter: {len(df_a_completer)} | ❌ Incomplets: {len(df_incomplet)}")
    print(f"   TOTAL: {total} | Taux: {taux:.1f}%")

# ── MAIN ─────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Scraper conciergeries luxe Paris")
    parser.add_argument("--no-email", action="store_true", help="Désactiver le scraping d'email (beaucoup plus rapide)")
    parser.add_argument("--workers", type=int, default=15, help="Nombre de workers parallèles pour email (défaut: 15)")
    parser.add_argument("--pages", type=int, default=3, help="Pages max par requête SerpAPI (défaut: 3)")
    parser.add_argument("--timeout", type=int, default=8, help="Timeout scraping email en secondes (défaut: 8)")
    args = parser.parse_args()

    if args.no_email:
        CONFIG.EMAIL_SCRAPING_ENABLED = False
        print("⚡ Mode RAPIDE: scraping d'email désactivé")

    CONFIG.EMAIL_MAX_WORKERS = args.workers
    CONFIG.MAX_PAGES_PAR_REQUETE = args.pages
    CONFIG.EMAIL_TIMEOUT = args.timeout

    print("" + "=" * 75)
    print("  SCRAPER CONCIERGERIES LUXE - PARIS [V2 ULTRA-RAPIDE]")
    print("=" * 75)
    print(f"⚙️  Configuration:")
    print(f"   • Requêtes: {len(REQUETES_CONCIERGERIE)}")
    print(f"   • Pages/requête: {CONFIG.MAX_PAGES_PAR_REQUETE}")
    print(f"   • Pause requêtes: {CONFIG.PAUSE_ENTRE_REQUETES}s")
    print(f"   • Email scraping: {'OUI' if CONFIG.EMAIL_SCRAPING_ENABLED else 'NON'}")
    if CONFIG.EMAIL_SCRAPING_ENABLED:
        print(f"   • Workers email: {CONFIG.EMAIL_MAX_WORKERS}")
        print(f"   • Timeout email: {CONFIG.EMAIL_TIMEOUT}s")
    print(f"   • Cache: {CONFIG.CACHE_DIR}")

    reponse = input("Lancer ? (o/n) : ").strip().lower()
    if reponse not in ('o', 'oui', 'y', 'yes'):
        print("Annulé.")
        return

    debut_total = time.time()

    # === PHASE 1: SerpAPI ===
    df = phase1_scraper_serpapi()

    if df.empty:
        print("❌ Aucun résultat.")
        return

    # === PHASE 2: Email parallèle ===
    if CONFIG.EMAIL_SCRAPING_ENABLED:
        df = phase2_scraper_email_parallele(df)

    # === PHASE 3: Analyse + Export ===
    print("🔍 Analyse complétude...")
    df = analyser_completude(df)

    fichier = f"conciergeries_luxe1_paris_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    export_excel(df, fichier)

    duree_totale = time.time() - debut_total
    print(f"" + "=" * 75)
    print(f"✅ TERMINÉ EN {duree_totale:.1f} SECONDES")
    print("=" * 75)
    print(f"📁 Fichier: {fichier}")

if __name__ == "__main__":
    main()