import asyncio
import html as html_module
import json
import re
import time
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ─── Configuration ─────────────────────────────────────────────────────────────

DEFAULT_INPUT   = "familinparis_activites_enrichi_V1.xlsx"
OUTPUT_SUFFIX   = "_enrichi_V2"
CONCURRENCY     = 20
REQUEST_TIMEOUT = 15

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "fr-FR,fr;q=0.9,en;q=0.8",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Referer": "https://www.familinparis.fr/",
}

# ─── Extraction 1 : Astro Island props= (feverup) ─────────────────────────────

RE_PROPS = re.compile(
    r'<astro-island\b[^>]*?'
    r'opts="[^"]*PlanLocationInfo[^"]*"'
    r'[^>]*?'
    r'props="([^"]*)"',
    re.DOTALL
)
RE_PROPS2 = re.compile(
    r'<astro-island\b[^>]*?'
    r'props="([^"]*)"'
    r'[^>]*?'
    r'opts="[^"]*PlanLocationInfo[^"]*"',
    re.DOTALL
)

def _parse_astro_props(raw_html: str) -> tuple[str, str]:
    """Extrait name + address depuis <astro-island props=> (feverup)."""
    for pattern in (RE_PROPS, RE_PROPS2):
        m = pattern.search(raw_html)
        if m:
            try:
                props = json.loads(html_module.unescape(m.group(1)))
                place   = props["place"][1]
                name    = place.get("name",    [0, ""])[1]
                address = place.get("address", [0, ""])[1]
                return str(name or ""), str(address or "")
            except Exception:
                pass
    return "", ""


# ─── Extraction 2 : <ul class="post-meta"> (familinparis) ─────────────────────

# Capture toute la liste post-meta
RE_POST_META = re.compile(
    r'<ul[^>]*class=["\'][^"\']*post-meta[^"\']*["\'][^>]*>(.*?)</ul>',
    re.DOTALL | re.IGNORECASE
)

# Capture chaque <li> dans la liste
RE_LI = re.compile(r'<li[^>]*>(.*?)</li>', re.DOTALL | re.IGNORECASE)

# Nettoie toutes les balises HTML
RE_TAGS = re.compile(r'<[^>]+>')

# Clé "Adresse" avec variations (espaces insécables, accents, deux-points)
RE_KEY_ADRESSE = re.compile(r'adresse\s*[:\u00a0]', re.IGNORECASE)

# Clé "Lieu" pour récupérer le nom du lieu
RE_KEY_LIEU = re.compile(r'lieu\s*[:\u00a0]', re.IGNORECASE)


def _strip_html(s: str) -> str:
    """Supprime les balises HTML, decode les entités et normalise les espaces."""
    s = html_module.unescape(s)
    return " ".join(RE_TAGS.sub(" ", s).split()).strip()


def _parse_post_meta(raw_html: str) -> tuple[str, str]:
    """
    Extrait location_name (Lieu) et location_address (Adresse) 
    depuis <ul class="post-meta">.
    
    Structure attendue :
      <ul class="post-meta">
        <li><span class="post-meta-key">Adresse&nbsp;:</span> 8 bis rue …</li>
        <li><span class="post-meta-key">Lieu&nbsp;:</span> Nom du lieu</li>
      </ul>
    """
    location_name    = ""
    location_address = ""

    m = RE_POST_META.search(raw_html)
    if not m:
        return "", ""

    ul_content = m.group(1)

    for li_m in RE_LI.finditer(ul_content):
        li_text = _strip_html(li_m.group(1))

        # Cherche "Adresse :"
        if RE_KEY_ADRESSE.search(li_text) and not location_address:
            # Retire la clé pour garder uniquement la valeur
            value = RE_KEY_ADRESSE.sub("", li_text, count=1).strip(" :–-\u00a0")
            if value:
                location_address = value

        # Cherche "Lieu :" pour le nom du lieu
        elif RE_KEY_LIEU.search(li_text) and not location_name:
            value = RE_KEY_LIEU.sub("", li_text, count=1).strip(" :–-\u00a0")
            if value:
                location_name = value

    return location_name, location_address


# ─── Extraction 3 : JSON-LD schema.org ────────────────────────────────────────

RE_JSONLD = re.compile(
    r'<script[^>]+type=["\']application/ld\+json["\'][^>]*>(.*?)</script>',
    re.DOTALL | re.IGNORECASE
)

def _parse_jsonld(raw_html: str) -> tuple[str, str]:
    """Extrait name + adresse depuis les blocs JSON-LD schema.org."""
    for m in RE_JSONLD.finditer(raw_html):
        try:
            data = json.loads(m.group(1).strip())
        except Exception:
            continue

        # Peut être un objet ou une liste
        items = data if isinstance(data, list) else [data]

        for item in items:
            # Types pertinents : Place, LocalBusiness, Event, etc.
            schema_type = item.get("@type", "")
            if not isinstance(schema_type, str):
                schema_type = " ".join(schema_type) if schema_type else ""

            name    = ""
            address = ""

            # Nom
            raw_name = item.get("name", "")
            if isinstance(raw_name, str):
                name = raw_name.strip()

            # Adresse — peut être une string ou un objet PostalAddress
            loc = item.get("location", "")
            raw_addr = item.get("address", loc if isinstance(loc, dict) else "")
            if isinstance(raw_addr, str):
                address = raw_addr.strip()
            elif isinstance(raw_addr, dict):
                parts = []
                for field in ("streetAddress", "addressLocality", "postalCode", "addressRegion"):
                    val = raw_addr.get(field, "")
                    if val:
                        parts.append(str(val).strip())
                address = ", ".join(parts)

            if name or address:
                return name, address

    return "", ""


# ─── Extraction 4 : itemprop / meta fallback ──────────────────────────────────

RE_ITEMPROP_ADDR = re.compile(
    r'itemprop=["\'](?:streetAddress|address)["\'][^>]*>([^<]+)',
    re.IGNORECASE
)
RE_ITEMPROP_NAME = re.compile(
    r'itemprop=["\']name["\'][^>]*>([^<]+)',
    re.IGNORECASE
)
RE_META_PLACE = re.compile(
    r'<meta[^>]+name=["\']geo\.placename["\'][^>]+content=["\']([^"\']+)["\']',
    re.IGNORECASE
)

def _parse_microdata(raw_html: str) -> tuple[str, str]:
    """Fallback : itemprop=address et itemprop=name."""
    name = address = ""

    m_addr = RE_ITEMPROP_ADDR.search(raw_html)
    if m_addr:
        address = _strip_html(m_addr.group(1))

    m_name = RE_ITEMPROP_NAME.search(raw_html)
    if m_name:
        name = _strip_html(m_name.group(1))

    if not address:
        m_meta = RE_META_PLACE.search(raw_html)
        if m_meta:
            address = _strip_html(m_meta.group(1))

    return name, address


# ─── Pipeline d'extraction (multi-stratégie) ──────────────────────────────────

def extract_location(raw_html: str) -> tuple[str, str, str]:
    """
    Tente les 4 stratégies dans l'ordre.
    Retourne (location_name, location_address, méthode_utilisée).
    """

    # 1. Astro Island (feverup)
    name, address = _parse_astro_props(raw_html)
    if name or address:
        return name, address, "astro-props"

    # 2. post-meta list (familinparis)
    name, address = _parse_post_meta(raw_html)
    if name or address:
        return name, address, "post-meta"

    # 3. JSON-LD
    name, address = _parse_jsonld(raw_html)
    if name or address:
        return name, address, "json-ld"

    # 4. Microdata / itemprop
    name, address = _parse_microdata(raw_html)
    if name or address:
        return name, address, "microdata"

    return "", "", "non trouvé"


# ─── HTTP async ────────────────────────────────────────────────────────────────

async def fetch_location(session, url: str, semaphore: asyncio.Semaphore) -> dict:
    async with semaphore:
        try:
            resp = await session.get(url, timeout=REQUEST_TIMEOUT, follow_redirects=True)
            resp.raise_for_status()
            raw_html = resp.text

            name, address, method = extract_location(raw_html)
            return {
                "location_name":    name,
                "location_address": address,
                "method":           method,
                "error":            "" if (name or address) else "aucune donnée trouvée",
            }
        except Exception as e:
            return {
                "location_name":    "",
                "location_address": "",
                "method":           "erreur",
                "error":            str(e)[:80],
            }


# ─── Lecture Excel ─────────────────────────────────────────────────────────────

def clean(t: str) -> str:
    return " ".join(t.split()).strip()

def read_excel(filepath: str):
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    header_row = url_col = name_col = None

    # Cherche la ligne d'en-têtes
    for ri in range(1, min(7, ws.max_row + 1)):
        for ci in range(1, ws.max_column + 1):
            val = str(ws.cell(row=ri, column=ci).value or "").strip().upper()
            if val in ("URL", "LIEN", "URL FEVERUP", "URL FAMILINPARIS"):
                header_row = ri
                url_col    = ci
            if "NOM" in val and any(x in val for x in ("GÉN", "GEN", "GÉNÉ", "GENERAL", "TITRE")):
                name_col = ci

    # Fallback : première cellule contenant une URL reconnue
    if not url_col:
        for ri in range(1, ws.max_row + 1):
            for ci in range(1, ws.max_column + 1):
                val = str(ws.cell(row=ri, column=ci).value or "")
                if any(domain in val for domain in ("feverup.com/m/", "familinparis.fr/")):
                    url_col    = ci
                    header_row = max(1, ri - 1)
                    break
            if url_col:
                break

    if not url_col:
        raise ValueError("Colonne URL introuvable.")

    activities = []
    for ri in range((header_row or 1) + 1, ws.max_row + 1):
        cell = ws.cell(row=ri, column=url_col)
        url  = cell.value
        if not url and cell.hyperlink:
            url = getattr(cell.hyperlink, "target", str(cell.hyperlink))
        url = str(url or "").strip()
        if not url or not url.startswith("http"):
            continue
        name = clean(str(ws.cell(row=ri, column=name_col).value or "")) if name_col else ""
        activities.append({"row": ri, "url": url, "name": name})

    print(f"    📋  {len(activities)} URLs détectées (en-têtes L{header_row}, URL col {url_col})")
    return wb, ws, activities, header_row or 3


# ─── Colonnes Excel ────────────────────────────────────────────────────────────

def get_col_letter(n: int) -> str:
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s

def add_columns(ws, header_row: int) -> tuple[int, int, int]:
    last = ws.max_column
    HDR  = PatternFill("solid", fgColor="2E7D32")   # vert foncé
    thin = Side(style="thin", color="81C784")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
    col_n, col_a, col_m = last + 1, last + 2, last + 3

    for col, label, w in [
        (col_n, "Location Name",    38),
        (col_a, "Location Address", 55),
        (col_m, "Méthode",          18),
    ]:
        c = ws.cell(row=header_row, column=col, value=label)
        c.font      = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        c.fill      = HDR
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = brd
        ws.column_dimensions[get_col_letter(col)].width = w

    return col_n, col_a, col_m

def write_row(ws, row: int, col_n: int, col_a: int, col_m: int,
              name: str, addr: str, method: str):
    ODD  = PatternFill("solid", fgColor="F1F8E9")
    EVEN = PatternFill("solid", fgColor="FFFFFF")
    thin = Side(style="thin", color="A5D6A7")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill = ODD if row % 2 == 0 else EVEN

    for col, val in [(col_n, name or "—"), (col_a, addr or "—"), (col_m, method)]:
        c = ws.cell(row=row, column=col, value=val)
        c.font      = Font(name="Arial", size=10)
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.fill      = fill
        c.border    = brd


# ─── Pipeline principal ────────────────────────────────────────────────────────

async def enrich(input_file: str, output_file: str):
    print(f"\n📂  Lecture de {input_file}...")
    wb, ws, activities, header_row = read_excel(input_file)
    total = len(activities)
    if not total:
        print("⚠️  Aucune URL trouvée.")
        return

    col_n, col_a, col_m = add_columns(ws, header_row)

    try:
        import httpx
    except ImportError:
        print("❌  httpx manquant — pip install httpx")
        return

    sem        = asyncio.Semaphore(CONCURRENCY)
    done = found = 0
    stats: dict[str, int] = {}
    start_time = time.time()

    print(f"\n🚀  {total} pages  |  {CONCURRENCY} requêtes parallèles  |  sans navigateur\n")
    print(f"  {'#':>5}  {'Nom':<38}  {'Méthode':<12}  {'Location Name':<28}  Adresse")
    print("  " + "─" * 115)

    BATCH = 50
    async with httpx.AsyncClient(headers=HEADERS, verify=False) as session:
        tasks = [fetch_location(session, a["url"], sem) for a in activities]

        for b in range(0, total, BATCH):
            batch_acts  = activities[b : b + BATCH]
            batch_tasks = tasks[b : b + BATCH]
            results     = await asyncio.gather(*batch_tasks, return_exceptions=True)

            for act, res in zip(batch_acts, results):
                done += 1
                if isinstance(res, Exception):
                    ln, la, meth = "", "", "exception"
                elif res.get("error") and not res.get("location_name") and not res.get("location_address"):
                    ln, la, meth = "", "", res.get("method", "?")
                else:
                    ln   = res["location_name"]
                    la   = res["location_address"]
                    meth = res["method"]
                    if ln or la:
                        found += 1
                    stats[meth] = stats.get(meth, 0) + 1

                write_row(ws, act["row"], col_n, col_a, col_m, ln, la, meth)

                print(
                    f"  {done:>5}/{total}  {act['name'][:36]:<38}  "
                    f"{meth:<12}  {(ln or '—')[:26]:<28}  {(la or '—')[:42]}"
                )

            wb.save(output_file)
            pct     = done * 100 // total
            elapsed = int(time.time() - start_time)
            print(f"\n  💾  {done}/{total} ({pct}%) — {elapsed}s → {output_file}\n")

    elapsed = int(time.time() - start_time)
    print(f"\n{'='*65}")
    print(f"  ✅  {found}/{total} localisations trouvées en {elapsed}s")
    print(f"\n  📊  Répartition par méthode :")
    for meth, count in sorted(stats.items(), key=lambda x: -x[1]):
        pct = count * 100 // total
        print(f"       {meth:<15} : {count:>4} ({pct}%)")
    print(f"\n  💾  {output_file}")
    print(f"{'='*65}")


# ─── Main ──────────────────────────────────────────────────────────────────────

async def main():
    print("=" * 65)
    print("  Enrichissement Localisation v4  (multi-sources)")
    print(f"  Stratégies : astro-props → post-meta → json-ld → microdata")
    print("=" * 65)

    inp = input(f"\n  Fichier Excel [{DEFAULT_INPUT}] : ").strip() or DEFAULT_INPUT
    if not Path(inp).exists():
        print(f"❌  Introuvable : {inp}")
        return

    p   = Path(inp)
    out = str(p.parent / (p.stem + OUTPUT_SUFFIX + p.suffix))
    print(f"  Sortie : {out}\n")

    await enrich(inp, out)

if __name__ == "__main__":
    import warnings
    warnings.filterwarnings("ignore")
    asyncio.run(main())