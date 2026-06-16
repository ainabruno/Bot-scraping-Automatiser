import time, re, requests, pandas as pd
from datetime import datetime

# ── CONFIGURATION ──────────────────────────────────────────────────────────────
try:
    from configB2b import SERP_API_KEY, PAUSE_ENTRE_REQUETES, REQUEST_TIMEOUT
except ImportError:
    SERP_API_KEY          = "d99b32948181e4132cbd1c3891fb971ab722c80e15fbebbd34a82d92ea9d76ac"   # ← remplacer
    PAUSE_ENTRE_REQUETES  = 5
    REQUEST_TIMEOUT       = 30

MAX_PAGES_PAR_REQUETE = 5

# ── SECTEURS CIBLES (vrais besoins, paient vite pour des leads) ────────────────
# Choisis UN secteur à la fois pour un ciblage précis

MARCHES = {

    # ── AGENCES IMMOBILIÈRES (France) ─────────────────────────────────────────
    # Besoin : trouver des vendeurs/acheteurs. Paient bien pour des contacts.
    "Immobilier_France": [
        "agence immobilière Paris",
        "agence immobilière Lyon",
        "agence immobilière Marseille",
        "agence immobilière Bordeaux",
        "agence immobilière Toulouse",
        "agence immobilière Nantes",
        "agence immobilière Nice",
        "agence immobilière Lille",
        "agence immobilière Strasbourg",
        "agence immobilière Montpellier",
        "agence immobilière Rennes",
        "promoteur immobilier Paris",
        "promoteur immobilier Lyon",
        "chasseur immobilier France",
        "gestionnaire locatif Paris",
    ],

    # ── AGENCES MARKETING & COMMUNICATION (France) ────────────────────────────
    # Besoin : trouver des clients PME. Budget élevé, décision rapide.
    "Agences_Marketing_France": [
        "agence marketing digital Paris",
        "agence communication Lyon",
        "agence SEO Paris",
        "agence publicité Marseille",
        "agence web Bordeaux",
        "agence marketing digital Lyon",
        "agence réseaux sociaux Paris",
        "agence inbound marketing France",
        "agence growth hacking Paris",
        "agence content marketing France",
        "agence emailing France",
        "agence marketing digital Toulouse",
        "agence branding Paris",
        "agence marketing digital Nantes",
        "agence marketing digital Lille",
    ],

    # ── COACHS & CONSULTANTS (France) ─────────────────────────────────────────
    # Besoin : remplir leur agenda. Achètent des leads en urgence.
    # "Coachs_Consultants_France": [
    #     "coach business France",
    #     "consultant marketing France",
    #     "coach entrepreneur Paris",
    #     "consultant formation Paris",
    #     "coach life business France",
    #     "consultant RH Paris",
    #     "coach vente France",
    #     "formateur professionnel Paris",
    #     "consultant stratégie Paris",
    #     "coach développement personnel France",
    #     "consultant digital France",
    #     "coach commercial France",
    #     "consultant management Paris",
    #     "formateur vente France",
    #     "consultant croissance startup France",
    # ],

    # # ── STARTUPS & SaaS (France) ───────────────────────────────────────────────
    # # Besoin : acquisition clients B2B. Budget marketing disponible.
    # "Startups_SaaS_France": [
    #     "startup SaaS Paris",
    #     "startup logiciel Lyon",
    #     "éditeur logiciel Paris",
    #     "startup fintech France",
    #     "startup RH logiciel France",
    #     "startup CRM France",
    #     "startup e-commerce solution France",
    #     "startup marketing automation France",
    #     "éditeur SaaS Bordeaux",
    #     "startup B2B Paris",
    #     "startup facturation logiciel France",
    #     "startup gestion PME France",
    #     "startup comptabilité logiciel France",
    #     "solution logicielle PME France",
    #     "startup growth Paris",
    # ],

    # # ── CABINETS COMPTABLES & AVOCATS (France) ────────────────────────────────
    # # Besoin : nouveaux clients PME. Secteur très rentable, peu prospecté.
    # "Cabinets_Experts_France": [
    #     "cabinet comptable Paris",
    #     "expert comptable Lyon",
    #     "cabinet expertise comptable Marseille",
    #     "cabinet avocat affaires Paris",
    #     "avocat droit des affaires Lyon",
    #     "cabinet conseil fiscal Paris",
    #     "expert comptable PME Bordeaux",
    #     "cabinet comptable Toulouse",
    #     "expert comptable Nantes",
    #     "cabinet comptable Lille",
    #     "avocat droit commercial Paris",
    #     "cabinet audit Paris",
    #     "cabinet conseil entreprise France",
    #     "expert comptable startup France",
    #     "cabinet comptable Strasbourg",
    # ],

    # # ── E-COMMERCE (France) ────────────────────────────────────────────────────
    # # Besoin : fournisseurs, clients B2B, partenaires logistiques.
    # "Ecommerce_France": [
    #     "boutique en ligne France",
    #     "e-commerce mode France",
    #     "e-commerce cosmétique France",
    #     "dropshipping France",
    #     "marketplace vendeur France",
    #     "e-commerce alimentation France",
    #     "boutique Shopify France",
    #     "e-commerce maison décoration France",
    #     "e-commerce sport France",
    #     "e-commerce bio France",
    #     "pure player France",
    #     "e-commerce enfant France",
    #     "e-commerce bijoux France",
    #     "e-commerce électronique France",
    #     "e-commerce luxe France",
    # ],

    # # ── BELGIQUE — AGENCES & PME ───────────────────────────────────────────────
    # "PME_Belgique": [
    #     "agence marketing Bruxelles",
    #     "agence immobilière Bruxelles",
    #     "coach business Belgique",
    #     "consultant PME Bruxelles",
    #     "agence communication Bruxelles",
    #     "startup SaaS Belgique",
    #     "agence web Bruxelles",
    #     "expert comptable Bruxelles",
    #     "agence SEO Belgique",
    #     "cabinet conseil Bruxelles",
    # ],

    # # ── CANADA FRANCOPHONE ─────────────────────────────────────────────────────
    # "PME_Canada_FR": [
    #     "agence marketing Montréal",
    #     "agence immobilière Québec",
    #     "coach business Montréal",
    #     "startup SaaS Montréal",
    #     "agence web Montréal",
    #     "consultant PME Québec",
    #     "agence communication Montréal",
    #     "expert comptable Montréal",
    #     "agence SEO Québec",
    #     "cabinet conseil Montréal",
    # ],
}

# ── HELPERS ────────────────────────────────────────────────────────────────────

def normaliser_telephone(tel: str) -> str:
    return re.sub(r"[^\d+]", "", tel) if tel else ""

def extraire_email_texte(texte: str) -> str:
    if not texte:
        return ""
    m = re.findall(r"\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}\b", texte)
    # Filtrer les emails génériques/inutiles
    exclusions = ["noreply", "no-reply", "donotreply", "support@google", "example"]
    for email in m:
        if not any(ex in email.lower() for ex in exclusions):
            return email
    return ""

def deviner_email(nom_entreprise: str, site_web: str) -> str:
    """Génère des emails probables basés sur le domaine du site"""
    if not site_web:
        return ""
    domaine = site_web.replace("https://", "").replace("http://", "").replace("www.", "").split("/")[0].strip()
    if not domaine:
        return ""
    # Formats d'emails les plus courants
    emails_probables = [
        f"contact@{domaine}",
        f"info@{domaine}",
        f"bonjour@{domaine}",
        f"hello@{domaine}",
    ]
    return emails_probables[0]  # On retourne le plus probable

def extraire_prenom_nom(nom_entreprise: str) -> tuple:
    """Essaie d'extraire prénom/nom si c'est un indépendant"""
    mots = nom_entreprise.strip().split()
    if len(mots) >= 2 and len(mots) <= 3:
        # Possible nom propre
        return mots[0], " ".join(mots[1:])
    return "", ""

# ── SCRAPING GOOGLE MAPS ───────────────────────────────────────────────────────

def scrape_requete(query: str, start: int = 0) -> tuple:
    params = {
        "engine":  "google_maps",
        "q":       query,
        "api_key": SERP_API_KEY,
        "hl":      "fr",
        "gl":      "fr",
        "type":    "search",
        "start":   start,
    }
    try:
        resp = requests.get("https://serpapi.com/search", params=params, timeout=REQUEST_TIMEOUT)
        resp.raise_for_status()
        data = resp.json()

        if "error" in data:
            print(f"      ⚠ SerpAPI : {data['error']}")
            return [], False

        resultats = []
        for place in data.get("local_results", []):
            gps = place.get("gps_coordinates", {})
            pid = place.get("place_id", "")
            site = place.get("website", "")
            email_auto = extraire_email_texte(place.get("description", ""))
            email_devine = deviner_email(place.get("title", ""), site) if not email_auto else ""
            prenom, nom = extraire_prenom_nom(place.get("title", ""))

            resultats.append({
                # Colonnes essentielles pour email en masse
                "email":              email_auto,
                "email_probable":     email_devine,
                "nom_entreprise":     place.get("title", ""),
                "prenom_contact":     prenom,
                "nom_contact":        nom,
                "telephone":          normaliser_telephone(place.get("phone", "")),
                "site_web":           site,
                # Données contextuelles
                "secteur":            query.split()[0:3],
                "adresse":            place.get("address", ""),
                "ville":              place.get("address", "").split(",")[-1].strip() if place.get("address") else "",
                "note":               place.get("rating", ""),
                "nb_avis":            place.get("reviews", 0),
                "categorie":          place.get("type", ""),
                "latitude":           gps.get("latitude"),
                "longitude":          gps.get("longitude"),
                "place_id":           pid,
                "lien_google_maps":   f"https://www.google.com/maps/place/?q=place_id:{pid}" if pid else "",
                "requete_source":     query,
                "date_collecte":      datetime.now().strftime("%d/%m/%Y"),
                # Colonnes pour suivi campagne
                "email_envoye":       "NON",
                "date_envoi":         "",
                "reponse":            "",
                "statut_prospect":    "NOUVEAU",
                "notes_commercial":   "",
            })

        pagination = data.get("serpapi_pagination", {})
        a_encore = "next" in pagination or "next_page_token" in pagination
        return resultats, a_encore

    except requests.exceptions.RequestException as e:
        print(f"      ⚠ Réseau : {e}")
        return [], False

# ── SCRAPING PAR SECTEUR ───────────────────────────────────────────────────────

def scraper_secteur(nom_secteur: str, requetes: list) -> pd.DataFrame:
    print(f"\n{'='*60}")
    print(f"  SECTEUR : {nom_secteur}  ({len(requetes)} requêtes)")
    print(f"{'='*60}")

    tous = []
    vus  = set()

    for idx, requete in enumerate(requetes, 1):
        print(f"\n[{idx:02d}/{len(requetes)}] {requete}")

        for page in range(MAX_PAGES_PAR_REQUETE):
            start = page * 20
            print(f"   → page {page+1} ...", end=" ", flush=True)

            resultats, a_encore = scrape_requete(requete, start)

            nouveaux = 0
            for h in resultats:
                cle = (h["nom_entreprise"].lower().strip(),
                       h["adresse"][:30].lower().strip() if h["adresse"] else "")
                if cle not in vus and h["nom_entreprise"]:
                    vus.add(cle)
                    tous.append(h)
                    nouveaux += 1

            print(f"{len(resultats)} résultats | {nouveaux} nouveaux | total {len(tous)}")

            if not a_encore or not resultats:
                break
            time.sleep(PAUSE_ENTRE_REQUETES)

        time.sleep(PAUSE_ENTRE_REQUETES)

    print(f"\n✅ {nom_secteur} : {len(tous)} prospects uniques")
    return pd.DataFrame(tous) if tous else pd.DataFrame()

# ── ANALYSE & SCORING ──────────────────────────────────────────────────────────

def scorer_prospects(df: pd.DataFrame) -> pd.DataFrame:
    """Score chaque prospect selon la qualité des données disponibles"""
    if df.empty:
        return df

    df = df.copy()
    scores = []
    priorites = []
    statuts_email = []

    for _, row in df.iterrows():
        score = 0
        statut = ""

        # Email
        if row.get("email"):
            score += 40
            statut = "EMAIL DIRECT"
        elif row.get("email_probable"):
            score += 20
            statut = "EMAIL PROBABLE"
        else:
            statut = "EMAIL MANQUANT"

        # Téléphone
        if row.get("telephone"):
            score += 20

        # Site web (permet de trouver l'email)
        if row.get("site_web"):
            score += 15

        # Note Google (crédibilité du prospect)
        try:
            note = float(row.get("note", 0) or 0)
            if note >= 4.5:
                score += 25
            elif note >= 4.0:
                score += 15
            elif note >= 3.5:
                score += 10
        except:
            pass

        # Nombre d'avis (taille de l'entreprise)
        try:
            avis = int(row.get("nb_avis", 0) or 0)
            if avis >= 100:
                score += 20
            elif avis >= 50:
                score += 15
            elif avis >= 10:
                score += 10
        except:
            pass

        scores.append(score)
        statuts_email.append(statut)

        if score >= 80:
            priorites.append("A — TOP")
        elif score >= 50:
            priorites.append("B — BON")
        elif score >= 30:
            priorites.append("C — MOYEN")
        else:
            priorites.append("D — FAIBLE")

    df["score"]        = scores
    df["priorite"]     = priorites
    df["statut_email"] = statuts_email

    # Trier par score décroissant
    df = df.sort_values("score", ascending=False).reset_index(drop=True)
    return df

# ── EXPORT CSV (prêt pour script envoi en masse) ───────────────────────────────

def export_csv_emailing(df: pd.DataFrame, fichier: str):
    """
    Export CSV optimisé pour script d'envoi email en masse.
    Colonnes dans l'ordre attendu par la plupart des outils d'emailing.
    """
    if df.empty:
        print("⚠ Aucune donnée à exporter.")
        return

    # Colonnes pour le script d'envoi
    cols_emailing = [
        "email",           # email principal
        "email_probable",  # email de secours si email vide
        "nom_entreprise",  # pour personnalisation {nom_entreprise}
        "prenom_contact",  # pour personnalisation {prenom}
        "nom_contact",     # pour personnalisation {nom}
        "telephone",
        "site_web",
        "ville",
        "secteur",
        "note",
        "priorite",
        "statut_email",
        "score",
        "lien_google_maps",
        "date_collecte",
        "email_envoye",
        "date_envoi",
        "reponse",
        "statut_prospect",
        "notes_commercial",
    ]

    # Garder seulement les colonnes disponibles
    cols_dispo = [c for c in cols_emailing if c in df.columns]
    df_export = df[cols_dispo].copy()

    # Colonne email_final : email direct si dispo, sinon email probable
    df_export.insert(0, "email_final",
        df_export.apply(
            lambda r: r["email"] if r.get("email") else r.get("email_probable", ""),
            axis=1
        )
    )

    df_export.to_csv(fichier, index=False, encoding="utf-8-sig")
    print(f"\n✅ CSV exporté : {fichier}")
    print(f"   Total prospects : {len(df_export)}")
    print(f"   Avec email direct  : {df_export['email'].notna().sum() and (df_export['email'] != '').sum()}")
    print(f"   Avec email probable: {(df_export['email_probable'] != '').sum()}")
    print(f"   Sans email         : {(df_export['email_final'] == '').sum()}")

# ── EXPORT EXCEL (pour suivi et complétion manuelle) ──────────────────────────

def export_excel_suivi(resultats_par_secteur: dict, fichier: str):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)

    couleurs = {
        "Immobilier_France":      "DEEBF7",
        "Agences_Marketing_France": "E2EFDA",
        "Coachs_Consultants_France": "FFF2CC",
        "Startups_SaaS_France":   "FCE4D6",
        "Cabinets_Experts_France": "F4E1D2",
        "Ecommerce_France":       "EAD1DC",
        "PME_Belgique":           "D9EAD3",
        "PME_Canada_FR":          "D0E4F5",
    }

    COLS = [
        ("priorite",        "Priorité",        10),
        ("score",           "Score",            8),
        ("statut_email",    "Email dispo ?",   18),
        ("email",           "Email direct",    30),
        ("email_probable",  "Email probable",  30),
        ("nom_entreprise",  "Entreprise",      35),
        ("telephone",       "Téléphone",       16),
        ("site_web",        "Site web",        32),
        ("ville",           "Ville",           18),
        ("note",            "Note /5",          8),
        ("nb_avis",         "Nb avis",          9),
        ("lien_google_maps","Google Maps",     28),
        ("email_envoye",    "Email envoyé ?",  14),
        ("date_envoi",      "Date envoi",      14),
        ("reponse",         "Réponse",         20),
        ("statut_prospect", "Statut",          16),
        ("notes_commercial","Notes",           30),
    ]

    thin = Side(style="thin", color="D0D0D0")
    bord = Border(bottom=thin, right=thin)

    for secteur, df in resultats_par_secteur.items():
        if df is None or df.empty:
            continue

        ws = wb.create_sheet(secteur.replace("_", " ")[:31])
        fill_h = PatternFill("solid", start_color="1F4E78", end_color="1F4E78")
        fill_d = PatternFill("solid",
                             start_color=couleurs.get(secteur, "F2F2F2"),
                             end_color=couleurs.get(secteur, "F2F2F2"))

        # Titre
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLS))
        c = ws.cell(1, 1, f"PROSPECTS — {secteur.replace('_', ' ')} — {len(df)} contacts")
        c.font      = Font(name="Arial", bold=True, size=12, color="FFFFFF")
        c.fill      = fill_h
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 26

        # En-têtes
        for ci, (col, header, largeur) in enumerate(COLS, 1):
            c = ws.cell(2, ci, header)
            c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
            c.fill      = fill_h
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(ci)].width = largeur
        ws.row_dimensions[2].height = 36

        # Données
        for ri, (_, row) in enumerate(df.iterrows(), 3):
            for ci, (col, _, _) in enumerate(COLS, 1):
                val = row.get(col, "")
                c   = ws.cell(ri, ci)

                if col in ("lien_google_maps", "site_web") and val:
                    c.value     = str(val)
                    c.font      = Font(name="Arial", color="0563C1", underline="single", size=9)
                    c.hyperlink = str(val)
                else:
                    c.value = val if pd.notna(val) and val != "" else ""
                    # Colorer selon priorité
                    if col == "priorite":
                        if "A" in str(val):
                            c.font = Font(name="Arial", bold=True, color="006400", size=10)
                        elif "B" in str(val):
                            c.font = Font(name="Arial", bold=True, color="FF8C00", size=10)
                        else:
                            c.font = Font(name="Arial", size=10)
                    else:
                        c.font = Font(name="Arial", size=9)

                c.fill      = fill_d
                c.alignment = Alignment(vertical="top")
                c.border    = bord
            ws.row_dimensions[ri].height = 22

        ws.freeze_panes = "A3"
        ws.auto_filter.ref = f"A2:{get_column_letter(len(COLS))}2"

    wb.save(fichier)
    total = sum(len(df) for df in resultats_par_secteur.values() if df is not None and not df.empty)
    print(f"\n✅ Excel créé : {fichier} | {total} prospects total")

# ── RÉSUMÉ FINAL ───────────────────────────────────────────────────────────────

def afficher_resume(resultats_par_secteur: dict):
    print("\n" + "="*65)
    print("  RÉSUMÉ DE LA COLLECTE")
    print("="*65)
    total = 0
    total_email = 0
    total_probable = 0

    for secteur, df in resultats_par_secteur.items():
        if df is None or df.empty:
            continue
        nb = len(df)
        avec_email = (df["email"] != "").sum() if "email" in df else 0
        avec_probable = (df["email_probable"] != "").sum() if "email_probable" in df else 0
        total += nb
        total_email += avec_email
        total_probable += avec_probable
        print(f"  {secteur:<35} {nb:>4} prospects | {avec_email:>3} emails directs | {avec_probable:>3} probables")

    print("-"*65)
    print(f"  {'TOTAL':<35} {total:>4} prospects | {total_email:>3} emails directs | {total_probable:>3} probables")
    print("="*65)

# ── MAIN ───────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("\n" + "="*65)
    print("  SCRAPER LEADS B2B — PROSPECTION EMAIL EN MASSE")
    print("="*65)

    print("\n  Secteurs disponibles :")
    secteurs_liste = list(MARCHES.keys())
    for i, s in enumerate(secteurs_liste, 1):
        print(f"  [{i}] {s}")
    print(f"  [0] Tous les secteurs")

    choix = input("\n▶️  Choisir un secteur (numéro) : ").strip()

    if choix == "0":
        secteurs_selectionnes = MARCHES
    elif choix.isdigit() and 1 <= int(choix) <= len(secteurs_liste):
        nom = secteurs_liste[int(choix) - 1]
        secteurs_selectionnes = {nom: MARCHES[nom]}
    else:
        print("Choix invalide.")
        exit()

    print(f"\n⚙️  Clé SerpAPI : {SERP_API_KEY[:10]}...")
    rep = input("▶️  Lancer le scraping ? (o/n) : ").strip().lower()
    if rep != "o":
        print("Annulé.")
        exit()

    resultats_par_secteur = {}

    for secteur, requetes in secteurs_selectionnes.items():
        df = scraper_secteur(secteur, requetes)
        if not df.empty:
            df = scorer_prospects(df)
        resultats_par_secteur[secteur] = df

    horodatage = datetime.now().strftime("%Y%m%d_%H%M")

    # Export 1 : CSV pour script envoi en masse
    # Un fichier CSV par secteur + un fichier global
    tous_df = []
    for secteur, df in resultats_par_secteur.items():
        if df is not None and not df.empty:
            df["secteur_nom"] = secteur
            # CSV par secteur
            csv_secteur = f"leads_{secteur}_{horodatage}.csv"
            export_csv_emailing(df, csv_secteur)
            tous_df.append(df)

    # CSV global (tous secteurs combinés)
    if tous_df:
        df_global = pd.concat(tous_df, ignore_index=True)
        df_global = df_global.sort_values("score", ascending=False)
        export_csv_emailing(df_global, f"leads_TOUS_{horodatage}.csv")

    # Export 2 : Excel pour suivi et complétion manuelle
    export_excel_suivi(resultats_par_secteur, f"suivi_prospects_{horodatage}.xlsx")

    # Résumé
    afficher_resume(resultats_par_secteur)

    print("\n📋 FICHIERS CRÉÉS :")
    print(f"   → leads_TOUS_{horodatage}.csv  ← À donner à ton script d'envoi")
    print(f"   → suivi_prospects_{horodatage}.xlsx  ← Pour suivi et complétion")
    print("\n📌 CONSEIL : Utilise la colonne 'email_final' dans ton script d'envoi.")
    print("   Si email_final est vide → visiter le site_web pour trouver l'email.")
    print("\n✅ TERMINÉ")