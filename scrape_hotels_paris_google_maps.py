import asyncio
import re
import random
import time
import urllib.request
from datetime import datetime
from pathlib import Path
from urllib.parse import quote

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from playwright.async_api import async_playwright, TimeoutError as PWTimeout, expect

# ═══════════════════════════════════════════════════════════════════════════════
# CONFIGURATION
# ═══════════════════════════════════════════════════════════════════════════════

CDP_URL       = "http://127.0.0.1:9222"   # Port Chrome DevTools
MAX_WORKERS   = 4                         # Réduit pour Google Maps (4 max)
SAVE_EVERY    = 50                        # Sauvegarde tous les N hôtels
PAGE_TIMEOUT  = 30_000                    # 30s timeout navigation
SCROLL_PAUSE  = 1.5                     # Pause entre scrolls
MAX_RETRIES   = 3                       # Tentatives en cas d'échec

# Recherche Google Maps
SEARCH_QUERY  = "hôtels Paris"
SEARCH_URL    = f"https://www.google.com/maps/search/{quote(SEARCH_QUERY)}"

# Fichiers
OUTPUT_FILE   = "hotels_paris_google_maps.xlsx"

# ═══════════════════════════════════════════════════════════════════════════════
# UTILITAIRES
# ═══════════════════════════════════════════════════════════════════════════════

def wait_for_chrome(max_attempts: int = 15) -> bool:
    """Vérifie que Chrome est accessible sur le port 9222."""
    print("🔍 Vérification Chrome sur port 9222...")
    for i in range(1, max_attempts + 1):
        try:
            urllib.request.urlopen(f"{CDP_URL}/json/version", timeout=2)
            print(f"✅ Chrome détecté !")
            return True
        except Exception:
            print(f"   ⏳ Tentative {i}/{max_attempts}...")
            time.sleep(2)
    return False


def load_existing_results(filepath: str) -> dict:
    """Charge les résultats déjà collectés pour reprise automatique."""
    if not Path(filepath).exists():
        return {}
    try:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        done = {}
        for row in ws.iter_rows(min_row=4, values_only=True):
            # Colonne B = nom, on l'utilise comme clé
            name = str(row[1] or "").strip()
            if name:
                done[name] = {
                    "name":    name,
                    "address": str(row[2] or ""),
                    "phone":   str(row[3] or ""),
                    "website": str(row[4] or ""),
                    "email":   str(row[5] or ""),
                }
        print(f"♻️  {len(done)} hôtels déjà traités — reprise activée")
        return done
    except Exception as e:
        print(f"⚠️  Impossible de lire {filepath} : {e}")
        return {}


def normalize_phone(phone: str) -> str:
    """Normalise un numéro de téléphone français."""
    if not phone:
        return ""
    # Supprime espaces, points, tirets
    cleaned = re.sub(r"[\s\.\-]", "", phone)
    # 0033 → +33
    cleaned = re.sub(r"^0033", "+33", cleaned)
    # 0x xx xx xx xx → +33 x xx xx xx xx
    if cleaned.startswith("0") and not cleaned.startswith("00"):
        cleaned = "+33" + cleaned[1:]
    return cleaned


# ═══════════════════════════════════════════════════════════════════════════════
# NAVIGATION GOOGLE MAPS — SCROLL INFINI
# ═══════════════════════════════════════════════════════════════════════════════

async def scroll_results_list(page, max_scrolls: int = 200) -> list[dict]:
    """
    Fait défiler la liste des résultats Google Maps et collecte les liens.
    Google Maps charge les résultats au scroll (infinite scroll).
    """
    hotels = []
    seen_urls = set()

    print(f"📜 Défilement de la liste des résultats...")
    print(f"   Objectif : ~1 629 hôtels | Workers : {MAX_WORKERS}")

    # Attendre que la liste des résultats apparaisse
    try:
        # Sélecteur pour la liste des résultats (container scrollable)
        await page.wait_for_selector("[role='main']", timeout=15000)
        await asyncio.sleep(2)
    except PWTimeout:
        print("❌ Impossible de charger la liste des résultats")
        return []

    # Le container scrollable est souvent le premier div avec overflow-y: auto
    # ou un div avec role="feed" ou data-async-context
    scroll_container = None

    for attempt in range(max_scrolls):
        # Récupérer tous les liens d'hôtels visibles
        # Google Maps utilise des liens avec href contenant /maps/place/
        links = await page.query_selector_all("a[href*='/maps/place/']")

        new_found = 0
        for link in links:
            try:
                href = await link.get_attribute("href") or ""
                name = await link.get_attribute("aria-label") or ""

                # Extraire le nom depuis le texte si aria-label vide
                if not name:
                    name_el = await link.query_selector(".fontHeadlineSmall, .qBF1Pd, h3")
                    if name_el:
                        name = await name_el.inner_text()

                if href and "/maps/place/" in href and name:
                    # Nettoyer l'URL (enlever les params inutiles)
                    clean_url = href.split("?")[0]
                    if clean_url not in seen_urls:
                        seen_urls.add(clean_url)
                        hotels.append({"name": name.strip(), "url": clean_url})
                        new_found += 1
            except Exception:
                continue

        if new_found > 0:
            print(f"   Scroll {attempt+1:>3} | +{new_found:>3} nouveaux | Total : {len(hotels)}")

        # Scroll dans le container
        try:
            # Méthode 1 : scroll sur le body ou un container spécifique
            await page.evaluate("""
                () => {
                    const containers = document.querySelectorAll('div[role="main"] > div > div > div > div');
                    for (let c of containers) {
                        if (c.scrollHeight > c.clientHeight) {
                            c.scrollTop = c.scrollHeight;
                            return true;
                        }
                    }
                    // Fallback : scroll sur window
                    window.scrollTo(0, document.body.scrollHeight);
                    return true;
                }
            """)
        except Exception:
            pass

        # Pause pour charger les nouveaux résultats
        await asyncio.sleep(SCROLL_PAUSE + random.uniform(0, 1))

        # Arrêt si on atteint ~1650 ou si plus de nouveaux résultats après 5 scrolls
        if len(hotels) >= 1650:
            print(f"✅ Objectif atteint : {len(hotels)} hôtels trouvés")
            break

        if attempt > 10 and new_found == 0:
            # Vérifier si c'est la fin de la liste
            end_indicator = await page.query_selector("text=/fin de la liste|end of the list|plus de résultats/i")
            if end_indicator:
                print(f"🏁 Fin de la liste atteinte : {len(hotels)} hôtels trouvés")
                break

    print(f"📋 {len(hotels)} hôtels uniques collectés")
    return hotels


# ═══════════════════════════════════════════════════════════════════════════════
# EXTRACTION DES DÉTAILS D'UN HÔTEL
# ═══════════════════════════════════════════════════════════════════════════════

async def extract_hotel_details(page, url: str, name: str, retries: int = 0) -> dict:
    """
    Ouvre la fiche d'un hôtel et extrait : adresse, téléphone, site web.
    """
    result = {
        "name": name,
        "address": "",
        "phone": "",
        "website": "",
        "email": "",
    }

    try:
        # Navigation vers la fiche hôtel
        await page.goto(url, wait_until="domcontentloaded", timeout=PAGE_TIMEOUT)
        await asyncio.sleep(1.5 + random.uniform(0, 1))

        # Attendre que le panneau latéral charge
        await page.wait_for_selector("[role='main'], [data-panel-id], .m6QErb", timeout=10000)
        await asyncio.sleep(0.5)

        # ── ADRESSE ──
        try:
            # Google Maps affiche l'adresse dans un bouton avec data-item-id="address"
            # ou dans un span avec classe spécifique
            address_selectors = [
                "button[data-item-id='address']",
                "[data-item-id*='address']",
                "[aria-label*='Adresse']",
                "[aria-label*='Address']",
                ".Io6YTe:has-text('Paris')",  # Classe courante pour les infos
            ]
            for sel in address_selectors:
                el = await page.query_selector(sel)
                if el:
                    text = await el.inner_text()
                    if "Paris" in text or any(d in text for d in "0123456789"):
                        result["address"] = text.strip()
                        break

            # Fallback : chercher dans tout le panneau
            if not result["address"]:
                all_spans = await page.query_selector_all(".Io6YTe, .fontBodyMedium")
                for sp in all_spans:
                    text = await sp.inner_text()
                    if "Paris" in text and len(text) > 10:
                        result["address"] = text.strip()
                        break
        except Exception:
            pass

        # ── TÉLÉPHONE ──
        try:
            # Le téléphone est dans un lien tel: ou un bouton avec data-item-id="phone"
            phone_selectors = [
                "a[href^='tel:']",
                "button[data-item-id='phone']",
                "[data-item-id*='phone']",
                "[aria-label*='Téléphone']",
                "[aria-label*='Phone']",
            ]
            for sel in phone_selectors:
                el = await page.query_selector(sel)
                if el:
                    # Essayer href d'abord
                    href = await el.get_attribute("href") or ""
                    if href.startswith("tel:"):
                        result["phone"] = normalize_phone(href.replace("tel:", ""))
                        break
                    # Sinon inner text
                    text = await el.inner_text()
                    if re.search(r"\d", text):
                        result["phone"] = normalize_phone(text)
                        break
        except Exception:
            pass

        # ── SITE WEB ──
        try:
            # Le site web est dans un lien avec data-item-id="authority"
            # ou un bouton qui ouvre un lien externe
            web_selectors = [
                "a[data-item-id='authority']",
                "[data-item-id*='authority']",
                "[aria-label*='Site web']",
                "[aria-label*='Website']",
                "a[href^='http']:not([href*='google']):not([href*='tripadvisor'])",
            ]
            for sel in web_selectors:
                els = await page.query_selector_all(sel)
                for el in els:
                    href = await el.get_attribute("href") or ""
                    text = await el.inner_text() or ""
                    # Filtrer les liens Google et garder les vrais sites
                    if href and not any(x in href for x in ["google.com", "google.fr", "tripadvisor"]):
                        if href.startswith("http"):
                            result["website"] = href
                            break
                        # Parfois c'est un lien relatif de redirection
                        if "/url?q=" in href:
                            # Extraire l'URL réelle du paramètre q
                            match = re.search(r"[?&]q=([^&]+)", href)
                            if match:
                                result["website"] = match.group(1)
                                break
                if result["website"]:
                    break
        except Exception:
            pass

        # ── EMAIL (via le site web si disponible) ──
        if result["website"]:
            try:
                email = await extract_email_from_website(page, result["website"])
                if email:
                    result["email"] = email
            except Exception:
                pass

    except PWTimeout:
        if retries < MAX_RETRIES:
            await asyncio.sleep(2 ** retries)  # Backoff exponentiel
            return await extract_hotel_details(page, url, name, retries + 1)
        result["phone"] = "TIMEOUT"
    except Exception as e:
        if retries < MAX_RETRIES:
            await asyncio.sleep(2 ** retries)
            return await extract_hotel_details(page, url, name, retries + 1)
        result["phone"] = f"ERREUR: {str(e)[:25]}"

    return result


async def extract_email_from_website(page, website_url: str) -> str:
    """
    Visite le site web de l'hôtel pour trouver un email de contact.
    Cherche dans la page contact, mentions légales, ou footer.
    """
    email = ""

    try:
        # Ouvrir une nouvelle page pour ne pas perdre la fiche Google Maps
        new_page = await page.context.new_page()

        try:
            await new_page.goto(website_url, wait_until="domcontentloaded", timeout=15000)
            await asyncio.sleep(1)

            # Chercher un pattern email dans le HTML
            content = await new_page.content()

            # Pattern email standard
            email_pattern = r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}'
            emails = re.findall(email_pattern, content)

            # Filtrer les emails communs (pas les noreply, support génériques)
            valid_emails = [e for e in emails 
                          if not any(x in e.lower() for x in 
                          ["noreply", "no-reply", "donotreply", "example", "test@"])]

            if valid_emails:
                # Prendre le premier email trouvé (souvent le contact principal)
                email = valid_emails[0]

            # Si pas d'email, chercher la page contact
            if not email:
                contact_links = await new_page.query_selector_all("a[href*='contact'], a[href*='Contact']")
                for link in contact_links:
                    href = await link.get_attribute("href") or ""
                    if href:
                        # Visiter la page contact
                        if not href.startswith("http"):
                            from urllib.parse import urljoin
                            href = urljoin(website_url, href)
                        try:
                            await new_page.goto(href, wait_until="domcontentloaded", timeout=10000)
                            await asyncio.sleep(0.5)
                            content = await new_page.content()
                            emails = re.findall(email_pattern, content)
                            valid_emails = [e for e in emails 
                                          if not any(x in e.lower() for x in 
                                          ["noreply", "no-reply", "donotreply", "example", "test@"])]
                            if valid_emails:
                                email = valid_emails[0]
                                break
                        except Exception:
                            continue

        finally:
            try:
                await new_page.close()
            except Exception:
                pass

    except Exception:
        pass

    return email


# ═══════════════════════════════════════════════════════════════════════════════
# WORKER CONCURRENT
# ═══════════════════════════════════════════════════════════════════════════════

async def worker(sem: asyncio.Semaphore, context,
                 hotel: dict, results: dict,
                 counter: list, total: int,
                 lock: asyncio.Lock) -> None:
    """Traite un hôtel : ouvre sa fiche et extrait les contacts."""
    url = hotel["url"]
    name = hotel["name"]

    async with sem:
        page = await context.new_page()
        try:
            data = await extract_hotel_details(page, url, name)

            async with lock:
                results[name] = data
                counter[0] += 1
                pct = counter[0] / total * 100

                # Affichage compact
                phone = data["phone"][:22] if data["phone"] else "—"
                web = "✓" if data["website"] else "✗"
                mail = "✓" if data["email"] else "✗"
                addr = data["address"][:30] if data["address"] else "—"

                print(
                    f"  [{counter[0]:>4}/{total}] {pct:5.1f}%"
                    f"  {name[:35]:<35}"
                    f"  📞 {phone:<24}"
                    f"  🌐 {web}"
                    f"  📧 {mail}"
                    f"  📍 {addr:<30}"
                )

        except Exception as e:
            async with lock:
                results[name] = {
                    "name": name, "address": "", "phone": f"ERREUR: {str(e)[:20]}",
                    "website": "", "email": "",
                }
                counter[0] += 1
                print(f"  [{counter[0]:>4}/{total}] ❌ {name[:40]}  {e}")

        finally:
            try:
                await page.close()
            except Exception:
                pass


# ═══════════════════════════════════════════════════════════════════════════════
# EXPORT EXCEL
# ═══════════════════════════════════════════════════════════════════════════════

def save_excel(hotels: list[dict], results: dict, filename: str) -> None:
    """Sauvegarde les résultats dans un fichier Excel formaté."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hôtels Paris - Google Maps"

    # Couleurs
    GREEN = PatternFill("solid", fgColor="34E0A1")
    ODD   = PatternFill("solid", fgColor="F0FBF7")
    EVEN  = PatternFill("solid", fgColor="FFFFFF")
    RED   = PatternFill("solid", fgColor="FFCCCC")
    thin  = Side(style="thin", color="B2DFD0")
    brd   = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Titre ──
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "Google Maps — Hôtels Paris (1 629 hôtels)"
    c.font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    c.fill = GREEN
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # ── Sous-titre ──
    ws.merge_cells("A2:F2")
    c = ws["A2"]
    done = sum(1 for h in hotels if results.get(h["name"], {}).get("phone") not in ("", "TIMEOUT", None))
    websites = sum(1 for h in hotels if results.get(h["name"], {}).get("website"))
    emails = sum(1 for h in hotels if results.get(h["name"], {}).get("email"))
    c.value = (
        f"Extrait le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
        f"  |  {len(results)} traités  |  {done} téléphones"
        f"  |  {websites} sites web  |  {emails} emails"
    )
    c.font = Font(name="Arial", italic=True, size=10, color="555555")
    c.fill = PatternFill("solid", fgColor="E8F8F3")
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    # ── En-têtes ──
    COLS = [
        ("#", 6),
        ("Nom de l'hôtel", 45),
        ("Adresse", 55),
        ("Téléphone", 22),
        ("Site Web", 45),
        ("Email", 35),
    ]
    for col, (h, w) in enumerate(COLS, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        c.fill = GREEN
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = brd
        ws.column_dimensions[chr(64 + col)].width = w
    ws.row_dimensions[3].height = 22

    # ── Données ──
    for i, hotel in enumerate(hotels, 1):
        name = hotel["name"]
        r = results.get(name, {})
        row = i + 3
        fill = ODD if i % 2 else EVEN

        vals = [i, name, r.get("address", ""), r.get("phone", ""),
                r.get("website", ""), r.get("email", "")]

        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = Font(name="Arial", size=10)
            c.alignment = Alignment(vertical="center", wrap_text=(col in (2, 3)))
            c.fill = fill
            c.border = brd

            # Site web en hyperlien
            if col == 5 and val and val.startswith("http"):
                c.hyperlink = val
                c.font = Font(name="Arial", size=10, color="0066CC", underline="single")

            # Email en hyperlien mailto
            if col == 6 and val and "@" in val:
                c.hyperlink = f"mailto:{val}"
                c.font = Font(name="Arial", size=10, color="007A4D", underline="single")

            # Téléphone en vert
            if col == 4 and val and val not in ("TIMEOUT", "", None) and not str(val).startswith("ERREUR"):
                c.font = Font(name="Arial", size=10, color="006400", bold=True)

            # Erreur en rouge
            if col == 4 and str(val).startswith("ERREUR"):
                c.fill = RED

        ws.row_dimensions[row].height = 18

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:F{len(hotels) + 3}"

    wb.save(filename)
    print(f"\n💾  Sauvegardé → {filename}  ({len(results)}/{len(hotels)} traités)")


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════

async def main():
    print("=" * 80)
    print("  Google Maps — Scraping Hôtels Paris (Chrome CDP)")
    print(f"  Workers : {MAX_WORKERS} | Sauvegarde tous les {SAVE_EVERY} hôtels")
    print("=" * 80 + "\n")

    # Vérifier Chrome
    if not wait_for_chrome():
        print("\n❌ Chrome non détecté ! Lance Chrome avec :")
        print('   Windows : "C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe" --remote-debugging-port=9222')
        print("   Mac     : /Applications/Google\\ Chrome.app/Contents/MacOS/Google\\ Chrome --remote-debugging-port=9222")
        print("   Linux   : google-chrome --remote-debugging-port=9222")
        return

    async with async_playwright() as p:
        browser = await p.chromium.connect_over_cdp(CDP_URL)
        print(f"✅ Connecté à Chrome ({browser.version})\n")

        context = (browser.contexts[0] if browser.contexts
                   else await browser.new_context(
                       viewport={"width": 1366, "height": 900},
                       user_agent=(
                           "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                           "AppleWebKit/537.36 (KHTML, like Gecko) "
                           "Chrome/124.0.0.0 Safari/537.36"
                       )
                   ))

        # Charger les résultats existants
        existing = load_existing_results(OUTPUT_FILE)

        # Étape 1 : Collecter les URLs des hôtels via scroll infini
        print(f"\n🔍 Navigation vers Google Maps...")
        page = await context.new_page()
        await page.goto(SEARCH_URL, wait_until="domcontentloaded", timeout=30000)
        await asyncio.sleep(3)

        # Vérifier si Google demande une confirmation (bouton "Tout accepter")
        try:
            accept_btn = await page.query_selector("button:has-text('Tout accepter'), button:has-text('Accept all')")
            if accept_btn:
                await accept_btn.click()
                await asyncio.sleep(1)
        except Exception:
            pass

        # Collecter les hôtels
        hotels = await scroll_results_list(page)
        await page.close()

        if not hotels:
            print("❌ Aucun hôtel trouvé. Vérifie que Google Maps est bien chargé.")
            return

        # Filtrer ceux déjà traités
        to_do = [h for h in hotels if h["name"] not in existing]
        already = len(hotels) - len(to_do)
        print(f"\n📊 Total : {len(hotels)}  |  Déjà traités : {already}  |  Restant : {len(to_do)}\n")

        if not to_do:
            print("✅ Tout est déjà traité !")
            save_excel(hotels, existing, OUTPUT_FILE)
            return

        # Étape 2 : Extraction des détails (parallèle)
        results = {**existing}  # Fusion avec existants
        sem = asyncio.Semaphore(MAX_WORKERS)
        counter = [already]
        lock = asyncio.Lock()
        total = len(hotels)

        start_time = time.time()

        # Traitement par batch
        for batch_start in range(0, len(to_do), SAVE_EVERY):
            batch = to_do[batch_start: batch_start + SAVE_EVERY]
            tasks = [
                worker(sem, context, h, results, counter, total, lock)
                for h in batch
            ]
            await asyncio.gather(*tasks)

            # Sauvegarde intermédiaire
            save_excel(hotels, results, OUTPUT_FILE)

            elapsed = time.time() - start_time
            done_so_far = counter[0] - already
            if done_so_far > 0:
                rate = done_so_far / elapsed
                remaining = len(to_do) - done_so_far
                eta_sec = remaining / rate if rate > 0 else 0
                eta_min = int(eta_sec // 60)
                eta_s = int(eta_sec % 60)
                print(f"\n⏱️  Vitesse : {rate:.1f} hôtels/s  |  ETA : {eta_min}m{eta_s:02d}s\n")

        # Sauvegarde finale
        elapsed_total = int(time.time() - start_time)
        print(f"\n{'=' * 80}")
        print(f"🎯  Terminé en {elapsed_total // 60}m{elapsed_total % 60:02d}s")
        print(f"    Total traités : {len(results)} / {len(hotels)}")

        phones = sum(1 for h in hotels if results.get(h["name"], {}).get("phone") not in ("", "TIMEOUT", None))
        websites = sum(1 for h in hotels if results.get(h["name"], {}).get("website"))
        emails = sum(1 for h in hotels if results.get(h["name"], {}).get("email"))
        print(f"    📞 Téléphones : {phones} | 🌐 Sites web : {websites} | 📧 Emails : {emails}")
        print(f"{'=' * 80}")

        save_excel(hotels, results, OUTPUT_FILE)

        # Aperçu
        print("\n📋 Aperçu (10 premiers avec téléphone) :")
        shown = 0
        for h in hotels:
            r = results.get(h["name"], {})
            if r.get("phone") and r["phone"] not in ("TIMEOUT", "", None):
                print(f"  {r['name'][:40]:<40}  {r['phone']:<22}  {r.get('website', '')[:40]}")
                shown += 1
                if shown >= 10:
                    break


if __name__ == "__main__":
    asyncio.run(main())