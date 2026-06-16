"""
Mariages.net Paris — Scraper v2 (CDP / vrai Chrome)
=====================================================
Même stratégie que funbooker :
  → On se connecte au Chrome RÉEL via CDP (port 9222)
  → Le site ne peut pas détecter l'automatisation
"""

import asyncio
import re
import time
import urllib.request
from collections import Counter
from datetime import datetime

import openpyxl
from bs4 import BeautifulSoup
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from playwright.async_api import async_playwright, TimeoutError as PWTimeout

# ── CONFIGURATION ──────────────────────────────────────────────────────────────

BASE_URL    = "https://www.mariages.net"
URL_PATTERN = "https://www.mariages.net/real-Home.php?region=paris&NumPage={page}"
TOTAL_PAGES = 47
OUTPUT_FILE = "mariages_paris.xlsx"
CDP_URL     = "http://127.0.0.1:9222"

PAUSE_ENTRE_PAGES = 2.5   # secondes entre pages (poli)
WAIT_PAGE_LOAD    = 4.0   # attente JS après navigation
MAX_RETRIES       = 3

# ── HELPERS ────────────────────────────────────────────────────────────────────

def clean(text) -> str:
    return " ".join(str(text).split()).strip() if text else ""

# ── VÉRIFICATION CHROME ────────────────────────────────────────────────────────

def wait_for_chrome(max_attempts: int = 15) -> bool:
    print("🔍  Vérification du port 9222 (Chrome)...")
    for attempt in range(1, max_attempts + 1):
        try:
            urllib.request.urlopen(f"{CDP_URL}/json/version", timeout=2)
            print(f"✅  Chrome détecté sur {CDP_URL}")
            return True
        except Exception:
            print(f"    ⏳ Tentative {attempt}/{max_attempts}...")
            time.sleep(2)
    return False

# ── FERMETURE DES POPUPS ───────────────────────────────────────────────────────

async def dismiss_overlays(page) -> None:
    for sel in [
        "#onetrust-accept-btn-handler",
        "button:has-text('Accepter')",
        "button:has-text('Tout accepter')",
        "button:has-text('Accept All')",
        "button:has-text('J\\'accepte')",
        ".cookie-accept",
        "[aria-label='Close']",
        "[aria-label='Fermer']",
    ]:
        try:
            btn = page.locator(sel).first
            if await btn.is_visible(timeout=1_500):
                await btn.click(timeout=2_000)
                await asyncio.sleep(0.5)
        except Exception:
            pass

# ── PARSING D'UNE CARTE ────────────────────────────────────────────────────────

def parse_wedding_card(div, page_num: int) -> dict:
    result = {
        "page":         page_num,
        "url":          "",
        "noms":         "",
        "ville":        "",
        "lieu":         "",
        "style":        "",
        "saison":       "",
        "couleur":      "",
        "photographe":  "",
        "nb_photos":    "",
        "image_url":    "",
        "id_reportage": "",
    }

    # URL + ID
    app_link = div.find(class_="app-link")
    if app_link:
        href = app_link.get("data-href", "") or app_link.get("href", "")
        if href:
            result["url"] = href if href.startswith("http") else BASE_URL + href
            m = re.search(r"rw(\d+)", href)
            if m:
                result["id_reportage"] = m.group(1)

    # Photo principale
    figure = div.find("figure")
    if figure:
        img = figure.find("img")
        if img:
            result["image_url"] = img.get("src", "") or img.get("data-src", "")
        counter = figure.find(class_="gallery-box-item-counter")
        if counter:
            nb = re.search(r"(\d+)", counter.get_text())
            result["nb_photos"] = nb.group(1) if nb else ""

    # Description
    desc = div.find(class_="gallery-box-description")
    if desc:
        names_tag = desc.find(class_="gallery-box-description-names")
        if names_tag:
            result["noms"] = clean(names_tag.get_text())

        paragraphs = desc.find_all("p")
        for p in paragraphs:
            text    = clean(p.get_text())
            classes = p.get("class", [])
            if "upper" in classes and not result["ville"]:
                result["ville"] = text
            elif "," in text and not result["style"]:
                parts = [s.strip() for s in text.split(",")]
                if len(parts) >= 2:
                    result["style"]  = parts[0]
                    result["saison"] = parts[1]
            elif text and not result["lieu"] and "upper" not in classes and "," not in text:
                result["lieu"] = text

        # Couleurs
        color_spans = desc.find_all(class_="realWeddingHero__weddingColor")
        colors = []
        for span in color_spans:
            cls = [c for c in span.get("class", []) if c != "realWeddingHero__weddingColor"]
            if cls:
                colors.append(cls[0])
        result["couleur"] = ", ".join(colors)

    # Photographe
    footer = div.find(class_="gallery-box-footer")
    if footer:
        owner = footer.find(class_="gallery-box-owner-name")
        if owner:
            for tag in owner.find_all(["svg", "span", "i"]):
                if "icon" in " ".join(tag.get("class", [])):
                    tag.decompose()
            result["photographe"] = clean(owner.get_text())

    return result

# ── SCRAPING D'UNE PAGE ────────────────────────────────────────────────────────

async def scraper_page(page, page_num: int) -> list[dict]:
    url = URL_PATTERN.format(page=page_num)

    for attempt in range(1, MAX_RETRIES + 1):
        try:
            await page.goto(url, wait_until="domcontentloaded", timeout=30_000)
            await asyncio.sleep(WAIT_PAGE_LOAD)

            if page_num == 1:
                await dismiss_overlays(page)
                await asyncio.sleep(1.0)

            try:
                await page.wait_for_selector(".app-rw-item", timeout=6_000)
            except PWTimeout:
                pass

            html  = await page.content()
            soup  = BeautifulSoup(html, "lxml")
            cards = soup.select(".app-rw-item")

            if cards:
                weddings = []
                for card in cards:
                    w = parse_wedding_card(card, page_num)
                    if w["noms"] or w["url"]:
                        weddings.append(w)
                return weddings

            title = await page.title()
            print(f"    ⚠  Aucune carte (tentative {attempt}/{MAX_RETRIES}) | title='{title}'")
            if attempt < MAX_RETRIES:
                await asyncio.sleep(3 * attempt)

        except PWTimeout:
            print(f"    ⚠  Timeout (tentative {attempt}/{MAX_RETRIES})")
            if attempt < MAX_RETRIES:
                await asyncio.sleep(3 * attempt)
        except Exception as e:
            print(f"    ⚠  Erreur : {e} (tentative {attempt}/{MAX_RETRIES})")
            if attempt < MAX_RETRIES:
                await asyncio.sleep(3 * attempt)

    return []

# ── SCRAPING COMPLET ───────────────────────────────────────────────────────────

async def scraper_toutes_les_pages() -> list[dict]:
    print("\n" + "=" * 65)
    print(f"  Mariages.net Paris — Scraper ({TOTAL_PAGES} pages) via CDP")
    print("=" * 65)

    # ── Instructions ──
    print("\n  INSTRUCTIONS :")
    print("  1. Lance Chrome en mode debug :")
    print('     chrome.exe --remote-debugging-port=9222 --user-data-dir="C:\\ChromeDebug"')
    print("  2. Navigue manuellement vers :")
    print(f"     {URL_PATTERN.format(page=1)}")
    print("  3. Attends que la page charge complètement")
    print("  4. Reviens ici et appuie sur Entrée")
    print("=" * 65 + "\n")
    input("  → Prêt ? Appuie sur Entrée : ")

    if not wait_for_chrome():
        print("\n❌ Chrome ne répond pas sur le port 9222.")
        print("   Lance Chrome avec la commande ci-dessus et réessaie.")
        return []

    all_weddings = []

    async with async_playwright() as p:
        try:
            browser = await p.chromium.connect_over_cdp(CDP_URL)
            print(f"✅  Connecté à Chrome ({browser.version})\n")
        except Exception as e:
            print(f"❌  Connexion CDP échouée : {e}")
            return []

        # Réutilise le contexte existant (avec cookies/session du vrai Chrome)
        context = (
            browser.contexts[0] if browser.contexts
            else await browser.new_context(viewport={"width": 1440, "height": 900})
        )
        page = await context.new_page()

        for page_num in range(1, TOTAL_PAGES + 1):
            print(f"[{page_num:02d}/{TOTAL_PAGES}] Page {page_num}...")

            weddings = await scraper_page(page, page_num)

            if weddings:
                all_weddings.extend(weddings)
                print(f"    ✅  {len(weddings)} reportages | total : {len(all_weddings)}")
            else:
                print(f"    ❌  0 reportage — page ignorée")

            if page_num < TOTAL_PAGES:
                await asyncio.sleep(PAUSE_ENTRE_PAGES)

        await page.close()

    print(f"\n{'=' * 65}")
    print(f"  🎯 TOTAL : {len(all_weddings)} reportages sur {TOTAL_PAGES} pages")
    print(f"{'=' * 65}\n")
    return all_weddings

# ── EXPORT EXCEL ───────────────────────────────────────────────────────────────

COLONNES = [
    ("#",           5),
    ("Page",        6),
    ("ID",          10),
    ("Noms",        35),
    ("Ville",       18),
    ("Lieu",        30),
    ("Style",       18),
    ("Saison",      14),
    ("Couleur",     14),
    ("Photographe", 28),
    ("Nb Photos",   10),
    ("URL",         60),
    ("Image URL",   60),
]

def style_fill(couleur: str) -> PatternFill:
    return PatternFill("solid", start_color=couleur, end_color=couleur)

def save_excel(weddings: list[dict], filename: str) -> None:
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "Reportages"

    HDR_BG = style_fill("C9184A")
    ODD    = style_fill("FFF0F5")
    EVEN   = style_fill("FFFFFF")
    LINK_C = "8B0000"
    thin   = Side(style="thin", color="FFCCD5")
    brd    = Border(left=thin, right=thin, top=thin, bottom=thin)
    n_cols = len(COLONNES)

    ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
    c = ws["A1"]
    c.value     = "Mariages.net — Reportages Paris (toutes pages)"
    c.font      = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    c.fill      = HDR_BG
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells(f"A2:{get_column_letter(n_cols)}2")
    c = ws["A2"]
    c.value = (
        f"Extrait le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
        f"  |  {len(weddings)} reportages  |  {TOTAL_PAGES} pages"
    )
    c.font      = Font(name="Arial", italic=True, size=10, color="444444")
    c.fill      = style_fill("FFE8EA")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    for ci, (header, width) in enumerate(COLONNES, 1):
        c = ws.cell(row=3, column=ci, value=header)
        c.font      = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        c.fill      = HDR_BG
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = brd
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[3].height = 24

    for i, w in enumerate(weddings, 1):
        row_num = i + 3
        fill    = ODD if i % 2 else EVEN

        def cell(col, value, hyperlink="", link=False):
            c = ws.cell(row=row_num, column=col, value=value)
            c.font = Font(
                name="Arial", size=10,
                color=LINK_C if link else "000000",
                underline="single" if link else None,
            )
            c.alignment = Alignment(vertical="top", wrap_text=(col in (4, 5, 6)))
            c.fill   = fill
            c.border = brd
            if hyperlink:
                c.hyperlink = hyperlink

        cell(1,  i)
        cell(2,  w.get("page", ""))
        cell(3,  w.get("id_reportage", ""))
        cell(4,  w.get("noms", ""))
        cell(5,  w.get("ville", ""))
        cell(6,  w.get("lieu", ""))
        cell(7,  w.get("style", ""))
        cell(8,  w.get("saison", ""))
        cell(9,  w.get("couleur", ""))
        cell(10, w.get("photographe", ""))
        cell(11, w.get("nb_photos", ""))
        cell(12, w.get("url", ""),       hyperlink=w.get("url",""),       link=bool(w.get("url")))
        cell(13, w.get("image_url", ""), hyperlink=w.get("image_url",""), link=bool(w.get("image_url")))
        ws.row_dimensions[row_num].height = 22

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(n_cols)}3"

    # Onglet Stats
    ws2 = wb.create_sheet("📊 Stats")
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 18

    styles  = Counter(w["style"]  for w in weddings if w.get("style"))
    saisons = Counter(w["saison"] for w in weddings if w.get("saison"))
    villes  = Counter(w["ville"]  for w in weddings if w.get("ville"))

    stats = [
        ("Total reportages",   len(weddings)),
        ("Pages scrapées",     TOTAL_PAGES),
        ("Avec photographe",   sum(1 for w in weddings if w.get("photographe"))),
        ("Avec lieu",          sum(1 for w in weddings if w.get("lieu"))),
        ("Avec style",         sum(1 for w in weddings if w.get("style"))),
        ("Avec couleur",       sum(1 for w in weddings if w.get("couleur"))),
        ("", ""),
        ("── Top Styles ──",  ""),
    ]
    for style, count in styles.most_common(10):
        stats.append((f"  {style}", count))
    stats += [("", ""), ("── Top Saisons ──", "")]
    for saison, count in saisons.most_common():
        stats.append((f"  {saison}", count))
    stats += [("", ""), ("── Top Villes ──", "")]
    for ville, count in villes.most_common(5):
        stats.append((f"  {ville}", count))
    stats.append(("Date extraction", datetime.now().strftime("%d/%m/%Y %H:%M")))

    for ri, (label, val) in enumerate(stats, 1):
        ws2.cell(ri, 1, label).font = Font(name="Arial", bold=("──" in str(label)), size=11)
        ws2.cell(ri, 2, val).font   = Font(name="Arial", size=11)
        ws2.row_dimensions[ri].height = 20

    wb.save(filename)
    print(f"💾  Fichier Excel : {filename}  ({len(weddings)} reportages)")

# ── MAIN ──────────────────────────────────────────────────────────────────────

async def main():
    print("=" * 65)
    print("  Mariages.net Paris — Scraper v2 (CDP / vrai Chrome)")
    print("=" * 65)

    weddings = await scraper_toutes_les_pages()

    if not weddings:
        print("\n⚠️  Aucun reportage collecté.")
        return

    save_excel(weddings, OUTPUT_FILE)

    print("\n📋  Aperçu des 5 premiers :")
    print(f"  {'#':>3}  {'Noms':<35}  {'Ville':<15}  {'Photographe'}")
    print("  " + "-" * 85)
    for i, w in enumerate(weddings[:5], 1):
        print(
            f"  {i:>3}  {w.get('noms','')[:33]:<35}  "
            f"{w.get('ville','')[:13]:<15}  "
            f"{w.get('photographe','') or '(N/A)'}"
        )
    if len(weddings) > 5:
        print(f"  … et {len(weddings)-5} autres dans {OUTPUT_FILE}")
    print(f"\n✅  Terminé !\n")

if __name__ == "__main__":
    asyncio.run(main())