"""
╔══════════════════════════════════════════════════════════════════════════════╗
║     GOOGLE TRAVEL — Scraper Hôtels Paris v3 (Sélecteurs Réels)             ║
║     Sélecteurs extraits du vrai HTML Google Travel                          ║
╚══════════════════════════════════════════════════════════════════════════════╝

LANCEMENT :
    # Chrome doit tourner avec :
    # chrome.exe --remote-debugging-port=9222 --user-data-dir=C:\\chrome_scrape
    python scraper_v3.py --mode cdp

    # Ou mode visible pour debug :
    python scraper_v3.py --mode visible
"""

import argparse
import asyncio
import csv
import json
import random
import re
import time
import urllib.request
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

try:
    from playwright_stealth import stealth_async
    HAS_STEALTH = True
except ImportError:
    HAS_STEALTH = False

from playwright.async_api import async_playwright, TimeoutError as PWTimeout

# ══════════════════════════════════════════════════════════════════════════════
# CONFIG
# ══════════════════════════════════════════════════════════════════════════════

CDP_URL         = "http://127.0.0.1:9222"
TARGET_URL      = "https://www.google.com/travel/search?q=hotels+in+Paris&hl=en&gl=us"
OUTPUT_XLSX     = "hotels_google_travel.xlsx"
OUTPUT_CSV      = "hotels_google_travel.csv"
CHECKPOINT_JSON = "hotels_checkpoint.json"

PAGE_LOAD_TIMEOUT   = 20_000
DELAY_AFTER_NEXT    = (1200, 2000)   # attente après clic Next (ms)
DELAY_EXTRACT       = (300, 600)     # courte pause avant extraction

# ══════════════════════════════════════════════════════════════════════════════
# SÉLECTEURS RÉELS (extraits du HTML fourni)
# ══════════════════════════════════════════════════════════════════════════════

# ─── Carte hôtel ────────────────────────────────────────────────────────────
# Chaque carte est un div avec jsname="mutHjb" et data-hveid
# Le lien principal est a.PVOOXe avec aria-label = nom de l'hôtel
# Le nom est aussi dans h2.BgYkof à l'intérieur

CARD_SELECTORS = [
    # Sélecteur principal : lien ancre PVOOXe (contient aria-label = nom)
    "a.PVOOXe",
    # Fallback : h2 avec classe BgYkof (titre de la carte hôtel)
    "h2.BgYkof",
    # Fallback 2 : div jsname="mutHjb" (conteneur carte)
    'div[jsname="mutHjb"]',
    # Fallback 3 : data-hveid sur le div parent
    "div[data-hveid] a[aria-label]",
]

# ─── Bouton "Next" (page suivante) ──────────────────────────────────────────
# Dans le HTML : button[aria-label="Next"] avec jsname="ATPzgb"
# Aussi : button[aria-label="Next"] dans le carousel photo (à éviter)
# On cible spécifiquement le bouton de pagination principal

NEXT_BTN_SELECTORS = [
    # Sélecteur le plus précis : jsname ATPzgb = bouton Next pagination
    'button[jsname="ATPzgb"][aria-label="Next"]',
    # Variante sans jsname fixe
    'button[aria-label="Next"]:not([tabindex="-1"])',
    # Dernier recours
    'button[aria-label="Next"]',
]

# ─── Bouton précédent (pour l'éviter) ───────────────────────────────────────
PREV_BTN_JSNAME = "l0FUab"   # jsname du bouton "Back"

# ══════════════════════════════════════════════════════════════════════════════
# UTILITAIRES
# ══════════════════════════════════════════════════════════════════════════════

def rand_ms(lo: int, hi: int) -> float:
    return random.randint(lo, hi) / 1000

def wait_for_chrome() -> bool:
    print("🔍 Recherche Chrome sur port 9222...")
    for i in range(20):
        try:
            urllib.request.urlopen(f"{CDP_URL}/json/version", timeout=2)
            print("✅ Chrome détecté !")
            return True
        except Exception:
            time.sleep(1)
    return False

def load_checkpoint() -> list[dict]:
    if Path(CHECKPOINT_JSON).exists():
        try:
            with open(CHECKPOINT_JSON, encoding="utf-8") as f:
                data = json.load(f)
            print(f"♻️  Checkpoint : {len(data)} hôtels déjà collectés")
            return data
        except Exception:
            pass
    return []

def save_checkpoint(hotels: list[dict]) -> None:
    with open(CHECKPOINT_JSON, "w", encoding="utf-8") as f:
        json.dump(hotels, f, ensure_ascii=False, indent=2)

# ══════════════════════════════════════════════════════════════════════════════
# EXTRACTION DES HÔTELS — LOGIQUE PRINCIPALE
# ══════════════════════════════════════════════════════════════════════════════

async def extract_hotels_from_page(page) -> list[dict]:
    """
    Extrait nom + lien de tous les hôtels visibles sur la page.

    Structure HTML réelle observée :
        <a class="PVOOXe" aria-label="Hôtel Mercure Paris Porte d'Orléans"
           href="/travel/search?...&qs=CAEy...">
        <h2 class="BgYkof ogfYpf ykx2he">Hôtel Mercure Paris Porte d'Orléans</h2>
    """
    hotels = []
    seen   = set()

    await asyncio.sleep(rand_ms(*DELAY_EXTRACT))

    # ── Stratégie 1 : via a.PVOOXe (lien principal de chaque carte) ──────────
    try:
        anchors = await page.query_selector_all("a.PVOOXe")
        print(f"     → a.PVOOXe trouvés : {len(anchors)}")

        for a in anchors:
            try:
                name = (await a.get_attribute("aria-label") or "").strip()
                href = (await a.get_attribute("href") or "").strip()

                # Ignorer les liens sans nom ou trop courts
                if not name or len(name) < 3:
                    continue

                # Ignorer les doublons
                key = name.lower()
                if key in seen:
                    continue
                seen.add(key)

                # Construire URL complète
                if href.startswith("/"):
                    href = "https://www.google.com" + href

                hotels.append({"name": name, "url": href})
            except Exception:
                continue

    except Exception as e:
        print(f"     ⚠️  Erreur a.PVOOXe : {e}")

    # ── Stratégie 2 : via h2.BgYkof (fallback si PVOOXe échoue) ─────────────
    if not hotels:
        try:
            h2s = await page.query_selector_all("h2.BgYkof")
            print(f"     → h2.BgYkof trouvés : {len(h2s)}")

            for h2 in h2s:
                try:
                    name = (await h2.inner_text()).strip()
                    if not name or len(name) < 3:
                        continue

                    key = name.lower()
                    if key in seen:
                        continue
                    seen.add(key)

                    # Remonter au parent pour trouver le lien
                    href = ""
                    try:
                        # Le lien est souvent un ancêtre direct
                        parent_link = await h2.evaluate("""
                            el => {
                                let p = el.parentElement;
                                for (let i = 0; i < 5; i++) {
                                    if (!p) break;
                                    const a = p.querySelector('a.PVOOXe');
                                    if (a) return a.href;
                                    p = p.parentElement;
                                }
                                return '';
                            }
                        """)
                        href = parent_link or ""
                    except Exception:
                        pass

                    hotels.append({"name": name, "url": href})
                except Exception:
                    continue

        except Exception as e:
            print(f"     ⚠️  Erreur h2.BgYkof : {e}")

    # ── Stratégie 3 : extraction JS pure (ultra-robuste) ─────────────────────
    if not hotels:
        try:
            result = await page.evaluate("""
                () => {
                    const hotels = [];
                    const seen = new Set();

                    // Méthode A : via a.PVOOXe
                    document.querySelectorAll('a.PVOOXe').forEach(a => {
                        const name = (a.getAttribute('aria-label') || '').trim();
                        const href = a.href || '';
                        if (name && name.length > 2 && !seen.has(name.toLowerCase())) {
                            seen.add(name.toLowerCase());
                            hotels.push({ name, url: href });
                        }
                    });

                    // Méthode B : via h2.BgYkof si A échoue
                    if (hotels.length === 0) {
                        document.querySelectorAll('h2.BgYkof').forEach(h2 => {
                            const name = h2.innerText.trim();
                            if (!name || name.length < 3 || seen.has(name.toLowerCase())) return;
                            seen.add(name.toLowerCase());
                            // Chercher lien parent
                            let href = '';
                            let p = h2.parentElement;
                            for (let i = 0; i < 6; i++) {
                                if (!p) break;
                                const a = p.querySelector('a[href*="/travel/"]');
                                if (a) { href = a.href; break; }
                                p = p.parentElement;
                            }
                            hotels.push({ name, url: href });
                        });
                    }

                    // Méthode C : data-hveid containers
                    if (hotels.length === 0) {
                        document.querySelectorAll('[data-hveid][jsname="mutHjb"]').forEach(card => {
                            const h2 = card.querySelector('h2');
                            const a  = card.querySelector('a[aria-label]');
                            const name = (a?.getAttribute('aria-label') || h2?.innerText || '').trim();
                            const href = a?.href || '';
                            if (name && !seen.has(name.toLowerCase())) {
                                seen.add(name.toLowerCase());
                                hotels.push({ name, url: href });
                            }
                        });
                    }

                    return hotels;
                }
            """)
            if result:
                print(f"     → JS extraction : {len(result)} hôtels")
                seen_js = {h["name"].lower() for h in hotels}
                for h in result:
                    if h["name"].lower() not in seen_js:
                        hotels.append(h)
                        seen_js.add(h["name"].lower())

        except Exception as e:
            print(f"     ⚠️  Erreur JS extraction : {e}")

    # ── Debug : afficher le DOM si toujours vide ──────────────────────────────
    if not hotels:
        print("     ⚠️  Aucun hôtel trouvé — analyse du DOM...")
        try:
            # Compter les éléments courants pour debug
            counts = await page.evaluate("""
                () => ({
                    pvooXe:   document.querySelectorAll('a.PVOOXe').length,
                    bgYkof:   document.querySelectorAll('h2.BgYkof').length,
                    mutHjb:   document.querySelectorAll('[jsname="mutHjb"]').length,
                    dataHveid: document.querySelectorAll('[data-hveid]').length,
                    allH2:    document.querySelectorAll('h2').length,
                    allLinks: document.querySelectorAll('a[href*="travel"]').length,
                    title:    document.title,
                })
            """)
            print(f"     DOM : {counts}")

            # Si pas d'éléments attendus, screenshot
            await page.screenshot(path="debug_page.png", full_page=False)
            print("     📸 Screenshot → debug_page.png")

        except Exception as e:
            print(f"     Debug error: {e}")

    return hotels


# ══════════════════════════════════════════════════════════════════════════════
# CLIC BOUTON "NEXT" — PAGINATION
# ══════════════════════════════════════════════════════════════════════════════

async def click_next_button(page) -> bool:
    """
    Clique sur le bouton Next de pagination.
    Retourne True si cliqué avec succès, False si fin ou introuvable.

    Note : il y a DEUX types de boutons "Next" dans le DOM :
      1. Le Next du carousel de photos (jsname="KpyLEe") → À ÉVITER
      2. Le Next de pagination globale (jsname="ATPzgb") → CIBLE
    """

    # ── Tentative 1 : bouton pagination principal (jsname="ATPzgb") ──────────
    try:
        # Ce sélecteur cible précisément le bouton de pagination
        # (pas le carousel photos)
        btn = await page.query_selector('button[jsname="ATPzgb"]')
        if btn:
            is_visible  = await btn.is_visible()
            is_disabled = await btn.get_attribute("disabled")
            aria_label  = await btn.get_attribute("aria-label") or ""

            print(f"     🔘 Bouton jsname=ATPzgb | visible={is_visible} | disabled={is_disabled} | label={aria_label}")

            if is_disabled is not None:
                print("     🏁 Bouton Next désactivé → fin pagination")
                return False

            if is_visible:
                await btn.scroll_into_view_if_needed()
                await asyncio.sleep(rand_ms(100, 300))
                await btn.click()
                await asyncio.sleep(rand_ms(*DELAY_AFTER_NEXT))
                return True
    except Exception as e:
        print(f"     ⚠️  Erreur clic jsname=ATPzgb : {e}")

    # ── Tentative 2 : aria-label="Next" non désactivé ────────────────────────
    try:
        # Récupérer TOUS les boutons Next et prendre celui qui n'est pas
        # dans un carousel (pas tabindex="-1")
        btns = await page.query_selector_all('button[aria-label="Next"]')
        print(f"     🔘 Boutons aria-label=Next trouvés : {len(btns)}")

        for btn in btns:
            tabindex    = await btn.get_attribute("tabindex")
            is_disabled = await btn.get_attribute("disabled")
            is_visible  = await btn.is_visible()

            # Le bouton carousel a tabindex="-1", celui de pagination non
            if tabindex == "-1":
                continue
            if is_disabled is not None:
                print("     🏁 Bouton Next disabled → fin")
                return False
            if is_visible:
                await btn.scroll_into_view_if_needed()
                await asyncio.sleep(rand_ms(100, 200))
                await btn.click()
                await asyncio.sleep(rand_ms(*DELAY_AFTER_NEXT))
                print("     ✅ Clic Next via aria-label (tabindex OK)")
                return True

    except Exception as e:
        print(f"     ⚠️  Erreur clic aria-label=Next : {e}")

    # ── Tentative 3 : JavaScript direct ──────────────────────────────────────
    try:
        result = await page.evaluate("""
            () => {
                // Chercher le bouton Next de PAGINATION (pas carousel)
                // Le bouton pagination a jsname="ATPzgb"
                const byJsname = document.querySelector('button[jsname="ATPzgb"]');
                if (byJsname) {
                    if (byJsname.disabled) return 'disabled';
                    byJsname.click();
                    return 'clicked_jsname';
                }

                // Fallback : bouton Next avec tabindex != -1
                const allNext = document.querySelectorAll('button[aria-label="Next"]');
                for (const btn of allNext) {
                    if (btn.getAttribute('tabindex') === '-1') continue;
                    if (btn.disabled) return 'disabled';
                    btn.click();
                    return 'clicked_fallback';
                }

                return 'not_found';
            }
        """)
        print(f"     🔘 JS click résultat : {result}")
        if result in ("clicked_jsname", "clicked_fallback"):
            await asyncio.sleep(rand_ms(*DELAY_AFTER_NEXT))
            return True
        elif result == "disabled":
            return False

    except Exception as e:
        print(f"     ⚠️  Erreur JS click Next : {e}")

    print("     ℹ️  Bouton Next introuvable → fin probable")
    return False


# ══════════════════════════════════════════════════════════════════════════════
# ATTENTE CHARGEMENT NOUVELLE PAGE
# ══════════════════════════════════════════════════════════════════════════════

async def wait_for_new_results(page, previous_first_name: str) -> bool:
    """
    Attend que les résultats changent après un clic Next.
    Compare le premier nom visible avec l'ancien.
    """
    for attempt in range(20):  # max 10s
        await asyncio.sleep(0.5)
        try:
            first = await page.query_selector("a.PVOOXe")
            if first:
                name = (await first.get_attribute("aria-label") or "").strip()
                if name and name != previous_first_name:
                    return True
        except Exception:
            pass
    return False  # timeout


# ══════════════════════════════════════════════════════════════════════════════
# BOUCLE PRINCIPALE
# ══════════════════════════════════════════════════════════════════════════════

async def scrape_all(page) -> list[dict]:
    """Parcourt toutes les pages en cliquant Next."""
    all_hotels   = load_checkpoint()
    known_names  = {h["name"].lower() for h in all_hotels}
    page_num     = 0
    consec_empty = 0
    t0           = time.time()

    print(f"\n{'━'*72}")
    print(f"  {'Page':>5}  {'Nouveaux':>8}  {'Total':>7}  {'Durée':>8}  Dernier hôtel")
    print(f"{'━'*72}")

    while True:
        page_num += 1

        # ── Extraire les hôtels de la page courante ──────────────────────────
        hotels_page = await extract_hotels_from_page(page)
        new_hotels  = []

        for h in hotels_page:
            k = h["name"].lower()
            if k not in known_names:
                known_names.add(k)
                all_hotels.append(h)
                new_hotels.append(h)

        elapsed   = time.time() - t0
        last_name = new_hotels[-1]["name"][:30] if new_hotels else "—"

        print(
            f"  {page_num:>5}  {len(new_hotels):>8}  {len(all_hotels):>7}"
            f"  {elapsed:>6.1f}s  {last_name}"
        )

        # Checkpoint toutes les 5 pages
        if page_num % 5 == 0:
            save_checkpoint(all_hotels)

        # Gestion pages vides
        if not new_hotels:
            consec_empty += 1
            if consec_empty >= 30:
                print("\n  🏁 3 pages vides consécutives → fin")
                break
        else:
            consec_empty = 0

        # ── Mémoriser le premier nom pour détecter le chargement suivant ─────
        first_name = hotels_page[0]["name"] if hotels_page else ""

        # ── Cliquer sur Next ──────────────────────────────────────────────────
        has_next = await click_next_button(page)
        if not has_next:
            print("\n  ✅ Fin de pagination")
            break

        # ── Attendre que la nouvelle page charge ──────────────────────────────
        changed = await wait_for_new_results(page, first_name)
        if not changed:
            print("  ⚠️  Résultats inchangés après Next")

    save_checkpoint(all_hotels)
    print(f"{'━'*72}")
    print(f"  🎯 {len(all_hotels)} hôtels | {page_num} pages | {time.time()-t0:.1f}s")
    return all_hotels


# ══════════════════════════════════════════════════════════════════════════════
# EXPORT XLSX + CSV
# ══════════════════════════════════════════════════════════════════════════════

def export(hotels: list[dict]) -> None:
    # CSV
    with open(OUTPUT_CSV, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=["#", "name", "url"])
        w.writeheader()
        for i, h in enumerate(hotels, 1):
            w.writerow({"#": i, "name": h["name"], "url": h["url"]})
    print(f"  💾 CSV  → {OUTPUT_CSV}")

    # XLSX
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hôtels Paris"

    GREEN = PatternFill("solid", fgColor="00897B")
    ODD   = PatternFill("solid", fgColor="E8F5E9")
    EVEN  = PatternFill("solid", fgColor="FFFFFF")
    thin  = Side(style="thin", color="B2DFDB")
    brd   = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Titre
    ws.merge_cells("A1:C1")
    c = ws["A1"]
    c.value = f"🏨  Hôtels Paris — Google Travel  |  {len(hotels)} établissements  |  {datetime.now():%d/%m/%Y %H:%M}"
    c.font      = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    c.fill      = GREEN
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # En-têtes
    for col, (h, w) in enumerate([("#", 7), ("Nom de l'hôtel", 55), ("Lien Google Travel", 72)], 1):
        cell = ws.cell(2, col, h)
        cell.font      = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        cell.fill      = GREEN
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = brd
        ws.column_dimensions[chr(64+col)].width = w
    ws.row_dimensions[2].height = 22

    # Données
    for i, h in enumerate(hotels, 1):
        row  = i + 2
        fill = ODD if i % 2 else EVEN

        c0 = ws.cell(row, 1, i)
        c0.font = Font(name="Arial", size=10, color="888888")
        c0.fill = fill ; c0.alignment = Alignment(horizontal="center") ; c0.border = brd

        c1 = ws.cell(row, 2, h["name"])
        c1.font = Font(name="Arial", size=10, bold=True)
        c1.fill = fill ; c1.border = brd ; c1.alignment = Alignment(vertical="center")

        c2 = ws.cell(row, 3, h["url"] or "—")
        if h.get("url", "").startswith("http"):
            c2.hyperlink = h["url"]
            c2.font = Font(name="Arial", size=9, color="0277BD", underline="single")
        else:
            c2.font = Font(name="Arial", size=9, color="999999")
        c2.fill = fill ; c2.border = brd ; c2.alignment = Alignment(vertical="center")

        ws.row_dimensions[row].height = 17

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:C{len(hotels)+2}"
    wb.save(OUTPUT_XLSX)
    print(f"  💾 XLSX → {OUTPUT_XLSX}")


# ══════════════════════════════════════════════════════════════════════════════
# SETUP ANTI-DÉTECTION
# ══════════════════════════════════════════════════════════════════════════════

async def patch_fingerprint(page) -> None:
    await page.add_init_script("""
        Object.defineProperty(navigator, 'webdriver', { get: () => undefined });
        Object.defineProperty(navigator, 'plugins',   { get: () => [1,2,3,4,5] });
        Object.defineProperty(navigator, 'languages', { get: () => ['en-US','en','fr'] });
        window.chrome = { runtime: {} };
        const orig = navigator.permissions.query.bind(navigator.permissions);
        navigator.permissions.query = p =>
            p.name === 'notifications'
                ? Promise.resolve({ state: Notification.permission })
                : orig(p);
    """)


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

async def main(mode: str):
    print("╔══════════════════════════════════════════════════════════════╗")
    print("║   Google Travel Scraper v3 — Sélecteurs Réels               ║")
    print(f"║   Mode : {mode:<52}║")
    print("╚══════════════════════════════════════════════════════════════╝\n")

    async with async_playwright() as p:

        if mode == "cdp":
            if not wait_for_chrome():
                print("❌ Chrome non détecté sur port 9222")
                print("\nLance Chrome avec :")
                print('  chrome.exe --remote-debugging-port=9222 --user-data-dir=C:\\chrome_scrape')
                return
            browser = await p.chromium.connect_over_cdp(CDP_URL)
            print(f"✅ Connecté Chrome {browser.version}")
            context = browser.contexts[0] if browser.contexts else await browser.new_context()

        else:  # visible ou headless
            browser = await p.chromium.launch(
                headless=(mode == "headless"),
                args=[
                    "--no-sandbox",
                    "--disable-blink-features=AutomationControlled",
                    "--window-size=1366,900",
                ],
            )
            context = await browser.new_context(
                viewport={"width": 1366, "height": 900},
                user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
                locale="en-US",
                timezone_id="Europe/Paris",
            )

        page = await context.new_page()
        await patch_fingerprint(page)
        if HAS_STEALTH:
            await stealth_async(page)
            print("  🛡️  playwright-stealth actif")

        # Navigation
        print(f"\n🌐 Navigation vers Google Travel...")
        try:
            await page.goto(TARGET_URL, wait_until="domcontentloaded", timeout=PAGE_LOAD_TIMEOUT)
        except PWTimeout:
            print("  ⚠️  Timeout navigation (on continue)")
        except Exception as e:
            print(f"  ❌ {e}") ; return

        await asyncio.sleep(rand_ms(1500, 2500))

        # Accepter cookies
        for txt in ["Accept all", "Tout accepter", "I agree", "Agree"]:
            try:
                btn = await page.query_selector(f"button:has-text('{txt}')")
                if btn and await btn.is_visible():
                    await btn.click()
                    print(f"  🍪 Cookies acceptés ({txt})")
                    await asyncio.sleep(rand_ms(800, 1200))
                    break
            except Exception:
                continue

        # Screenshot initial si mode visible
        if mode == "visible":
            await page.screenshot(path="debug_init.png")
            print("  📸 debug_init.png")

        # Scraping
        all_hotels = await scrape_all(page)

        # Export
        print(f"\n📦 Export...")
        if all_hotels:
            export(all_hotels)
            print(f"\n✅ {len(all_hotels)} hôtels exportés")
        else:
            print("❌ Aucun hôtel — vérifie debug_page.png")

        if mode != "cdp":
            await browser.close()


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--mode", choices=["cdp", "visible", "headless"], default="cdp")
    args = ap.parse_args()
    asyncio.run(main(args.mode))