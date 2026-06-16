import asyncio
import os
import re
import json
from datetime import datetime
from playwright.async_api import async_playwright
import PyPDF2
from openpyxl import load_workbook

# ============================================================
# CONFIGURATION
# ============================================================
DOWNLOAD_DIR = r"C:\Users\A\Downloads"
PDF_FILENAME = "Profile.pdf"
EXCEL_FILE = "linkedin_profiles_formation_ia.xlsx"
SHEET_NAME = "Sheet1"
COOKIES_SAVE_PATH = "linkedin_cookies_new.json"

class LinkedInContactExtractor:
    def __init__(self):
        self.playwright = None
        self.browser = None
        self.context = None
        self.page = None
        
    async def init_browser(self):
        """Initialise Playwright"""
        try:
            self.playwright = await async_playwright().start()
            
            self.browser = await self.playwright.chromium.launch(
                headless=False,
                args=[
                    '--disable-blink-features=AutomationControlled',
                    '--no-sandbox',
                ]
            )
            
            self.context = await self.browser.new_context(
                viewport={'width': 1920, 'height': 1080},
                user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
                locale='fr-FR',
                accept_downloads=True
            )
            
            self.page = await self.context.new_page()
            print("✅ Navigateur initialisé")
            return True
            
        except Exception as e:
            print(f"❌ Erreur initialisation: {e}")
            return False
    
    async def manual_login(self):
        """Connexion manuelle à LinkedIn"""
        try:
            print("\n" + "="*70)
            print("🔐 CONNEXION MANUELLE À LINKEDIN")
            print("="*70)
            print("1. Une fenêtre de navigateur va s'ouvrir")
            print("2. Connectez-vous à LinkedIn manuellement")
            print("3. Une fois connecté et sur votre feed, revenez ici")
            print("="*70)
            
            await self.page.goto("https://www.linkedin.com", wait_until="domcontentloaded")
            await asyncio.sleep(2)
            
            input("\n⏸️  Connectez-vous à LinkedIn, puis appuyez sur ENTRÉE ici...")
            
            await asyncio.sleep(2)
            
            if "feed" in self.page.url or "mynetwork" in self.page.url or "in/" in self.page.url:
                print("\n✅ Connexion réussie !")
                await self.save_cookies()
                return True
            else:
                print("\n⚠️  Vous ne semblez pas être connecté")
                retry = input("Voulez-vous réessayer ? (o/n) : ")
                if retry.lower() == 'o':
                    return await self.manual_login()
                return False
                
        except Exception as e:
            print(f"❌ Erreur connexion: {e}")
            return False
    
    async def save_cookies(self):
        """Sauvegarde les cookies pour la prochaine session"""
        try:
            cookies = await self.context.cookies()
            with open(COOKIES_SAVE_PATH, "w", encoding="utf-8") as f:
                json.dump(cookies, f, indent=2)
            print(f"✅ Cookies sauvegardés dans {COOKIES_SAVE_PATH}")
        except Exception as e:
            print(f"⚠️  Erreur sauvegarde cookies: {e}")
    
    def load_excel_data(self):
        """Charge les données depuis Excel"""
        try:
            print("\n📊 Chargement des données Excel...")
            
            if not os.path.exists(EXCEL_FILE):
                print(f"❌ Fichier Excel introuvable: {EXCEL_FILE}")
                return None, None
            
            workbook = load_workbook(EXCEL_FILE)
            sheet = workbook[SHEET_NAME] if SHEET_NAME in workbook.sheetnames else workbook.active
            
            data = []
            headers = []
            
            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(cell).strip() if cell else f"Column_{j}" for j, cell in enumerate(row)]
                else:
                    row_data = {}
                    for j, cell in enumerate(row):
                        if j < len(headers):
                            row_data[headers[j]] = cell if cell else ""
                    data.append(row_data)
            
            print(f"✅ {len(data)} profils chargés")
            return data, workbook
            
        except Exception as e:
            print(f"❌ Erreur chargement Excel: {e}")
            return None, None
    
    async def download_profile_pdf(self, profile_url):
        """Télécharge le PDF d'un profil - Version optimisée"""
        try:
            print(f"   🔗 Accès à {profile_url}")
            await self.page.goto(profile_url, wait_until="domcontentloaded", timeout=20000)
            await asyncio.sleep(2)
            
            # Scroll vers le haut
            await self.page.evaluate("window.scrollTo(0, 0)")
            await asyncio.sleep(1)
            
            # ============================================================
            # ÉTAPE 1: Trouver et cliquer sur le bouton "Plus"
            # ============================================================
            print("   🔍 Recherche du bouton 'Plus'...")
            
            more_button = None
            
            # Stratégie 1: Par ID pattern (le plus fiable selon le HTML)
            try:
                print("   🔄 Tentative de clic JavaScript direct...")
                clicked = await self.page.evaluate("""() => {
                    const buttons = Array.from(document.querySelectorAll('button'));
                    for (let btn of buttons) {
                        const text = btn.textContent.trim();
                        const ariaLabel = btn.getAttribute('aria-label') || '';
                        if ((text === 'Plus' || text === 'More' || 
                                ariaLabel.includes('Plus d') || 
                                ariaLabel.includes('More actions')) &&
                            (btn.closest('.pv-top-card') || btn.id.includes('profile-overflow'))) {
                            btn.scrollIntoView({behavior: 'smooth', block: 'center'});
                            btn.click();
                            return true;
                        }
                    }
                    return false;
                }""")
                
                if clicked:
                    await asyncio.sleep(1)
                    print("   ✅ Menu 'Plus' ouvert via JavaScript direct")
                    # Créer un faux objet pour continuer
                    more_button = "clicked_via_js"
            except Exception as e:
                print(f"   ⚠️  Clic JavaScript échoué: {e}")
            
            # Scroller vers le bouton et cliquer (sauf si déjà cliqué via JS)
            if more_button != "clicked_via_js":
                await more_button.scroll_into_view_if_needed()
                await asyncio.sleep(0.5)
                await more_button.click()
                await asyncio.sleep(1)
                print("   ✅ Menu 'Plus' ouvert")
            
            # ============================================================
            # ÉTAPE 2: Trouver et cliquer sur "Enregistrer au format PDF"
            # ============================================================
            print("   🔍 Recherche de l'option PDF...")
            
            pdf_option = None
            
            # Stratégie 1: Par aria-label exact (le plus fiable)
            try:
                pdf_option = await self.page.wait_for_selector(
                    'div[aria-label="Enregistrer au format PDF"]',
                    state="visible",
                    timeout=5000
                )
                print("   ✓ Option PDF trouvée via aria-label FR")
            except:
                pass
            
            # Télécharger le PDF
            print("   ⏳ Démarrage du téléchargement...")
            async with self.page.expect_download(timeout=30000) as download_info:
                await pdf_option.click()
                download = await download_info.value
            
            pdf_path = os.path.join(DOWNLOAD_DIR, PDF_FILENAME)
            await download.save_as(pdf_path)
            print(f"   ✅ PDF téléchargé: {PDF_FILENAME}")
            
            return pdf_path
            
        except Exception as e:
            print(f"   ❌ Erreur téléchargement: {e}")
            try:
                await self.page.screenshot(path="debug_error.png")
                print("   📸 Screenshot sauvegardé: debug_error.png")
            except:
                pass
            return None
    
    def extract_contact_from_pdf(self, pdf_path):
        """Extrait email et téléphone du PDF"""
        try:
            with open(pdf_path, 'rb') as f:
                reader = PyPDF2.PdfReader(f)
                text = ""
                for page in reader.pages:
                    text += page.extract_text() or ""
                
                # Email
                email = None
                email_pattern = r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}'
                emails = re.findall(email_pattern, text)
                for e in emails:
                    if not any(x in e.lower() for x in ['linkedin', 'example', 'test']):
                        email = e
                        break
                
                # Téléphone
                phone = None
                phone_patterns = [
                    r'\+?\d{1,3}[\s.-]?\(?\d{1,4}\)?[\s.-]?\d{1,4}[\s.-]?\d{1,4}[\s.-]?\d{1,9}',
                    r'0[1-9](?:[\s.-]?\d{2}){4}',
                ]
                
                for pattern in phone_patterns:
                    phones = re.findall(pattern, text)
                    if phones:
                        p = phones[0].strip()
                        if len(re.sub(r'\D', '', p)) >= 10:
                            phone = p
                            break
                
                return email, phone
                
        except Exception as e:
            print(f"   ⚠️  Erreur extraction: {e}")
            return None, None
    
    async def update_excel(self, workbook, row_index, email, phone):
        """Met à jour Excel"""
        try:
            sheet = workbook[SHEET_NAME] if SHEET_NAME in workbook.sheetnames else workbook.active
            actual_row = row_index + 2
            
            headers = [cell.value for cell in sheet[1]]
            
            for i, header in enumerate(headers, start=1):
                if header and 'email' in str(header).lower() and email:
                    sheet.cell(row=actual_row, column=i, value=email)
                if header and 'phone' in str(header).lower() and phone:
                    sheet.cell(row=actual_row, column=i, value=phone)
            
            workbook.save(EXCEL_FILE)
            print(f"   💾 Excel mis à jour (ligne {actual_row})")
            
        except Exception as e:
            print(f"   ⚠️  Erreur mise à jour: {e}")
    
    async def process_profile(self, row_index, profile_url, current_email, current_phone):
        """Traite un profil"""
        print(f"\n{'='*70}")
        print(f"📋 Profil {row_index + 1}: {profile_url}")
        print(f"{'='*70}")
        
        if current_email and current_phone:
            print("   ✓ Données déjà complètes")
            return {"email": current_email, "phone": current_phone}
        
        pdf_path = await self.download_profile_pdf(profile_url)
        
        if not pdf_path:
            return {"email": current_email or "", "phone": current_phone or ""}
        
        await asyncio.sleep(2)
        
        email, phone = self.extract_contact_from_pdf(pdf_path)
        
        result = {
            "email": email or current_email or "",
            "phone": phone or current_phone or ""
        }
        
        if email:
            print(f"   ✅ Email: {email}")
        if phone:
            print(f"   ✅ Téléphone: {phone}")
        
        try:
            if os.path.exists(pdf_path):
                os.remove(pdf_path)
                print("   🗑️  PDF supprimé")
        except:
            pass
        
        await asyncio.sleep(3)
        
        return result
    
    async def run(self):
        """Lance l'extraction"""
        try:
            if not await self.init_browser():
                return
            
            if not await self.manual_login():
                print("❌ Connexion échouée")
                return
            
            data, workbook = self.load_excel_data()
            if not data:
                return
            
            print("\n" + "="*70)
            print(f"🎯 EXTRACTION: {len(data)} profils")
            print("="*70)
            
            for index, row in enumerate(data):
                profile_url = row.get('profile_url', '')
                current_email = row.get('emails', '')
                current_phone = row.get('phone', '')
                
                if not profile_url:
                    continue
                
                extracted = await self.process_profile(
                    index, profile_url, current_email, current_phone
                )
                
                if extracted['email'] or extracted['phone']:
                    await self.update_excel(
                        workbook, index, extracted['email'], extracted['phone']
                    )
            
        except KeyboardInterrupt:
            print("\n⚠️  Arrêt manuel")
        finally:
            if self.browser:
                await self.browser.close()
            if self.playwright:
                await self.playwright.stop()
            print("\n✅ TERMINÉ")

async def main():
    extractor = LinkedInContactExtractor()
    await extractor.run()

if __name__ == "__main__":
    print("""
╔══════════════════════════════════════════════════════════════════╗
║     📧 LINKEDIN CONTACT EXTRACTOR - VERSION OPTIMISÉE 📞        ║
║                                                                  ║
║  ✅ Sélecteurs améliorés basés sur le HTML réel                 ║
║  🔐 Connexion manuelle sécurisée                                ║
║  📄 Extraction via PDF avec multi-stratégies                    ║
║  💾 Mise à jour Excel automatique                               ║
╚══════════════════════════════════════════════════════════════════╝
    """)
    
    asyncio.run(main())