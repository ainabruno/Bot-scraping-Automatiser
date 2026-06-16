"""
Fever Up — Activités Paris Scraper v2
=======================================
PROBLÈME v1 : Fever Up utilise un "virtual scroll" qui DÉTRUIT les cartes
hors-écran → on ne voyait que 48 éléments même en scrollant jusqu'en bas.

SOLUTION v2 : Double stratégie
  1. INTERCEPTION API (principale) — On capture les requêtes XHR/fetch que
     Fever Up envoie à son backend pour charger les activités par lots.
     → Récupère TOUTES les activités (250+) sans limite de DOM.

  2. FALLBACK DOM — Si l'API change, on reprend le scroll mais en
     sauvegardant chaque carte AVANT qu'elle soit détruite par le virtual
     scroll, via un MutationObserver JavaScript injecté dans la page.

Données collectées :
  - URL              : https://feverup.com/m/<id>
  - Nom_Prestataire  : ex. La Grande Halle de la Villette
  - Nom_General      : ex. La Légende du Titanic : L'exposition ultime
  - Prix             : ex. 24.25 EUR
  - Note             : ex. 4.3
  - Nb_Avis          : ex. 4664
  - Dates            : ex. 09 juin - 31 août
  - Image_URL        : URL de l'image principale

ÉTAPES :
  1. Lance Chrome en mode debug :
     chrome.exe --remote-debugging-port=9222 --user-data-dir="C:\\ChromeDebug"
  2. Navigue vers https://feverup.com/fr/paris/choses-a-faire
  3. Attends le chargement complet
  4. Lance ce script : python feverup_paris_scraper.py

Installation :
    pip install playwright openpyxl
    playwright install chromium
"""

import asyncio
import json
import random
import re
import time
import urllib.request
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from playwright.async_api import async_playwright, TimeoutError as PWTimeout

# ─── Configuration ─────────────────────────────────────────────────────────────

BASE_URL     = "https://feverup.com"
URL_MAIN     = f"{BASE_URL}/fr/paris/choses-a-faire"
OUTPUT       = "feverup_paris_activites.xlsx"
PAGE_TIMEOUT = 40_000
CDP_URL      = "http://127.0.0.1:9222"

# Délai entre scrolls (secondes) — augmenter si connexion lente
PAUSE_SCROLL = 1.5
# Nombre de scrolls sans nouveaux éléments avant d'arrêter
MAX_NO_CHANGE = 5

# ─── Sélecteurs DOM (fallback) ─────────────────────────────────────────────────

SEL_CARD      = "li[data-plan-id]"
SEL_NAME      = "[data-testid='fv-plan-card-title']"
SEL_LOCATION  = "[data-testid='fv-plan-location__name']"
SEL_RATING    = ".fv-rating-plan .fw-bold"
SEL_NB_AVIS   = ".fv-rating-plan__num"
SEL_DATES     = "[data-testid='fv-plan-card-v2__date-range']"
SEL_IMG       = "figure img"

# Mots-clés pour identifier les appels API Fever Up
API_KEYWORDS = [
    "api.feverup.com",
    "feverup.com/api",
    "/plans",
    "/experiences",
    "/events",
    "what_to_do",
    "plan_list",
    "wpf",         # "What's Popular Feed"
]

# ─── Helpers ───────────────────────────────────────────────────────────────────

def clean(text: str) -> str:
    return " ".join(text.split()).strip()

def build_url(plan_id) -> str:
    return f"{BASE_URL}/m/{plan_id}"

def parse_price(raw) -> str:
    if isinstance(raw, (int, float)):
        return str(raw)
    if isinstance(raw, str):
        m = re.search(r"[\d,\.]+", raw)
        return m.group(0) if m else raw.strip()
    return ""

def extract_nb_avis(text: str) -> str:
    m = re.search(r"\(?([\d,\s]+)\)?", text)
    return m.group(1).replace(",", "").replace(" ", "").strip() if m else ""

# ─── Chrome connection ─────────────────────────────────────────────────────────

def wait_for_chrome(max_attempts: int = 15) -> bool:
    print("🔍  Vérification du port 9222 (Chrome)...")
    for attempt in range(1, max_attempts + 1):
        try:
            urllib.request.urlopen(f"{CDP_URL}/json/version", timeout=2)
            print(f"✅  Chrome détecté sur {CDP_URL}")
            return True
        except Exception:
            print(f"    ⏳ Tentative {attempt}/{max_attempts}...")
            time.sleep(2)
    return False

async def dismiss_overlays(page) -> None:
    for sel in [
        "#onetrust-accept-btn-handler",
        "button:has-text('Accepter')",
        "button:has-text('Accept All')",
        "button:has-text('Tout accepter')",
        "[aria-label='Close']",
    ]:
        try:
            btn = page.locator(sel).first
            if await btn.is_visible(timeout=1_500):
                await btn.click(timeout=3_000)
                await asyncio.sleep(0.6)
        except Exception:
            pass

async def is_blocked(page) -> bool:
    try:
        content = await page.content()
        return any(p.lower() in content.lower() for p in [
            "access is temporarily restricted", "unusual activity",
            "automated (bot) activity", "403 forbidden",
        ])
    except Exception:
        return False

# ─── STRATÉGIE 1 : Interception des requêtes API ──────────────────────────────

def is_api_call(url: str) -> bool:
    """Détermine si une URL est un appel API Fever Up contenant des activités."""
    url_lower = url.lower()
    return any(kw in url_lower for kw in API_KEYWORDS)

def parse_api_response(data: dict | list) -> list[dict]:
    """
    Tente d'extraire les activités depuis une réponse API JSON.
    Fever Up peut retourner les données sous plusieurs formats.
    """
    activities = []

    # Chercher récursivement les plans/activités dans la réponse
    def find_plans(obj, depth=0):
        if depth > 6:
            return
        if isinstance(obj, list):
            for item in obj:
                find_plans(item, depth + 1)
        elif isinstance(obj, dict):
            # Chercher les champs caractéristiques d'un plan Fever Up
            plan_id = (
                obj.get("id") or obj.get("plan_id") or
                obj.get("planId") or obj.get("_id")
            )
            title = (
                obj.get("title") or obj.get("name") or
                obj.get("plan_name") or obj.get("planName") or ""
            )

            if plan_id and title:
                # C'est probablement un plan !
                act = extract_from_api_obj(obj)
                if act:
                    activities.append(act)
            else:
                # Chercher dans les sous-champs
                for key in ["plans", "results", "data", "items",
                            "experiences", "events", "feed", "list",
                            "content", "payload"]:
                    if key in obj:
                        find_plans(obj[key], depth + 1)

    find_plans(data)
    return activities

def extract_from_api_obj(obj: dict) -> dict | None:
    """Extrait les champs utiles d'un objet plan de l'API."""
    plan_id = (
        obj.get("id") or obj.get("plan_id") or
        obj.get("planId") or obj.get("_id") or ""
    )
    if not plan_id:
        return None

    title = (
        obj.get("title") or obj.get("name") or
        obj.get("plan_name") or obj.get("planName") or ""
    )
    if not title:
        return None

    # Lieu / prestataire
    venue = obj.get("venue") or obj.get("location") or obj.get("place") or {}
    if isinstance(venue, dict):
        prestataire = (
            venue.get("name") or venue.get("title") or
            venue.get("venue_name") or ""
        )
    else:
        prestataire = str(venue)

    # Prix
    price_obj = obj.get("price") or obj.get("min_price") or {}
    if isinstance(price_obj, dict):
        prix_val  = price_obj.get("amount") or price_obj.get("value") or ""
        prix_curr = price_obj.get("currency") or "EUR"
        prix = f"{prix_val} {prix_curr}".strip() if prix_val else ""
    else:
        prix = parse_price(price_obj)

    # Note
    rating_obj = obj.get("rating") or obj.get("review") or {}
    if isinstance(rating_obj, dict):
        note    = str(rating_obj.get("score") or rating_obj.get("average") or "")
        nb_avis = str(rating_obj.get("count") or rating_obj.get("total") or "")
    else:
        note    = str(obj.get("rating_score") or obj.get("score") or "")
        nb_avis = str(obj.get("review_count") or obj.get("reviews_count") or "")

    # Dates
    dates = ""
    date_from = obj.get("date_from") or obj.get("start_date") or obj.get("from") or ""
    date_to   = obj.get("date_to")   or obj.get("end_date")   or obj.get("to")   or ""
    if date_from and date_to:
        dates = f"{date_from} - {date_to}"
    elif date_from:
        dates = str(date_from)

    # Image
    img = ""
    images = obj.get("images") or obj.get("photos") or []
    if isinstance(images, list) and images:
        first = images[0]
        img = first.get("url") or first.get("src") or str(first)
    elif isinstance(images, str):
        img = images
    if not img:
        img = obj.get("image") or obj.get("image_url") or obj.get("photo") or ""

    return {
        "url":             build_url(plan_id),
        "nom_prestataire": clean(str(prestataire)),
        "nom_general":     clean(str(title)),
        "prix":            prix,
        "note":            note,
        "nb_avis":         nb_avis,
        "dates":           dates,
        "image_url":       img,
    }

# ─── STRATÉGIE 2 : MutationObserver DOM (fallback) ────────────────────────────

MUTATION_OBSERVER_JS = """
() => {
    // Stockage global des cartes capturées avant destruction
    if (!window.__feverup_cards) {
        window.__feverup_cards = {};
    }

    const harvest = (node) => {
        if (!node || node.nodeType !== 1) return;

        // Chercher les li[data-plan-id] dans ce nœud et ses enfants
        const cards = node.matches && node.matches('li[data-plan-id]')
            ? [node]
            : Array.from(node.querySelectorAll('li[data-plan-id]'));

        for (const card of cards) {
            const planId = card.getAttribute('data-plan-id');
            if (!planId || window.__feverup_cards[planId]) continue;

            const nameEl    = card.querySelector('[data-testid="fv-plan-card-title"]');
            const locEl     = card.querySelector('[data-testid="fv-plan-location__name"]');
            const ratingEl  = card.querySelector('.fv-rating-plan .fw-bold');
            const avisEl    = card.querySelector('.fv-rating-plan__num');
            const datesEl   = card.querySelector('[data-testid="fv-plan-card-v2__date-range"]');
            const imgEl     = card.querySelector('figure img');

            window.__feverup_cards[planId] = {
                plan_id:         planId,
                plan_name:       card.getAttribute('data-plan-name') || (nameEl ? nameEl.innerText : ''),
                plan_price:      card.getAttribute('data-plan-price') || '',
                plan_currency:   card.getAttribute('data-plan-currency') || 'EUR',
                nom_prestataire: locEl    ? locEl.innerText.trim()    : '',
                note:            ratingEl ? ratingEl.innerText.trim() : '',
                nb_avis:         avisEl   ? avisEl.innerText.trim()   : '',
                dates:           datesEl  ? datesEl.innerText.trim()  : '',
                image_url:       imgEl    ? (imgEl.src || imgEl.getAttribute('data-src') || '') : '',
            };
        }
    };

    // Observer les ajouts ET suppressions dans tout le body
    const observer = new MutationObserver((mutations) => {
        for (const m of mutations) {
            // Cartes ajoutées
            for (const node of m.addedNodes) harvest(node);
            // Cartes supprimées (virtual scroll) — on les a déjà capturées à l'ajout
        }
    });

    observer.observe(document.body, { childList: true, subtree: true });

    // Harvester les cartes déjà présentes
    harvest(document.body);

    return 'MutationObserver installé';
}
"""

GET_CAPTURED_JS = "() => Object.values(window.__feverup_cards || {})"

# ─── Scroll robuste pour virtual scroll ───────────────────────────────────────

async def scroll_and_capture(page) -> list[dict]:
    """
    Installe le MutationObserver, puis scrolle pour déclencher
    le chargement de tous les lots. Retourne les cartes capturées.
    """
    print("    🔧  Installation du MutationObserver...")
    result = await page.evaluate(MUTATION_OBSERVER_JS)
    print(f"    ✅  {result}")

    last_count   = 0
    no_change    = 0
    scroll_step  = 600   # px par scroll
    total_height = await page.evaluate("document.body.scrollHeight")

    print("    🔄  Scroll progressif (virtual scroll)...")

    current_pos = 0
    while True:
        # Scroll d'un pas
        current_pos += scroll_step
        await page.evaluate(f"window.scrollTo(0, {current_pos})")
        await asyncio.sleep(PAUSE_SCROLL)

        # Mettre à jour la hauteur totale (peut grandir avec lazy load)
        total_height = await page.evaluate("document.body.scrollHeight")

        # Compter les cartes capturées
        captured = await page.evaluate(GET_CAPTURED_JS)
        count = len(captured)

        if count != last_count:
            print(f"        📦 {count} activités capturées | scroll={current_pos}px/{total_height}px")
            no_change = 0
            last_count = count
        else:
            no_change += 1

        # Condition d'arrêt
        if current_pos >= total_height and no_change >= 2:
            # Un dernier scroll tout en bas
            await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            await asyncio.sleep(PAUSE_SCROLL * 1.5)
            captured = await page.evaluate(GET_CAPTURED_JS)
            count = len(captured)
            print(f"\n    ✅  Fin du scroll — {count} activités capturées au total.")
            return captured

        if no_change >= MAX_NO_CHANGE:
            # Essai : scroll to bottom direct
            await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            await asyncio.sleep(PAUSE_SCROLL)
            new_height = await page.evaluate("document.body.scrollHeight")
            if new_height == total_height:
                captured = await page.evaluate(GET_CAPTURED_JS)
                print(f"\n    ✅  Fin du scroll — {len(captured)} activités capturées.")
                return captured
            else:
                total_height = new_height
                no_change = 0

    return await page.evaluate(GET_CAPTURED_JS)

def dom_cards_to_activities(cards: list[dict]) -> list[dict]:
    """Convertit les cartes DOM capturées en format activité standard."""
    activities = []
    for c in cards:
        plan_id = c.get("plan_id", "")
        if not plan_id:
            continue
        name = clean(c.get("plan_name", ""))
        if not name:
            continue

        prix_val  = c.get("plan_price", "")
        prix_curr = c.get("plan_currency", "EUR")
        prix = f"{prix_val} {prix_curr}".strip() if prix_val else ""

        activities.append({
            "url":             build_url(plan_id),
            "nom_prestataire": clean(c.get("nom_prestataire", "")),
            "nom_general":     name,
            "prix":            prix,
            "note":            c.get("note", ""),
            "nb_avis":         extract_nb_avis(c.get("nb_avis", "")),
            "dates":           clean(c.get("dates", "")),
            "image_url":       c.get("image_url", ""),
        })
    return activities

# ─── Scraping principal ────────────────────────────────────────────────────────

async def scrape() -> list[dict]:
    print("\n" + "="*65)
    print("  INSTRUCTIONS AVANT DE CONTINUER :")
    print("  1. Lance Chrome en mode debug :")
    print('     chrome.exe --remote-debugging-port=9222 --user-data-dir="C:\\ChromeDebug"')
    print("  2. Navigue vers :")
    print(f"     {URL_MAIN}")
    print("  3. Attends que la page se charge")
    print("  4. Reviens ici et appuie sur Entrée")
    print("="*65 + "\n")
    input("  → Prêt ? Appuie sur Entrée : ")

    if not wait_for_chrome():
        print("\n❌ Chrome ne répond pas sur le port 9222.")
        return []

    api_activities: list[dict] = []
    api_seen_ids:   set        = set()

    async with async_playwright() as p:
        try:
            browser = await p.chromium.connect_over_cdp(CDP_URL)
            print(f"✅  Connecté à Chrome ({browser.version})\n")
        except Exception as e:
            print(f"❌ Connexion CDP échouée : {e}")
            return []

        context = (
            browser.contexts[0] if browser.contexts
            else await browser.new_context(viewport={"width": 1366, "height": 900})
        )
        page = await context.new_page()

        # ── Intercepteur de réponses API ───────────────────────────────────
        print("🌐  Activation de l'intercepteur de réponses API...")

        async def handle_response(response):
            url = response.url
            if not is_api_call(url):
                return
            try:
                ct = response.headers.get("content-type", "")
                if "json" not in ct:
                    return
                data = await response.json()
                found = parse_api_response(data)
                new = 0
                for act in found:
                    act_id = act["url"]
                    if act_id not in api_seen_ids:
                        api_seen_ids.add(act_id)
                        api_activities.append(act)
                        new += 1
                if new:
                    print(f"    🌐  API [{new} nouveaux] depuis {url[:80]}")
            except Exception:
                pass

        page.on("response", handle_response)

        # ── Navigation ─────────────────────────────────────────────────────
        print(f"📄  Chargement de {URL_MAIN}")
        await page.goto(URL_MAIN, wait_until="domcontentloaded", timeout=60_000)
        await asyncio.sleep(2.5)
        await dismiss_overlays(page)

        if await is_blocked(page):
            print("⚠️  Page bloquée — recharge dans Chrome (F5) puis appuie sur Entrée ici.")
            input("  → Entrée quand débloqué : ")

        # ── Scroll pour déclencher le lazy loading (et capturer via DOM) ───
        dom_raw = await scroll_and_capture(page)

        # Attendre un peu que les derniers appels API arrivent
        await asyncio.sleep(2.0)

        await page.close()

    # ── Fusion API + DOM ───────────────────────────────────────────────────
    print(f"\n{'='*65}")
    print(f"  Résultats bruts :")
    print(f"    🌐 API   : {len(api_activities)} activités")
    print(f"    🖥️  DOM  : {len(dom_raw)} cartes capturées")

    # Priorité à l'API (données plus riches), compléter avec DOM
    all_activities: list[dict] = list(api_activities)
    seen_urls = {a["url"] for a in all_activities}

    dom_activities = dom_cards_to_activities(dom_raw)
    dom_new = 0
    for act in dom_activities:
        if act["url"] not in seen_urls:
            seen_urls.add(act["url"])
            all_activities.append(act)
            dom_new += 1

    print(f"    ➕ DOM supplémentaires : {dom_new}")
    print(f"    🎯 TOTAL FINAL         : {len(all_activities)}")
    print(f"{'='*65}\n")

    return all_activities

# ─── Export Excel ──────────────────────────────────────────────────────────────

def save_excel(activities: list[dict], filename: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fever Up Paris"

    HDR_BG = PatternFill("solid", fgColor="FF6B35")
    ODD    = PatternFill("solid", fgColor="FFF5F0")
    EVEN   = PatternFill("solid", fgColor="FFFFFF")
    LINK_C = "CC3300"
    thin   = Side(style="thin", color="FFB89A")
    brd    = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Titre
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value     = "Fever Up — Activités à Paris"
    c.font      = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    c.fill      = HDR_BG
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Sous-titre
    ws.merge_cells("A2:H2")
    c = ws["A2"]
    c.value = (
        f"Extrait le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
        f"  |  {len(activities)} activités"
    )
    c.font      = Font(name="Arial", italic=True, size=10, color="555555")
    c.fill      = PatternFill("solid", fgColor="FFF0EA")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # En-têtes
    headers = [
        ("#",               5),
        ("URL",            42),
        ("Nom Prestataire",40),
        ("Nom Général",    50),
        ("Prix",           14),
        ("Note",            8),
        ("Nb Avis",        10),
        ("Dates",          22),
    ]
    for col, (h, w) in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font      = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        c.fill      = HDR_BG
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = brd
        ws.column_dimensions[chr(64 + col)].width = w
    ws.row_dimensions[3].height = 22

    # Données
    for i, act in enumerate(activities, 1):
        row  = i + 3
        fill = ODD if i % 2 == 1 else EVEN

        def cell(col, value, hyperlink=None, link=False):
            c = ws.cell(row=row, column=col, value=value)
            c.font = Font(
                name="Arial", size=10,
                color=LINK_C if link else "000000",
                underline="single" if link else None
            )
            c.alignment = Alignment(vertical="center", wrap_text=(col in [3, 4]))
            c.fill   = fill
            c.border = brd
            if hyperlink:
                c.hyperlink = hyperlink

        cell(1, i)
        cell(2, act["url"],             hyperlink=act["url"], link=True)
        cell(3, act["nom_prestataire"])
        cell(4, act["nom_general"])
        cell(5, act["prix"])
        cell(6, act["note"])
        cell(7, act["nb_avis"])
        cell(8, act["dates"])
        ws.row_dimensions[row].height = 20

    ws.freeze_panes = "A4"
    wb.save(filename)
    print(f"💾  Excel sauvegardé : {filename}  ({len(activities)} activités)")

# ─── Main ──────────────────────────────────────────────────────────────────────

async def main():
    print("=" * 65)
    print("  Fever Up Paris — Scraper v2 (API intercept + MutationObserver)")
    print(f"  Source : {URL_MAIN}")
    print("=" * 65)

    activities = await scrape()
    if not activities:
        print("\n⚠️  Aucune activité collectée.")
        return

    save_excel(activities, OUTPUT)

    print("\n📋  Aperçu des 10 premières :")
    print(f"  {'#':>3}  {'Prestataire':<28}  {'Nom':<40}  {'Prix':<12}  Note")
    print("  " + "-"*100)
    for i, a in enumerate(activities[:10], 1):
        print(
            f"  {i:>3}  {a['nom_prestataire'][:26]:<28}  "
            f"{a['nom_general'][:38]:<40}  {a['prix']:<12}  {a['note']}"
        )
    if len(activities) > 10:
        print(f"  … et {len(activities)-10} autres dans {OUTPUT}")

if __name__ == "__main__":
    asyncio.run(main())