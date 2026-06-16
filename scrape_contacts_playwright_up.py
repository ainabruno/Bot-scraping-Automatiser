# """
# scrape_contacts_playwright.py  v2
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Corrections v2 :
#   • Plus de wait_for_selector("#search") → lit le HTML brut (anti-timeout)
#   • Un contexte de navigateur isolé PAR worker (anti-détection Google)
#   • Fallback DuckDuckGo si Google échoue
#   • URL encodée correctement (urllib.parse.quote_plus)
#   • Timeout réduit à 12s + domcontentloaded
#   • Concurrence = 2 par défaut (safe)

# UTILISATION :
#   pip install playwright openpyxl pandas
#   playwright install chromium
#   python scrape_contacts_playwright.py

# Reprend là où il s'est arrêté (checkpoint.json).
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# """

# import asyncio, json, logging, random, re
# from pathlib import Path
# from urllib.parse import quote_plus

# import pandas as pd
# from playwright.async_api import async_playwright, Browser

# # ── Config ───────────────────────────────────────────────────
# INPUT_FILE  = "airbnb_paris_contacts_sans_dup.xlsx"
# OUTPUT_FILE = "airbnb_paris_contacts_ENRICHI.xlsx"
# CHECKPOINT  = "checkpoint.json"

# HEADLESS     = True   # False = voir le navigateur (debug)
# CONCURRENCY  = 2      # workers parallèles (garder ≤ 3)
# DELAY_MIN    = 2.0    # secondes entre requêtes par worker
# DELAY_MAX    = 4.0
# PAGE_TIMEOUT = 12_000 # ms

# IGNORE_DOMAINS = re.compile(
#     r"(airbnb|google|facebook|instagram|twitter|youtube|"
#     r"tripadvisor|booking|yelp|maps|wikipedia|linkedin|"
#     r"pages-jaunes|pagesjaunes|lafourchette|thefork)\.",
#     re.I,
# )

# logging.basicConfig(
#     level=logging.INFO,
#     format="%(asctime)s %(levelname)s %(message)s",
#     datefmt="%H:%M:%S",
# )
# log = logging.getLogger(__name__)

# # ── Regex ────────────────────────────────────────────────────
# PHONE_RE = re.compile(
#     r"(?<!\d)(\+33[\s.\-]?[1-9](?:[\s.\-]?\d{2}){4}|0[1-9](?:[\s.\-]?\d{2}){4})(?!\d)"
# )
# EMAIL_RE = re.compile(r"[a-zA-Z0-9._%+\-]{2,}@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,6}")
# URL_RE   = re.compile(r"https?://(?:www\.)?([a-zA-Z0-9\-]+\.[a-z]{2,6})(?:/[^\s\"<>]*)?")

# BAD_EMAILS = ("noreply","no-reply","example","test@","support@google",
#               "privacy@","legal@","info@google","schema.org")
# FIRST_NAMES = {
#     "dawit","loik","matthieu","adrian","béatrice","robin","benoît",
#     "danièle","tiss","jacques","nathaniel","candice","celine","fatoumata",
#     "isabelle","nicolas","pierre","marie","thomas","julien","sophie",
# }

# # ── Helpers ──────────────────────────────────────────────────
# def normalize_phone(raw: str) -> str:
#     d = re.sub(r"[^\d+]", "", raw)
#     if d.startswith("+33"):
#         pass
#     elif d.startswith("33") and len(d) == 11:
#         d = "+" + d
#     elif d.startswith("0") and len(d) == 10:
#         d = "+33" + d[1:]
#     else:
#         return ""
#     return d if len(d) >= 11 else ""


# def extract(text: str) -> dict:
#     out = {"telephone": "", "email": "", "site_web": ""}
#     for m in PHONE_RE.findall(text):
#         p = normalize_phone(m)
#         if p:
#             out["telephone"] = p
#             break
#     for e in EMAIL_RE.findall(text):
#         if not any(b in e.lower() for b in BAD_EMAILS) and len(e) < 80:
#             out["email"] = e.lower()
#             break
#     for m in URL_RE.finditer(text):
#         domain = m.group(1)
#         if not IGNORE_DOMAINS.search(domain):
#             url = m.group(0).rstrip(".,);\"'")
#             parts = url.split("/")
#             out["site_web"] = "/".join(parts[:3])
#             break
#     return out


# def build_query(row: pd.Series) -> str:
#     name = str(row.get("Fournisseur", "")).strip()
#     exp  = str(row.get("Nom de l'expérience", "")).strip()
#     is_firstname = len(name.split()) <= 1 or name.lower() in FIRST_NAMES
#     if is_firstname:
#         return f'{name} Paris "{exp[:50]}" contact téléphone'
#     return f'"{name}" Paris téléphone email site officiel contact'


# # ── Checkpoint ───────────────────────────────────────────────
# def load_cp() -> dict:
#     if Path(CHECKPOINT).exists():
#         with open(CHECKPOINT, encoding="utf-8") as f:
#             return json.load(f)
#     return {}


# def save_cp(data: dict) -> None:
#     with open(CHECKPOINT, "w", encoding="utf-8") as f:
#         json.dump(data, f, ensure_ascii=False, indent=2)


# # ── Page helpers ──────────────────────────────────────────────
# async def get_text(page, url: str) -> str:
#     """
#     Charge une URL et retourne le texte brut du body.
#     N'attend AUCUN sélecteur spécifique → robuste même si Google change sa structure.
#     """
#     try:
#         await page.goto(url, timeout=PAGE_TIMEOUT, wait_until="domcontentloaded")
#         await page.wait_for_timeout(700)
#         return await page.evaluate("document.body.innerText || ''")
#     except Exception as e:
#         log.debug(f"get_text [{url[:55]}]: {e}")
#         return ""


# async def accept_cookies(page):
#     for label in ["Tout accepter", "Accept all", "Accepter tout", "J'accepte", "Agree"]:
#         try:
#             btn = page.locator(f'button:has-text("{label}")')
#             if await btn.count():
#                 await btn.first.click(timeout=2000)
#                 await page.wait_for_timeout(500)
#                 return
#         except Exception:
#             pass


# async def google(page, query: str) -> str:
#     url  = f"https://www.google.com/search?q={quote_plus(query)}&hl=fr&gl=fr&num=6"
#     text = await get_text(page, url)
#     # Gestion page consent Google
#     if not text or "Avant de continuer" in text or "consent.google" in page.url:
#         await accept_cookies(page)
#         await page.wait_for_timeout(600)
#         text = await page.evaluate("document.body.innerText || ''")
#     return text


# async def duckduckgo(page, query: str) -> str:
#     url = f"https://html.duckduckgo.com/html/?q={quote_plus(query)}&kl=fr-fr"
#     return await get_text(page, url)


# async def visit_site(page, site_url: str) -> dict:
#     result = {"telephone": "", "email": ""}
#     if not site_url:
#         return result
#     for path in ["", "/contact", "/nous-contacter", "/contact-us"]:
#         text = await get_text(page, site_url.rstrip("/") + path)
#         ex   = extract(text)
#         result["telephone"] = result["telephone"] or ex["telephone"]
#         result["email"]     = result["email"]     or ex["email"]
#         if result["telephone"] and result["email"]:
#             break
#         await asyncio.sleep(0.4)
#     return result


# # ── Worker ────────────────────────────────────────────────────
# async def worker(wid: int, queue: asyncio.Queue, browser: Browser,
#                  df: pd.DataFrame, checkpoint: dict, results: dict):
#     """Chaque worker a son propre contexte → Google ne les lie pas."""
#     ctx = await browser.new_context(
#         user_agent=(
#             "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
#             "AppleWebKit/537.36 (KHTML, like Gecko) "
#             f"Chrome/12{random.randint(0,4)}.0.0.0 Safari/537.36"
#         ),
#         locale="fr-FR",
#         timezone_id="Europe/Paris",
#         viewport={"width": 1280, "height": 900},
#     )
#     await ctx.add_init_script(
#         "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
#     )
#     page = await ctx.new_page()

#     try:
#         while True:
#             try:
#                 idx = queue.get_nowait()
#             except asyncio.QueueEmpty:
#                 break

#             key = str(idx)
#             row = df.iloc[idx]
#             name = str(row["Fournisseur"])

#             # Déjà dans le checkpoint → skip
#             if key in checkpoint:
#                 log.info(f"[W{wid}] #{idx+1} SKIP — {name}")
#                 results[idx] = checkpoint[key]
#                 queue.task_done()
#                 continue

#             r = {"index": idx, "telephone": "", "email": "", "site_web": ""}
#             query = build_query(row)
#             log.info(f"[W{wid}] #{idx+1} {name}")

#             # 1) Google
#             text = await google(page, query)
#             ex   = extract(text)
#             for k, v in ex.items():
#                 if v: r[k] = v

#             # 2) Fallback DuckDuckGo si rien trouvé
#             if not r["telephone"] and not r["site_web"]:
#                 text2 = await duckduckgo(page, query)
#                 ex2   = extract(text2)
#                 for k, v in ex2.items():
#                     if v and not r[k]: r[k] = v

#             # 3) Visiter le site pour compléter tel/email
#             if r["site_web"] and (not r["telephone"] or not r["email"]):
#                 details = await visit_site(page, r["site_web"])
#                 r["telephone"] = r["telephone"] or details["telephone"]
#                 r["email"]     = r["email"]     or details["email"]

#             log.info(
#                 f"[W{wid}] #{idx+1} ✅ {name}  "
#                 f"☎ {r['telephone'] or '—'}  "
#                 f"✉ {r['email'] or '—'}  "
#                 f"🌐 {r['site_web'] or '—'}"
#             )

#             results[idx]    = r
#             checkpoint[key] = r
#             save_cp(checkpoint)
#             queue.task_done()

#             await asyncio.sleep(random.uniform(DELAY_MIN, DELAY_MAX))

#     finally:
#         await page.close()
#         await ctx.close()


# # ── Export ────────────────────────────────────────────────────
# def export_excel(df: pd.DataFrame, results: dict):
#     from openpyxl import load_workbook
#     from openpyxl.styles import Font, PatternFill, Alignment
#     from openpyxl.utils import get_column_letter

#     for idx, r in results.items():
#         if r.get("telephone"): df.at[idx, "Telephone"] = r["telephone"]
#         if r.get("email"):     df.at[idx, "email"]     = r["email"]
#         if r.get("site_web"):  df.at[idx, "Site web"]  = r["site_web"]

#     df.to_excel(OUTPUT_FILE, index=False, engine="openpyxl")

#     wb = load_workbook(OUTPUT_FILE)
#     ws = wb.active

#     hf = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
#     for cell in ws[1]:
#         cell.fill = hf
#         cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
#         cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
#     ws.row_dimensions[1].height = 28

#     col_map = {c.value: c.column for c in ws[1]}
#     green  = PatternFill("solid", start_color="D9EAD3", end_color="D9EAD3")
#     yellow = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")

#     for col_name in ["Telephone", "email", "Site web"]:
#         c = col_map.get(col_name)
#         if not c: continue
#         for row in range(2, ws.max_row + 1):
#             cell = ws.cell(row=row, column=c)
#             cell.font = Font(name="Arial", size=9)
#             cell.fill = green if cell.value else yellow

#     for col, w in {1:5,2:42,3:28,4:35,5:18,6:50,7:44,8:10,9:18,10:30,11:35}.items():
#         ws.column_dimensions[get_column_letter(col)].width = w

#     ws.freeze_panes = "A2"
#     ws.auto_filter.ref = ws.dimensions
#     wb.save(OUTPUT_FILE)

#     df2   = pd.read_excel(OUTPUT_FILE)
#     total = len(df2)
#     tel   = df2["Telephone"].notna().sum()
#     mail  = df2["email"].notna().sum()
#     site  = df2["Site web"].notna().sum()
#     print("\n" + "━"*55)
#     print(f"  ✅  {OUTPUT_FILE}")
#     print(f"  📞  Téléphone : {tel}/{total}  ({tel/total*100:.1f}%)")
#     print(f"  ✉️   Email     : {mail}/{total}  ({mail/total*100:.1f}%)")
#     print(f"  🌐  Site web  : {site}/{total}  ({site/total*100:.1f}%)")
#     print("━"*55)


# # ── Main ──────────────────────────────────────────────────────
# async def main():
#     df         = pd.read_excel(INPUT_FILE)
#     checkpoint = load_cp()
#     results: dict = {}

#     todo = [i for i in range(len(df)) if str(i) not in checkpoint]
#     log.info(f"📋 Total:{len(df)}  Checkpoint:{len(checkpoint)}  À faire:{len(todo)}")

#     for key, val in checkpoint.items():
#         results[int(key)] = val

#     queue: asyncio.Queue = asyncio.Queue()
#     for i in todo:
#         await queue.put(i)

#     async with async_playwright() as pw:
#         browser = await pw.chromium.launch(
#             headless=HEADLESS,
#             args=[
#                 "--no-sandbox",
#                 "--disable-blink-features=AutomationControlled",
#                 "--disable-dev-shm-usage",
#                 "--disable-gpu",
#             ],
#         )
#         await asyncio.gather(*[
#             worker(wid, queue, browser, df, checkpoint, results)
#             for wid in range(1, CONCURRENCY + 1)
#         ])
#         await browser.close()

#     export_excel(df, results)


# if __name__ == "__main__":
#     asyncio.run(main())
"""
scrape_contacts_playwright.py  v3
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Stratégie multi-sources par type de fournisseur :

  ÉTAPE 1 — Scrape page Airbnb (URL déjà dans le fichier)
            → nom complet hôte, site web, Instagram, description
  ÉTAPE 2 — Google avec le NOM COMPLET trouvé sur Airbnb
            (bien meilleur que juste "Gladys Paris")
  ÉTAPE 3 — Pages Jaunes (pour entreprises)
  ÉTAPE 4 — DuckDuckGo fallback
  ÉTAPE 5 — Visite du site officiel trouvé

UTILISATION :
  pip install playwright openpyxl pandas
  playwright install chromium
  python scrape_contacts_playwright.py
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
"""

import asyncio, json, logging, random, re
from pathlib import Path
from urllib.parse import quote_plus

import pandas as pd
from playwright.async_api import async_playwright, Browser

# ── Config ───────────────────────────────────────────────────
INPUT_FILE  = "airbnb_paris_contacts_sans_dup1.xlsx"
OUTPUT_FILE = "airbnb_paris_contacts_ENRICHI1.xlsx"
CHECKPOINT  = "checkpoint1.json"

HEADLESS     = True
CONCURRENCY  = 2      # workers (2 = safe, 3 = plus rapide mais risque CAPTCHA)
DELAY_MIN    = 1.8
DELAY_MAX    = 3.5
PAGE_TIMEOUT = 14_000

IGNORE_DOMAINS = re.compile(
    r"(airbnb|google|facebook|instagram|twitter|youtube|"
    r"tripadvisor|booking|yelp|maps\.google|wikipedia|linkedin|"
    r"pages-jaunes|pagesjaunes|lafourchette|thefork|"
    r"viator|getyourguide|expedia|klook)\.",
    re.I,
)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# ── Regex ────────────────────────────────────────────────────
PHONE_RE = re.compile(
    r"(?<!\d)(\+33[\s.\-]?[1-9](?:[\s.\-]?\d{2}){4}"
    r"|0[1-9](?:[\s.\-]?\d{2}){4})(?!\d)"
)
EMAIL_RE = re.compile(r"[a-zA-Z0-9._%+\-]{2,}@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,6}")
URL_RE   = re.compile(r"https?://(?:www\.)?([a-zA-Z0-9\-]+\.[a-z]{2,6})(?:/[^\s\"<>\"]*)?")
INSTA_RE = re.compile(r"instagram\.com/([a-zA-Z0-9_.]+)")
BAD_EMAILS = ("noreply","no-reply","example","test@","support@google",
              "privacy@","legal@","schema.org","sentry","wixpress")

# ── Helpers ──────────────────────────────────────────────────
def normalize_phone(raw: str) -> str:
    d = re.sub(r"[^\d+]", "", raw)
    if d.startswith("+33"):
        pass
    elif d.startswith("33") and len(d) == 11:
        d = "+" + d
    elif d.startswith("0") and len(d) == 10:
        d = "+33" + d[1:]
    else:
        return ""
    return d if len(d) >= 11 else ""


def extract(text: str) -> dict:
    out = {"telephone": "", "email": "", "site_web": "", "instagram": ""}
    for m in PHONE_RE.findall(text):
        p = normalize_phone(m)
        if p:
            out["telephone"] = p
            break
    for e in EMAIL_RE.findall(text):
        if not any(b in e.lower() for b in BAD_EMAILS) and len(e) < 80:
            out["email"] = e.lower()
            break
    for m in URL_RE.finditer(text):
        domain = m.group(1)
        if not IGNORE_DOMAINS.search(domain):
            url = m.group(0).rstrip(".,);\"'")
            out["site_web"] = "/".join(url.split("/")[:3])
            break
    ig = INSTA_RE.search(text)
    if ig and ig.group(1) not in ("p", "explore", "reel", "stories", "accounts"):
        out["instagram"] = f"https://instagram.com/{ig.group(1)}"
    return out


def is_company(name: str) -> bool:
    keywords = ["SAS","SARL","SRL","LLC","LTD","TOURS","CLUB","EXPERIENCE",
                "EXPERIENCES","PARIS","FOOD","RIVER","BOAT","COMPANY",
                "GROUP","AGENCY","STUDIO","SCHOOL","ATELIER","ASSOCIATION"]
    n = name.upper()
    return any(k in n for k in keywords)


def build_queries(row: pd.Series, airbnb_data: dict) -> list[str]:
    """
    Construit les requêtes selon le contexte.
    Utilise le nom complet trouvé sur Airbnb si disponible.
    """
    name     = str(row.get("Fournisseur", "")).strip()
    exp      = str(row.get("Nom de l'expérience", "")).strip()
    full_name = airbnb_data.get("host_name", "").strip()
    site      = airbnb_data.get("site_web", "")

    queries = []

    # Si on a déjà un site Airbnb → chercher avec ce nom complet
    if full_name and full_name.lower() != name.lower():
        queries.append(f'"{full_name}" Paris contact téléphone email')
        queries.append(f'"{full_name}" site officiel Paris')

    # Entreprise
    if is_company(name):
        queries.append(f'"{name}" Paris téléphone email site officiel')
        queries.append(f'"{name}" pages jaunes Paris contact')
    # Prénom seul ou multi-personnes → utiliser expérience
    elif len(name.split()) <= 2 or " ou " in name.lower():
        queries.append(f'"{name}" "{exp[:45]}" Paris contact')
        queries.append(f'{name} Paris "{exp[:35]}" site instagram')
    else:
        queries.append(f'"{name}" Paris téléphone email site officiel contact')

    return queries[:3]  # max 3 requêtes par fournisseur


# ── Checkpoint ───────────────────────────────────────────────
def load_cp() -> dict:
    if Path(CHECKPOINT).exists():
        with open(CHECKPOINT, encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_cp(data: dict) -> None:
    with open(CHECKPOINT, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


# ── Page navigation ───────────────────────────────────────────
async def get_text(page, url: str, wait_ms: int = 800) -> str:
    try:
        await page.goto(url, timeout=PAGE_TIMEOUT, wait_until="domcontentloaded")
        await page.wait_for_timeout(wait_ms)
        return await page.evaluate("document.body.innerText || ''")
    except Exception as e:
        log.debug(f"get_text error [{url[:55]}]: {e}")
        return ""


async def get_html(page, url: str) -> str:
    try:
        await page.goto(url, timeout=PAGE_TIMEOUT, wait_until="domcontentloaded")
        await page.wait_for_timeout(1000)
        return await page.content()
    except Exception as e:
        log.debug(f"get_html error [{url[:55]}]: {e}")
        return ""


async def accept_cookies(page):
    for label in ["Tout accepter","Accept all","Accepter tout","J'accepte","Agree","Fermer"]:
        try:
            btn = page.locator(f'button:has-text("{label}")')
            if await btn.count():
                await btn.first.click(timeout=2000)
                await page.wait_for_timeout(400)
                return
        except Exception:
            pass


# ── Source 1 : Page Airbnb ────────────────────────────────────
async def scrape_airbnb(page, airbnb_url: str) -> dict:
    """
    Scrape la page d'expérience Airbnb pour extraire :
    - Nom complet de l'hôte
    - Site web (si mentionné dans description)
    - Instagram (si mentionné)
    - Email (rare mais parfois dans description)
    """
    result = {"host_name": "", "site_web": "", "instagram": "", "email": "", "telephone": ""}
    if not airbnb_url or "airbnb" not in airbnb_url:
        return result

    html = await get_html(page, airbnb_url)
    if not html:
        return result

    # Nom de l'hôte : chercher dans les balises JSON-LD ou meta
    host_patterns = [
        r'"hostName"\s*:\s*"([^"]+)"',
        r'"name"\s*:\s*"([^"]{3,50})".*?"@type"\s*:\s*"Person"',
        r'Hébergé par\s+([A-ZÀÂÄÉÈÊËÎÏÔÖÙÛÜ][a-zàâäéèêëîïôöùûü\s\-]{2,30})',
        r'Hosted by\s+([A-Z][a-zA-Z\s\-]{2,30})',
        r'"givenName"\s*:\s*"([^"]+)"',
    ]
    for pat in host_patterns:
        m = re.search(pat, html)
        if m:
            candidate = m.group(1).strip()
            if 2 < len(candidate) < 50 and candidate.lower() not in ("airbnb","paris"):
                result["host_name"] = candidate
                break

    # Extraire tout le texte pour chercher contacts
    text = re.sub(r'<[^>]+>', ' ', html)
    text = re.sub(r'\s+', ' ', text)
    ex = extract(text)
    result["site_web"]   = result["site_web"]   or ex["site_web"]
    result["instagram"]  = result["instagram"]  or ex["instagram"]
    result["email"]      = result["email"]      or ex["email"]
    result["telephone"]  = result["telephone"]  or ex["telephone"]

    log.debug(f"  Airbnb → host='{result['host_name']}' site='{result['site_web']}' ig='{result['instagram']}'")
    return result


# ── Source 2 : Google ─────────────────────────────────────────
async def google(page, query: str) -> str:
    url  = f"https://www.google.com/search?q={quote_plus(query)}&hl=fr&gl=fr&num=6"
    text = await get_text(page, url, wait_ms=700)
    if not text or "Avant de continuer" in text or "consent.google" in page.url:
        await accept_cookies(page)
        await page.wait_for_timeout(500)
        text = await page.evaluate("document.body.innerText || ''")
    return text


# ── Source 3 : DuckDuckGo ─────────────────────────────────────
async def duckduckgo(page, query: str) -> str:
    url = f"https://html.duckduckgo.com/html/?q={quote_plus(query)}&kl=fr-fr"
    return await get_text(page, url, wait_ms=600)


# ── Source 4 : Pages Jaunes ───────────────────────────────────
async def pages_jaunes(page, name: str, ville: str = "Paris") -> dict:
    result = {"telephone": "", "site_web": "", "email": ""}
    try:
        url = f"https://www.pagesjaunes.fr/recherche/{quote_plus(name)}/{quote_plus(ville)}"
        text = await get_text(page, url, wait_ms=1000)
        ex   = extract(text)
        result["telephone"] = ex["telephone"]
        result["site_web"]  = ex["site_web"]
        result["email"]     = ex["email"]
    except Exception:
        pass
    return result


# ── Source 5 : Visit site ─────────────────────────────────────
async def visit_site(page, site_url: str) -> dict:
    result = {"telephone": "", "email": ""}
    if not site_url:
        return result
    for path in ["", "/contact", "/nous-contacter", "/contact-us", "/about"]:
        text = await get_text(page, site_url.rstrip("/") + path, wait_ms=600)
        ex   = extract(text)
        result["telephone"] = result["telephone"] or ex["telephone"]
        result["email"]     = result["email"]     or ex["email"]
        if result["telephone"] and result["email"]:
            break
        await asyncio.sleep(0.3)
    return result


# ── Source 6 : Instagram (extraire email/site du bio) ─────────
async def scrape_instagram(page, ig_url: str) -> dict:
    result = {"telephone": "", "email": "", "site_web": ""}
    if not ig_url:
        return result
    try:
        text = await get_text(page, ig_url, wait_ms=1500)
        ex   = extract(text)
        result.update({k: v for k, v in ex.items() if v})
    except Exception:
        pass
    return result


# ── Worker ────────────────────────────────────────────────────
async def worker(wid: int, queue: asyncio.Queue, browser: Browser,
                 df: pd.DataFrame, checkpoint: dict, results: dict):
    ctx = await browser.new_context(
        user_agent=(
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            f"Chrome/12{random.randint(0,4)}.0.0.0 Safari/537.36"
        ),
        locale="fr-FR",
        timezone_id="Europe/Paris",
        viewport={"width": 1280, "height": 900},
    )
    await ctx.add_init_script(
        "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
    )
    page = await ctx.new_page()

    try:
        while True:
            try:
                idx = queue.get_nowait()
            except asyncio.QueueEmpty:
                break

            key = str(idx)
            row = df.iloc[idx]
            name      = str(row["Fournisseur"]).strip()
            airbnb_url = str(row.get("URL", "")).strip()

            if key in checkpoint:
                log.info(f"[W{wid}] #{idx+1} SKIP — {name}")
                results[idx] = checkpoint[key]
                queue.task_done()
                continue

            r = {"index": idx, "telephone": "", "email": "", "site_web": "", "instagram": "", "host_name": ""}
            log.info(f"[W{wid}] #{idx+1} {name}")

            # ── ÉTAPE 1 : Scraper la page Airbnb ──────────────
            airbnb_data = await scrape_airbnb(page, airbnb_url)
            r["host_name"] = airbnb_data.get("host_name", "")
            for k in ("telephone", "email", "site_web", "instagram"):
                if airbnb_data.get(k): r[k] = airbnb_data[k]

            # ── ÉTAPE 2 : Google (avec nom complet si dispo) ──
            if not (r["telephone"] and r["email"]):
                queries = build_queries(row, airbnb_data)
                for q in queries:
                    text = await google(page, q)
                    ex   = extract(text)
                    for k, v in ex.items():
                        if v and not r.get(k): r[k] = v
                    if r["telephone"] and r["email"] and r["site_web"]:
                        break
                    await asyncio.sleep(random.uniform(0.8, 1.5))

            # ── ÉTAPE 3 : DuckDuckGo fallback ─────────────────
            if not r["telephone"] and not r["site_web"]:
                q2   = f'"{r["host_name"] or name}" Paris contact téléphone'
                text = await duckduckgo(page, q2)
                ex   = extract(text)
                for k, v in ex.items():
                    if v and not r.get(k): r[k] = v

            # ── ÉTAPE 4 : Pages Jaunes (entreprises) ──────────
            if is_company(name) and not r["telephone"]:
                pj = await pages_jaunes(page, name)
                for k, v in pj.items():
                    if v and not r.get(k): r[k] = v

            # ── ÉTAPE 5 : Visiter le site officiel ────────────
            if r["site_web"] and (not r["telephone"] or not r["email"]):
                details = await visit_site(page, r["site_web"])
                r["telephone"] = r["telephone"] or details["telephone"]
                r["email"]     = r["email"]     or details["email"]

            # ── ÉTAPE 6 : Instagram → extraire email/site ─────
            if r["instagram"] and (not r["email"] or not r["site_web"]):
                ig_data = await scrape_instagram(page, r["instagram"])
                r["site_web"] = r["site_web"] or ig_data.get("site_web","")
                r["email"]    = r["email"]    or ig_data.get("email","")

            log.info(
                f"[W{wid}] #{idx+1} ✅ {name}"
                + (f" ({r['host_name']})" if r['host_name'] and r['host_name'].lower() != name.lower() else "")
                + f"  ☎ {r['telephone'] or '—'}"
                + f"  ✉ {r['email'] or '—'}"
                + f"  🌐 {r['site_web'] or '—'}"
                + (f"  📸 {r['instagram']}" if r['instagram'] else "")
            )

            results[idx]    = r
            checkpoint[key] = r
            save_cp(checkpoint)
            queue.task_done()

            await asyncio.sleep(random.uniform(DELAY_MIN, DELAY_MAX))

    finally:
        await page.close()
        await ctx.close()


# ── Export Excel ──────────────────────────────────────────────
def export_excel(df: pd.DataFrame, results: dict):
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    # Ajouter colonne Instagram et Nom hôte si pas déjà là
    if "Instagram" not in df.columns:
        df.insert(df.columns.get_loc("Site web") + 1, "Instagram", "")
    if "Nom hôte Airbnb" not in df.columns:
        df.insert(df.columns.get_loc("Fournisseur") + 1, "Nom hôte Airbnb", "")

    for idx, r in results.items():
        if r.get("telephone"):   df.at[idx, "Telephone"]       = r["telephone"]
        if r.get("email"):       df.at[idx, "email"]           = r["email"]
        if r.get("site_web"):    df.at[idx, "Site web"]        = r["site_web"]
        if r.get("instagram"):   df.at[idx, "Instagram"]       = r["instagram"]
        if r.get("host_name"):   df.at[idx, "Nom hôte Airbnb"] = r["host_name"]

    df.to_excel(OUTPUT_FILE, index=False, engine="openpyxl")

    wb = load_workbook(OUTPUT_FILE)
    ws = wb.active

    hf   = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
    hfnt = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    for cell in ws[1]:
        cell.fill = hf
        cell.font = hfnt
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30

    col_map = {c.value: c.column for c in ws[1]}
    green  = PatternFill("solid", start_color="D9EAD3", end_color="D9EAD3")
    yellow = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")
    blue   = PatternFill("solid", start_color="D0E4F7", end_color="D0E4F7")

    contact_cols  = ["Telephone", "email", "Site web", "Instagram"]
    enriched_cols = ["Nom hôte Airbnb"]

    for col_name in contact_cols:
        c = col_map.get(col_name)
        if not c: continue
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = Font(name="Arial", size=9)
            cell.fill = green if cell.value else yellow

    for col_name in enriched_cols:
        c = col_map.get(col_name)
        if not c: continue
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = Font(name="Arial", size=9, italic=True)
            if cell.value:
                cell.fill = blue

    # Largeurs
    widths = {"#":5, "Nom de l'expérience":40, "Fournisseur":26, "Nom hôte Airbnb":22,
              "Adresse (rue)":30, "Ville / CP":15, "Adresse complète":40,
              "URL":40, "Statut":9, "Telephone":18, "email":28, "Site web":32, "Instagram":30}
    for i, col in enumerate(df.columns, 1):
        w = widths.get(col, 18)
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(OUTPUT_FILE)

    df2   = pd.read_excel(OUTPUT_FILE)
    total = len(df2)
    tel   = df2["Telephone"].notna().sum()
    mail  = df2["email"].notna().sum()
    site  = df2["Site web"].notna().sum()
    insta = df2["Instagram"].notna().sum() if "Instagram" in df2.columns else 0
    host  = df2["Nom hôte Airbnb"].notna().sum() if "Nom hôte Airbnb" in df2.columns else 0

    print("\n" + "━"*58)
    print(f"  ✅  {OUTPUT_FILE}")
    print(f"  📞  Téléphone      : {tel}/{total}  ({tel/total*100:.1f}%)")
    print(f"  ✉️   Email          : {mail}/{total}  ({mail/total*100:.1f}%)")
    print(f"  🌐  Site web       : {site}/{total}  ({site/total*100:.1f}%)")
    print(f"  📸  Instagram      : {insta}/{total}  ({insta/total*100:.1f}%)")
    print(f"  👤  Nom hôte trouvé: {host}/{total}  ({host/total*100:.1f}%)")
    print("━"*58)


# ── Main ──────────────────────────────────────────────────────
async def main():
    df         = pd.read_excel(INPUT_FILE)
    checkpoint = load_cp()
    results: dict = {}

    todo = [i for i in range(len(df)) if str(i) not in checkpoint]
    log.info(f"📋 Total:{len(df)}  Checkpoint:{len(checkpoint)}  À faire:{len(todo)}")

    for key, val in checkpoint.items():
        results[int(key)] = val

    queue: asyncio.Queue = asyncio.Queue()
    for i in todo:
        await queue.put(i)

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(
            headless=HEADLESS,
            args=[
                "--no-sandbox",
                "--disable-blink-features=AutomationControlled",
                "--disable-dev-shm-usage",
                "--disable-gpu",
            ],
        )
        await asyncio.gather(*[
            worker(wid, queue, browser, df, checkpoint, results)
            for wid in range(1, CONCURRENCY + 1)
        ])
        await browser.close()

    export_excel(df, results)


if __name__ == "__main__":
    asyncio.run(main())