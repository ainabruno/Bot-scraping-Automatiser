"""
hotels_paris_scraper.py
=======================
Scraper Google Maps pour récupérer TOUS les hôtels de Paris
Objectif : ~1629 hôtels avec nom, adresse, téléphone, email, site web
Identifie automatiquement les données incomplètes à compléter manuellement
"""

import time
import re
import requests
import pandas as pd
from urllib.parse import quote
from datetime import datetime

# ── CONFIGURATION ──────────────────────────────────────────────────────────────
# Importer la configuration depuis config.py
try:
    from config import *
    print("✅ Configuration chargée depuis config.py")
except ImportError:
    print("⚠️  Fichier config.py non trouvé, utilisation de la configuration par défaut")
    SERP_API_KEY = "b4b185517e867b6f276ee8d9ff7b8986a3e90fda5854d9f614200ee5d8084823"
    PAUSE_ENTRE_REQUETES = 2.0
    MAX_PAGES_PAR_REQUETE = 5
    REQUEST_TIMEOUT = 30
    SCRAPER_PAR_ARRONDISSEMENT = True
    AJOUTER_REQUETES_SUPPLEMENTAIRES = True
    RECUPERER_EMAIL_DEPUIS_SITE = False
    TIMEOUT_SCRAPING_EMAIL = 10
    
    ARRONDISSEMENTS_PARIS = [
        "Paris 1er", "Paris 2e", "Paris 3e", "Paris 4e", "Paris 5e",
        "Paris 6e", "Paris 7e", "Paris 8e", "Paris 9e", "Paris 10e",
        "Paris 11e", "Paris 12e", "Paris 13e", "Paris 14e", "Paris 15e",
        "Paris 16e", "Paris 17e", "Paris 18e", "Paris 19e", "Paris 20e",
    ]
    
    REQUETES_SUPPLEMENTAIRES = [
        "hotel Paris",
        "hôtel Paris centre",
        "boutique hotel Paris",
        "hotel luxe Paris",
        "hotel economique Paris",
        "auberge Paris",
        "résidence hoteliere Paris",
    ]

# ── FONCTIONS UTILITAIRES ──────────────────────────────────────────────────────

def extraire_email_depuis_texte(texte: str) -> str:
    """Tente d'extraire un email depuis du texte."""
    if not texte:
        return ""
    pattern = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
    matches = re.findall(pattern, texte)
    return matches[0] if matches else ""


def scraper_site_web_pour_email(url: str, timeout: int = 10) -> str:
    """
    Tente de récupérer un email depuis le site web de l'hôtel.
    ATTENTION : peut être lent, à utiliser avec parcimonie.
    """
    if not url:
        return ""
    
    try:
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        }
        response = requests.get(url, headers=headers, timeout=timeout)
        if response.status_code == 200:
            return extraire_email_depuis_texte(response.text)
    except:
        pass
    return ""


def normaliser_telephone(tel: str) -> str:
    """Normalise le format du téléphone."""
    if not tel:
        return ""
    # Garde uniquement les chiffres et le +
    tel = re.sub(r'[^\d+]', '', tel)
    return tel


def scrape_hotels_requete(query: str, start: int = 0) -> tuple[list[dict], bool]:
    """
    Interroge SerpAPI Google Maps pour une requête.
    Retourne (liste de résultats, a_encore_des_resultats)
    """
    params = {
        "engine":   "google_maps",
        "q":        query,
        "api_key":  SERP_API_KEY,
        "hl":       "fr",
        "gl":       "fr",
        "type":     "search",
        "start":    start,  # Pour la pagination
    }

    try:
        response = requests.get(
            "https://serpapi.com/search",
            params=params,
            timeout=30
        )
        response.raise_for_status()
        data = response.json()

        if "error" in data:
            print(f"      ⚠ Erreur API SerpAPI : {data['error']}")
            return [], False

        resultats = []
        for place in data.get("local_results", []):
            gps   = place.get("gps_coordinates", {})
            lat   = gps.get("latitude")
            lon   = gps.get("longitude")
            pid   = place.get("place_id", "")
            tel   = place.get("phone", "")
            site  = place.get("website", "")
            
            # Informations de base
            hotel = {
                "nom":              place.get("title", ""),
                "adresse":          place.get("address", ""),
                "telephone":        normaliser_telephone(tel),
                "site_web":         site,
                "email":            "",  # À récupérer si possible
                "note":             place.get("rating", ""),
                "nb_avis":          place.get("reviews", 0),
                "categorie":        place.get("type", ""),
                "latitude":         lat,
                "longitude":        lon,
                "place_id":         pid,
                "lien_google_maps": f"https://www.google.com/maps/place/?q=place_id:{pid}" if pid else "",
                "requete_source":   query,
            }
            
            resultats.append(hotel)

        # Vérifier s'il y a encore des résultats (pour pagination)
        serpapi_pagination = data.get("serpapi_pagination", {})
        a_encore = "next" in serpapi_pagination or "next_page_token" in serpapi_pagination
        
        return resultats, a_encore

    except requests.exceptions.RequestException as e:
        print(f"      ⚠ Erreur réseau : {e}")
        return [], False


def scraper_hotels_paris_complet(
    inclure_arrondissements: bool = True,
    inclure_requetes_supp: bool = True,
    max_pages_par_requete: int = 5,
    tenter_email_depuis_site: bool = False
) -> pd.DataFrame:
    """
    Lance un scraping complet de tous les hôtels de Paris.
    
    Args:
        inclure_arrondissements: Scraper par arrondissement
        inclure_requetes_supp: Ajouter les requêtes supplémentaires
        max_pages_par_requete: Nombre max de pages à parcourir par requête
        tenter_email_depuis_site: Tenter de récupérer l'email depuis le site web (LENT)
    """
    
    print("=" * 70)
    print("SCRAPING HOTELS PARIS - RECUPERATION COMPLETE")
    print(f"Objectif : ~1629 hôtels avec contact complet")
    print("=" * 70)
    
    tous_resultats = []
    vus = set()  # Pour déduplication (nom + adresse)
    
    # Construire la liste des requêtes
    requetes = []
    
    if inclure_arrondissements:
        for arrond in ARRONDISSEMENTS_PARIS:
            requetes.append(f"hotel {arrond}")
    
    if inclure_requetes_supp:
        requetes.extend(REQUETES_SUPPLEMENTAIRES)
    
    print(f"\n📊 {len(requetes)} requêtes à effectuer\n")
    
    # Scraper chaque requête
    for idx, requete in enumerate(requetes, 1):
        print(f"[{idx:02d}/{len(requetes)}] '{requete}'")
        
        page = 0
        while page < max_pages_par_requete:
            start = page * 20  # SerpAPI retourne ~20 résultats par page
            
            if page > 0:
                print(f"      → Page {page + 1}", end=" ... ")
            else:
                print(f"      → ", end="")
            
            resultats, a_encore = scrape_hotels_requete(requete, start=start)
            
            nouveaux = 0
            for hotel in resultats:
                # Clé de déduplication : nom + début adresse
                nom_norm = hotel["nom"].lower().strip()
                addr_norm = hotel["adresse"][:30].lower().strip() if hotel["adresse"] else ""
                cle = (nom_norm, addr_norm)
                
                if cle not in vus and hotel["nom"]:
                    vus.add(cle)
                    
                    # Tenter de récupérer l'email depuis le site web si demandé
                    if tenter_email_depuis_site and hotel["site_web"] and not hotel["email"]:
                        email = scraper_site_web_pour_email(hotel["site_web"])
                        hotel["email"] = email
                    
                    tous_resultats.append(hotel)
                    nouveaux += 1
            
            print(f"{len(resultats)} trouvés, {nouveaux} nouveaux → Total: {len(tous_resultats)}")
            
            # Si plus de résultats ou limite de pages atteinte
            if not a_encore or len(resultats) == 0:
                break
            
            page += 1
            time.sleep(PAUSE_ENTRE_REQUETES)
        
        # Pause entre requêtes
        if idx < len(requetes):
            time.sleep(PAUSE_ENTRE_REQUETES)
    
    print(f"\n✅ Scraping terminé : {len(tous_resultats)} hôtels uniques récupérés")
    
    return pd.DataFrame(tous_resultats) if tous_resultats else pd.DataFrame()


def analyser_completude(df: pd.DataFrame) -> pd.DataFrame:
    """
    Analyse la complétude des données et ajoute les colonnes de statut.
    """
    if df.empty:
        return df
    
    df = df.copy()
    
    # Vérifier la complétude pour chaque ligne
    statuts = []
    champs_manquants = []
    priorites = []
    
    for _, row in df.iterrows():
        manquants = []
        
        if not row.get("telephone"):
            manquants.append("téléphone")
        if not row.get("email"):
            manquants.append("email")
        if not row.get("site_web"):
            manquants.append("site_web")
        if not row.get("adresse"):
            manquants.append("adresse")
        
        # Statut
        if len(manquants) == 0:
            statut = "✅ COMPLET"
            priorite = "A"
        elif "email" in manquants and len(manquants) == 1:
            statut = "⚠️ EMAIL MANQUANT"
            priorite = "B"
        elif "telephone" in manquants or "adresse" in manquants:
            statut = "❌ INCOMPLET - PRIORITAIRE"
            priorite = "C"
        else:
            statut = "⚠️ À COMPLÉTER"
            priorite = "B"
        
        statuts.append(statut)
        champs_manquants.append(", ".join(manquants) if manquants else "Aucun")
        priorites.append(priorite)
    
    df["statut_completude"] = statuts
    df["champs_manquants"] = champs_manquants
    df["priorite_completion"] = priorites
    df["email_manuel"] = ""  # Colonne vide pour saisie manuelle
    df["telephone_manuel"] = ""  # Colonne vide pour saisie manuelle
    df["notes"] = ""  # Colonne pour commentaires
    
    # Trier par priorité de complétion
    ordre = {"C": 0, "B": 1, "A": 2}
    df["_ordre"] = df["priorite_completion"].map(ordre)
    df = df.sort_values(["_ordre", "nom"], ascending=[True, True])
    df = df.drop(columns=["_ordre"])
    
    return df.reset_index(drop=True)


def export_excel_hotels(df: pd.DataFrame, fichier: str):
    """
    Exporte en Excel avec mise en forme et séparation par statut de complétude.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    # Séparer par statut
    df_incomplet = df[df["priorite_completion"] == "C"].copy()
    df_a_completer = df[df["priorite_completion"] == "B"].copy()
    df_complet = df[df["priorite_completion"] == "A"].copy()
    
    # Colonnes à afficher
    cols_principales = [
        "priorite_completion", "statut_completude", "champs_manquants",
        "nom", "adresse", "telephone", "telephone_manuel",
        "email", "email_manuel", "site_web",
        "note", "nb_avis", "lien_google_maps", "notes"
    ]
    
    # Styles
    fills = {
        "header":    PatternFill("solid", start_color="1F4E78", end_color="1F4E78"),
        "incomplet": PatternFill("solid", start_color="FCE4D6", end_color="FCE4D6"),
        "a_completer": PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC"),
        "complet":   PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA"),
    }
    
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    bold_font = Font(name="Arial", bold=True, size=10)
    normal_font = Font(name="Arial", size=10)
    link_font = Font(name="Arial", size=10, color="0563C1", underline="single")
    thin_border = Side(style="thin", color="D0D0D0")
    border = Border(bottom=thin_border, right=thin_border)
    
    wb = Workbook()
    wb.remove(wb.active)
    
    def ecrire_onglet(wb, df_onglet, titre, fill_data):
        ws = wb.create_sheet(titre)
        
        if df_onglet.empty:
            ws.cell(1, 1, f"Aucun hôtel dans cette catégorie").font = normal_font
            return
        
        df_out = df_onglet[cols_principales].copy()
        
        # En-têtes
        headers = {
            "priorite_completion": "Priorité",
            "statut_completude": "Statut",
            "champs_manquants": "Champs manquants",
            "nom": "Nom de l'hôtel",
            "adresse": "Adresse complète",
            "telephone": "Téléphone (auto)",
            "telephone_manuel": "Téléphone (manuel)",
            "email": "Email (auto)",
            "email_manuel": "Email (manuel)",
            "site_web": "Site web",
            "note": "Note /5",
            "nb_avis": "Nb avis",
            "lien_google_maps": "🔗 Fiche Google Maps",
            "notes": "Notes / Commentaires",
        }
        
        # Titre
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols_principales))
        titre_cell = ws.cell(1, 1, f"HOTELS PARIS — {titre}")
        titre_cell.font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
        titre_cell.fill = fills["header"]
        titre_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28
        
        # En-têtes colonnes
        for col_idx, col_name in enumerate(cols_principales, 1):
            cell = ws.cell(2, col_idx, headers.get(col_name, col_name))
            cell.font = header_font
            cell.fill = fills["header"]
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[2].height = 40
        
        # Données
        for row_idx, (_, data_row) in enumerate(df_out.iterrows(), 3):
            for col_idx, col_name in enumerate(cols_principales, 1):
                val = data_row[col_name]
                cell = ws.cell(row_idx, col_idx)
                
                # Liens cliquables
                if col_name in ("lien_google_maps", "site_web") and val:
                    cell.value = str(val)
                    cell.font = link_font
                    cell.hyperlink = str(val)
                else:
                    cell.value = val if pd.notna(val) and val != "" else ""
                    cell.font = bold_font if col_name == "nom" else normal_font
                
                cell.fill = fill_data
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=(col_name in ("adresse", "champs_manquants", "notes"))
                )
                cell.border = border
            
            ws.row_dimensions[row_idx].height = 30
        
        # Largeurs colonnes
        largeurs = {
            "priorite_completion": 8,
            "statut_completude": 22,
            "champs_manquants": 20,
            "nom": 35,
            "adresse": 40,
            "telephone": 16,
            "telephone_manuel": 16,
            "email": 28,
            "email_manuel": 28,
            "site_web": 32,
            "note": 8,
            "nb_avis": 8,
            "lien_google_maps": 24,
            "notes": 30,
        }
        for col_idx, col_name in enumerate(cols_principales, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = largeurs.get(col_name, 15)
        
        # Figer les 2 premières lignes
        ws.freeze_panes = "A3"
    
    # Créer les onglets
    ecrire_onglet(wb, df_incomplet, "1_INCOMPLETS_PRIORITAIRE", fills["incomplet"])
    ecrire_onglet(wb, df_a_completer, "2_A_COMPLETER", fills["a_completer"])
    ecrire_onglet(wb, df_complet, "3_COMPLETS", fills["complet"])
    
    # Onglet statistiques
    ws_stats = wb.create_sheet("STATISTIQUES", 0)
    ws_stats.column_dimensions["A"].width = 35
    ws_stats.column_dimensions["B"].width = 15
    
    stats_title = ws_stats.cell(1, 1, "STATISTIQUES DE COMPLETUDE")
    stats_title.font = Font(name="Arial", bold=True, size=14)
    stats_title.fill = fills["header"]
    ws_stats.cell(1, 1).font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    ws_stats.cell(1, 1).fill = fills["header"]
    ws_stats.merge_cells("A1:B1")
    ws_stats.row_dimensions[1].height = 28
    
    stats = [
        ("Total hôtels récupérés", len(df)),
        ("", ""),
        ("✅ Complets (tous les champs)", len(df_complet)),
        ("⚠️ À compléter (email manquant)", len(df_a_completer)),
        ("❌ Incomplets (téléphone/adresse)", len(df_incomplet)),
        ("", ""),
        ("Taux de complétude", f"{len(df_complet)/len(df)*100:.1f}%"),
        ("", ""),
        ("Hotels sans téléphone", len(df[df["telephone"] == ""])),
        ("Hotels sans email", len(df[df["email"] == ""])),
        ("Hotels sans site web", len(df[df["site_web"] == ""])),
    ]
    
    for i, (label, valeur) in enumerate(stats, 2):
        c1 = ws_stats.cell(i, 1, label)
        c1.font = bold_font if label else normal_font
        c1.alignment = Alignment(vertical="center")
        
        c2 = ws_stats.cell(i, 2, valeur)
        c2.font = bold_font if label else normal_font
        c2.alignment = Alignment(horizontal="right", vertical="center")
        
        if label:
            c1.fill = PatternFill("solid", start_color="F2F2F2", end_color="F2F2F2")
            c2.fill = PatternFill("solid", start_color="F2F2F2", end_color="F2F2F2")
        
        ws_stats.row_dimensions[i].height = 22
    
    # Instructions
    ws_inst = wb.create_sheet("MODE_EMPLOI", 0)
    ws_inst.column_dimensions["A"].width = 80
    
    inst_title = ws_inst.cell(1, 1, "MODE D'EMPLOI - COMPLETION MANUELLE DES DONNEES")
    inst_title.font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    inst_title.fill = fills["header"]
    inst_title.alignment = Alignment(horizontal="center", vertical="center")
    ws_inst.row_dimensions[1].height = 28
    
    instructions = [
        "",
        "📋 ORDRE DE TRAITEMENT :",
        "",
        "1. Commencer par l'onglet '1_INCOMPLETS_PRIORITAIRE' (téléphone ou adresse manquants)",
        "   → Cliquer sur le lien Google Maps pour chaque hôtel",
        "   → Compléter les champs 'telephone_manuel' et 'email_manuel'",
        "",
        "2. Continuer avec '2_A_COMPLETER' (principalement emails manquants)",
        "   → Visiter le site web de l'hôtel",
        "   → Chercher la page Contact ou Réservation",
        "   → Compléter le champ 'email_manuel'",
        "",
        "3. Vérifier '3_COMPLETS' pour s'assurer que les données auto-récupérées sont correctes",
        "",
        "💡 ASTUCES :",
        "",
        "- Les colonnes 'telephone_manuel' et 'email_manuel' sont prévues pour la saisie manuelle",
        "- Utiliser la colonne 'notes' pour tout commentaire (ex: 'fermé définitivement', 'fusionné avec...')",
        "- Le lien Google Maps permet de vérifier toutes les infos et voir les horaires",
        "- Pour l'email : chercher sur le site web, dans les avis Google, ou appeler l'hôtel",
        "",
        f"📊 Date du scraping : {datetime.now().strftime('%d/%m/%Y à %H:%M')}",
    ]
    
    for i, texte in enumerate(instructions, 2):
        cell = ws_inst.cell(i, 1, texte)
        if texte.startswith("📋") or texte.startswith("💡"):
            cell.font = Font(name="Arial", bold=True, size=12)
            cell.fill = PatternFill("solid", start_color="E7E6E6", end_color="E7E6E6")
        elif texte.startswith("📊"):
            cell.font = Font(name="Arial", italic=True, size=10)
        else:
            cell.font = Font(name="Arial", size=10)
        
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws_inst.row_dimensions[i].height = 20 if texte else 10
    
    wb.save(fichier)
    
    print(f"\n✅ Fichier Excel créé : {fichier}")
    print(f"\n📊 RÉSUMÉ :")
    print(f"   ✅ Complets :             {len(df_complet):4d} hôtels")
    print(f"   ⚠️  À compléter :          {len(df_a_completer):4d} hôtels")
    print(f"   ❌ Incomplets prioritaires : {len(df_incomplet):4d} hôtels")
    print(f"   ─────────────────────────────────")
    print(f"   📊 TOTAL :                {len(df):4d} hôtels")
    print(f"\n   Taux de complétude : {len(df_complet)/len(df)*100:.1f}%")


# ── SCRIPT PRINCIPAL ───────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("\n" + "=" * 70)
    print("  SCRAPER HOTELS PARIS - RECUPERATION COMPLETE")
    print("=" * 70)
    
    print("\n⚙️  CONFIGURATION :")
    print("   • Scraping par arrondissement : OUI")
    print("   • Requêtes supplémentaires : OUI")
    print("   • Pages par requête : 5 (max ~100 hôtels par requête)")
    print("   • Récupération email depuis site web : NON (trop lent)")
    print("\n💡 Pour l'email, il faudra compléter manuellement après le scraping")
    
    reponse = input("\n▶️  Lancer le scraping ? (o/n) : ").strip().lower()
    
    if reponse != 'o':
        print("Annulé.")
        exit()
    
    # Lancer le scraping complet
    df_hotels = scraper_hotels_paris_complet(
        inclure_arrondissements=True,
        inclure_requetes_supp=True,
        max_pages_par_requete=5,
        tenter_email_depuis_site=False  # Mettre True si vous voulez essayer (LENT)
    )
    
    if df_hotels.empty:
        print("\n❌ Aucun résultat récupéré.")
        exit()
    
    # Analyser la complétude
    print("\n🔍 Analyse de la complétude des données...")
    df_hotels = analyser_completude(df_hotels)
    
    # Export Excel
    fichier_output = f"hotels_paris_complet_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    export_excel_hotels(df_hotels, fichier_output)
    
    print("\n" + "=" * 70)
    print("✅ SCRAPING TERMINÉ")
    print("=" * 70)
    print(f"\n📁 Fichier créé : {fichier_output}")
    print("\n📋 PROCHAINES ÉTAPES :")
    print("   1. Ouvrir le fichier Excel")
    print("   2. Commencer par l'onglet '1_INCOMPLETS_PRIORITAIRE'")
    print("   3. Compléter manuellement les colonnes 'telephone_manuel' et 'email_manuel'")
    print("   4. Utiliser les liens Google Maps pour vérifier/récupérer les infos")
    print("\n💡 L'email n'est généralement pas disponible via Google Maps,")
    print("   il faudra le récupérer en visitant le site web de chaque hôtel.")