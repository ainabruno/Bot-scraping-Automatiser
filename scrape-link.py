import asyncio
import os
import re
import json
import time
from datetime import datetime
from pathlib import Path
from playwright.async_api import async_playwright, TimeoutError as PlaywrightTimeout
import PyPDF2
from openpyxl import load_workbook

# ============================================================
# CONFIGURATION
# ============================================================
COOKIES_PATH = "linkedin_cookies.json"
DOWNLOAD_DIR = r"C:\Users\A\Downloads"
PDF_FILENAME = "Profile.pdf"
EXCEL_FILE = "linkedin_profiles_formation_ia.xlsx"  # Nom de votre fichier Excel
SHEET_NAME = "Sheet1"  # Nom de la feuille (adapter si différent)

class LinkedInContactExtractor:
    def __init__(self):
        self.playwright = None
        self.browser = None
        self.context = None
        self.page = None
        self.profiles_data = []
        
    async def init_browser(self):
        """Initialise Playwright avec téléchargement activé"""
        try:
            self.playwright = await async_playwright().start()
            
            self.browser = await self.playwright.chromium.launch(
                headless=False,
                args=[
                    '--disable-blink-features=AutomationControlled',
                    '--disable-dev-shm-usage',
                    '--no-sandbox',
                ]
            )
            
            # Context avec téléchargement configuré
            self.context = await self.browser.new_context(
                viewport={'width': 1920, 'height': 1080},
                user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
                locale='fr-FR',
                timezone_id='Europe/Paris',
                accept_downloads=True
            )
            
            await self.context.add_init_script("""
                Object.defineProperty(navigator, 'webdriver', {
                    get: () => undefined
                });
            """)
            
            self.page = await self.context.new_page()
            print("✅ Navigateur initialisé avec support téléchargement")
            return True
            
        except Exception as e:
            print(f"❌ Erreur initialisation: {e}")
            return False
    
    async def load_cookies(self):
        """Charge les cookies LinkedIn"""
        if not os.path.exists(COOKIES_PATH):
            print(f"❌ Fichier cookies introuvable: {COOKIES_PATH}")
            return False
        
        try:
            with open(COOKIES_PATH, "r", encoding="utf-8") as f:
                cookies = json.load(f)
                
                playwright_cookies = []
                for cookie in cookies:
                    same_site = str(cookie.get("sameSite", "")).lower()
                    if same_site == "no_restriction":
                        same_site = "None"
                    elif same_site not in ["strict", "lax", "none"]:
                        same_site = "Lax"
                    else:
                        same_site = same_site.capitalize()
                    
                    playwright_cookie = {
                        "name": cookie["name"],
                        "value": cookie["value"],
                        "domain": cookie["domain"],
                        "path": cookie["path"],
                        "secure": cookie.get("secure", False),
                        "httpOnly": cookie.get("httpOnly", False),
                        "sameSite": same_site
                    }
                    
                    if "expirationDate" in cookie:
                        playwright_cookie["expires"] = int(cookie["expirationDate"])
                    
                    playwright_cookies.append(playwright_cookie)
                
                await self.context.add_cookies(playwright_cookies)
                print("✅ Cookies chargés")
                return True
                
        except Exception as e:
            print(f"❌ Erreur chargement cookies: {e}")
            return False
    
    async def check_logged_in(self):
        """Vérifie la session LinkedIn"""
        try:
            await self.page.goto("https://www.linkedin.com/feed/", wait_until="domcontentloaded", timeout=15000)
            await asyncio.sleep(2)
            
            if "feed" in self.page.url:
                print("✅ Session LinkedIn active\n")
                return True
            else:
                print("❌ Session expirée - Cookies invalides")
                return False
                
        except Exception as e:
            print(f"❌ Erreur vérification session: {e}")
            return False
    
    def load_excel_data(self):
        """Charge les données depuis le fichier Excel"""
        try:
            print("📊 Chargement des données Excel...")
            
            if not os.path.exists(EXCEL_FILE):
                print(f"❌ Fichier Excel introuvable: {EXCEL_FILE}")
                print("💡 Assurez-vous que le fichier est dans le même dossier que le script")
                return None, None
            
            # Charger le workbook
            workbook = load_workbook(EXCEL_FILE)
            
            # Sélectionner la feuille
            if SHEET_NAME in workbook.sheetnames:
                sheet = workbook[SHEET_NAME]
            else:
                sheet = workbook.active
                print(f"⚠️  Feuille '{SHEET_NAME}' non trouvée, utilisation de la feuille active: {sheet.title}")
            
            # Lire les données
            data = []
            headers = []
            
            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                if i == 0:
                    # Première ligne = en-têtes
                    headers = [str(cell).strip() if cell else f"Column_{j}" for j, cell in enumerate(row)]
                else:
                    # Créer un dictionnaire pour chaque ligne
                    row_data = {}
                    for j, cell in enumerate(row):
                        if j < len(headers):
                            row_data[headers[j]] = cell if cell else ""
                    data.append(row_data)
            
            print(f"✅ {len(data)} profils chargés depuis Excel")
            print(f"📋 Colonnes détectées: {', '.join(headers[:5])}...\n")
            
            return data, workbook
            
        except Exception as e:
            print(f"❌ Erreur chargement Excel: {e}")
            return None, None
    
    async def download_profile_pdf(self, profile_url):
        """Télécharge le PDF du profil LinkedIn"""
        try:
            print(f"   🔗 Accès à {profile_url}")
            await self.page.goto(profile_url, wait_until="domcontentloaded", timeout=20000)
            await asyncio.sleep(3)
            
            # Vérifier si on est bien sur le profil
            if "linkedin.com/in/" not in self.page.url:
                print("   ⚠️  Pas un profil LinkedIn valide")
                return None
            
            # Cliquer sur le bouton "Plus" (More)
            try:
                more_button = await self.page.wait_for_selector(
                    "button[aria-label*='Plus'][aria-label*='More']", 
                    timeout=5000
                )
                await more_button.click()
                await asyncio.sleep(1)
                print("   ✓ Bouton 'Plus' cliqué")
            except:
                # Alternative: chercher via texte
                try:
                    more_button = await self.page.wait_for_selector(
                        "button:has-text('Plus'), button:has-text('More')",
                        timeout=5000
                    )
                    await more_button.click()
                    await asyncio.sleep(1)
                except Exception as e:
                    print(f"   ❌ Bouton 'Plus' non trouvé: {e}")
                    return None
            
            # Cliquer sur "Enregistrer au format PDF" / "Save to PDF"
            try:
                # Attendre le menu déroulant
                await asyncio.sleep(1)
                
                # Chercher le lien PDF
                pdf_link = await self.page.wait_for_selector(
                    "a[href*='view-pdf'], div[role='menuitem']:has-text('PDF')",
                    timeout=5000
                )
                
                # Déclencher le téléchargement
                async with self.page.expect_download() as download_info:
                    await pdf_link.click()
                    download = await download_info.value
                
                # Sauvegarder le fichier
                pdf_path = os.path.join(DOWNLOAD_DIR, PDF_FILENAME)
                await download.save_as(pdf_path)
                print(f"   ✅ PDF téléchargé: {PDF_FILENAME}")
                
                return pdf_path
                
            except Exception as e:
                print(f"   ❌ Erreur téléchargement PDF: {e}")
                return None
            
        except Exception as e:
            print(f"   ❌ Erreur accès profil: {e}")
            return None
    
    def extract_email_from_pdf(self, pdf_path):
        """Extrait l'email depuis le PDF"""
        try:
            with open(pdf_path, 'rb') as f:
                reader = PyPDF2.PdfReader(f)
                text = ""
                for page in reader.pages:
                    text += page.extract_text() or ""
                
                # Recherche d'emails
                email_pattern = r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}'
                emails = re.findall(email_pattern, text)
                
                if emails:
                    # Nettoyer et retourner le premier email valide
                    for email in emails:
                        # Éviter les faux positifs
                        if not any(x in email.lower() for x in ['linkedin', 'example', 'test']):
                            return email
                
        except Exception as e:
            print(f"   ⚠️  Erreur extraction email: {e}")
        
        return None
    
    def extract_phone_from_pdf(self, pdf_path):
        """Extrait le numéro de téléphone depuis le PDF"""
        try:
            with open(pdf_path, 'rb') as f:
                reader = PyPDF2.PdfReader(f)
                text = ""
                for page in reader.pages:
                    text += page.extract_text() or ""
                
                # Patterns pour téléphones français et internationaux
                phone_patterns = [
                    r'\+?\d{1,3}[\s.-]?\(?\d{1,4}\)?[\s.-]?\d{1,4}[\s.-]?\d{1,4}[\s.-]?\d{1,9}',
                    r'0[1-9](?:[\s.-]?\d{2}){4}',  # Format FR: 06 12 34 56 78
                    r'\(\d{3}\)\s?\d{3}-\d{4}',  # Format US: (555) 123-4567
                ]
                
                for pattern in phone_patterns:
                    phones = re.findall(pattern, text)
                    if phones:
                        # Nettoyer le numéro
                        phone = phones[0].strip()
                        # Vérifier longueur minimale
                        if len(re.sub(r'\D', '', phone)) >= 10:
                            return phone
                
        except Exception as e:
            print(f"   ⚠️  Erreur extraction téléphone: {e}")
        
        return None
    
    def delete_pdf(self, pdf_path):
        """Supprime le fichier PDF"""
        try:
            if os.path.exists(pdf_path):
                os.remove(pdf_path)
                print(f"   🗑️  PDF supprimé")
        except Exception as e:
            print(f"   ⚠️  Erreur suppression PDF: {e}")
    
    async def process_profile(self, row_index, profile_url, current_email, current_phone):
        """Traite un profil individuel"""
        print(f"\n{'='*70}")
        print(f"📋 Profil {row_index + 1}")
        print(f"{'='*70}")
        
        # Vérifier si les données existent déjà
        if current_email and current_phone:
            print("   ✓ Email et téléphone déjà présents, passage au suivant")
            return {"email": current_email, "phone": current_phone}
        
        # Télécharger le PDF
        pdf_path = await self.download_profile_pdf(profile_url)
        
        if not pdf_path:
            return {"email": current_email or "", "phone": current_phone or ""}
        
        # Attendre que le fichier soit complètement écrit
        await asyncio.sleep(2)
        
        # Extraire les informations
        extracted_data = {"email": current_email or "", "phone": current_phone or ""}
        
        if not current_email:
            email = self.extract_email_from_pdf(pdf_path)
            if email:
                print(f"   ✅ Email trouvé: {email}")
                extracted_data["email"] = email
            else:
                print("   ⚠️  Aucun email trouvé")
        
        if not current_phone:
            phone = self.extract_phone_from_pdf(pdf_path)
            if phone:
                print(f"   ✅ Téléphone trouvé: {phone}")
                extracted_data["phone"] = phone
            else:
                print("   ⚠️  Aucun téléphone trouvé")
        
        # Supprimer le PDF
        self.delete_pdf(pdf_path)
        
        # Pause pour éviter les limites
        await asyncio.sleep(3)
        
        return extracted_data
    
    async def update_excel_file(self, workbook, row_index, email, phone):
        """Met à jour le fichier Excel avec les nouvelles données"""
        try:
            sheet = workbook[SHEET_NAME] if SHEET_NAME in workbook.sheetnames else workbook.active
            
            # Row index + 2 (1 pour l'index Python qui commence à 0, +1 pour l'en-tête)
            actual_row = row_index + 2
            
            # Trouver les colonnes Email et Phone
            headers = [cell.value for cell in sheet[1]]
            
            email_col = None
            phone_col = None
            
            for i, header in enumerate(headers, start=1):
                if header and 'email' in str(header).lower():
                    email_col = i
                if header and 'phone' in str(header).lower():
                    phone_col = i
            
            # Mettre à jour les cellules
            if email and email_col:
                sheet.cell(row=actual_row, column=email_col, value=email)
            
            if phone and phone_col:
                sheet.cell(row=actual_row, column=phone_col, value=phone)
            
            # Sauvegarder le fichier
            workbook.save(EXCEL_FILE)
            print(f"   💾 Excel mis à jour (ligne {actual_row})")
            
        except Exception as e:
            print(f"   ⚠️  Erreur mise à jour Excel: {e}")
    
    async def run(self):
        """Lance l'extraction complète"""
        try:
            # Initialiser le navigateur
            if not await self.init_browser():
                return
            
            # Charger les cookies
            if not await self.load_cookies():
                print("\n💡 Créez un fichier 'linkedin_cookies.json' avec vos cookies")
                return
            
            # Vérifier la connexion
            if not await self.check_logged_in():
                return
            
            # Charger les données Excel
            data, workbook = self.load_excel_data()
            if not data:
                return
            
            print("="*70)
            print(f"🎯 EXTRACTION DES CONTACTS")
            print(f"📊 {len(data)} profils à traiter")
            print("="*70)
            
            # Traiter chaque profil
            for index, row in enumerate(data):
                profile_url = row.get('profile_url', '')
                current_email = row.get('emails', '')  # Colonne "emails" dans votre Excel
                current_phone = row.get('phone', '')  # Colonne "phone" dans votre Excel
                
                if not profile_url:
                    print(f"\n⚠️  Profil {index + 1}: URL manquante")
                    continue
                
                # Traiter le profil
                extracted = await self.process_profile(
                    index, 
                    profile_url, 
                    current_email, 
                    current_phone
                )
                
                # Mettre à jour Excel
                if extracted['email'] or extracted['phone']:
                    await self.update_excel_file(
                        workbook, 
                        index, 
                        extracted['email'], 
                        extracted['phone']
                    )
            
        except KeyboardInterrupt:
            print("\n⚠️  Arrêt manuel")
        except Exception as e:
            print(f"\n❌ Erreur: {e}")
        finally:
            await self.cleanup()
    
    async def cleanup(self):
        """Ferme le navigateur"""
        if self.browser:
            await self.browser.close()
        if self.playwright:
            await self.playwright.stop()
        print("\n🔒 Navigateur fermé")
        print("\n✅ EXTRACTION TERMINÉE")

async def main():
    extractor = LinkedInContactExtractor()
    await extractor.run()

if __name__ == "__main__":
    print("""
╔══════════════════════════════════════════════════════════════════╗
║     📧 LINKEDIN CONTACT EXTRACTOR (via PDF) 📞                  ║
║                                                                  ║
║  🔍 Extraction: Email + Téléphone                               ║
║  📄 Méthode: Téléchargement PDF des profils                    ║
║  💾 Mise à jour: Fichier Excel automatique                      ║
╚══════════════════════════════════════════════════════════════════╝
    """)
    
    asyncio.run(main())