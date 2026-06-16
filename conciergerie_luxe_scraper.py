"""
conciergerie_luxe_scraper.py
============================
Scraper Google Maps via SerpAPI pour récupérer les conciergeries de luxe à Paris
Objectif : Données complètes (Nom, Adresse, Téléphone, Email, Site Web)
Gestion robuste des erreurs 429, backoff exponentiel, cache, et extraction email
"""

import os
import time
import re
import hashlib
import json
import random
import logging
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional, List, Dict, Tuple, Set
from dataclasses import dataclass, asdict
from urllib.parse import quote, urljoin, urlparse

import requests
import pandas as pd
from dotenv import load_dotenv

# ── CHARGEMENT CONFIGURATION ───────────────────────────────────────────────────
load_dotenv()

# Logging professionnel
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s | %(levelname)-8s | %(message)s',
    datefmt='%H:%M:%S'
)
logger = logging.getLogger(__name__)

# ── CONFIGURATION ──────────────────────────────────────────────────────────────
@dataclass
class Config:
    """Configuration centralisée du scraper."""
    SERP_API_KEY: str = os.getenv("SERP_API_KEY", "")
    PAUSE_ENTRE_REQUETES: float = 3.5          # Secondes entre requêtes (augmenté)
    PAUSE_ENTRE_PAGES: float = 2.0              # Secondes entre pages d'une même requête
    MAX_PAGES_PAR_REQUETE: int = 5
    REQUEST_TIMEOUT: int = 30
    MAX_RETRIES: int = 5                        # Nombre max de retries sur 429
    BASE_BACKOFF: float = 2.0                   # Délai de base pour backoff (secondes)
    MAX_BACKOFF: float = 120.0                  # Délai max entre retries
    JITTER_FACTOR: float = 0.3                  # 30% de jitter aléatoire
    CACHE_DIR: str = "./cache_serpapi"
    CACHE_TTL_HOURS: int = 24                   # Durée de vie du cache
    EMAIL_SCRAPING_ENABLED: bool = True
    EMAIL_SCRAPING_TIMEOUT: int = 15
    CONCURRENCY_DELAY: float = 0.5              # Délai supplémentaire entre requêtes similaires
    
    # Validation
    def __post_init__(self):
        if not self.SERP_API_KEY:
            raise ValueError("❌ SERP_API_KEY manquant. Définissez-la dans .env")

CONFIG = Config()

# ── USER-AGENTS ROTATION ───────────────────────────────────────────────────────
USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.0.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.0.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.0.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.0.36",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.0.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.0.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:121.0) Gecko/20100101 Firefox/121.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.1 Safari/605.1.15",
]

def get_random_headers() -> dict:
    """Retourne des headers HTTP réalistes avec User-Agent aléatoire."""
    return {
        'User-Agent': random.choice(USER_AGENTS),
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
        'Accept-Language': 'fr-FR,fr;q=0.9,en-US;q=0.8,en;q=0.7',
        'Accept-Encoding': 'gzip, deflate, br',
        'DNT': '1',
        'Connection': 'keep-alive',
        'Upgrade-Insecure-Requests': '1',
        'Sec-Fetch-Dest': 'document',
        'Sec-Fetch-Mode': 'navigate',
        'Sec-Fetch-Site': 'none',
        'Cache-Control': 'max-age=0',
    }

# ── CACHE DISQUE ───────────────────────────────────────────────────────────────
class DiskCache:
    """Cache disque simple avec TTL pour éviter les requêtes redondantes."""
    
    def __init__(self, cache_dir: str, ttl_hours: int):
        self.cache_dir = Path(cache_dir)
        self.cache_dir.mkdir(parents=True, exist_ok=True)
        self.ttl = timedelta(hours=ttl_hours)
    
    def _get_cache_path(self, key: str) -> Path:
        """Génère un chemin de cache à partir d'une clé."""
        hash_key = hashlib.md5(key.encode()).hexdigest()
        return self.cache_dir / f"{hash_key}.json"
    
    def get(self, key: str) -> Optional[dict]:
        """Récupère une valeur du cache si elle existe et n'est pas expirée."""
        path = self._get_cache_path(key)
        if not path.exists():
            return None
        
        try:
            with open(path, 'r', encoding='utf-8') as f:
                entry = json.load(f)
            
            cached_time = datetime.fromisoformat(entry['timestamp'])
            if datetime.now() - cached_time > self.ttl:
                path.unlink()  # Supprimer l'entrée expirée
                return None
            
            logger.debug(f"💾 Cache HIT pour: {key[:60]}...")
            return entry['data']
        except (json.JSONDecodeError, KeyError, OSError):
            return None
    
    def set(self, key: str, data: dict) -> None:
        """Stocke une valeur dans le cache."""
        path = self._get_cache_path(key)
        entry = {
            'timestamp': datetime.now().isoformat(),
            'data': data
        }
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(entry, f, ensure_ascii=False)

CACHE = DiskCache(CONFIG.CACHE_DIR, CONFIG.CACHE_TTL_HOURS)

# ── FONCTIONS UTILITAIRES ──────────────────────────────────────────────────────

def normaliser_nom(nom: str) -> str:
    """Normalise un nom pour la déduplication."""
    if not nom:
        return ""
    # Supprimer les espaces multiples, convertir en minuscules
    nom = re.sub(r'\s+', ' ', nom.lower().strip())
    # Supprimer les caractères spéciaux courants
    nom = re.sub(r'[^\w\s]', '', nom)
    return nom

def normaliser_adresse(adresse: str) -> str:
    """Normalise une adresse pour la déduplication."""
    if not adresse:
        return ""
    # Garder les 40 premiers caractères significatifs
    addr = re.sub(r'\s+', ' ', adresse.lower().strip())
    addr = re.sub(r'[^\w\s]', '', addr)
    return addr[:40]

def generer_cle_dedup(nom: str, adresse: str) -> str:
    """Génère une clé de déduplication robuste."""
    nom_norm = normaliser_nom(nom)
    addr_norm = normaliser_adresse(adresse)
    combined = f"{nom_norm}|{addr_norm}"
    return hashlib.md5(combined.encode()).hexdigest()

def normaliser_telephone(tel: str) -> str:
    """Normalise le format du téléphone français/international."""
    if not tel:
        return ""
    # Garde uniquement les chiffres et le +
    tel = re.sub(r'[^\d+]', '', tel)
    # Format français : ajouter +33 si commence par 0
    if tel.startswith('0') and len(tel) == 10:
        tel = '+33' + tel[1:]
    return tel

def extraire_email(texte: str) -> Optional[str]:
    """Extrait le premier email valide trouvé dans un texte."""
    if not texte:
        return None
    # Pattern email robuste
    pattern = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
    matches = re.findall(pattern, texte)
    
    # Filtrer les emails suspects (noreply, no-reply, etc.)
    emails_filtrés = [
        e for e in matches 
        if not any(x in e.lower() for x in ['noreply', 'no-reply', 'donotreply', 'example', 'test@'])
    ]
    
    return emails_filtrés[0] if emails_filtrés else (matches[0] if matches else None)

def extraire_email_depuis_site(url: str, timeout: int = 15) -> Optional[str]:
    """
    Tente de récupérer un email depuis le site web d'une conciergerie.
    Stratégie : page d'accueil → page contact → page à propos.
    """
    if not url:
        return None
    
    parsed = urlparse(url)
    base_url = f"{parsed.scheme}://{parsed.netloc}"
    
    pages_a_tester = [
        url,                                    # Page d'accueil
        urljoin(base_url, "/contact"),          # Page contact
        urljoin(base_url, "/contact-us"),       # Anglais
        urljoin(base_url, "/a-propos"),         # Français
        urljoin(base_url, "/about"),            # Anglais
        urljoin(base_url, "/nous-contacter"),   # Variante
    ]
    
    headers = get_random_headers()
    
    for page_url in pages_a_tester:
        try:
            response = requests.get(
                page_url, 
                headers=headers, 
                timeout=timeout,
                allow_redirects=True
            )
            if response.status_code == 200:
                email = extraire_email(response.text)
                if email:
                    logger.info(f"   📧 Email trouvé sur {page_url}: {email}")
                    return email
            time.sleep(0.5)  # Petite pause entre les pages
        except requests.exceptions.RequestException:
            continue
    
    return None

def calculer_backoff(attempt: int) -> float:
    """
    Calcule le délai de backoff exponentiel avec jitter.
    Formule: min(base * 2^attempt, max) * (1 + random * jitter)
    """
    exponential = CONFIG.BASE_BACKOFF * (2 ** attempt)
    capped = min(exponential, CONFIG.MAX_BACKOFF)
    jitter = 1 + (random.random() * CONFIG.JITTER_FACTOR)
    delay = capped * jitter
    return delay

# ── FONCTION PRINCIPALE SERPAPI ────────────────────────────────────────────────

def scrape_serpapi(query: str, start: int = 0) -> Tuple[List[dict], bool, Optional[str]]:
    """
    Interroge SerpAPI Google Maps avec gestion robuste des erreurs.
    
    Returns:
        (liste_resultats, a_encore_des_resultats, erreur_message)
    """
    cache_key = f"serpapi:{query}:{start}"
    
    # Vérifier le cache d'abord
    cached = CACHE.get(cache_key)
    if cached is not None:
        logger.debug(f"   💾 Utilisation du cache pour '{query}' (start={start})")
        return cached.get("results", []), cached.get("has_more", False), None
    
    params = {
        "engine": "google_maps",
        "q": query,
        "api_key": CONFIG.SERP_API_KEY,
        "hl": "fr",
        "gl": "fr",
        "type": "search",
        "start": start,
    }
    
    last_error = None
    
    for attempt in range(CONFIG.MAX_RETRIES):
        try:
            logger.debug(f"   🌐 Requête SerpAPI (tentative {attempt + 1}/{CONFIG.MAX_RETRIES})")
            
            response = requests.get(
                "https://serpapi.com/search",
                params=params,
                timeout=CONFIG.REQUEST_TIMEOUT
            )
            
            # Gestion du rate limit 429
            if response.status_code == 429:
                # Vérifier le header Retry-After
                retry_after = response.headers.get('Retry-After')
                if retry_after:
                    try:
                        wait_time = int(retry_after)
                    except ValueError:
                        wait_time = int(calculer_backoff(attempt))
                else:
                    wait_time = int(calculer_backoff(attempt))
                
                logger.warning(f"   ⏳ 429 reçu. Attente de {wait_time}s avant retry...")
                time.sleep(wait_time)
                last_error = "429 Too Many Requests"
                continue  # Retry
            
            # Autres erreurs HTTP
            if response.status_code >= 400:
                logger.error(f"   ❌ Erreur HTTP {response.status_code}: {response.text[:200]}")
                last_error = f"HTTP {response.status_code}"
                if response.status_code in [401, 403]:
                    return [], False, "Clé API invalide ou quota épuisé"
                time.sleep(calculer_backoff(attempt))
                continue
            
            response.raise_for_status()
            data = response.json()
            
            # Vérifier les erreurs dans la réponse JSON
            if "error" in data:
                error_msg = data["error"]
                logger.error(f"   ❌ Erreur API SerpAPI: {error_msg}")
                if "rate limit" in str(error_msg).lower():
                    time.sleep(calculer_backoff(attempt))
                    continue
                return [], False, str(error_msg)
            
            # Extraction des résultats
            resultats = []
            for place in data.get("local_results", []):
                gps = place.get("gps_coordinates", {})
                pid = place.get("place_id", "")
                tel = place.get("phone", "")
                site = place.get("website", "")
                
                resultats.append({
                    "nom": place.get("title", ""),
                    "adresse": place.get("address", ""),
                    "telephone": normaliser_telephone(tel),
                    "site_web": site,
                    "email": "",  # Sera rempli plus tard si possible
                    "note": place.get("rating", ""),
                    "nb_avis": place.get("reviews", 0),
                    "categorie": place.get("type", ""),
                    "latitude": gps.get("latitude"),
                    "longitude": gps.get("longitude"),
                    "place_id": pid,
                    "lien_google_maps": f"https://www.google.com/maps/place/?q=place_id:{pid}" if pid else "",
                    "requete_source": query,
                })
            
            # Vérifier pagination
            serpapi_pagination = data.get("serpapi_pagination", {})
            has_more = "next" in serpapi_pagination or "next_page_token" in serpapi_pagination
            
            # Stocker dans le cache
            CACHE.set(cache_key, {
                "results": resultats,
                "has_more": has_more,
                "timestamp": datetime.now().isoformat()
            })
            
            return resultats, has_more, None
            
        except requests.exceptions.Timeout:
            wait_time = calculer_backoff(attempt)
            logger.warning(f"   ⏱️ Timeout. Attente de {wait_time:.1f}s...")
            time.sleep(wait_time)
            last_error = "Timeout"
            
        except requests.exceptions.ConnectionError as e:
            wait_time = calculer_backoff(attempt)
            logger.warning(f"   🔌 Erreur connexion: {e}. Attente de {wait_time:.1f}s...")
            time.sleep(wait_time)
            last_error = f"Connection error: {e}"
            
        except requests.exceptions.RequestException as e:
            wait_time = calculer_backoff(attempt)
            logger.error(f"   🌐 Erreur réseau: {e}")
            time.sleep(wait_time)
            last_error = str(e)
    
    # Toutes les tentatives échouées
    logger.error(f"   💥 Échec après {CONFIG.MAX_RETRIES} tentatives pour '{query}'")
    return [], False, last_error

# ── REQUÊTES OPTIMISÉES POUR CONCIERGERIES ───────────────────────────────────

REQUETES_CONCIERGERIE = [
    # Requêtes principales (français)
    "conciergerie de luxe Paris",
    "conciergerie privée Paris",
    "conciergerie haut de gamme Paris",
    "concierge service Paris luxe",
    "concierge personnel Paris",
    "conciergerie exclusive Paris",
    "conciergerie premium Paris",
    "conciergerie de prestige Paris",
    "conciergerie sur mesure Paris",
    
    # Requêtes par arrondissement ciblé (zones luxe)
    "conciergerie luxe Paris 1er",
    "conciergerie luxe Paris 2e",
    "conciergerie luxe Paris 8e",
    "conciergerie luxe Paris 16e",
    "conciergerie luxe Paris 7e",
    "conciergerie luxe Paris 6e",
    "conciergerie luxe Paris 4e",
    
    # Services spécifiques
    "gestion location luxe Paris conciergerie",
    "conciergerie Airbnb luxe Paris",
    "property management luxury Paris",
    "conciergerie immobilière luxe Paris",
    
    # Requêtes anglaises (internationales)
    "luxury concierge service Paris",
    "private concierge Paris France",
    "VIP concierge service Paris",
    "high end concierge Paris",
    "elite concierge Paris",
    "luxury lifestyle management Paris",
    "personal concierge service Paris",
    
    # Services associés
    "personal assistant Paris luxe",
    "lifestyle management Paris",
    "luxury travel concierge Paris",
    "VIP travel service Paris",
    "chauffeur privé Paris luxe",
    "shopping personal shopper Paris luxe",
    "reservation restaurant Michelin Paris concierge",
]

# ── SCRAPER PRINCIPAL ───────────────────────────────────────────────────────────

def scraper_conciergeries_complet(
    max_pages_par_requete: int = None,
    tenter_email: bool = None
) -> pd.DataFrame:
    """
    Lance un scraping complet des conciergeries de luxe à Paris.
    
    Returns:
        DataFrame avec toutes les conciergeries trouvées
    """
    max_pages = max_pages_par_requete or CONFIG.MAX_PAGES_PAR_REQUETE
    scraper_email = tenter_email if tenter_email is not None else CONFIG.EMAIL_SCRAPING_ENABLED
    
    print("=" * 75)
    print("  SCRAPER CONCIERGERIES DE LUXE - PARIS")
    print("  Objectif : Données complètes (Nom, Adresse, Téléphone, Email, Site Web)")
    print("=" * 75)
    print(f"\n⚙️  Configuration:")
    print(f"   • Requêtes à effectuer : {len(REQUETES_CONCIERGERIE)}")
    print(f"   • Pages max par requête : {max_pages}")
    print(f"   • Pause entre requêtes : {CONFIG.PAUSE_ENTRE_REQUETES}s")
    print(f"   • Max retries sur 429 : {CONFIG.MAX_RETRIES}")
    print(f"   • Backoff max : {CONFIG.MAX_BACKOFF}s")
    print(f"   • Cache activé : OUI ({CONFIG.CACHE_DIR})")
    print(f"   • Scraping email depuis site : {'OUI' if scraper_email else 'NON'}")
    print(f"   • Clé API : {'✅ Configurée' if CONFIG.SERP_API_KEY else '❌ MANQUANTE'}")
    print()
    
    tous_resultats: List[dict] = []
    vus: Set[str] = set()  # Set de hashes MD5 pour déduplication
    stats = {
        "total_requetes": 0,
        "requetes_reussies": 0,
        "requetes_echouees": 0,
        "requetes_from_cache": 0,
        "total_trouves": 0,
        "nouveaux_ajoutes": 0,
        "emails_trouves": 0,
    }
    
    for idx, requete in enumerate(REQUETES_CONCIERGERIE, 1):
        logger.info(f"[{idx:02d}/{len(REQUETES_CONCIERGERIE)}] '{requete}'")
        stats["total_requetes"] += 1
        
        page = 0
        requete_reussie = False
        
        while page < max_pages:
            start = page * 20
            
            if page > 0:
                logger.info(f"      → Page {page + 1}")
            
            resultats, a_encore, erreur = scrape_serpapi(requete, start=start)
            
            if erreur and not resultats:
                logger.error(f"      💥 Échec complet: {erreur}")
                break
            
            if resultats:
                requete_reussie = True
            
            nouveaux = 0
            for place in resultats:
                cle = generer_cle_dedup(place["nom"], place["adresse"])
                
                if cle not in vus and place["nom"]:
                    vus.add(cle)
                    stats["total_trouves"] += 1
                    
                    # Tenter de récupérer l'email depuis le site web
                    if scraper_email and place["site_web"] and not place["email"]:
                        logger.info(f"      🔍 Recherche email pour: {place['nom'][:40]}...")
                        email = extraire_email_depuis_site(place["site_web"])
                        if email:
                            place["email"] = email
                            stats["emails_trouves"] += 1
                            logger.info(f"      ✅ Email trouvé: {email}")
                        time.sleep(1)  # Pause respectueuse entre les sites
                    
                    tous_resultats.append(place)
                    nouveaux += 1
            
            stats["nouveaux_ajoutes"] += nouveaux
            logger.info(f"      📊 {len(resultats)} trouvés, {nouveaux} nouveaux → Total unique: {len(tous_resultats)}")
            
            if not a_encore or len(resultats) == 0:
                break
            
            page += 1
            time.sleep(CONFIG.PAUSE_ENTRE_PAGES)
        
        if requete_reussie:
            stats["requetes_reussies"] += 1
        else:
            stats["requetes_echouees"] += 1
        
        # Pause entre requêtes (avec jitter aléatoire)
        if idx < len(REQUETES_CONCIERGERIE):
            pause = CONFIG.PAUSE_ENTRE_REQUETES + random.uniform(0, 1.5)
            logger.debug(f"   ⏳ Pause de {pause:.1f}s avant la prochaine requête...")
            time.sleep(pause)
    
    # ── RÉSUMÉ ──
    print("\n" + "=" * 75)
    print("  RÉSULTATS DU SCRAPING")
    print("=" * 75)
    print(f"\n📊 Statistiques:")
    print(f"   • Requêtes totales : {stats['total_requetes']}")
    print(f"   • Requêtes réussies : {stats['requetes_reussies']}")
    print(f"   • Requêtes échouées : {stats['requetes_echouees']}")
    print(f"   • Résultats bruts trouvés : {stats['total_trouves']}")
    print(f"   • Entrées uniques ajoutées : {stats['nouveaux_ajoutes']}")
    print(f"   • Emails trouvés via scraping : {stats['emails_trouves']}")
    print(f"\n✅ Scraping terminé : {len(tous_resultats)} conciergeries uniques récupérées")
    
    if not tous_resultats:
        logger.warning("⚠️ Aucun résultat récupéré!")
        return pd.DataFrame()
    
    return pd.DataFrame(tous_resultats)

# ── ANALYSE DE COMPLÉTUDE ──────────────────────────────────────────────────────

def analyser_completude(df: pd.DataFrame) -> pd.DataFrame:
    """
    Analyse la complétude des données et ajoute les colonnes de statut.
    """
    if df.empty:
        return df
    
    df = df.copy()
    
    statuts = []
    champs_manquants = []
    priorites = []
    scores = []
    
    for _, row in df.iterrows():
        manquants = []
        score = 0
        
        # Vérifier chaque champ
        champs = {
            "telephone": row.get("telephone"),
            "email": row.get("email"),
            "site_web": row.get("site_web"),
            "adresse": row.get("adresse"),
        }
        
        for champ, valeur in champs.items():
            if not valeur or str(valeur).strip() == "":
                manquants.append(champ)
            else:
                score += 25  # 25 points par champ
        
        # Déterminer le statut
        if score == 100:
            statut = "✅ COMPLET"
            priorite = "A"
        elif score >= 75:
            statut = "⚠️ QUASI COMPLET"
            priorite = "B"
        elif score >= 50:
            statut = "⚡ MOYENNEMENT COMPLET"
            priorite = "B"
        elif "telephone" in manquants or "adresse" in manquants:
            statut = "❌ INCOMPLET - PRIORITAIRE"
            priorite = "C"
        else:
            statut = "⚠️ À COMPLÉTER"
            priorite = "C"
        
        statuts.append(statut)
        champs_manquants.append(", ".join(manquants) if manquants else "Aucun")
        priorites.append(priorite)
        scores.append(score)
    
    df["statut_completude"] = statuts
    df["champs_manquants"] = champs_manquants
    df["priorite_completion"] = priorites
    df["score_completude"] = scores
    df["email_manuel"] = ""
    df["telephone_manuel"] = ""
    df["notes"] = ""
    
    # Trier par priorité puis par score
    ordre_priorite = {"C": 0, "B": 1, "A": 2}
    df["_ordre"] = df["priorite_completion"].map(ordre_priorite)
    df = df.sort_values(["_ordre", "score_completude", "nom"], ascending=[True, False, True])
    df = df.drop(columns=["_ordre"])
    
    return df.reset_index(drop=True)

# ── EXPORT EXCEL ───────────────────────────────────────────────────────────────

def export_excel(df: pd.DataFrame, fichier: str):
    """
    Exporte en Excel avec mise en forme professionnelle.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.warning("openpyxl non installé. Export CSV uniquement.")
        df.to_csv(fichier.replace('.xlsx', '.csv'), index=False, encoding='utf-8-sig')
        return
    
    # Séparer par statut
    df_incomplet = df[df["priorite_completion"] == "C"].copy()
    df_a_completer = df[df["priorite_completion"] == "B"].copy()
    df_complet = df[df["priorite_completion"] == "A"].copy()
    
    cols_principales = [
        "priorite_completion", "statut_completude", "score_completude", "champs_manquants",
        "nom", "adresse", "telephone", "telephone_manuel",
        "email", "email_manuel", "site_web",
        "note", "nb_avis", "categorie", "lien_google_maps", "notes"
    ]
    
    # Styles
    fills = {
        "header": PatternFill("solid", fgColor="1F4E78"),
        "incomplet": PatternFill("solid", fgColor="FCE4D6"),
        "a_completer": PatternFill("solid", fgColor="FFF2CC"),
        "complet": PatternFill("solid", fgColor="E2EFDA"),
    }
    
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    bold_font = Font(name="Calibri", bold=True, size=10)
    normal_font = Font(name="Calibri", size=10)
    link_font = Font(name="Calibri", size=10, color="0563C1", underline="single")
    thin_border = Side(style="thin", color="D0D0D0")
    border = Border(bottom=thin_border, right=thin_border)
    
    wb = Workbook()
    wb.remove(wb.active)
    
    def creer_onglet(wb, df_onglet, titre, fill_data):
        ws = wb.create_sheet(titre)
        
        if df_onglet.empty:
            ws.cell(1, 1, "Aucune entrée dans cette catégorie").font = normal_font
            return
        
        df_out = df_onglet[cols_principales].copy()
        
        headers = {
            "priorite_completion": "Priorité",
            "statut_completude": "Statut",
            "score_completude": "Score %",
            "champs_manquants": "Champs manquants",
            "nom": "Nom",
            "adresse": "Adresse",
            "telephone": "Téléphone (auto)",
            "telephone_manuel": "Téléphone (manuel)",
            "email": "Email (auto)",
            "email_manuel": "Email (manuel)",
            "site_web": "Site Web",
            "note": "Note /5",
            "nb_avis": "Nb avis",
            "categorie": "Catégorie",
            "lien_google_maps": "🔗 Google Maps",
            "notes": "Notes",
        }
        
        # Titre
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols_principales))
        titre_cell = ws.cell(1, 1, f"CONCIERGERIES LUXE PARIS — {titre}")
        titre_cell.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
        titre_cell.fill = fills["header"]
        titre_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28
        
        # En-têtes
        for col_idx, col_name in enumerate(cols_principales, 1):
            cell = ws.cell(2, col_idx, headers.get(col_name, col_name))
            cell.font = header_font
            cell.fill = fills["header"]
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[2].height = 40
        
        # Données
        for row_idx, (_, data_row) in enumerate(df_out.iterrows(), 3):
            for col_idx, col_name in enumerate(cols_principales, 1):
                val = data_row[col_name]
                cell = ws.cell(row_idx, col_idx)
                
                if col_name in ("lien_google_maps", "site_web") and val:
                    cell.value = str(val)
                    cell.font = link_font
                    cell.hyperlink = str(val)
                else:
                    cell.value = val if pd.notna(val) and val != "" else ""
                    cell.font = bold_font if col_name == "nom" else normal_font
                
                cell.fill = fill_data
                cell.alignment = Alignment(vertical="top", wrap_text=(col_name in ("adresse", "champs_manquants", "notes")))
                cell.border = border
            
            ws.row_dimensions[row_idx].height = 30
        
        # Largeurs
        largeurs = {
            "priorite_completion": 8, "statut_completude": 22, "score_completude": 8,
            "champs_manquants": 20, "nom": 35, "adresse": 40,
            "telephone": 16, "telephone_manuel": 16,
            "email": 30, "email_manuel": 30,
            "site_web": 32, "note": 8, "nb_avis": 8,
            "categorie": 20, "lien_google_maps": 24, "notes": 30,
        }
        for col_idx, col_name in enumerate(cols_principales, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = largeurs.get(col_name, 15)
        
        ws.freeze_panes = "A3"
    
    # Créer les onglets
    creer_onglet(wb, df_incomplet, "1_INCOMPLETS_PRIORITAIRE", fills["incomplet"])
    creer_onglet(wb, df_a_completer, "2_A_COMPLETER", fills["a_completer"])
    creer_onglet(wb, df_complet, "3_COMPLETS", fills["complet"])
    
    # Onglet statistiques
    ws_stats = wb.create_sheet("STATISTIQUES", 0)
    ws_stats.column_dimensions["A"].width = 40
    ws_stats.column_dimensions["B"].width = 15
    
    stats_title = ws_stats.cell(1, 1, "STATISTIQUES DE COMPLÉTUDE")
    stats_title.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    stats_title.fill = fills["header"]
    ws_stats.merge_cells("A1:B1")
    ws_stats.row_dimensions[1].height = 28
    
    total = len(df)
    complet = len(df_complet)
    taux = (complet / total * 100) if total > 0 else 0
    
    stats = [
        ("Total conciergeries", total),
        ("", ""),
        ("✅ Complets (100%)", len(df_complet)),
        ("⚠️ Quasi complets (75-99%)", len(df[df["score_completude"] >= 75]) - len(df_complet)),
        ("⚡ Moyennement complets (50-74%)", len(df[(df["score_completude"] >= 50) & (df["score_completude"] < 75)])),
        ("❌ Incomplets (<50%)", len(df_incomplet)),
        ("", ""),
        ("Taux de complétude", f"{taux:.1f}%"),
        ("Score moyen", f"{df['score_completude'].mean():.1f}%"),
        ("", ""),
        ("Sans téléphone", len(df[df["telephone"] == ""])),
        ("Sans email", len(df[df["email"] == ""])),
        ("Sans site web", len(df[df["site_web"] == ""])),
        ("Sans adresse", len(df[df["adresse"] == ""])),
    ]
    
    for i, (label, valeur) in enumerate(stats, 2):
        c1 = ws_stats.cell(i, 1, label)
        c1.font = bold_font if label else normal_font
        c2 = ws_stats.cell(i, 2, valeur)
        c2.font = bold_font if label else normal_font
        c2.alignment = Alignment(horizontal="right")
        if label:
            c1.fill = PatternFill("solid", fgColor="F2F2F2")
            c2.fill = PatternFill("solid", fgColor="F2F2F2")
        ws_stats.row_dimensions[i].height = 22
    
    # Mode d'emploi
    ws_inst = wb.create_sheet("MODE_EMPLOI", 0)
    ws_inst.column_dimensions["A"].width = 85
    
    inst_title = ws_inst.cell(1, 1, "MODE D'EMPLOI - COMPLÉTION MANUELLE")
    inst_title.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    inst_title.fill = fills["header"]
    inst_title.alignment = Alignment(horizontal="center", vertical="center")
    ws_inst.row_dimensions[1].height = 28
    
    instructions = [
        "",
        "📋 ORDRE DE TRAITEMENT :",
        "",
        "1. Commencer par l'onglet '1_INCOMPLETS_PRIORITAIRE'",
        "   → Cliquer sur le lien Google Maps pour chaque conciergerie",
        "   → Compléter 'telephone_manuel' et 'email_manuel'",
        "",
        "2. Continuer avec '2_A_COMPLETER'",
        "   → Visiter le site web de la conciergerie",
        "   → Chercher la page Contact / Réservation",
        "   → Compléter 'email_manuel'",
        "",
        "3. Vérifier '3_COMPLETS' pour valider les données auto-récupérées",
        "",
        "💡 ASTUCES :",
        "",
        "- Les colonnes '_manuel' sont prévues pour la saisie manuelle",
        "- Utiliser 'notes' pour tout commentaire",
        "- Le lien Google Maps permet de vérifier toutes les infos",
        "- Pour l'email : chercher sur le site, dans les avis Google, ou appeler",
        "",
        f"📊 Date du scraping : {datetime.now().strftime('%d/%m/%Y à %H:%M')}",
        f"🛡️ Cache activé : {CONFIG.CACHE_DIR}",
    ]
    
    for i, texte in enumerate(instructions, 2):
        cell = ws_inst.cell(i, 1, texte)
        if texte.startswith(("📋", "💡")):
            cell.font = Font(name="Calibri", bold=True, size=12)
            cell.fill = PatternFill("solid", fgColor="E7E6E6")
        elif texte.startswith("📊"):
            cell.font = Font(name="Calibri", italic=True, size=10)
        else:
            cell.font = Font(name="Calibri", size=10)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws_inst.row_dimensions[i].height = 20 if texte else 10
    
    wb.save(fichier)
    
    print(f"\n✅ Fichier Excel créé : {fichier}")
    print(f"\n📊 RÉSUMÉ :")
    print(f"   ✅ Complets :              {len(df_complet):4d}")
    print(f"   ⚠️  À compléter :           {len(df_a_completer):4d}")
    print(f"   ❌ Incomplets prioritaires : {len(df_incomplet):4d}")
    print(f"   ─────────────────────────────────")
    print(f"   📊 TOTAL :                {total:4d}")
    print(f"\n   Taux de complétude : {taux:.1f}%")

# ── SCRIPT PRINCIPAL ───────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("\n" + "=" * 75)
    print("  SCRAPER CONCIERGERIES LUXE - PARIS")
    print("=" * 75)
    
    # Vérifier la configuration
    try:
        _ = CONFIG.SERP_API_KEY
    except ValueError as e:
        print(f"\n❌ {e}")
        print("\nCréez un fichier .env avec:")
        print('   SERP_API_KEY="votre_cle_api_ici"')
        exit(1)
    
    print("\n⚙️  CONFIGURATION :")
    print(f"   • Requêtes : {len(REQUETES_CONCIERGERIE)} requêtes optimisées")
    print(f"   • Pages par requête : {CONFIG.MAX_PAGES_PAR_REQUETE}")
    print(f"   • Pause entre requêtes : {CONFIG.PAUSE_ENTRE_REQUETES}s (+ jitter)")
    print(f"   • Retry sur 429 : {CONFIG.MAX_RETRIES} tentatives avec backoff exponentiel")
    print(f"   • Cache : {CONFIG.CACHE_DIR} (TTL: {CONFIG.CACHE_TTL_HOURS}h)")
    print(f"   • Scraping email : {'OUI' if CONFIG.EMAIL_SCRAPING_ENABLED else 'NON'}")
    print("\n💡 Le cache évite de re-consommer des crédits API sur les requêtes récentes.")
    
    reponse = input("\n▶️  Lancer le scraping ? (o/n) : ").strip().lower()
    
    if reponse not in ('o', 'oui', 'y', 'yes'):
        print("Annulé.")
        exit()
    
    # Lancer le scraping
    df = scraper_conciergeries_complet()
    
    if df.empty:
        print("\n❌ Aucun résultat récupéré. Vérifiez votre clé API et votre connexion.")
        exit()
    
    # Analyser la complétude
    print("\n🔍 Analyse de la complétude des données...")
    df = analyser_completude(df)
    
    # Export
    fichier_output = f"conciergeries_luxe_paris_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    export_excel(df, fichier_output)
    
    print("\n" + "=" * 75)
    print("✅ SCRAPING TERMINÉ")
    print("=" * 75)
    print(f"\n📁 Fichier créé : {fichier_output}")
    print("\n📋 PROCHAINES ÉTAPES :")
    print("   1. Ouvrir le fichier Excel")
    print("   2. Commencer par l'onglet '1_INCOMPLETS_PRIORITAIRE'")
    print("   3. Compléter manuellement les colonnes 'telephone_manuel' et 'email_manuel'")
    print("   4. Utiliser les liens Google Maps pour vérifier/récupérer les infos")