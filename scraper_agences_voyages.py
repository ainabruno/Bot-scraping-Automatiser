import time, re, requests, pandas as pd
from datetime import datetime
from urllib.parse import quote

# ── CONFIGURATION ──────────────────────────────────────────────────────────────
try:
    from config import SERP_API_KEY, PAUSE_ENTRE_REQUETES, REQUEST_TIMEOUT
except ImportError:
    SERP_API_KEY          = "f77905421b96292fedf81fab00d96982701c9696f6d33e31b5f04c715a4414fd"   # ← remplacer
    PAUSE_ENTRE_REQUETES  = 2.5
    REQUEST_TIMEOUT       = 30

MAX_PAGES_PAR_REQUETE = 5   # ~100 résultats par requête

# ── REQUÊTES PAR MARCHÉ ────────────────────────────────────────────────────────

MARCHES = {

    # ── ÉTATS-UNIS ─────────────────────────────────────────────────────────────
    "USA": [
        "travel agency Paris France tours New York",
        "travel agency Paris France tours Los Angeles",
        "travel agency Paris France tours Chicago",
        "travel agency Paris France tours Miami",
        "travel agency Paris France tours San Francisco",
        "luxury travel agency France Europe USA",
        "honeymoon travel agency Paris France",
        "France tour operator United States",
        "Europe travel agency Paris package USA",
        "travel agency Paris France tours Houston",
        "travel agency Paris France tours Boston",
        "travel agency Paris France tours Washington DC",
        "travel agency Paris France tours Seattle",
        "French travel specialist USA",
        "Paris travel agency group tours America",
    ],

    # ── CHINE ──────────────────────────────────────────────────────────────────
    "Chine": [
        "旅行社 巴黎 法国 北京",          # Agence Paris France Pékin
        "旅行社 巴黎 法国 上海",          # Shanghai
        "旅行社 巴黎 法国 广州",          # Guangzhou
        "travel agency Paris France China Beijing",
        "travel agency Paris France Shanghai",
        "France tour operator China",
        "Europe luxury tour China Paris",
        "法国旅行社 巴黎 欧洲",
        "travel agency Paris Guangzhou China",
        "China outbound tour operator Paris France",
    ],

    # ── JAPON ──────────────────────────────────────────────────────────────────
    "Japon": [
        "旅行会社 パリ フランス 東京",     # Agence Paris France Tokyo
        "旅行会社 パリ フランス 大阪",     # Osaka
        "travel agency Paris France Tokyo Japan",
        "travel agency Paris France Osaka Japan",
        "France tour operator Japan",
        "Europe travel agency Japan Paris",
        "luxury France tour Japan honeymoon",
        "日本 フランス ツアー 旅行会社",
    ],

    # ── CORÉE DU SUD ───────────────────────────────────────────────────────────
    "Coree_du_Sud": [
        "여행사 파리 프랑스 서울",         # Agence Paris France Séoul
        "여행사 파리 프랑스 부산",         # Busan
        "travel agency Paris France Seoul Korea",
        "France tour operator Korea",
        "Europe travel package Korea Paris",
        "프랑스 여행사 파리 유럽",
    ],

    # ── INDE ───────────────────────────────────────────────────────────────────
    "Inde": [
        "travel agency Paris France Mumbai India",
        "travel agency Paris France Delhi India",
        "travel agency Paris France Bangalore India",
        "France tour operator India",
        "Europe luxury tour India Paris",
        "Paris honeymoon package India travel agency",
        "France travel agent India",
    ],

    # ── SINGAPOUR ──────────────────────────────────────────────────────────────
    "Singapour": [
        "travel agency Paris France Singapore",
        "France tour operator Singapore",
        "Europe luxury travel Singapore Paris",
        "Paris package tour Singapore",
        "French travel specialist Singapore",
    ],

    # ── ARABIE SAOUDITE ────────────────────────────────────────────────────────
    "Arabie_Saoudite": [
        "وكالة سفر باريس فرنسا الرياض",   # Agence Paris France Riyad
        "وكالة سفر باريس فرنسا جدة",       # Djeddah
        "travel agency Paris France Riyadh Saudi Arabia",
        "travel agency Paris France Jeddah Saudi Arabia",
        "France tour operator Saudi Arabia",
        "Europe luxury travel Saudi Arabia Paris",
        "Paris vacation package Saudi travel agency",
    ],

    # ── ÉMIRATS ARABES UNIS ────────────────────────────────────────────────────
    "Emirats_Arabes_Unis": [
        "وكالة سفر باريس فرنسا دبي",       # Dubaï
        "وكالة سفر باريس فرنسا أبوظبي",    # Abu Dhabi
        "travel agency Paris France Dubai UAE",
        "travel agency Paris France Abu Dhabi",
        "France tour operator Dubai",
        "luxury travel agency Dubai Paris France",
        "Paris honeymoon package Dubai travel",
    ],

    # ── QATAR ──────────────────────────────────────────────────────────────────
    "Qatar": [
        "وكالة سفر باريس فرنسا الدوحة",    # Doha
        "travel agency Paris France Doha Qatar",
        "France tour operator Qatar",
        "Europe luxury tour Qatar Paris",
    ],

    # ── AGENTS INDÉPENDANTS USA (via Google/LinkedIn) ──────────────────────────
    "Agents_Independants_USA": [
        "independent travel agent Paris France specialist USA",
        "home based travel agent Paris France USA",
        "luxury travel advisor France Paris USA",
        "freelance travel agent Paris specialist America",
        "travel consultant France Paris USA independent",
        "Paris France travel advisor host agency USA",
        "independent travel advisor France honeymoon USA",
        "travel agent France specialist site:linkedin.com",
        "travel advisor Paris France site:instagram.com",
        "France travel specialist independent agent",
    ],
}

# ── HELPERS ────────────────────────────────────────────────────────────────────

def normaliser_telephone(tel: str) -> str:
    return re.sub(r"[^\d+]", "", tel) if tel else ""


def extraire_email(texte: str) -> str:
    if not texte:
        return ""
    m = re.findall(r"\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}\b", texte)
    return m[0] if m else ""


def scrape_requete(query: str, start: int = 0) -> tuple[list[dict], bool]:
    params = {
        "engine":  "google_maps",
        "q":       query,
        "api_key": SERP_API_KEY,
        "hl":      "fr",
        "gl":      "fr",
        "type":    "search",
        "start":   start,
    }
    try:
        resp = requests.get("https://serpapi.com/search", params=params,
                            timeout=REQUEST_TIMEOUT)
        resp.raise_for_status()
        data = resp.json()
        if "error" in data:
            print(f"      ⚠ SerpAPI : {data['error']}")
            return [], False

        resultats = []
        for place in data.get("local_results", []):
            gps = place.get("gps_coordinates", {})
            pid = place.get("place_id", "")
            resultats.append({
                "nom":              place.get("title", ""),
                "adresse":          place.get("address", ""),
                "telephone":        normaliser_telephone(place.get("phone", "")),
                "email":            "",
                "site_web":         place.get("website", ""),
                "note":             place.get("rating", ""),
                "nb_avis":          place.get("reviews", 0),
                "categorie":        place.get("type", ""),
                "latitude":         gps.get("latitude"),
                "longitude":        gps.get("longitude"),
                "place_id":         pid,
                "lien_google_maps": f"https://www.google.com/maps/place/?q=place_id:{pid}" if pid else "",
                "requete_source":   query,
            })

        pagination = data.get("serpapi_pagination", {})
        a_encore = "next" in pagination or "next_page_token" in pagination
        return resultats, a_encore

    except requests.exceptions.RequestException as e:
        print(f"      ⚠ Réseau : {e}")
        return [], False


# ── SCRAPING PAR MARCHÉ ────────────────────────────────────────────────────────

def scraper_marche(nom_marche: str, requetes: list[str]) -> pd.DataFrame:
    print(f"\n{'='*60}")
    print(f"  MARCHÉ : {nom_marche}  ({len(requetes)} requêtes)")
    print(f"{'='*60}")

    tous = []
    vus  = set()

    for idx, requete in enumerate(requetes, 1):
        print(f"\n[{idx:02d}/{len(requetes)}] {requete}")

        for page in range(MAX_PAGES_PAR_REQUETE):
            start = page * 20
            label = f"page {page+1}" if page > 0 else "page 1"
            print(f"   → {label} ...", end=" ", flush=True)

            resultats, a_encore = scrape_requete(requete, start)

            nouveaux = 0
            for h in resultats:
                cle = (h["nom"].lower().strip(),
                       h["adresse"][:30].lower().strip() if h["adresse"] else "")
                if cle not in vus and h["nom"]:
                    vus.add(cle)
                    tous.append(h)
                    nouveaux += 1

            print(f"{len(resultats)} résultats | {nouveaux} nouveaux | total {len(tous)}")

            if not a_encore or not resultats:
                break
            time.sleep(PAUSE_ENTRE_REQUETES)

        time.sleep(PAUSE_ENTRE_REQUETES)

    print(f"\n✅ {nom_marche} : {len(tous)} agences uniques")
    return pd.DataFrame(tous) if tous else pd.DataFrame()


# ── ANALYSE COMPLÉTUDE ─────────────────────────────────────────────────────────

def analyser_completude(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    df = df.copy()
    statuts, manquants, priorites = [], [], []

    for _, row in df.iterrows():
        m = []
        if not row.get("telephone"):  m.append("téléphone")
        if not row.get("email"):      m.append("email")
        if not row.get("site_web"):   m.append("site_web")
        if not row.get("adresse"):    m.append("adresse")

        if not m:
            statuts.append("✅ COMPLET"); priorites.append("A")
        elif m == ["email"]:
            statuts.append("⚠️ EMAIL MANQUANT"); priorites.append("B")
        elif "telephone" in m or "adresse" in m:
            statuts.append("❌ INCOMPLET"); priorites.append("C")
        else:
            statuts.append("⚠️ À COMPLÉTER"); priorites.append("B")

        manquants.append(", ".join(m) if m else "Aucun")

    df["statut"]           = statuts
    df["champs_manquants"] = manquants
    df["priorite"]         = priorites
    df["email_manuel"]     = ""
    df["tel_manuel"]       = ""
    df["notes"]            = ""

    ordre = {"C": 0, "B": 1, "A": 2}
    df["_o"] = df["priorite"].map(ordre)
    df = df.sort_values(["_o", "nom"]).drop(columns=["_o"]).reset_index(drop=True)
    return df


# ── EXPORT EXCEL ───────────────────────────────────────────────────────────────

COLS_AFFICHAGE = [
    "priorite", "statut", "champs_manquants",
    "nom", "adresse", "telephone", "tel_manuel",
    "email", "email_manuel", "site_web",
    "note", "nb_avis", "lien_google_maps", "notes",
]

HEADERS_FR = {
    "priorite":         "Priorité",
    "statut":           "Statut",
    "champs_manquants": "Champs manquants",
    "nom":              "Nom de l'agence",
    "adresse":          "Adresse",
    "telephone":        "Téléphone (auto)",
    "tel_manuel":       "Téléphone (manuel)",
    "email":            "Email (auto)",
    "email_manuel":     "Email (manuel)",
    "site_web":         "Site web",
    "note":             "/5",
    "nb_avis":          "Nb avis",
    "lien_google_maps": "🔗 Google Maps",
    "notes":            "Notes",
}

LARGEURS = {
    "priorite": 8, "statut": 22, "champs_manquants": 20,
    "nom": 38, "adresse": 40, "telephone": 16, "tel_manuel": 16,
    "email": 30, "email_manuel": 30, "site_web": 35,
    "note": 7, "nb_avis": 8, "lien_google_maps": 26, "notes": 30,
}


def _style_fill(couleur: str):
    from openpyxl.styles import PatternFill
    return PatternFill("solid", start_color=couleur, end_color=couleur)


def ecrire_onglet(wb, df_in: pd.DataFrame, titre: str, couleur_data: str):
    from openpyxl.styles import Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet(titre)

    fill_header = _style_fill("1F4E78")
    fill_data   = _style_fill(couleur_data)
    fill_titre  = _style_fill("1F4E78")
    thin = Side(style="thin", color="D0D0D0")
    bord = Border(bottom=thin, right=thin)

    font_h    = Font(name="Arial", bold=True,  color="FFFFFF", size=10)
    font_bold = Font(name="Arial", bold=True,  size=10)
    font_norm = Font(name="Arial",             size=10)
    font_link = Font(name="Arial", color="0563C1", underline="single", size=10)

    n_cols = len(COLS_AFFICHAGE)

    # Titre
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c = ws.cell(1, 1, f"AGENCES DE VOYAGES — {titre}")
    c.font      = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    c.fill      = fill_titre
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    if df_in.empty:
        ws.cell(2, 1, "Aucun résultat pour ce marché.").font = font_norm
        return

    df = df_in.copy()
    for col in COLS_AFFICHAGE:
        if col not in df.columns:
            df[col] = ""

    # En-têtes
    for ci, col in enumerate(COLS_AFFICHAGE, 1):
        c = ws.cell(2, ci, HEADERS_FR.get(col, col))
        c.font = font_h; c.fill = fill_header
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 38

    # Données
    for ri, (_, row) in enumerate(df[COLS_AFFICHAGE].iterrows(), 3):
        for ci, col in enumerate(COLS_AFFICHAGE, 1):
            val = row[col]
            c   = ws.cell(ri, ci)
            if col in ("lien_google_maps", "site_web") and val:
                c.value     = str(val)
                c.font      = font_link
                c.hyperlink = str(val)
            else:
                c.value = val if pd.notna(val) and val != "" else ""
                c.font  = font_bold if col == "nom" else font_norm
            c.fill      = fill_data
            c.alignment = Alignment(vertical="top",
                                    wrap_text=(col in ("adresse", "champs_manquants", "notes")))
            c.border = bord
        ws.row_dimensions[ri].height = 28

    # Largeurs colonnes
    for ci, col in enumerate(COLS_AFFICHAGE, 1):
        ws.column_dimensions[get_column_letter(ci)].width = LARGEURS.get(col, 15)

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(n_cols)}2"


def onglet_stats(wb, resultats_par_marche: dict):
    from openpyxl.styles import Font, PatternFill, Alignment
    ws = wb.create_sheet("📊 STATISTIQUES", 0)
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 14

    fill_h = _style_fill("1F4E78")
    fill_r = _style_fill("F2F2F2")
    font_w = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    font_b = Font(name="Arial", bold=True, size=11)
    font_n = Font(name="Arial", size=10)

    ws.merge_cells("A1:B1")
    c = ws.cell(1, 1, "STATISTIQUES — AGENCES VOYAGES CIBLANT PARIS")
    c.font = font_w; c.fill = fill_h
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    row = 2
    total_global = 0
    for marche, df in resultats_par_marche.items():
        if df is None or df.empty:
            continue
        nb = len(df)
        total_global += nb
        c1 = ws.cell(row, 1, f"🌍 {marche.replace('_', ' ')}")
        c2 = ws.cell(row, 2, nb)
        for c in (c1, c2):
            c.font = font_b; c.fill = fill_r
            c.alignment = Alignment(vertical="center")
        ws.row_dimensions[row].height = 22
        row += 1

    ws.cell(row, 1, "TOTAL").font = Font(name="Arial", bold=True, size=12)
    ws.cell(row, 2, total_global).font = Font(name="Arial", bold=True, size=12)
    ws.row_dimensions[row].height = 24


def onglet_mode_emploi(wb):
    from openpyxl.styles import Font, Alignment, PatternFill
    ws = wb.create_sheet("📋 MODE D'EMPLOI", 0)
    ws.column_dimensions["A"].width = 85

    fill_h = _style_fill("1F4E78")
    fill_s = _style_fill("E7E6E6")
    ws.merge_cells("A1:A1")
    c = ws.cell(1, 1, "MODE D'EMPLOI — AGENCES DE VOYAGES PARIS")
    c.font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    c.fill = fill_h
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    lignes = [
        "",
        "🗂️  STRUCTURE DU FICHIER",
        "",
        "  • Un onglet par marché géographique (USA, Chine, Japon, Corée, Inde, Singapour,",
        "    Arabie Saoudite, Émirats Arabes Unis, Qatar, Agents Indépendants USA)",
        "  • Chaque onglet contient : nom, adresse, téléphone, email, site web, note, lien Maps",
        "",
        "📋  ORDRE DE TRAITEMENT",
        "",
        "  1. Trier par colonne 'Priorité' (C = Incomplet prioritaire)",
        "  2. Cliquer sur le lien Google Maps pour accéder à la fiche complète",
        "  3. Compléter 'Téléphone (manuel)' et 'Email (manuel)' si manquants",
        "  4. Visiter le site web de l'agence pour trouver l'email de contact",
        "",
        "💡  ASTUCES EMAIL",
        "",
        "  • Chercher la page 'Contact', 'About', 'Team' ou 'Réservation'",
        "  • Utiliser Hunter.io ou FindThatEmail si le site ne l'affiche pas",
        "  • Pour les agences USA : consulter ASTA (asta.org) ou ARTA",
        "  • Pour les agences asiatiques : chercher sur Alibaba, WeChat, Kakao",
        "  • Pour le Golfe : annuaires locaux (yellow pages, kompass)",
        "",
        "📊  SOURCES DE DONNÉES",
        "",
        "  • Google Maps via SerpAPI (données publiques)",
        "  • Recherches spécifiques par ville et langue locale",
        "  • Recherches LinkedIn/Instagram pour les agents indépendants",
        "",
        f"  📅 Date du scraping : {datetime.now().strftime('%d/%m/%Y à %H:%M')}",
    ]

    for i, texte in enumerate(lignes, 2):
        c = ws.cell(i, 1, texte)
        if texte.startswith("🗂️") or texte.startswith("📋") or texte.startswith("💡") or texte.startswith("📊"):
            c.font = Font(name="Arial", bold=True, size=11)
            c.fill = fill_s
        elif texte.startswith("  •") or texte.startswith("  1") or texte.startswith("  2") or texte.startswith("  3") or texte.startswith("  4"):
            c.font = Font(name="Arial", size=10)
        elif texte.startswith("  📅"):
            c.font = Font(name="Arial", italic=True, size=10)
        else:
            c.font = Font(name="Arial", size=10)
        c.alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[i].height = 18 if texte else 8


# Couleurs par onglet
COULEURS_MARCHE = {
    "USA":                 "DEEBF7",   # bleu clair
    "Chine":               "FCE4D6",   # saumon
    "Japon":               "FFF2CC",   # jaune pâle
    "Coree_du_Sud":        "E2EFDA",   # vert pâle
    "Inde":                "F4E1D2",   # orange pâle
    "Singapour":           "D9EAD3",   # vert clair
    "Arabie_Saoudite":     "EAD1DC",   # rose
    "Emirats_Arabes_Unis": "D0E4F5",   # bleu pâle
    "Qatar":               "F5E6FA",   # lavande
    "Agents_Independants_USA": "E8F5E9",  # menthe
}


def export_excel(resultats_par_marche: dict, fichier: str):
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)

    onglet_mode_emploi(wb)
    onglet_stats(wb, resultats_par_marche)

    for marche, df in resultats_par_marche.items():
        if df is None or df.empty:
            print(f"   ⚠ Onglet '{marche}' vide — ignoré")
            continue
        couleur = COULEURS_MARCHE.get(marche, "F2F2F2")
        titre   = marche.replace("_", " ")
        ecrire_onglet(wb, df, titre, couleur)
        print(f"   ✅ Onglet '{titre}' — {len(df)} agences")

    wb.save(fichier)
    total = sum(len(df) for df in resultats_par_marche.values() if df is not None and not df.empty)
    print(f"\n✅ Fichier créé : {fichier}")
    print(f"📊 Total : {total} agences")


# ── MAIN ───────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("\n" + "="*65)
    print("  SCRAPER — AGENCES DE VOYAGES PARIS (USA / ASIE / GOLFE)")
    print("="*65)
    print(f"\n⚙️  Clé SerpAPI : {SERP_API_KEY[:10]}...")
    print(f"   Pause entre requêtes : {PAUSE_ENTRE_REQUETES}s")
    print(f"   Pages max/requête   : {MAX_PAGES_PAR_REQUETE}")
    print(f"\n   Marchés à scraper  : {list(MARCHES.keys())}\n")

    rep = input("▶️  Lancer le scraping ? (o/n) : ").strip().lower()
    if rep != "o":
        print("Annulé.")
        exit()

    resultats_par_marche = {}

    for marche, requetes in MARCHES.items():
        df = scraper_marche(marche, requetes)
        if not df.empty:
            df = analyser_completude(df)
        resultats_par_marche[marche] = df

    fichier = f"agences_voyages_paris_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    export_excel(resultats_par_marche, fichier)

    print("\n" + "="*65)
    print("✅ SCRAPING TERMINÉ")
    print("="*65)
    print(f"\n📁 Fichier : {fichier}")
    print("\n📋 PROCHAINES ÉTAPES :")
    print("   1. Ouvrir le fichier Excel")
    print("   2. Travailler marché par marché (un onglet = un marché)")
    print("   3. Compléter emails/téléphones manquants via Google Maps")
    print("   4. Pour les emails : visiter le site web de chaque agence")