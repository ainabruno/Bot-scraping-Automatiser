"""
Scraper Dentalemploi — MedicoJob / Centres dentaires Paris IDF
Usage : python scraper_dentalemploi.py
Nécessite : pip install playwright openpyxl
            python -m playwright install chromium
"""

import asyncio
from datetime import datetime
from playwright.async_api import async_playwright
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── CONFIG ──────────────────────────────────────────────────────────────────
BASE_URL   = "https://dentalemploi.com/offres-emploi-dentaire"
PARAMS     = "?metier_id=17366&searchedAddress=paris&city=paris"
TOTAL_PAGES = 1
OUTPUT     = "dentalemploi_offres1.xlsx"
HEADLESS   = True   # Passer à False pour voir le navigateur (debug)
# ─────────────────────────────────────────────────────────────────────────────


def get_dept(loc):
    if "75" in loc: return "75 - Paris"
    if "92" in loc: return "92 - Hauts-de-Seine"
    if "93" in loc: return "93 - Seine-Saint-Denis"
    if "94" in loc: return "94 - Val-de-Marne"
    if "77" in loc: return "77 - Seine-et-Marne"
    if "78" in loc: return "78 - Yvelines"
    if "91" in loc: return "91 - Essonne"
    if "95" in loc: return "95 - Val-d'Oise"
    return "Autre"


def classify_priority(societe, resume):
    s = societe.lower()
    if any(x in s for x in ["centre dentaire", "cpam", "mutualit", "dentelia",
                              "dentego", "dentylis", "samsic", "cabinet dentaire"]):
        return "A"
    if any(x in s for x in ["dn8", "advisory", "reseau talents", "capijob",
                              "solution médic", "annonces médic"]):
        return "B"
    return "B"


async def scrape_page(page, page_num):
    url = f"{BASE_URL}{PARAMS}&page={page_num}"
    print(f"  Page {page_num} → {url}")
    await page.goto(url, wait_until="networkidle", timeout=45000)

    # Attente adaptative — certains sites chargent en JS
    try:
        await page.wait_for_selector(".offers_emploi_item", timeout=20000)
    except Exception:
        print(f"    Sélecteur non trouvé page {page_num} — tentative récupération HTML brut")

    items = await page.query_selector_all(".offers_emploi_item")
    if not items:
        print(f"    Aucune offre détectée page {page_num}")
        return []

    results = []
    for item in items:
        async def txt(sel):
            el = await item.query_selector(sel)
            return (await el.inner_text()).strip() if el else ""

        contrat   = await txt(".offers_item_contract_type")
        posted    = (await txt(".offers_item_posted_in")).replace("Posté", "").strip()
        titre     = await txt("h2")
        loc       = await txt(".offers_item_localisation")
        societe   = await txt(".offers_item_society_name")
        resume    = await txt(".offers_item_resume")
        lien_el   = await item.query_selector("a.btn")
        lien      = await lien_el.get_attribute("href") if lien_el else ""

        results.append({
            "Date extraction":        datetime.today().strftime("%d/%m/%Y"),
            "Source":                 "Dentalemploi",
            "Lien source":            url,
            "Page":                   page_num,
            "Nom du centre / cabinet": societe,
            "Type":                   "centre" if any(x in societe.lower() for x in ["centre","cpam","mutualit","samsic"]) else "cabinet",
            "Titre poste":            titre,
            "Type contrat":           contrat,
            "Ville":                  loc.split(",")[0].strip() if loc else "",
            "Département":            get_dept(loc),
            "Date de l'offre":        posted,
            "Résumé offre":           resume,
            "Lien offre":             lien,
            "Nom du contact":         "",
            "Fonction contact":       "",
            "Téléphone contact":      "",
            "Email / formulaire":     "",
            "Urgence":                "forte" if "1 semaine" in posted else ("moyenne" if "2 semaine" in posted else "faible"),
            "Priorité A/B/C":         classify_priority(societe, resume),
            "Commentaire":            "",
        })

    return results


def build_excel(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Offres centres"
    headers = list(rows[0].keys())

    RED = "C0392B"
    header_fill = PatternFill("solid", start_color=RED, end_color=RED)
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin   = Side(style="thin", color="E5E5E5")
    brd    = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = header_font; c.fill = header_fill
        c.alignment = center; c.border = brd
    ws.row_dimensions[1].height = 32

    fill_w = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
    fill_l = PatternFill("solid", start_color="FEF9F9", end_color="FEF9F9")
    fill_a = PatternFill("solid", start_color="FDEDEC", end_color="FDEDEC")
    prio_ci = headers.index("Priorité A/B/C") + 1

    for ri, row in enumerate(rows, 2):
        prio = row.get("Priorité A/B/C", "")
        fill = fill_a if prio == "A" else (fill_w if ri % 2 == 0 else fill_l)
        for ci, key in enumerate(headers, 1):
            c = ws.cell(row=ri, column=ci, value=row[key])
            c.font = Font(name="Arial", size=9)
            c.fill = fill; c.border = brd; c.alignment = left
        # Color priority cell
        pc = ws.cell(row=ri, column=prio_ci)
        pc.alignment = Alignment(horizontal="center", vertical="center")
        if prio == "A":
            pc.font = Font(name="Arial", size=9, bold=True, color="922B21")
        elif prio == "B":
            pc.font = Font(name="Arial", size=9, bold=True, color="7D6608")
        ws.row_dimensions[ri].height = 28

    col_widths = {
        "Date extraction": 13, "Source": 12, "Lien source": 18, "Page": 6,
        "Nom du centre / cabinet": 30, "Type": 10, "Titre poste": 26,
        "Type contrat": 10, "Ville": 22, "Département": 22,
        "Date de l'offre": 14, "Résumé offre": 52, "Lien offre": 50,
        "Nom du contact": 20, "Fonction contact": 18,
        "Téléphone contact": 16, "Email / formulaire": 24,
        "Urgence": 10, "Priorité A/B/C": 12, "Commentaire": 28,
    }
    for ci, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(h, 15)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # Résumé
    ws2 = wb.create_sheet("Résumé")
    ws2["A1"] = "Résumé extraction — Dentalemploi / MedicoJob"
    ws2["A1"].font = Font(bold=True, name="Arial", size=13, color="C0392B")
    summary = [
        ("Date d'extraction",    datetime.today().strftime("%d/%m/%Y")),
        ("Source",               "Dentalemploi.com"),
        ("Recherche",            "Chirurgien-dentiste Paris / IDF"),
        ("Pages scrapées",       TOTAL_PAGES),
        ("Total offres",         len(rows)),
        ("Paris 75",             sum(1 for r in rows if "75" in r["Département"])),
        ("Priorité A",           sum(1 for r in rows if r["Priorité A/B/C"] == "A")),
        ("Priorité B",           sum(1 for r in rows if r["Priorité A/B/C"] == "B")),
        ("Offres urgentes",      sum(1 for r in rows if r["Urgence"] == "forte")),
    ]
    for i, (k, v) in enumerate(summary, 3):
        ws2[f"A{i}"] = k; ws2[f"B{i}"] = v
        ws2[f"A{i}"].font = Font(bold=True, name="Arial", size=10)
        ws2[f"B{i}"].font = Font(name="Arial", size=10)
    ws2.column_dimensions["A"].width = 26
    ws2.column_dimensions["B"].width = 50

    wb.save(OUTPUT)
    print(f"\nFichier sauvegardé : {OUTPUT}")
    print(f"Total offres : {len(rows)}")
    print(f"Priorité A   : {sum(1 for r in rows if r['Priorité A/B/C'] == 'A')}")
    print(f"Priorité B   : {sum(1 for r in rows if r['Priorité A/B/C'] == 'B')}")


async def main():
    all_data = []
    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=HEADLESS,
            args=["--no-sandbox", "--disable-blink-features=AutomationControlled"]
        )
        context = await browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
            viewport={"width": 1280, "height": 900},
            locale="fr-FR",
        )
        # Masquer la détection Playwright
        await context.add_init_script("""
            Object.defineProperty(navigator, 'webdriver', {get: () => undefined})
        """)
        page = await context.new_page()

        for i in range(1, TOTAL_PAGES + 1):
            try:
                data = await scrape_page(page, i)
                all_data.extend(data)
                print(f"    → {len(data)} offres récupérées")
                await asyncio.sleep(2)
            except Exception as e:
                print(f"    Erreur page {i}: {e}")

        await browser.close()

    print(f"\nTotal : {len(all_data)} offres")
    if all_data:
        build_excel(all_data)
    else:
        print("Aucune donnée — le site a peut-être bloqué le scraping.")
        print("Essaie avec HEADLESS = False pour voir ce qui se passe.")


if __name__ == "__main__":
    asyncio.run(main())