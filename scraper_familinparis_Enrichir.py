import json
import os
import re
import time
import urllib.parse
import urllib.request
from datetime import datetime
from pathlib import Path

import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from playwright.sync_api import sync_playwright

# ── CONFIGURATION ──────────────────────────────────────────────────────────────

INPUT_FILE      = "familinparis_activites_enrichi.xlsx"
OUTPUT_FILE     = "familinparis_enrichi_contacts.xlsx"
CHECKPOINT_F    = "checkpoint_contacts.json"

# Connexion à un vrai Chrome (lancé avec --remote-debugging-port=9222).
# Si CDP_URL est joignable, on s'y connecte (anti-blocage Google).
# Sinon on retombe sur un Chromium headless lancé par Playwright.
CDP_URL         = "http://127.0.0.1:9222"

PAUSE           = 1.5    # entre appels Google (maps / search)
PAUSE_SCRAPE    = 0.8    # entre scrapes web
REQUEST_TIMEOUT = 20
CHECKPOINT_EVERY= 50

HEADERS_WEB = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "fr-FR,fr;q=0.9,en;q=0.8",
    "Accept": "text/html,application/xhtml+xml,*/*;q=0.8",
}

UA_PLAYWRIGHT = HEADERS_WEB["User-Agent"]

# ── REGEX ──────────────────────────────────────────────────────────────────────

PHONE_RE = re.compile(
    r"""(?:(?:\+|00)\s*\d{1,3}[\s.\-]?)?"""   # préfixe international
    r"""(?:\(?\d{1,4}\)?[\s.\-]?)?"""           # indicatif zone
    r"""(?:\d[\s.\-]?){7,12}\d""",
    re.VERBOSE,
)

# Numéros français : 0X XX XX XX XX ou +33 X XX XX XX XX
PHONE_FR_RE = re.compile(
    r"""(?:(?:\+33|0033)[\s.\-]?[1-9](?:[\s.\-]?\d{2}){4}"""   # +33 / 0033
    r"""|0[1-9](?:[\s.\-]?\d{2}){4})""",                          # 0X…
    re.VERBOSE,
)

EMAIL_RE = re.compile(
    r"\b[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}\b"
)

BLACKLIST_EMAIL_DOMAINS = {
    "example.com", "test.com", "domain.com", "email.com",
    "wixpress.com", "sentry.io", "wordpress.org",
    "schema.org", "google.com", "facebook.com",
    "w3.org", "acquia.com", "amazonaws.com",
}
BLACKLIST_EMAIL_PATTERNS = ["pixel", "track", "sentry", "analytics", "noreply", "no-reply"]

RE_TAGS  = re.compile(r'<[^>]+>')
RE_NBSP  = re.compile(r'&nbsp;|&#160;', re.IGNORECASE)
RE_ENTIT = re.compile(r'&\w+;')

PAGES_CONTACT = [
    "contact", "nous-contacter", "contactez-nous", "contact-us",
    "a-propos", "about", "informations", "equipe", "qui-sommes-nous",
]

# ── UTILITAIRES TEXTE ──────────────────────────────────────────────────────────

def strip_html(s: str) -> str:
    import html as html_mod
    s = html_mod.unescape(s)
    return " ".join(RE_TAGS.sub(" ", s).split()).strip()


def nettoyer_telephone(tel: str) -> str:
    """Normalise un numéro de téléphone — retourne "" si invalide."""
    if not tel:
        return ""
    # Garder chiffres, +, espaces, tirets, points, parenthèses
    t = re.sub(r"[^\d+\s.\-()]", "", tel).strip()
    digits_only = re.sub(r"[^\d]", "", t)
    return t if 7 <= len(digits_only) <= 15 else ""


def nettoyer_email(email: str) -> str:
    if not email:
        return ""
    e = email.lower().strip()
    domain = e.split("@")[-1]
    if domain in BLACKLIST_EMAIL_DOMAINS:
        return ""
    if any(pat in domain for pat in BLACKLIST_EMAIL_PATTERNS):
        return ""
    return e


def extraire_emails(texte: str) -> list[str]:
    found = [nettoyer_email(e) for e in EMAIL_RE.findall(texte or "")]
    return list(dict.fromkeys(e for e in found if e))


def extraire_telephones(texte: str) -> list[str]:
    found = []
    for pat in (PHONE_FR_RE, PHONE_RE):
        for m in pat.finditer(texte or ""):
            t = nettoyer_telephone(m.group(0))
            if t:
                found.append(t)
    return list(dict.fromkeys(found))


def nettoyer_url_base(url: str) -> str:
    if not url:
        return ""
    if not url.startswith("http"):
        url = "https://" + url
    p = urllib.parse.urlparse(url)
    return f"{p.scheme}://{p.netloc}"


def cellval(x) -> str:
    """Convertit une valeur de cellule (NaN inclus) en chaîne propre."""
    if x is None:
        return ""
    try:
        if pd.isna(x):
            return ""
    except (TypeError, ValueError):
        pass
    s = str(x).strip()
    if s.lower() in ("nan", "none", "null", "<na>"):
        return ""
    return s


def ville_depuis_adresse(adresse: str) -> str:
    """Extrait la partie 'ville' d'une adresse (dernier segment non-CP)."""
    if not adresse:
        return ""
    for part in reversed([p.strip() for p in adresse.split(",")]):
        if part and not re.fullmatch(r"\d{5}", part):
            return part
    return ""


# ── COUCHE 0 : Scraping page familinparis.fr ──────────────────────────────────

RE_POST_META = re.compile(
    r'<ul[^>]*class=["\'][^"\']*post-meta[^"\']*["\'][^>]*>(.*?)</ul>',
    re.DOTALL | re.IGNORECASE,
)
RE_LI = re.compile(r'<li[^>]*>(.*?)</li>', re.DOTALL | re.IGNORECASE)

RE_KEY = re.compile(
    r'^(adresse|lieu|t[ée]l[ée]phone?|t[ée]l\.?|email|e-mail|site\s*web?|web|contact)\s*[:\u00a0]?\s*',
    re.IGNORECASE,
)

# Liens tel: et mailto: dans le HTML
RE_TEL_HREF    = re.compile(r'href=["\']tel:([^"\']+)["\']',    re.IGNORECASE)
RE_MAILTO_HREF = re.compile(r'href=["\']mailto:([^"\'?\s]+)',   re.IGNORECASE)

# JSON-LD
RE_JSONLD = re.compile(
    r'<script[^>]+type=["\']application/ld\+json["\'][^>]*>(.*?)</script>',
    re.DOTALL | re.IGNORECASE,
)


def _scrape_post_meta(raw_html: str) -> dict:
    """Extrait les champs depuis <ul class='post-meta'>."""
    result = {}
    m = RE_POST_META.search(raw_html)
    if not m:
        return result
    for li_m in RE_LI.finditer(m.group(1)):
        text = strip_html(li_m.group(1))
        key_m = RE_KEY.match(text)
        if not key_m:
            continue
        key   = key_m.group(1).lower().strip().rstrip(".")
        value = text[key_m.end():].strip()
        if not value:
            continue

        if re.match(r't[ée]l', key, re.IGNORECASE):
            t = nettoyer_telephone(value)
            if t:
                result.setdefault("telephone", t)
        elif "email" in key or "e-mail" in key:
            e = nettoyer_email(value)
            if e:
                result.setdefault("email", e)
        elif "site" in key or "web" in key:
            result.setdefault("site_web", value)
        elif key == "adresse":
            result.setdefault("adresse", value)
        elif key == "lieu":
            result.setdefault("lieu", value)

    return result


def _scrape_jsonld(raw_html: str) -> dict:
    """Extrait téléphone / email / url depuis les blocs JSON-LD."""
    result = {}
    for m in RE_JSONLD.finditer(raw_html):
        try:
            data = json.loads(m.group(1).strip())
        except Exception:
            continue
        items = data if isinstance(data, list) else [data]
        for item in items:
            if not isinstance(item, dict):
                continue
            if not result.get("telephone"):
                t = nettoyer_telephone(str(item.get("telephone", "")))
                if t:
                    result["telephone"] = t
            if not result.get("email"):
                e = nettoyer_email(str(item.get("email", "")))
                if e:
                    result["email"] = e
            if not result.get("site_web"):
                url = item.get("url", item.get("sameAs", ""))
                if isinstance(url, list):
                    url = url[0] if url else ""
                if url and "familinparis" not in str(url):
                    result["site_web"] = str(url)
    return result


def _scrape_hrefs(raw_html: str) -> dict:
    """Extrait téléphone et email depuis href='tel:' et href='mailto:'."""
    result = {}
    m_tel = RE_TEL_HREF.search(raw_html)
    if m_tel:
        t = nettoyer_telephone(m_tel.group(1))
        if t:
            result["telephone"] = t
    m_mail = RE_MAILTO_HREF.search(raw_html)
    if m_mail:
        e = nettoyer_email(m_mail.group(1))
        if e:
            result["email"] = e
    return result


def _scrape_page_texte(raw_html: str) -> dict:
    """Cherche emails et téléphones dans tout le texte de la page."""
    result = {}
    # Supprimer scripts/styles d'abord
    clean = re.sub(r'<(script|style)[^>]*>.*?</\1>', '', raw_html, flags=re.DOTALL|re.IGNORECASE)
    texte = strip_html(clean)

    emails = extraire_emails(texte)
    if emails:
        result["email"] = emails[0]

    tels = extraire_telephones(texte)
    if tels:
        result["telephone"] = tels[0]

    return result


def scrape_familinparis(url: str, session: requests.Session) -> dict:
    """
    Scrape une page familinparis.fr pour extraire tel, email, site_web.
    Fusionne les résultats de toutes les stratégies.
    """
    result = {}
    try:
        r = session.get(url, headers=HEADERS_WEB, timeout=REQUEST_TIMEOUT,
                        allow_redirects=True)
        if r.status_code != 200:
            return {"_erreur": f"HTTP {r.status_code}"}
        raw_html = r.text

        # Fusionne dans l'ordre de fiabilité
        for extractor in [_scrape_post_meta, _scrape_jsonld, _scrape_hrefs, _scrape_page_texte]:
            partial = extractor(raw_html)
            for k, v in partial.items():
                if v and not result.get(k):
                    result[k] = v

        # Si site_web == url familinparis, on ignore
        site = result.get("site_web", "")
        if site and "familinparis.fr" in site:
            result.pop("site_web", None)

    except Exception as e:
        result["_erreur"] = str(e)[:80]

    return result


# ── COUCHE 1 : Scraping du site officiel ──────────────────────────────────────

def scrape_site_officiel(url_base: str, session: requests.Session) -> dict:
    """
    Visite la page d'accueil puis /contact du site officiel.
    Retourne : {"telephone": ..., "email": ...}
    """
    result = {}
    if not url_base:
        return result

    def _fetch(url: str) -> dict:
        try:
            r = session.get(url, headers=HEADERS_WEB, timeout=REQUEST_TIMEOUT,
                            allow_redirects=True)
            if r.status_code != 200:
                return {}
            raw = r.text
            partial = {}
            for extractor in [_scrape_hrefs, _scrape_jsonld, _scrape_page_texte]:
                for k, v in extractor(raw).items():
                    if v and not partial.get(k):
                        partial[k] = v
            return partial
        except Exception:
            return {}

    # Page d'accueil
    result = _fetch(url_base)
    if result.get("telephone") and result.get("email"):
        return result
    time.sleep(PAUSE_SCRAPE)

    # Pages contact
    for page in PAGES_CONTACT:
        if result.get("telephone") and result.get("email"):
            break
        contact_url = f"{url_base.rstrip('/')}/{page}"
        partial = _fetch(contact_url)
        for k, v in partial.items():
            if v and not result.get(k):
                result[k] = v
        time.sleep(PAUSE_SCRAPE * 0.4)

    return result


# ── COUCHE 2 : Google Maps via Playwright ─────────────────────────────────────

DOMAINES_EXCLUS = {
    "funbooker", "familinparis", "tripadvisor", "facebook", "instagram",
    "yelp", "google", "youtube", "linkedin", "twitter", "booking",
    "viator", "getyourguide", "billetreduc", "billetweb", "weezevent",
}


def playwright_maps(nom: str, adresse: str, page) -> dict:
    """
    Cherche '<nom> <ville>' sur Google Maps, ouvre la première fiche
    et extrait téléphone / site web / note / nb avis / coordonnées.
    """
    ville = ville_depuis_adresse(adresse)
    query = f"{nom} {ville}".strip() if ville else nom

    try:
        url = "https://www.google.com/maps/search/" + urllib.parse.quote(query)
        page.goto(url, timeout=REQUEST_TIMEOUT * 1000, wait_until="domcontentloaded")
        page.wait_for_timeout(2500)

        # Si plusieurs résultats : cliquer sur le premier de la liste
        try:
            first_link = page.locator('a.hfpxzc').first
            if first_link.count() > 0:
                first_link.click(timeout=5000)
                page.wait_for_timeout(2000)
        except Exception:
            pass

        # Le panneau de détail est dans un div role="main"
        try:
            page.wait_for_selector('div[role="main"]', timeout=8000)
        except Exception:
            pass

        result = {}

        # Téléphone
        try:
            tel_el = page.locator('button[data-item-id^="phone:"]').first
            if tel_el.count() > 0:
                aria = tel_el.get_attribute("aria-label") or ""
                tel = nettoyer_telephone(aria)
                if not tel:
                    tel = nettoyer_telephone(tel_el.inner_text())
                if tel:
                    result["telephone"] = tel
        except Exception:
            pass

        # Site web
        try:
            web_el = page.locator('a[data-item-id="authority"]').first
            if web_el.count() > 0:
                href = web_el.get_attribute("href")
                if href:
                    result["site_web"] = nettoyer_url_base(href)
        except Exception:
            pass

        # Adresse
        try:
            addr_el = page.locator('button[data-item-id="address"]').first
            if addr_el.count() > 0:
                aria = addr_el.get_attribute("aria-label") or ""
                result["adresse_maps"] = re.sub(r'^Adresse\s*:\s*', '', aria, flags=re.IGNORECASE).strip()
        except Exception:
            pass

        # Titre de la fiche
        try:
            title_el = page.locator('h1').first
            if title_el.count() > 0:
                result["nom_maps"] = title_el.inner_text().strip()
        except Exception:
            pass

        # Note + nombre d'avis
        try:
            rating_el = page.locator('div[role="main"] span[role="img"]').first
            if rating_el.count() > 0:
                aria = rating_el.get_attribute("aria-label") or ""
                m = re.search(r'([\d,.]+)\s*étoile.*?([\d\s]+)\s*avis', aria, re.IGNORECASE)
                if m:
                    result["note"] = m.group(1).replace(",", ".")
                    result["nb_avis"] = re.sub(r'\s', '', m.group(2))
        except Exception:
            pass

        # place_id / maps_url via l'URL actuelle
        try:
            cur_url = page.url
            result["maps_url"] = cur_url
            m = re.search(r'!1s([^!]+)', cur_url)
            if m:
                result["place_id"] = m.group(1)
        except Exception:
            pass

        return result

    except Exception as e:
        return {"_erreur": str(e)[:120]}


# ── COUCHE 3 : Google Search via Playwright → site officiel ──────────────────

def playwright_google_site(nom: str, adresse: str, page) -> dict:
    ville = ville_depuis_adresse(adresse)
    query = f"{nom} {ville} site officiel contact".strip()

    try:
        url = "https://www.google.com/search?q=" + urllib.parse.quote(query) + "&hl=fr&gl=fr&num=10"
        page.goto(url, timeout=REQUEST_TIMEOUT * 1000, wait_until="domcontentloaded")
        page.wait_for_timeout(1500)

        # Récupérer les liens des résultats organiques
        links = page.locator('div#search a[href^="http"]')
        n = links.count()
        for i in range(min(n, 15)):
            href = links.nth(i).get_attribute("href") or ""
            if not href.startswith("http"):
                continue
            dom = urllib.parse.urlparse(href).netloc.lower()
            if any(exclu in dom for exclu in DOMAINES_EXCLUS):
                continue
            if "google.com" in dom:
                continue
            return {"site_web": nettoyer_url_base(href), "source": href}

        return {}

    except Exception as e:
        return {"_erreur": str(e)[:120]}


# ── COUCHE 5 : Google Search via Playwright → email dans snippets ────────────

def playwright_google_email(nom: str, adresse: str, page) -> dict:
    ville = ville_depuis_adresse(adresse)
    query = f"{nom} {ville} email contact réservation".strip()

    try:
        url = "https://www.google.com/search?q=" + urllib.parse.quote(query) + "&hl=fr&gl=fr&num=10"
        page.goto(url, timeout=REQUEST_TIMEOUT * 1000, wait_until="domcontentloaded")
        page.wait_for_timeout(1500)

        # Récupérer tout le texte visible de la zone de résultats
        try:
            texte = page.locator("div#search").inner_text()
        except Exception:
            texte = page.locator("body").inner_text()

        emails = extraire_emails(texte)
        if emails:
            return {"email": emails[0]}

        return {}

    except Exception as e:
        return {"_erreur": str(e)[:120]}


# ── ORCHESTRATEUR PAR LIGNE ────────────────────────────────────────────────────

def enrichir_ligne(row: dict, session: requests.Session, page, idx: int) -> dict:
    """
    Multi-couche avec court-circuit : dès qu'on a tel+email+site, on s'arrête.
    """
    nom     = cellval(row.get("Nom", ""))
    adresse = cellval(row.get("Adresse", ""))
    url_fp  = cellval(row.get("URL", ""))   # URL familinparis de l'activité

    # Valeurs déjà présentes (ne pas écraser si déjà remplies)
    result = {
        "telephone":    cellval(row.get("Telephone", "")),
        "email":        cellval(row.get("Email", "")),
        "site_web":     cellval(row.get("Site web", "")),
        "note":         cellval(row.get("Note", "")),
        "nb_avis":      cellval(row.get("Nb avis", "")),
        "maps_url":     cellval(row.get("Maps URL", "")),
        "nom_maps":     cellval(row.get("Nom Maps", "")),
        "adresse_maps": cellval(row.get("Adresse Maps", "")),
        "methodes":     [],
        "statut":       "",
    }

    def _complet():
        return bool(result["telephone"] and result["email"] and result["site_web"])

    def _merge(partial: dict, methode: str):
        """Fusionne sans écraser les valeurs déjà remplies."""
        changed = False
        for k in ("telephone", "email", "site_web", "note", "nb_avis",
                   "maps_url", "nom_maps", "adresse_maps"):
            if partial.get(k) and not result.get(k):
                result[k] = partial[k]
                changed = True
        if changed:
            result["methodes"].append(methode)

    # ── C0 : Scraping page familinparis ──────────────────────────────────────
    if url_fp and not _complet():
        print(f"  [{idx}] C0-scrape familinparis...")
        c0 = scrape_familinparis(url_fp, session)
        if "_erreur" not in c0:
            _merge(c0, "C0-familinparis")
        time.sleep(PAUSE_SCRAPE)

    # ── C1 : Scraping du site officiel (si on a déjà un site_web) ────────────
    if result["site_web"] and not _complet():
        print(f"  [{idx}] C1-scrape site officiel : {result['site_web'][:45]}...")
        c1 = scrape_site_officiel(result["site_web"], session)
        _merge(c1, "C1-site-officiel")

    # ── C2 : Google Maps (Playwright) ─────────────────────────────────────────
    if not _complet():
        print(f"  [{idx}] C2-Maps (playwright) : {nom[:40]}...")
        c2 = playwright_maps(nom, adresse, page)
        time.sleep(PAUSE)
        if "_erreur" not in c2:
            _merge(c2, "C2-maps")

    # ── C1 bis : Scraping site officiel trouvé par Maps ───────────────────────
    if result["site_web"] and "C1-site-officiel" not in result["methodes"] and not _complet():
        print(f"  [{idx}] C1b-scrape site maps : {result['site_web'][:45]}...")
        c1b = scrape_site_officiel(result["site_web"], session)
        _merge(c1b, "C1b-site-maps")

    # ── C3 : Google Search (Playwright) → site_web ────────────────────────────
    if not result["site_web"] and not _complet():
        print(f"  [{idx}] C3-Google site (playwright) : {nom[:40]}...")
        c3 = playwright_google_site(nom, adresse, page)
        time.sleep(PAUSE)
        if "_erreur" not in c3:
            _merge(c3, "C3-google-site")

    # ── C4 : Scraping du site trouvé en C3 ───────────────────────────────────
    if result["site_web"] and "C1" not in " ".join(result["methodes"]) and not result.get("email"):
        print(f"  [{idx}] C4-scrape nouveau site : {result['site_web'][:45]}...")
        c4 = scrape_site_officiel(result["site_web"], session)
        _merge(c4, "C4-site-c3")

    # ── C5 : Google Search (Playwright) → email dans snippets ────────────────
    if not result["email"]:
        print(f"  [{idx}] C5-Google email (playwright) : {nom[:40]}...")
        c5 = playwright_google_email(nom, adresse, page)
        time.sleep(PAUSE)
        if "_erreur" not in c5:
            _merge(c5, "C5-google-email")

    # ── Statut final ──────────────────────────────────────────────────────────
    manquants = [
        f for f, k in [("tel", "telephone"), ("email", "email"), ("site", "site_web")]
        if not result[k]
    ]
    if not manquants:
        result["statut"] = "✅ COMPLET"
    elif len(manquants) == 1:
        result["statut"] = f"⚠️ manque: {manquants[0]}"
    else:
        result["statut"] = f"❌ manque: {', '.join(manquants)}"

    result["methode"] = " + ".join(result["methodes"]) or "—"
    return result


# ── CHECKPOINT ────────────────────────────────────────────────────────────────

# Chemins Chrome courants sur Windows (testés dans l'ordre)
CHROME_PATHS = [
    r"C:\Program Files\Google\Chrome\Application\chrome.exe",
    r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
    os.path.expandvars(r"%LOCALAPPDATA%\Google\Chrome\Application\chrome.exe"),
]
CHROME_USER_DATA_DIR = r"C:\Temp\gyg_debug"
CHROME_LAUNCH_WAIT   = 15  # secondes max d'attente après lancement


def trouver_chrome_exe() -> str:
    for path in CHROME_PATHS:
        if path and Path(path).exists():
            return path
    return ""


def lancer_chrome_debug() -> bool:
    """
    Lance Chrome avec --remote-debugging-port=9222 (équivalent du .bat).
    Si un Chrome de debug tourne déjà sur ce profil/port, on le réutilise.
    Retourne True si le port devient joignable.
    """
    import subprocess

    chrome_exe = trouver_chrome_exe()
    if not chrome_exe:
        print("⚠️  chrome.exe introuvable automatiquement.")
        return False

    Path(CHROME_USER_DATA_DIR).mkdir(parents=True, exist_ok=True)

    print(f"🚀 Lancement de Chrome (debug port 9222) : {chrome_exe}")
    try:
        subprocess.Popen(
            [
                chrome_exe,
                "--remote-debugging-port=9222",
                f'--user-data-dir={CHROME_USER_DATA_DIR}',
                "--no-first-run",
                "--no-default-browser-check",
            ],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
        )
    except Exception as e:
        print(f"⚠️  Impossible de lancer Chrome : {e}")
        return False

    # Attendre que le port 9222 réponde
    for _ in range(CHROME_LAUNCH_WAIT):
        if chrome_cdp_disponible():
            print("✅ Chrome de debug prêt.\n")
            return True
        time.sleep(1)

    print("⚠️  Chrome lancé mais port 9222 non joignable après attente.")
    return False


def chrome_cdp_disponible(max_attempts: int = 1) -> bool:
    """Vérifie si un vrai Chrome est lancé avec --remote-debugging-port=9222."""
    for i in range(max_attempts):
        try:
            urllib.request.urlopen(f"{CDP_URL}/json/version", timeout=2)
            return True
        except Exception:
            if i < max_attempts - 1:
                time.sleep(2)
    return False


def ouvrir_navigateur(p):
    """
    Tente de se connecter à un vrai Chrome (CDP, anti-blocage Google).
    Si indisponible, lance un Chromium headless intégré à Playwright.
    Retourne (browser, context, page, via_cdp).
    """
    if not chrome_cdp_disponible():
        lancer_chrome_debug()

    if chrome_cdp_disponible():
        print(f"✅ Vrai Chrome détecté sur {CDP_URL} (mode CDP)\n")
        browser = p.chromium.connect_over_cdp(CDP_URL)
        context = (browser.contexts[0] if browser.contexts
                   else browser.new_context(
                       user_agent=UA_PLAYWRIGHT,
                       locale="fr-FR",
                       viewport={"width": 1280, "height": 900},
                   ))
        page = context.new_page()
        return browser, context, page, True

    print("⚠️  Pas de Chrome CDP sur 9222 → Chromium headless intégré.")
    print("   (Pour utiliser ton vrai Chrome : lance-le avec")
    print('    chrome.exe --remote-debugging-port=9222 --user-data-dir=C:\\ChromeBot)\n')
    browser = p.chromium.launch(headless=True)
    context = browser.new_context(
        user_agent=UA_PLAYWRIGHT,
        locale="fr-FR",
        viewport={"width": 1280, "height": 900},
    )
    page = context.new_page()
    return browser, context, page, False


def charger_checkpoint() -> dict:
    if Path(CHECKPOINT_F).exists():
        try:
            with open(CHECKPOINT_F, "r", encoding="utf-8") as f:
                data = json.load(f)
            data = sanitize_checkpoint(data)
            print(f"✅ Checkpoint : {len(data)} lignes déjà traitées")
            return data
        except Exception:
            pass
    return {}


def sanitize_checkpoint(data: dict) -> dict:
    """
    Nettoie un checkpoint pollué par un ancien bug pandas (valeurs "nan"
    littérales). Toute entrée contenant "nan"/"none"/"" dans telephone,
    email ou site_web pour un statut "DÉJÀ COMPLET" est invalidée pour
    être retraitée.
    """
    cleaned = {}
    for num, res in data.items():
        if not isinstance(res, dict):
            continue
        res2 = {}
        for k, v in res.items():
            if isinstance(v, str) and v.strip().lower() in ("nan", "none", "null", "<na>"):
                res2[k] = ""
            else:
                res2[k] = v
        # Si plus rien d'utile (tout vide après nettoyage), on retire
        # l'entrée pour qu'elle soit retraitée.
        if not (res2.get("telephone") or res2.get("email") or res2.get("site_web")):
            continue
        cleaned[num] = res2
    return cleaned


def sauvegarder_checkpoint(data: dict):
    with open(CHECKPOINT_F, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


# ── EXPORT EXCEL ──────────────────────────────────────────────────────────────

def export_excel(df: pd.DataFrame, fichier: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Activités enrichies"

    COLS = [
        "#", "Nom", "Adresse", "Telephone", "Site web", "Email",
        "Note", "Nb avis", "Maps URL", "Nom Maps", "Adresse Maps",
        "Méthode", "Statut", "URL",
    ]
    LARGEURS = {
        "#": 5, "Nom": 45, "Adresse": 38, "Telephone": 16,
        "Site web": 35, "Email": 32, "Note": 7, "Nb avis": 8,
        "Maps URL": 28, "Nom Maps": 32, "Adresse Maps": 35,
        "Méthode": 28, "Statut": 22, "URL": 55,
    }

    fill_h    = PatternFill("solid", fgColor="1F4E78")
    fill_ok   = PatternFill("solid", fgColor="E2EFDA")
    fill_warn = PatternFill("solid", fgColor="FFF2CC")
    fill_err  = PatternFill("solid", fgColor="FCE4D6")

    thin  = Side(style="thin", color="D0D0D0")
    bord  = Border(bottom=thin, right=thin)
    font_h = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    font_n = Font(name="Arial", size=10)
    font_b = Font(name="Arial", bold=True, size=10)
    font_l = Font(name="Arial", color="0563C1", underline="single", size=10)

    # Ligne titre
    ws.merge_cells(f"A1:{get_column_letter(len(COLS))}1")
    c = ws.cell(1, 1, f"FamilinParis — Contacts enrichis — {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    c.font      = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    c.fill      = fill_h
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # En-têtes
    for ci, col in enumerate(COLS, 1):
        c = ws.cell(2, ci, col)
        c.font      = font_h
        c.fill      = fill_h
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 36

    # Données
    for ri, (_, row) in enumerate(df.iterrows(), 3):
        statut  = str(row.get("Statut", ""))
        fill_row = fill_ok if "✅" in statut else (fill_warn if "⚠️" in statut else fill_err)

        for ci, col in enumerate(COLS, 1):
            key = col if col in df.columns else col.lower().replace(" ", "_")
            cell_text = cellval(row.get(col, ""))

            c = ws.cell(ri, ci)
            if col in ("Site web", "Maps URL", "URL") and cell_text:
                c.value     = cell_text
                c.font      = font_l
                c.hyperlink = cell_text
            elif col == "Nom":
                c.value = cell_text
                c.font  = font_b
            else:
                c.value = cell_text
                c.font  = font_n
            c.fill      = fill_row
            c.alignment = Alignment(vertical="top",
                                    wrap_text=(col in ("Nom", "Adresse", "Adresse Maps")))
            c.border = bord
        ws.row_dimensions[ri].height = 24

    for ci, col in enumerate(COLS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = LARGEURS.get(col, 15)

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(COLS))}2"

    # Onglet stats
    ws2 = wb.create_sheet("📊 Stats")
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 12

    total = len(df)

    def count_filled(col_name):
        if col_name not in df.columns:
            return 0
        return df[col_name].apply(lambda x: bool(str(x).strip()) if pd.notna(x) else False).sum()

    stats_data = [
        ("Champ",       "Remplis", "Manquants"),
        ("Téléphone",   count_filled("Telephone"),  total - count_filled("Telephone")),
        ("Site web",    count_filled("Site web"),   total - count_filled("Site web")),
        ("Email",       count_filled("Email"),      total - count_filled("Email")),
        ("TOTAL lignes", total, ""),
    ]

    # Répartition des méthodes
    if "Méthode" in df.columns:
        ws2.cell(len(stats_data) + 2, 1, "Répartition méthodes").font = Font(bold=True)
        methode_counts: dict[str, int] = {}
        for val in df["Méthode"].dropna():
            for part in str(val).split("+"):
                part = part.strip()
                if part and part != "—":
                    methode_counts[part] = methode_counts.get(part, 0) + 1
        for ri_offset, (k, v) in enumerate(sorted(methode_counts.items(), key=lambda x: -x[1]), 1):
            ws2.cell(len(stats_data) + 2 + ri_offset, 1, k)
            ws2.cell(len(stats_data) + 2 + ri_offset, 2, v)

    for ri, row_s in enumerate(stats_data, 1):
        for ci, val in enumerate(row_s, 1):
            c = ws2.cell(ri, ci, val)
            if ri == 1:
                c.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
                c.fill = fill_h
            else:
                c.font = Font(name="Arial", size=11)
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[ri].height = 22

    wb.save(fichier)

    tel_n  = count_filled("Telephone")
    site_n = count_filled("Site web")
    mail_n = count_filled("Email")
    print(f"\n✅  {fichier}")
    print(f"   Téléphone : {tel_n}/{total} ({tel_n/total*100:.1f}%)")
    print(f"   Site web  : {site_n}/{total} ({site_n/total*100:.1f}%)")
    print(f"   Email     : {mail_n}/{total} ({mail_n/total*100:.1f}%)")


# ── MAIN ──────────────────────────────────────────────────────────────────────

def main():
    print("\n" + "=" * 65)
    print("  Enrichissement Contacts v3 — Playwright (sans SerpAPI)")
    print("  Priorité : scraping direct → Google Maps → Google Search")
    print("=" * 65)

    inp = input(f"\n  Fichier source [{INPUT_FILE}] : ").strip() or INPUT_FILE
    if not Path(inp).exists():
        print(f"❌  Introuvable : {inp}")
        return
    out = input(f"  Fichier sortie [{OUTPUT_FILE}] : ").strip() or OUTPUT_FILE

    df = pd.read_excel(inp)
    print(f"\n📊  {len(df)} lignes — colonnes : {df.columns.tolist()}")

    # Normaliser les noms de colonnes attendus
    col_map = {}
    for col in df.columns:
        c = col.strip().lower()
        if c in ("nom", "titre", "name"):                     col_map[col] = "Nom"
        elif c in ("adresse", "adresse maps", "address"):     col_map[col] = "Adresse"
        elif c in ("url", "lien", "url feverup"):             col_map[col] = "URL"
        elif c in ("telephone", "téléphone", "tel", "tél"):   col_map[col] = "Telephone"
        elif c == "email":                                     col_map[col] = "Email"
        elif c in ("site web", "site_web", "website"):        col_map[col] = "Site web"
    df = df.rename(columns=col_map)

    # Ajouter colonnes manquantes (dtype object pour accepter du texte)
    for col in ["Telephone", "Email", "Site web", "Note", "Nb avis",
                "Maps URL", "Nom Maps", "Adresse Maps", "Méthode", "Statut"]:
        if col not in df.columns:
            df[col] = pd.Series([""] * len(df), dtype="object")
        else:
            df[col] = df[col].astype("object")

    # Les colonnes pré-existantes (Telephone, Email, Site web, URL, Adresse...)
    # peuvent être en float64 si entièrement vides dans le fichier source —
    # on les force en object pour éviter les FutureWarning de pandas.
    for col in ["Nom", "Adresse", "URL", "Telephone", "Email", "Site web"]:
        if col in df.columns:
            df[col] = df[col].astype("object")

    checkpoint = charger_checkpoint()
    total = len(df)

    # Compter lignes déjà complètes
    already_full = sum(
        1 for _, row in df.iterrows()
        if all(cellval(row.get(c, ""))
               for c in ("Telephone", "Email", "Site web"))
    )
    print(f"   Déjà complets : {already_full}/{total}")
    print(f"   À enrichir    : {total - already_full}/{total}\n")

    session = requests.Session()
    traite  = 0

    with sync_playwright() as p:
        browser, context, page, via_cdp = ouvrir_navigateur(p)

        try:
            for idx, row in df.iterrows():
                num = cellval(row.get("#", idx))

                # Déjà dans checkpoint → restaurer
                if num in checkpoint:
                    res = checkpoint[num]
                    for k, col in [("telephone","Telephone"), ("email","Email"),
                                   ("site_web","Site web"), ("note","Note"),
                                   ("nb_avis","Nb avis"), ("maps_url","Maps URL"),
                                   ("nom_maps","Nom Maps"), ("adresse_maps","Adresse Maps"),
                                   ("methode","Méthode"), ("statut","Statut")]:
                        if res.get(k):
                            df.at[idx, col] = res[k]
                    traite += 1
                    continue

                # Court-circuit : déjà complet dans le fichier source
                if all(cellval(row.get(c, "")) for c in ("Telephone", "Email", "Site web")):
                    checkpoint[num] = {
                        "telephone": cellval(row.get("Telephone", "")),
                        "email":     cellval(row.get("Email", "")),
                        "site_web":  cellval(row.get("Site web", "")),
                        "statut":    "✅ DÉJÀ COMPLET",
                        "methode":   "—",
                    }
                    traite += 1
                    continue

                nom = cellval(row.get("Nom", ""))
                print(f"\n[{idx+1}/{total}] #{num} — {nom[:50]}")

                res = enrichir_ligne(dict(row), session, page, idx + 1)

                for k, col in [("telephone","Telephone"), ("email","Email"),
                               ("site_web","Site web"), ("note","Note"),
                               ("nb_avis","Nb avis"), ("maps_url","Maps URL"),
                               ("nom_maps","Nom Maps"), ("adresse_maps","Adresse Maps"),
                               ("methode","Méthode"), ("statut","Statut")]:
                    if res.get(k):
                        df.at[idx, col] = res[k]

                print(f"  ✓ tel={res.get('telephone') or '—'} | "
                      f"site={str(res.get('site_web',''))[:30] or '—'} | "
                      f"email={res.get('email') or '—'} | "
                      f"{res.get('statut','')}")

                checkpoint[num] = res
                traite += 1

                if traite % CHECKPOINT_EVERY == 0:
                    sauvegarder_checkpoint(checkpoint)
                    export_excel(df, out)
                    print(f"\n💾  Checkpoint #{traite} sauvegardé\n")
        finally:
            try:
                page.close()
            except Exception:
                pass
            if not via_cdp:
                context.close()
                browser.close()
            # Si on est connecté à un vrai Chrome (CDP), on NE le ferme PAS :
            # c'est la fenêtre/navigateur de l'utilisateur.

    sauvegarder_checkpoint(checkpoint)
    export_excel(df, out)

    print("\n" + "=" * 65)
    print("✅  ENRICHISSEMENT TERMINÉ")
    print("=" * 65)


if __name__ == "__main__":
    main()

# import json
# import os
# import re
# import time
# import urllib.parse
# from datetime import datetime
# from pathlib import Path

# import pandas as pd
# import requests
# from bs4 import BeautifulSoup
# from openpyxl import Workbook
# from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
# from openpyxl.utils import get_column_letter

# # ── CONFIGURATION ──────────────────────────────────────────────────────────────

# SERP_API_KEY    = "d99b32948181e4132cbd1c3891fb971ab722c80e15fbebbd34a82d92ea9d76ac"
# INPUT_FILE      = "familinparis_activites_enrichi.xlsx"
# OUTPUT_FILE     = "familinparis_enrichi_contacts.xlsx"
# CHECKPOINT_F    = "checkpoint_contacts.json"

# PAUSE           = 1.5    # entre appels SerpAPI
# PAUSE_SCRAPE    = 0.8    # entre scrapes web
# REQUEST_TIMEOUT = 20
# CHECKPOINT_EVERY= 50

# HEADERS_WEB = {
#     "User-Agent": (
#         "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
#         "AppleWebKit/537.36 (KHTML, like Gecko) "
#         "Chrome/124.0.0.0 Safari/537.36"
#     ),
#     "Accept-Language": "fr-FR,fr;q=0.9,en;q=0.8",
#     "Accept": "text/html,application/xhtml+xml,*/*;q=0.8",
# }

# # ── REGEX ──────────────────────────────────────────────────────────────────────

# PHONE_RE = re.compile(
#     r"""(?:(?:\+|00)\s*\d{1,3}[\s.\-]?)?"""   # préfixe international
#     r"""(?:\(?\d{1,4}\)?[\s.\-]?)?"""           # indicatif zone
#     r"""(?:\d[\s.\-]?){7,12}\d""",
#     re.VERBOSE,
# )

# # Numéros français : 0X XX XX XX XX ou +33 X XX XX XX XX
# PHONE_FR_RE = re.compile(
#     r"""(?:(?:\+33|0033)[\s.\-]?[1-9](?:[\s.\-]?\d{2}){4}"""   # +33 / 0033
#     r"""|0[1-9](?:[\s.\-]?\d{2}){4})""",                          # 0X…
#     re.VERBOSE,
# )

# EMAIL_RE = re.compile(
#     r"\b[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}\b"
# )

# BLACKLIST_EMAIL_DOMAINS = {
#     "example.com", "test.com", "domain.com", "email.com",
#     "wixpress.com", "sentry.io", "wordpress.org",
#     "schema.org", "google.com", "facebook.com",
#     "w3.org", "acquia.com", "amazonaws.com",
# }
# BLACKLIST_EMAIL_PATTERNS = ["pixel", "track", "sentry", "analytics", "noreply", "no-reply"]

# RE_TAGS  = re.compile(r'<[^>]+>')
# RE_NBSP  = re.compile(r'&nbsp;|&#160;', re.IGNORECASE)
# RE_ENTIT = re.compile(r'&\w+;')

# PAGES_CONTACT = [
#     "contact", "nous-contacter", "contactez-nous", "contact-us",
#     "a-propos", "about", "informations", "equipe", "qui-sommes-nous",
# ]

# # ── UTILITAIRES TEXTE ──────────────────────────────────────────────────────────

# def strip_html(s: str) -> str:
#     import html as html_mod
#     s = html_mod.unescape(s)
#     return " ".join(RE_TAGS.sub(" ", s).split()).strip()


# def nettoyer_telephone(tel: str) -> str:
#     """Normalise un numéro de téléphone — retourne "" si invalide."""
#     if not tel:
#         return ""
#     # Garder chiffres, +, espaces, tirets, points, parenthèses
#     t = re.sub(r"[^\d+\s.\-()]", "", tel).strip()
#     digits_only = re.sub(r"[^\d]", "", t)
#     return t if 7 <= len(digits_only) <= 15 else ""


# def nettoyer_email(email: str) -> str:
#     if not email:
#         return ""
#     e = email.lower().strip()
#     domain = e.split("@")[-1]
#     if domain in BLACKLIST_EMAIL_DOMAINS:
#         return ""
#     if any(pat in domain for pat in BLACKLIST_EMAIL_PATTERNS):
#         return ""
#     return e


# def extraire_emails(texte: str) -> list[str]:
#     found = [nettoyer_email(e) for e in EMAIL_RE.findall(texte or "")]
#     return list(dict.fromkeys(e for e in found if e))


# def extraire_telephones(texte: str) -> list[str]:
#     found = []
#     for pat in (PHONE_FR_RE, PHONE_RE):
#         for m in pat.finditer(texte or ""):
#             t = nettoyer_telephone(m.group(0))
#             if t:
#                 found.append(t)
#     return list(dict.fromkeys(found))


# def nettoyer_url_base(url: str) -> str:
#     if not url:
#         return ""
#     if not url.startswith("http"):
#         url = "https://" + url
#     p = urllib.parse.urlparse(url)
#     return f"{p.scheme}://{p.netloc}"


# # ── COUCHE 0 : Scraping page familinparis.fr ──────────────────────────────────

# RE_POST_META = re.compile(
#     r'<ul[^>]*class=["\'][^"\']*post-meta[^"\']*["\'][^>]*>(.*?)</ul>',
#     re.DOTALL | re.IGNORECASE,
# )
# RE_LI = re.compile(r'<li[^>]*>(.*?)</li>', re.DOTALL | re.IGNORECASE)

# RE_KEY = re.compile(
#     r'^(adresse|lieu|t[ée]l[ée]phone?|t[ée]l\.?|email|e-mail|site\s*web?|web|contact)\s*[:\u00a0]?\s*',
#     re.IGNORECASE,
# )

# # Liens tel: et mailto: dans le HTML
# RE_TEL_HREF    = re.compile(r'href=["\']tel:([^"\']+)["\']',    re.IGNORECASE)
# RE_MAILTO_HREF = re.compile(r'href=["\']mailto:([^"\'?\s]+)',   re.IGNORECASE)

# # JSON-LD
# RE_JSONLD = re.compile(
#     r'<script[^>]+type=["\']application/ld\+json["\'][^>]*>(.*?)</script>',
#     re.DOTALL | re.IGNORECASE,
# )


# def _scrape_post_meta(raw_html: str) -> dict:
#     """Extrait les champs depuis <ul class='post-meta'>."""
#     result = {}
#     m = RE_POST_META.search(raw_html)
#     if not m:
#         return result
#     for li_m in RE_LI.finditer(m.group(1)):
#         text = strip_html(li_m.group(1))
#         key_m = RE_KEY.match(text)
#         if not key_m:
#             continue
#         key   = key_m.group(1).lower().strip().rstrip(".")
#         value = text[key_m.end():].strip()
#         if not value:
#             continue

#         if re.match(r't[ée]l', key, re.IGNORECASE):
#             t = nettoyer_telephone(value)
#             if t:
#                 result.setdefault("telephone", t)
#         elif "email" in key or "e-mail" in key:
#             e = nettoyer_email(value)
#             if e:
#                 result.setdefault("email", e)
#         elif "site" in key or "web" in key:
#             result.setdefault("site_web", value)
#         elif key == "adresse":
#             result.setdefault("adresse", value)
#         elif key == "lieu":
#             result.setdefault("lieu", value)

#     return result


# def _scrape_jsonld(raw_html: str) -> dict:
#     """Extrait téléphone / email / url depuis les blocs JSON-LD."""
#     result = {}
#     for m in RE_JSONLD.finditer(raw_html):
#         try:
#             data = json.loads(m.group(1).strip())
#         except Exception:
#             continue
#         items = data if isinstance(data, list) else [data]
#         for item in items:
#             if not isinstance(item, dict):
#                 continue
#             if not result.get("telephone"):
#                 t = nettoyer_telephone(str(item.get("telephone", "")))
#                 if t:
#                     result["telephone"] = t
#             if not result.get("email"):
#                 e = nettoyer_email(str(item.get("email", "")))
#                 if e:
#                     result["email"] = e
#             if not result.get("site_web"):
#                 url = item.get("url", item.get("sameAs", ""))
#                 if isinstance(url, list):
#                     url = url[0] if url else ""
#                 if url and "familinparis" not in str(url):
#                     result["site_web"] = str(url)
#     return result


# def _scrape_hrefs(raw_html: str) -> dict:
#     """Extrait téléphone et email depuis href='tel:' et href='mailto:'."""
#     result = {}
#     m_tel = RE_TEL_HREF.search(raw_html)
#     if m_tel:
#         t = nettoyer_telephone(m_tel.group(1))
#         if t:
#             result["telephone"] = t
#     m_mail = RE_MAILTO_HREF.search(raw_html)
#     if m_mail:
#         e = nettoyer_email(m_mail.group(1))
#         if e:
#             result["email"] = e
#     return result


# def _scrape_page_texte(raw_html: str) -> dict:
#     """Cherche emails et téléphones dans tout le texte de la page."""
#     result = {}
#     # Supprimer scripts/styles d'abord
#     clean = re.sub(r'<(script|style)[^>]*>.*?</\1>', '', raw_html, flags=re.DOTALL|re.IGNORECASE)
#     texte = strip_html(clean)

#     emails = extraire_emails(texte)
#     if emails:
#         result["email"] = emails[0]

#     tels = extraire_telephones(texte)
#     if tels:
#         result["telephone"] = tels[0]

#     return result


# def scrape_familinparis(url: str, session: requests.Session) -> dict:
#     """
#     Scrape une page familinparis.fr pour extraire tel, email, site_web.
#     Fusionne les résultats de toutes les stratégies.
#     """
#     result = {}
#     try:
#         r = session.get(url, headers=HEADERS_WEB, timeout=REQUEST_TIMEOUT,
#                         allow_redirects=True)
#         if r.status_code != 200:
#             return {"_erreur": f"HTTP {r.status_code}"}
#         raw_html = r.text

#         # Fusionne dans l'ordre de fiabilité
#         for extractor in [_scrape_post_meta, _scrape_jsonld, _scrape_hrefs, _scrape_page_texte]:
#             partial = extractor(raw_html)
#             for k, v in partial.items():
#                 if v and not result.get(k):
#                     result[k] = v

#         # Si site_web == url familinparis, on ignore
#         site = result.get("site_web", "")
#         if site and "familinparis.fr" in site:
#             result.pop("site_web", None)

#     except Exception as e:
#         result["_erreur"] = str(e)[:80]

#     return result


# # ── COUCHE 1 : Scraping du site officiel ──────────────────────────────────────

# def scrape_site_officiel(url_base: str, session: requests.Session) -> dict:
#     """
#     Visite la page d'accueil puis /contact du site officiel.
#     Retourne : {"telephone": ..., "email": ...}
#     """
#     result = {}
#     if not url_base:
#         return result

#     def _fetch(url: str) -> dict:
#         try:
#             r = session.get(url, headers=HEADERS_WEB, timeout=REQUEST_TIMEOUT,
#                             allow_redirects=True)
#             if r.status_code != 200:
#                 return {}
#             raw = r.text
#             partial = {}
#             for extractor in [_scrape_hrefs, _scrape_jsonld, _scrape_page_texte]:
#                 for k, v in extractor(raw).items():
#                     if v and not partial.get(k):
#                         partial[k] = v
#             return partial
#         except Exception:
#             return {}

#     # Page d'accueil
#     result = _fetch(url_base)
#     if result.get("telephone") and result.get("email"):
#         return result
#     time.sleep(PAUSE_SCRAPE)

#     # Pages contact
#     for page in PAGES_CONTACT:
#         if result.get("telephone") and result.get("email"):
#             break
#         contact_url = f"{url_base.rstrip('/')}/{page}"
#         partial = _fetch(contact_url)
#         for k, v in partial.items():
#             if v and not result.get(k):
#                 result[k] = v
#         time.sleep(PAUSE_SCRAPE * 0.4)

#     return result


# # ── COUCHE 2 : SerpAPI Google Maps ────────────────────────────────────────────

# def serpapi_maps(nom: str, adresse: str, session: requests.Session) -> dict:
#     query = nom
#     if adresse:
#         ville_parts = [p.strip() for p in adresse.split(",")]
#         # Cherche la partie "ville" (dernier morceau non vide qui n'est pas juste un CP)
#         for part in reversed(ville_parts):
#             if part and not re.fullmatch(r'\d{5}', part):
#                 query = f"{nom} {part}"
#                 break

#     params = {
#         "engine":  "google_maps",
#         "q":       query,
#         "hl":      "fr",
#         "gl":      "fr",
#         "type":    "search",
#         "api_key": SERP_API_KEY,
#     }
#     try:
#         r = session.get("https://serpapi.com/search", params=params,
#                         timeout=REQUEST_TIMEOUT)
#         r.raise_for_status()
#         data = r.json()
#         if "error" in data:
#             return {"_erreur": data["error"]}
#         places = data.get("local_results", [])
#         if not places:
#             return {}
#         p   = places[0]
#         gps = p.get("gps_coordinates", {})
#         pid = p.get("place_id", "")
#         return {
#             "telephone":    nettoyer_telephone(p.get("phone", "")),
#             "site_web":     p.get("website", ""),
#             "note":         p.get("rating", ""),
#             "nb_avis":      p.get("reviews", 0),
#             "latitude":     gps.get("latitude"),
#             "longitude":    gps.get("longitude"),
#             "place_id":     pid,
#             "maps_url":     f"https://www.google.com/maps/place/?q=place_id:{pid}" if pid else "",
#             "nom_maps":     p.get("title", ""),
#             "adresse_maps": p.get("address", ""),
#         }
#     except Exception as e:
#         return {"_erreur": str(e)[:80]}


# # ── COUCHE 3 : SerpAPI Google Search → site ───────────────────────────────────

# DOMAINES_EXCLUS = {
#     "funbooker", "familinparis", "tripadvisor", "facebook", "instagram",
#     "yelp", "google", "youtube", "linkedin", "twitter", "booking",
#     "viator", "getyourguide", "billetreduc", "billetweb", "weezevent",
# }

# def serpapi_google_site(nom: str, adresse: str, session: requests.Session) -> dict:
#     ville = ""
#     if adresse:
#         for part in reversed([p.strip() for p in adresse.split(",")]):
#             if part and not re.fullmatch(r'\d{5}', part):
#                 ville = part
#                 break
#     query = f"{nom} {ville} site officiel contact".strip()

#     params = {
#         "engine":  "google",
#         "q":       query,
#         "hl":      "fr",
#         "gl":      "fr",
#         "num":     5,
#         "api_key": SERP_API_KEY,
#     }
#     try:
#         r = session.get("https://serpapi.com/search", params=params,
#                         timeout=REQUEST_TIMEOUT)
#         r.raise_for_status()
#         data = r.json()
#         if "error" in data:
#             return {"_erreur": data["error"]}
#         for res in data.get("organic_results", []):
#             lien = res.get("link", "")
#             dom  = urllib.parse.urlparse(lien).netloc.lower()
#             if not any(exclu in dom for exclu in DOMAINES_EXCLUS):
#                 return {"site_web": nettoyer_url_base(lien), "source": lien}
#     except Exception as e:
#         return {"_erreur": str(e)[:80]}
#     return {}


# # ── COUCHE 5 : SerpAPI Google Search → email dans snippets ───────────────────

# def serpapi_google_email(nom: str, adresse: str, session: requests.Session) -> dict:
#     ville = ""
#     if adresse:
#         for part in reversed([p.strip() for p in adresse.split(",")]):
#             if part and not re.fullmatch(r'\d{5}', part):
#                 ville = part
#                 break
#     query = f"{nom} {ville} email contact réservation".strip()

#     params = {
#         "engine":  "google",
#         "q":       query,
#         "hl":      "fr",
#         "gl":      "fr",
#         "num":     5,
#         "api_key": SERP_API_KEY,
#     }
#     try:
#         r = session.get("https://serpapi.com/search", params=params,
#                         timeout=REQUEST_TIMEOUT)
#         r.raise_for_status()
#         data = r.json()
#         if "error" in data:
#             return {"_erreur": data["error"]}
#         for res in data.get("organic_results", []):
#             texte = res.get("snippet", "") + " " + res.get("title", "")
#             emails = extraire_emails(texte)
#             if emails:
#                 return {"email": emails[0], "source_email": res.get("link", "")}
#     except Exception as e:
#         return {"_erreur": str(e)[:80]}
#     return {}


# # ── ORCHESTRATEUR PAR LIGNE ────────────────────────────────────────────────────

# def enrichir_ligne(row: dict, session: requests.Session, idx: int) -> dict:
#     """
#     Multi-couche avec court-circuit : dès qu'on a tel+email+site, on s'arrête.
#     """
#     nom     = str(row.get("Nom", "")).strip()
#     adresse = str(row.get("Adresse", "")).strip()
#     url_fp  = str(row.get("URL", "")).strip()   # URL familinparis de l'activité

#     # Valeurs déjà présentes (ne pas écraser si déjà remplies)
#     result = {
#         "telephone":    str(row.get("Telephone", "") or "").strip(),
#         "email":        str(row.get("Email", "")     or "").strip(),
#         "site_web":     str(row.get("Site web", "")  or "").strip(),
#         "note":         str(row.get("Note", "")      or "").strip(),
#         "nb_avis":      str(row.get("Nb avis", "")   or "").strip(),
#         "maps_url":     str(row.get("Maps URL", "")  or "").strip(),
#         "nom_maps":     str(row.get("Nom Maps", "")  or "").strip(),
#         "adresse_maps": str(row.get("Adresse Maps", "") or "").strip(),
#         "methodes":     [],
#         "statut":       "",
#     }

#     def _complet():
#         return bool(result["telephone"] and result["email"] and result["site_web"])

#     def _merge(partial: dict, methode: str):
#         """Fusionne sans écraser les valeurs déjà remplies."""
#         changed = False
#         for k in ("telephone", "email", "site_web", "note", "nb_avis",
#                    "maps_url", "nom_maps", "adresse_maps"):
#             if partial.get(k) and not result.get(k):
#                 result[k] = partial[k]
#                 changed = True
#         if changed:
#             result["methodes"].append(methode)

#     # ── C0 : Scraping page familinparis ──────────────────────────────────────
#     if url_fp and not _complet():
#         print(f"  [{idx}] C0-scrape familinparis...")
#         c0 = scrape_familinparis(url_fp, session)
#         if "_erreur" not in c0:
#             _merge(c0, "C0-familinparis")
#         time.sleep(PAUSE_SCRAPE)

#     # ── C1 : Scraping du site officiel (si on a déjà un site_web) ────────────
#     if result["site_web"] and not _complet():
#         print(f"  [{idx}] C1-scrape site officiel : {result['site_web'][:45]}...")
#         c1 = scrape_site_officiel(result["site_web"], session)
#         _merge(c1, "C1-site-officiel")

#     # ── C2 : SerpAPI Maps ─────────────────────────────────────────────────────
#     if not _complet():
#         print(f"  [{idx}] C2-Maps : {nom[:40]}...")
#         c2 = serpapi_maps(nom, adresse, session)
#         time.sleep(PAUSE)
#         if "_erreur" not in c2:
#             _merge(c2, "C2-maps")

#     # ── C1 bis : Scraping site officiel trouvé par Maps ───────────────────────
#     if result["site_web"] and "C1-site-officiel" not in result["methodes"] and not _complet():
#         print(f"  [{idx}] C1b-scrape site maps : {result['site_web'][:45]}...")
#         c1b = scrape_site_officiel(result["site_web"], session)
#         _merge(c1b, "C1b-site-maps")

#     # ── C3 : SerpAPI Google Search → site_web ────────────────────────────────
#     if not result["site_web"] and not _complet():
#         print(f"  [{idx}] C3-Google site : {nom[:40]}...")
#         c3 = serpapi_google_site(nom, adresse, session)
#         time.sleep(PAUSE)
#         if "_erreur" not in c3:
#             _merge(c3, "C3-google-site")

#     # ── C4 : Scraping du site trouvé en C3 ───────────────────────────────────
#     if result["site_web"] and "C1" not in " ".join(result["methodes"]) and not result.get("email"):
#         print(f"  [{idx}] C4-scrape nouveau site : {result['site_web'][:45]}...")
#         c4 = scrape_site_officiel(result["site_web"], session)
#         _merge(c4, "C4-site-c3")

#     # ── C5 : SerpAPI Google Search → email dans snippets ─────────────────────
#     if not result["email"]:
#         print(f"  [{idx}] C5-Google email : {nom[:40]}...")
#         c5 = serpapi_google_email(nom, adresse, session)
#         time.sleep(PAUSE)
#         if "_erreur" not in c5:
#             _merge(c5, "C5-google-email")

#     # ── Statut final ──────────────────────────────────────────────────────────
#     manquants = [
#         f for f, k in [("tel", "telephone"), ("email", "email"), ("site", "site_web")]
#         if not result[k]
#     ]
#     if not manquants:
#         result["statut"] = "✅ COMPLET"
#     elif len(manquants) == 1:
#         result["statut"] = f"⚠️ manque: {manquants[0]}"
#     else:
#         result["statut"] = f"❌ manque: {', '.join(manquants)}"

#     result["methode"] = " + ".join(result["methodes"]) or "—"
#     return result


# # ── CHECKPOINT ────────────────────────────────────────────────────────────────

# def charger_checkpoint() -> dict:
#     if Path(CHECKPOINT_F).exists():
#         try:
#             with open(CHECKPOINT_F, "r", encoding="utf-8") as f:
#                 data = json.load(f)
#             print(f"✅ Checkpoint : {len(data)} lignes déjà traitées")
#             return data
#         except Exception:
#             pass
#     return {}

# def sauvegarder_checkpoint(data: dict):
#     with open(CHECKPOINT_F, "w", encoding="utf-8") as f:
#         json.dump(data, f, ensure_ascii=False, indent=2)


# # ── EXPORT EXCEL ──────────────────────────────────────────────────────────────

# def export_excel(df: pd.DataFrame, fichier: str):
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Activités enrichies"

#     COLS = [
#         "#", "Nom", "Adresse", "Telephone", "Site web", "Email",
#         "Note", "Nb avis", "Maps URL", "Nom Maps", "Adresse Maps",
#         "Méthode", "Statut", "URL",
#     ]
#     LARGEURS = {
#         "#": 5, "Nom": 45, "Adresse": 38, "Telephone": 16,
#         "Site web": 35, "Email": 32, "Note": 7, "Nb avis": 8,
#         "Maps URL": 28, "Nom Maps": 32, "Adresse Maps": 35,
#         "Méthode": 28, "Statut": 22, "URL": 55,
#     }

#     fill_h    = PatternFill("solid", fgColor="1F4E78")
#     fill_ok   = PatternFill("solid", fgColor="E2EFDA")
#     fill_warn = PatternFill("solid", fgColor="FFF2CC")
#     fill_err  = PatternFill("solid", fgColor="FCE4D6")

#     thin  = Side(style="thin", color="D0D0D0")
#     bord  = Border(bottom=thin, right=thin)
#     font_h = Font(name="Arial", bold=True, color="FFFFFF", size=10)
#     font_n = Font(name="Arial", size=10)
#     font_b = Font(name="Arial", bold=True, size=10)
#     font_l = Font(name="Arial", color="0563C1", underline="single", size=10)

#     # Ligne titre
#     ws.merge_cells(f"A1:{get_column_letter(len(COLS))}1")
#     c = ws.cell(1, 1, f"FamilinParis — Contacts enrichis — {datetime.now().strftime('%d/%m/%Y %H:%M')}")
#     c.font      = Font(name="Arial", bold=True, size=13, color="FFFFFF")
#     c.fill      = fill_h
#     c.alignment = Alignment(horizontal="center", vertical="center")
#     ws.row_dimensions[1].height = 28

#     # En-têtes
#     for ci, col in enumerate(COLS, 1):
#         c = ws.cell(2, ci, col)
#         c.font      = font_h
#         c.fill      = fill_h
#         c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
#     ws.row_dimensions[2].height = 36

#     # Données
#     for ri, (_, row) in enumerate(df.iterrows(), 3):
#         statut  = str(row.get("Statut", ""))
#         fill_row = fill_ok if "✅" in statut else (fill_warn if "⚠️" in statut else fill_err)

#         for ci, col in enumerate(COLS, 1):
#             key = col if col in df.columns else col.lower().replace(" ", "_")
#             val = str(row.get(col, "") or "").strip()

#             c = ws.cell(ri, ci)
#             if col in ("Site web", "Maps URL", "URL") and val:
#                 c.value     = val
#                 c.font      = font_l
#                 c.hyperlink = val
#             elif col == "Nom":
#                 c.value = val
#                 c.font  = font_b
#             else:
#                 c.value = val
#                 c.font  = font_n
#             c.fill      = fill_row
#             c.alignment = Alignment(vertical="top",
#                                     wrap_text=(col in ("Nom", "Adresse", "Adresse Maps")))
#             c.border = bord
#         ws.row_dimensions[ri].height = 24

#     for ci, col in enumerate(COLS, 1):
#         ws.column_dimensions[get_column_letter(ci)].width = LARGEURS.get(col, 15)

#     ws.freeze_panes = "A3"
#     ws.auto_filter.ref = f"A2:{get_column_letter(len(COLS))}2"

#     # Onglet stats
#     ws2 = wb.create_sheet("📊 Stats")
#     ws2.column_dimensions["A"].width = 30
#     ws2.column_dimensions["B"].width = 12
#     ws2.column_dimensions["C"].width = 12

#     total = len(df)

#     def count_filled(col_name):
#         if col_name not in df.columns:
#             return 0
#         return df[col_name].apply(lambda x: bool(str(x).strip()) if pd.notna(x) else False).sum()

#     stats_data = [
#         ("Champ",       "Remplis", "Manquants"),
#         ("Téléphone",   count_filled("Telephone"),  total - count_filled("Telephone")),
#         ("Site web",    count_filled("Site web"),   total - count_filled("Site web")),
#         ("Email",       count_filled("Email"),      total - count_filled("Email")),
#         ("TOTAL lignes", total, ""),
#     ]

#     # Répartition des méthodes
#     if "Méthode" in df.columns:
#         ws2.cell(len(stats_data) + 2, 1, "Répartition méthodes").font = Font(bold=True)
#         methode_counts: dict[str, int] = {}
#         for val in df["Méthode"].dropna():
#             for part in str(val).split("+"):
#                 part = part.strip()
#                 if part and part != "—":
#                     methode_counts[part] = methode_counts.get(part, 0) + 1
#         for ri_offset, (k, v) in enumerate(sorted(methode_counts.items(), key=lambda x: -x[1]), 1):
#             ws2.cell(len(stats_data) + 2 + ri_offset, 1, k)
#             ws2.cell(len(stats_data) + 2 + ri_offset, 2, v)

#     for ri, row_s in enumerate(stats_data, 1):
#         for ci, val in enumerate(row_s, 1):
#             c = ws2.cell(ri, ci, val)
#             if ri == 1:
#                 c.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
#                 c.fill = fill_h
#             else:
#                 c.font = Font(name="Arial", size=11)
#             c.alignment = Alignment(horizontal="center", vertical="center")
#         ws2.row_dimensions[ri].height = 22

#     wb.save(fichier)

#     tel_n  = count_filled("Telephone")
#     site_n = count_filled("Site web")
#     mail_n = count_filled("Email")
#     print(f"\n✅  {fichier}")
#     print(f"   Téléphone : {tel_n}/{total} ({tel_n/total*100:.1f}%)")
#     print(f"   Site web  : {site_n}/{total} ({site_n/total*100:.1f}%)")
#     print(f"   Email     : {mail_n}/{total} ({mail_n/total*100:.1f}%)")


# # ── MAIN ──────────────────────────────────────────────────────────────────────

# def main():
#     print("\n" + "=" * 65)
#     print("  Enrichissement Contacts v2 — Téléphone / Email / Site web")
#     print("  Priorité : scraping direct → SerpAPI Maps → Google Search")
#     print("=" * 65)

#     inp = input(f"\n  Fichier source [{INPUT_FILE}] : ").strip() or INPUT_FILE
#     if not Path(inp).exists():
#         print(f"❌  Introuvable : {inp}")
#         return
#     out = input(f"  Fichier sortie [{OUTPUT_FILE}] : ").strip() or OUTPUT_FILE

#     df = pd.read_excel(inp)
#     print(f"\n📊  {len(df)} lignes — colonnes : {df.columns.tolist()}")

#     # Normaliser les noms de colonnes attendus
#     col_map = {}
#     for col in df.columns:
#         c = col.strip().lower()
#         if c in ("nom", "titre", "name"):                     col_map[col] = "Nom"
#         elif c in ("adresse", "adresse maps", "address"):     col_map[col] = "Adresse"
#         elif c in ("url", "lien", "url feverup"):             col_map[col] = "URL"
#         elif c in ("telephone", "téléphone", "tel", "tél"):   col_map[col] = "Telephone"
#         elif c == "email":                                     col_map[col] = "Email"
#         elif c in ("site web", "site_web", "website"):        col_map[col] = "Site web"
#     df = df.rename(columns=col_map)

#     # Ajouter colonnes manquantes
#     for col in ["Telephone", "Email", "Site web", "Note", "Nb avis",
#                 "Maps URL", "Nom Maps", "Adresse Maps", "Méthode", "Statut"]:
#         if col not in df.columns:
#             df[col] = ""

#     checkpoint = charger_checkpoint()
#     total = len(df)

#     # Compter lignes déjà complètes
#     already_full = sum(
#         1 for _, row in df.iterrows()
#         if all(str(row.get(c, "") or "").strip()
#                for c in ("Telephone", "Email", "Site web"))
#     )
#     print(f"   Déjà complets : {already_full}/{total}")
#     print(f"   À enrichir    : {total - already_full}/{total}\n")

#     session = requests.Session()
#     traite  = 0

#     for idx, row in df.iterrows():
#         num = str(row.get("#", idx))

#         # Déjà dans checkpoint → restaurer
#         if num in checkpoint:
#             res = checkpoint[num]
#             for k, col in [("telephone","Telephone"), ("email","Email"),
#                            ("site_web","Site web"), ("note","Note"),
#                            ("nb_avis","Nb avis"), ("maps_url","Maps URL"),
#                            ("nom_maps","Nom Maps"), ("adresse_maps","Adresse Maps"),
#                            ("methode","Méthode"), ("statut","Statut")]:
#                 if res.get(k):
#                     df.at[idx, col] = res[k]
#             traite += 1
#             continue

#         # Court-circuit : déjà complet dans le fichier source
#         if all(str(row.get(c, "") or "").strip() for c in ("Telephone", "Email", "Site web")):
#             checkpoint[num] = {
#                 "telephone": str(row.get("Telephone", "")),
#                 "email":     str(row.get("Email", "")),
#                 "site_web":  str(row.get("Site web", "")),
#                 "statut":    "✅ DÉJÀ COMPLET",
#                 "methode":   "—",
#             }
#             traite += 1
#             continue

#         nom = str(row.get("Nom", "")).strip()
#         print(f"\n[{idx+1}/{total}] #{num} — {nom[:50]}")

#         res = enrichir_ligne(dict(row), session, idx + 1)

#         for k, col in [("telephone","Telephone"), ("email","Email"),
#                        ("site_web","Site web"), ("note","Note"),
#                        ("nb_avis","Nb avis"), ("maps_url","Maps URL"),
#                        ("nom_maps","Nom Maps"), ("adresse_maps","Adresse Maps"),
#                        ("methode","Méthode"), ("statut","Statut")]:
#             if res.get(k):
#                 df.at[idx, col] = res[k]

#         print(f"  ✓ tel={res.get('telephone') or '—'} | "
#               f"site={str(res.get('site_web',''))[:30] or '—'} | "
#               f"email={res.get('email') or '—'} | "
#               f"{res.get('statut','')}")

#         checkpoint[num] = res
#         traite += 1

#         if traite % CHECKPOINT_EVERY == 0:
#             sauvegarder_checkpoint(checkpoint)
#             export_excel(df, out)
#             print(f"\n💾  Checkpoint #{traite} sauvegardé\n")

#     sauvegarder_checkpoint(checkpoint)
#     export_excel(df, out)

#     print("\n" + "=" * 65)
#     print("✅  ENRICHISSEMENT TERMINÉ")
#     print("=" * 65)


# if __name__ == "__main__":
#     main()

