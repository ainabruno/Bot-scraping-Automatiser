import asyncio
import re
import time
from datetime import datetime

import openpyxl
from bs4 import BeautifulSoup
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from playwright.async_api import async_playwright, TimeoutError as PWTimeout

# ── CONFIGURATION ──────────────────────────────────────────────────────────────

BASE_URL    = "https://www.familinparis.fr"
URL_PATTERN = "https://www.familinparis.fr/page/{page}/?s&menu0&menu1&menu3"
TOTAL_PAGES = 29
OUTPUT_FILE = "familinparis_activites.xlsx"

PAUSE_ENTRE_PAGES = 1.5   # secondes entre pages
WAIT_PAGE_LOAD    = 5.0   # attente après navigation pour le JS
MAX_RETRIES       = 3

# ── HELPERS ────────────────────────────────────────────────────────────────────

def clean(text: str) -> str:
    return " ".join(str(text).split()).strip() if text else ""

def extract_image_url(style: str) -> str:
    if not style:
        return ""
    m = re.search(r"url\(['\"]?(.*?)['\"]?\)", style)
    return m.group(1).strip() if m else ""

def extraire_age(texte: str) -> str:
    if not texte:
        return ""
    m = re.search(r"(\d+)\s*ans?", texte)
    return f"{m.group(1)} ans" if m else clean(texte)

# ── PARSING D'UNE CARTE ────────────────────────────────────────────────────────

def parse_grid(div, page_num: int) -> dict:
    result = {
        "page":      page_num,
        "url":       "",
        "image_url": "",
        "date":      "",
        "age_min":   "",
        "lieu":      "",
        "titre":     "",
        "tags":      "",
        "duree":     "",
        "prix":      "",
    }

    # Lien
    a_tag = div.find("a")
    if a_tag:
        href = a_tag.get("href", "")
        result["url"] = href if href.startswith("http") else BASE_URL + href

    # Image + date + âge
    img_div = div.find("div", class_="image")
    if img_div:
        result["image_url"] = extract_image_url(img_div.get("style", ""))

        date_div = img_div.find("div", class_="dateOnImg")
        if date_div:
            result["date"] = clean(date_div.get_text())

        age_div = img_div.find("div", class_="rond-blanc-age")
        if age_div:
            result["age_min"] = extraire_age(age_div.get_text())

    # Texte : lieu, titre, tags, prix
    texte_div = div.find("div", class_="texte")
    if texte_div:
        h5 = texte_div.find("h5")
        if h5:
            result["lieu"] = clean(h5.get_text())

        h3 = texte_div.find("h3")
        if h3:
            result["titre"] = clean(h3.get_text())

        tags_div = texte_div.find("div", class_="tags")
        if tags_div:
            spans = [clean(s.get_text()) for s in tags_div.find_all("span") if s.get_text().strip()]
            result["tags"] = " | ".join(spans)
            durees = [s for s in spans if re.match(r"\d+h", s, re.I)]
            if durees:
                result["duree"] = durees[0]

        prix_div = texte_div.find("div", class_="prix")
        if prix_div:
            parts = [clean(s.get_text()) for s in prix_div.find_all("span") if s.get_text().strip()]
            result["prix"] = " ".join(parts)

    return result

# ── FERMETURE DES POPUPS ───────────────────────────────────────────────────────

async def dismiss_overlays(page) -> None:
    for sel in [
        "#onetrust-accept-btn-handler",
        "button:has-text('Accepter')",
        "button:has-text('Tout accepter')",
        "button:has-text('Accept All')",
        "button:has-text('OK')",
        ".cookie-accept",
        "[aria-label='Close']",
        "[aria-label='Fermer']",
    ]:
        try:
            btn = page.locator(sel).first
            if await btn.is_visible(timeout=1_500):
                await btn.click(timeout=2_000)
                await asyncio.sleep(0.5)
        except Exception:
            pass

# ── SCRAPING D'UNE PAGE ────────────────────────────────────────────────────────

async def scraper_page(page, page_num: int) -> list[dict]:
    url = URL_PATTERN.format(page=page_num)

    for attempt in range(1, MAX_RETRIES + 1):
        try:
            await page.goto(url, wait_until="domcontentloaded", timeout=30_000)

            # Attendre que le JS charge les cartes
            await asyncio.sleep(WAIT_PAGE_LOAD)

            # Fermer les popups (seulement page 1)
            if page_num == 1:
                await dismiss_overlays(page)
                await asyncio.sleep(1.0)

            # Essayer d'attendre un sélecteur de carte
            try:
                await page.wait_for_selector("div.grid", timeout=8_000)
            except PWTimeout:
                pass

            # Récupérer le HTML rendu
            html = await page.content()
            soup = BeautifulSoup(html, "lxml")

            # Chercher les grids
            grids = soup.select("div.col-12.grids div.grid")
            if not grids:
                grids = soup.select("div.grids div.grid")
            if not grids:
                grids = soup.select("div.grid")

            if grids:
                activities = []
                for grid in grids:
                    act = parse_grid(grid, page_num)
                    if act["titre"] or act["url"]:
                        activities.append(act)
                return activities

            # Si toujours rien — debug
            title = await page.title()
            print(f"    ⚠  Aucune carte (tentative {attempt}/{MAX_RETRIES}) | title='{title}'")

            if attempt < MAX_RETRIES:
                await asyncio.sleep(3 * attempt)

        except PWTimeout:
            print(f"    ⚠  Timeout (tentative {attempt}/{MAX_RETRIES})")
            if attempt < MAX_RETRIES:
                await asyncio.sleep(3 * attempt)
        except Exception as e:
            print(f"    ⚠  Erreur : {e} (tentative {attempt}/{MAX_RETRIES})")
            if attempt < MAX_RETRIES:
                await asyncio.sleep(3 * attempt)

    return []

# ── SCRAPING COMPLET ───────────────────────────────────────────────────────────

async def scraper_toutes_les_pages() -> list[dict]:
    print("\n" + "=" * 65)
    print("  FamilinParis — Scraper v2 Playwright (29 pages)")
    print("=" * 65)

    all_activities = []

    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=True,
            args=[
                "--no-sandbox",
                "--disable-blink-features=AutomationControlled",
            ]
        )
        context = await browser.new_context(
            viewport={"width": 1440, "height": 900},
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
            locale="fr-FR",
        )
        page = await context.new_page()

        # Masquer Playwright (anti-détection)
        await page.add_init_script("""
            Object.defineProperty(navigator, 'webdriver', { get: () => undefined });
            window.chrome = { runtime: {} };
        """)

        for page_num in range(1, TOTAL_PAGES + 1):
            print(f"\n[{page_num:02d}/{TOTAL_PAGES}] page {page_num}...")

            activities = await scraper_page(page, page_num)

            if activities:
                all_activities.extend(activities)
                print(f"    ✅  {len(activities)} activités | total : {len(all_activities)}")
            else:
                print(f"    ❌  0 activité — page ignorée")

            if page_num < TOTAL_PAGES:
                await asyncio.sleep(PAUSE_ENTRE_PAGES)

        await browser.close()

    print(f"\n{'=' * 65}")
    print(f"  🎯 TOTAL : {len(all_activities)} activités sur {TOTAL_PAGES} pages")
    print(f"{'=' * 65}\n")
    return all_activities

# ── EXPORT EXCEL ───────────────────────────────────────────────────────────────

COLONNES = [
    ("#",         5),
    ("Page",      6),
    ("Titre",     55),
    ("Lieu",      30),
    ("Date",      28),
    ("Âge min",   10),
    ("Durée",     10),
    ("Tags",      30),
    ("Prix",      14),
    ("URL",       55),
    ("Image URL", 55),
]

def style_fill(couleur: str) -> PatternFill:
    return PatternFill("solid", start_color=couleur, end_color=couleur)

def save_excel(activities: list[dict], filename: str) -> None:
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "FamilinParis"

    HDR_BG = style_fill("E63946")
    ODD    = style_fill("FFF0F0")
    EVEN   = style_fill("FFFFFF")
    LINK_C = "C1121F"
    thin   = Side(style="thin", color="FFCCD5")
    brd    = Border(left=thin, right=thin, top=thin, bottom=thin)
    n_cols = len(COLONNES)

    # Titre
    ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
    c = ws["A1"]
    c.value     = "FamilinParis — Activités Paris (toutes pages)"
    c.font      = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    c.fill      = HDR_BG
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Sous-titre
    ws.merge_cells(f"A2:{get_column_letter(n_cols)}2")
    c = ws["A2"]
    c.value = (
        f"Extrait le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
        f"  |  {len(activities)} activités  |  {TOTAL_PAGES} pages"
    )
    c.font      = Font(name="Arial", italic=True, size=10, color="444444")
    c.fill      = style_fill("FFE8EA")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # En-têtes
    for ci, (header, width) in enumerate(COLONNES, 1):
        c = ws.cell(row=3, column=ci, value=header)
        c.font      = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        c.fill      = HDR_BG
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = brd
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[3].height = 24

    # Données
    for i, act in enumerate(activities, 1):
        row_num = i + 3
        fill    = ODD if i % 2 else EVEN

        def cell(col, value, hyperlink="", link=False):
            c = ws.cell(row=row_num, column=col, value=value)
            c.font = Font(
                name="Arial", size=10,
                color=LINK_C if link else "000000",
                underline="single" if link else None,
            )
            c.alignment = Alignment(vertical="top", wrap_text=(col in (3, 4, 8)))
            c.fill   = fill
            c.border = brd
            if hyperlink:
                c.hyperlink = hyperlink

        cell(1,  i)
        cell(2,  act.get("page", ""))
        cell(3,  act.get("titre", ""))
        cell(4,  act.get("lieu", ""))
        cell(5,  act.get("date", ""))
        cell(6,  act.get("age_min", ""))
        cell(7,  act.get("duree", ""))
        cell(8,  act.get("tags", ""))
        cell(9,  act.get("prix", ""))
        cell(10, act.get("url", ""),       hyperlink=act.get("url",""),       link=bool(act.get("url")))
        cell(11, act.get("image_url", ""), hyperlink=act.get("image_url",""), link=bool(act.get("image_url")))
        ws.row_dimensions[row_num].height = 22

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(n_cols)}3"

    # Onglet stats
    ws2 = wb.create_sheet("📊 Stats")
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 14
    stats = [
        ("Total activités",   len(activities)),
        ("Pages scrapées",    TOTAL_PAGES),
        ("Avec prix",         sum(1 for a in activities if a.get("prix"))),
        ("Avec durée",        sum(1 for a in activities if a.get("duree"))),
        ("Avec âge minimum",  sum(1 for a in activities if a.get("age_min"))),
        ("Avec date",         sum(1 for a in activities if a.get("date"))),
        ("Date extraction",   datetime.now().strftime("%d/%m/%Y %H:%M")),
    ]
    for ri, (label, val) in enumerate(stats, 1):
        ws2.cell(ri, 1, label).font = Font(name="Arial", bold=True, size=11)
        ws2.cell(ri, 2, val).font   = Font(name="Arial", size=11)
        ws2.row_dimensions[ri].height = 22

    wb.save(filename)
    print(f"💾  Fichier Excel : {filename}  ({len(activities)} activités)")

# ── MAIN ──────────────────────────────────────────────────────────────────────

async def main():
    print("=" * 65)
    print("  FamilinParis Scraper v2 — Playwright (rendu JavaScript)")
    print("=" * 65)

    activities = await scraper_toutes_les_pages()

    if not activities:
        print("\n⚠️  Aucune activité collectée.")
        return

    save_excel(activities, OUTPUT_FILE)

    print("\n📋  Aperçu des 5 premières :")
    print(f"  {'#':>3}  {'Titre':<45}  {'Lieu':<25}  Prix")
    print("  " + "-" * 90)
    for i, a in enumerate(activities[:5], 1):
        print(
            f"  {i:>3}  {a.get('titre','')[:43]:<45}  "
            f"{a.get('lieu','')[:23]:<25}  "
            f"{a.get('prix','') or '(N/A)'}"
        )
    if len(activities) > 5:
        print(f"  … et {len(activities)-5} autres dans {OUTPUT_FILE}")
    print(f"\n✅  Terminé !\n")

if __name__ == "__main__":
    asyncio.run(main())