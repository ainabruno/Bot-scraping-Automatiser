#!/usr/bin/env python3
"""
╔══════════════════════════════════════════════════════════════════════════╗
║   SCRAPER AGENCES DE VOYAGES PARIS — Vrai Chrome CDP (v3)               ║
║                                                                          ║
║  ÉTAPE 1 — Lancer Chrome debug :                                        ║
║    chrome.exe --remote-debugging-port=9222 --user-data-dir=C:\ChromeBot ║
║    OU double-cliquer lancer_chrome.bat                                   ║
║  ÉTAPE 2 — python scraper_agences_v3.py                                 ║
╚══════════════════════════════════════════════════════════════════════════╝
"""

import asyncio, re, random, time, urllib.request, sys
from pathlib import Path
from datetime import datetime

import pandas as pd
from playwright.async_api import async_playwright, TimeoutError as PWTimeout

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────────────────────────────────────
CDP_URL        = "http://127.0.0.1:9222"
OUTPUT_FILE    = f"agences_voyages_paris_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
MAX_WORKERS    = 1      # 1 seul onglet pour éviter les blocages Google
PAUSE_MIN      = 3.0    # pause min entre requêtes (secondes)
PAUSE_MAX      = 6.0    # pause max
SAVE_EVERY     = 10     # sauvegarde intermédiaire toutes les N requêtes

# ─────────────────────────────────────────────────────────────────────────────
# REQUÊTES PAR MARCHÉ
# ─────────────────────────────────────────────────────────────────────────────
MARCHES = {
    "USA": [
        "travel agency Paris France tours New York",
        "travel agency Paris France tours Los Angeles",
        "travel agency Paris France tours Chicago",
        "travel agency Paris France tours Miami",
        "luxury travel agency France Europe USA",
        "honeymoon travel agency Paris France",
        "France tour operator United States",
        "Paris travel agency group tours America",
        "French travel specialist USA",
        "Europe travel agency Paris package USA",
        "Paris France travel agency Boston",
        "Paris France travel agency Seattle",
        "Paris France travel agency Washington DC",
    ],
    "Chine": [
        "旅行社 巴黎 法国 北京",
        "旅行社 巴黎 法国 上海",
        "travel agency Paris France China Beijing",
        "travel agency Paris France Shanghai",
        "France tour operator China",
        "Europe luxury tour China Paris",
        "法国旅行社 巴黎 欧洲",
    ],
    "Japon": [
        "旅行会社 パリ フランス 東京",
        "travel agency Paris France Tokyo Japan",
        "travel agency Paris France Osaka Japan",
        "France tour operator Japan",
        "luxury France tour Japan honeymoon",
    ],
    "Coree_du_Sud": [
        "여행사 파리 프랑스 서울",
        "travel agency Paris France Seoul Korea",
        "France tour operator Korea",
        "Europe travel package Korea Paris",
    ],
    "Inde": [
        "travel agency Paris France Mumbai India",
        "travel agency Paris France Delhi India",
        "France tour operator India",
        "Paris honeymoon package India travel agency",
    ],
    "Singapour": [
        "travel agency Paris France Singapore",
        "France tour operator Singapore",
        "luxury travel Paris package Singapore",
    ],
    "Arabie_Saoudite": [
        "وكالة سفر باريس فرنسا الرياض",
        "travel agency Paris France Riyadh Saudi Arabia",
        "travel agency Paris France Jeddah Saudi Arabia",
        "France tour operator Saudi Arabia",
    ],
    "Emirats_Arabes_Unis": [
        "وكالة سفر باريس فرنسا دبي",
        "travel agency Paris France Dubai UAE",
        "travel agency Paris France Abu Dhabi",
        "luxury travel agency Dubai Paris France",
    ],
    "Qatar": [
        "travel agency Paris France Doha Qatar",
        "وكالة سفر باريس فرنسا الدوحة",
        "France tour operator Qatar",
    ],
    "Agents_Independants_USA": [
        "independent travel agent Paris France specialist USA",
        "luxury travel advisor France Paris USA",
        "Paris France travel advisor host agency USA",
        "travel consultant France Paris USA independent",
        "France travel specialist independent agent",
    ],
}

# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def extraire_email(texte: str) -> str:
    if not texte:
        return ""
    m = re.findall(r"\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}\b", texte)
    return m[0] if m else ""

def extraire_domaine(url: str) -> str:
    if not url:
        return ""
    m = re.search(r"(?:https?://)?(?:www\.)?([^/\s]+)", url)
    return m.group(1).lower() if m else ""

def nettoyer_titre(titre: str) -> str:
    """Garde la partie la plus descriptive d'un titre Google."""
    for sep in [" - ", " | ", " – ", " — ", " · "]:
        if sep in titre:
            parties = [p.strip() for p in titre.split(sep)]
            titre = max(parties, key=len)
    return titre.strip()

def wait_for_chrome(max_attempts: int = 10) -> bool:
    print(f"🔍 Connexion à Chrome sur {CDP_URL}...")
    for i in range(1, max_attempts + 1):
        try:
            urllib.request.urlopen(f"{CDP_URL}/json/version", timeout=2)
            print("✅ Chrome détecté !\n")
            return True
        except Exception:
            print(f"   ⏳ Tentative {i}/{max_attempts}...")
            time.sleep(2)
    return False

async def is_blocked(page) -> bool:
    """Détecte Google reCAPTCHA / page sorry / 429."""
    try:
        url = page.url
        if any(x in url for x in ("google.com/sorry", "recaptcha", "/sorry/")):
            return True
        content = await page.inner_text("body")
        if "reCAPTCHA" in content or "Je ne suis pas un robot" in content:
            return True
        if "Our systems have detected unusual traffic" in content:
            return True
    except Exception:
        pass
    return False

# ─────────────────────────────────────────────────────────────────────────────
# EXTRACTION DES RÉSULTATS GOOGLE
# ─────────────────────────────────────────────────────────────────────────────

async def extraire_resultats_page(page, query: str) -> list[dict]:
    """Extrait tous les résultats organiques de la page Google ouverte."""
    resultats = []

    # Mots-clés pour filtrer les résultats hors-sujet
    mots_cles = ["travel", "tour", "agence", "voyage", "agency",
                 "旅行", "여행", "سفر", "holiday", "vacation", "trip"]

    try:
        # ── Résultats organiques ──────────────────────────────────────────────
        liens = await page.query_selector_all("div.g, div[data-hveid]")
        for lien in liens:
            try:
                # Titre
                titre_el = await lien.query_selector("h3")
                titre = (await titre_el.inner_text()).strip() if titre_el else ""
                if not titre:
                    continue

                # URL
                url_el = await lien.query_selector("a[href]")
                url = ""
                if url_el:
                    href = await url_el.get_attribute("href")
                    if href and href.startswith("http") and "google" not in href:
                        url = href

                # Snippet
                snippet_el = await lien.query_selector(
                    "div.VwiC3b, span.aCOpRe, div[data-sncf], .IsZvec"
                )
                snippet = ""
                if snippet_el:
                    snippet = (await snippet_el.inner_text()).strip()[:250]

                texte_total = (titre + " " + url + " " + snippet).lower()
                if not any(m.lower() in texte_total for m in mots_cles):
                    continue

                resultats.append({
                    "nom":              nettoyer_titre(titre),
                    "adresse":          "",
                    "telephone":        "",
                    "email":            extraire_email(snippet),
                    "site_web":         url,
                    "note":             "",
                    "nb_avis":          "",
                    "categorie":        "Agence de voyage",
                    "lien_google_maps": "",
                    "requete_source":   query,
                    "snippet":          snippet,
                })
            except Exception:
                continue

        # ── Local Pack (3 fiches Google Maps en haut) ─────────────────────────
        local_items = await page.query_selector_all("div.VkpGBb, div.cXedhc, .rllt__details")
        for item in local_items:
            try:
                txt = await item.inner_text()
                lines = [l.strip() for l in txt.split("\n") if l.strip()]
                if not lines:
                    continue
                nom   = lines[0]
                adres = lines[1] if len(lines) > 1 else ""
                note  = ""
                note_m = re.search(r"(\d[.,]\d)", txt)
                if note_m:
                    note = note_m.group(1)

                # Lien Maps
                lien_el = await item.query_selector("a[href*='maps']")
                lien = ""
                if lien_el:
                    lien = await lien_el.get_attribute("href") or ""

                resultats.append({
                    "nom":              nom,
                    "adresse":          adres,
                    "telephone":        "",
                    "email":            "",
                    "site_web":         "",
                    "note":             note,
                    "nb_avis":          "",
                    "categorie":        "Agence de voyage (Maps)",
                    "lien_google_maps": lien,
                    "requete_source":   query,
                    "snippet":          adres,
                })
            except Exception:
                continue

    except Exception as e:
        print(f"      ⚠ Extraction : {e}")

    return resultats

# ─────────────────────────────────────────────────────────────────────────────
# RECHERCHE — navigation par frappe dans la barre Google (anti-blocage)
# ─────────────────────────────────────────────────────────────────────────────

async def search_google(page, query: str) -> tuple[list[dict], bool]:
    """
    1. Va sur google.com
    2. Tape la requête dans la barre de recherche
    3. Extrait les résultats
    Retourne (liste_resultats, a_encore_des_pages)
    """
    print(f"   🔍 {query}")
    try:
        # Aller sur Google
        await page.goto("https://www.google.com",
                        wait_until="domcontentloaded", timeout=20_000)
        await asyncio.sleep(random.uniform(0.8, 1.5))

        # Trouver la barre de recherche
        search_bar = None
        for sel in ["textarea[name='q']", "input[name='q']", "textarea#APjFqb"]:
            try:
                el = await page.wait_for_selector(sel, timeout=5_000)
                if el:
                    search_bar = el
                    break
            except Exception:
                pass

        if not search_bar:
            print("   ⚠️  Barre de recherche introuvable")
            return [], False

        # Taper la requête
        await search_bar.click()
        await asyncio.sleep(random.uniform(0.2, 0.5))
        await search_bar.fill("")
        await search_bar.type(query, delay=random.randint(35, 80))
        await asyncio.sleep(random.uniform(0.3, 0.7))
        await search_bar.press("Enter")

        # Attendre les résultats
        try:
            await page.wait_for_selector("#rso, #search", timeout=15_000)
        except PWTimeout:
            print("   ⚠️  Résultats non chargés (timeout)")
            return [], False

        # Vérifier si bloqué
        if await is_blocked(page):
            print("   🚨 Bloqué par Google ! Pause 120s...")
            await asyncio.sleep(120)
            return [], False

        # Pause humaine avant extraction
        await asyncio.sleep(random.uniform(PAUSE_MIN, PAUSE_MAX))

        resultats = await extraire_resultats_page(page, query)

        # Vérifier s'il y a une page suivante
        suivant = await page.query_selector("a#pnnext, a[aria-label='Page suivante']")
        a_encore = suivant is not None

        return resultats, a_encore

    except PWTimeout:
        print("   ⚠️  Timeout navigation")
        return [], False
    except Exception as e:
        print(f"   ⚠️  Erreur : {e}")
        return [], False


async def search_page_suivante(page, query: str, page_num: int) -> tuple[list[dict], bool]:
    """Clique sur 'Suivant' pour aller à la page suivante."""
    try:
        suivant = await page.query_selector("a#pnnext, a[aria-label='Page suivante']")
        if not suivant:
            return [], False

        await suivant.click()
        await page.wait_for_selector("#rso, #search", timeout=15_000)

        if await is_blocked(page):
            print("   🚨 Bloqué ! Pause 120s...")
            await asyncio.sleep(120)
            return [], False

        await asyncio.sleep(random.uniform(PAUSE_MIN, PAUSE_MAX))
        resultats = await extraire_resultats_page(page, query)

        suivant2 = await page.query_selector("a#pnnext, a[aria-label='Page suivante']")
        return resultats, (suivant2 is not None)

    except Exception as e:
        print(f"   ⚠️  Page suivante erreur : {e}")
        return [], False

# ─────────────────────────────────────────────────────────────────────────────
# SCRAPING PAR MARCHÉ
# ─────────────────────────────────────────────────────────────────────────────

MAX_PAGES = 3  # pages Google par requête

async def scraper_marche_async(page, nom_marche: str, requetes: list[str]) -> pd.DataFrame:
    print(f"\n{'='*60}")
    print(f"  MARCHÉ : {nom_marche}  ({len(requetes)} requêtes)")
    print(f"{'='*60}")

    tous = []
    vus  = set()

    for idx, requete in enumerate(requetes, 1):
        print(f"\n[{idx:02d}/{len(requetes)}] {requete}")

        # Page 1 — nouvelle recherche
        print(f"   → page 1 ...", end=" ", flush=True)
        resultats, a_encore = await search_google(page, requete)

        nouveaux = _dedup(resultats, vus, tous)
        print(f"{len(resultats)} résultats | {nouveaux} nouveaux | total {len(tous)}")

        # Pages suivantes (clic sur Suivant)
        for p in range(2, MAX_PAGES + 1):
            if not a_encore:
                break
            print(f"   → page {p} ...", end=" ", flush=True)
            resultats, a_encore = await search_page_suivante(page, requete, p)
            nouveaux = _dedup(resultats, vus, tous)
            print(f"{len(resultats)} résultats | {nouveaux} nouveaux | total {len(tous)}")

            if not resultats:
                break

        # Pause entre requêtes
        await asyncio.sleep(random.uniform(PAUSE_MIN, PAUSE_MAX))

    print(f"\n✅ {nom_marche} : {len(tous)} agences uniques")
    return pd.DataFrame(tous) if tous else pd.DataFrame()


def _dedup(resultats: list, vus: set, tous: list) -> int:
    """Déduplique par (nom normalisé, domaine). Retourne le nb de nouveaux."""
    nouveaux = 0
    for h in resultats:
        domaine = extraire_domaine(h.get("site_web", ""))
        cle_nom = h["nom"].lower().strip()[:50]
        cle = (cle_nom, domaine)
        if cle not in vus and h["nom"]:
            vus.add(cle)
            tous.append(h)
            nouveaux += 1
    return nouveaux

# ─────────────────────────────────────────────────────────────────────────────
# ANALYSE COMPLÉTUDE
# ─────────────────────────────────────────────────────────────────────────────

def analyser_completude(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    df = df.copy()
    statuts, manquants, priorites = [], [], []

    for _, row in df.iterrows():
        m = []
        if not str(row.get("telephone", "")).strip(): m.append("téléphone")
        if not str(row.get("email", "")).strip():     m.append("email")
        if not str(row.get("site_web", "")).strip():  m.append("site_web")
        if not str(row.get("adresse", "")).strip():   m.append("adresse")

        if not m:
            statuts.append("✅ COMPLET");    priorites.append("A")
        elif m == ["email"]:
            statuts.append("⚠️ EMAIL MANQUANT"); priorites.append("B")
        elif "telephone" in m and "adresse" in m:
            statuts.append("❌ INCOMPLET");  priorites.append("C")
        else:
            statuts.append("⚠️ À COMPLÉTER"); priorites.append("B")

        manquants.append(", ".join(m) if m else "Aucun")

    df["statut"]           = statuts
    df["champs_manquants"] = manquants
    df["priorite"]         = priorites
    df["email_manuel"]     = ""
    df["tel_manuel"]       = ""
    df["notes"]            = ""

    ordre = {"C": 0, "B": 1, "A": 2}
    df["_o"] = df["priorite"].map(ordre)
    df = df.sort_values(["_o", "nom"]).drop(columns=["_o"]).reset_index(drop=True)
    return df

# ─────────────────────────────────────────────────────────────────────────────
# EXPORT EXCEL
# ─────────────────────────────────────────────────────────────────────────────

COLS = [
    "priorite", "statut", "champs_manquants",
    "nom", "adresse", "telephone", "tel_manuel",
    "email", "email_manuel", "site_web",
    "note", "nb_avis", "snippet", "lien_google_maps", "notes",
]

HEADERS_FR = {
    "priorite": "Priorité", "statut": "Statut",
    "champs_manquants": "Champs manquants", "nom": "Nom de l'agence",
    "adresse": "Adresse", "telephone": "Téléphone (auto)",
    "tel_manuel": "Téléphone (manuel)", "email": "Email (auto)",
    "email_manuel": "Email (manuel)", "site_web": "🌐 Site web",
    "note": "/5", "nb_avis": "Nb avis", "snippet": "Description",
    "lien_google_maps": "🔗 Google Maps", "notes": "Notes",
}

LARGEURS = {
    "priorite": 8, "statut": 22, "champs_manquants": 20,
    "nom": 38, "adresse": 35, "telephone": 16, "tel_manuel": 16,
    "email": 28, "email_manuel": 28, "site_web": 40,
    "note": 7, "nb_avis": 8, "snippet": 45, "lien_google_maps": 26, "notes": 30,
}

COULEURS = {
    "USA": "DEEBF7", "Chine": "FCE4D6", "Japon": "FFF2CC",
    "Coree_du_Sud": "E2EFDA", "Inde": "F4E1D2", "Singapour": "D9EAD3",
    "Arabie_Saoudite": "EAD1DC", "Emirats_Arabes_Unis": "D0E4F5",
    "Qatar": "F5E6FA", "Agents_Independants_USA": "E8F5E9",
}

def _fill(couleur):
    from openpyxl.styles import PatternFill
    return PatternFill("solid", start_color=couleur, end_color=couleur)

def ecrire_onglet(wb, df_in, titre, couleur):
    from openpyxl.styles import Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet(titre)
    thin = Side(style="thin", color="D0D0D0")
    bord = Border(bottom=thin, right=thin)
    font_h    = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    font_bold = Font(name="Arial", bold=True, size=10)
    font_norm = Font(name="Arial", size=10)
    font_link = Font(name="Arial", color="0563C1", underline="single", size=10)
    n = len(COLS)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n)
    c = ws.cell(1, 1, f"AGENCES DE VOYAGES — {titre}")
    c.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    c.fill = _fill("1F4E78")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    if df_in.empty:
        ws.cell(2, 1, "Aucun résultat.").font = font_norm
        return

    df = df_in.copy()
    for col in COLS:
        if col not in df.columns:
            df[col] = ""

    for ci, col in enumerate(COLS, 1):
        c = ws.cell(2, ci, HEADERS_FR.get(col, col))
        c.font = font_h; c.fill = _fill("1F4E78")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 38

    for ri, (_, row) in enumerate(df[COLS].iterrows(), 3):
        for ci, col in enumerate(COLS, 1):
            val = row[col]
            c   = ws.cell(ri, ci)
            if col in ("lien_google_maps", "site_web") and val:
                c.value = str(val); c.font = font_link; c.hyperlink = str(val)
            else:
                c.value = val if pd.notna(val) and val != "" else ""
                c.font  = font_bold if col == "nom" else font_norm
            c.fill = _fill(couleur)
            c.alignment = Alignment(vertical="top",
                wrap_text=(col in ("adresse", "champs_manquants", "notes", "snippet")))
            c.border = bord
        ws.row_dimensions[ri].height = 28

    for ci, col in enumerate(COLS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = LARGEURS.get(col, 15)

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(n)}2"


def export_excel(resultats_par_marche: dict, fichier: str):
    from openpyxl import Workbook
    wb = Workbook(); wb.remove(wb.active)

    # Onglet mode d'emploi
    ws0 = wb.create_sheet("📋 MODE D'EMPLOI")
    ws0.column_dimensions["A"].width = 90
    lignes = [
        "MODE D'EMPLOI — AGENCES DE VOYAGES PARIS",
        "", "🗂️  STRUCTURE",
        "  • Un onglet par marché (USA, Chine, Japon, etc.)",
        "  • Données extraites de Google via vrai Chrome (CDP)",
        "", "📋  UTILISATION",
        "  1. Trier par colonne Priorité (C = incomplet en priorité)",
        "  2. Cliquer sur 🌐 Site web pour visiter l'agence",
        "  3. Trouver l'email sur la page Contact / About",
        "  4. Remplir Email (manuel) et Téléphone (manuel)",
        "", "💡  ASTUCES",
        "  • Hunter.io pour trouver les emails",
        "  • Consulter ASTA (asta.org) pour les USA",
        "", f"  📅 Scraping : {datetime.now().strftime('%d/%m/%Y à %H:%M')}",
    ]
    from openpyxl.styles import Font, Alignment
    for i, t in enumerate(lignes, 1):
        c = ws0.cell(i, 1, t)
        bold = any(t.startswith(s) for s in ["MODE", "🗂️", "📋", "💡"])
        c.font = Font(name="Arial", bold=bold, size=11 if i == 1 else 10)
        c.alignment = Alignment(vertical="top", wrap_text=True)
        ws0.row_dimensions[i].height = 20

    # Onglet stats
    ws1 = wb.create_sheet("📊 STATISTIQUES")
    ws1.column_dimensions["A"].width = 35
    ws1.column_dimensions["B"].width = 12
    from openpyxl.styles import Font, Alignment
    ws1.merge_cells("A1:B1")
    c = ws1.cell(1, 1, "STATISTIQUES")
    c.font = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    c.fill = _fill("1F4E78")
    c.alignment = Alignment(horizontal="center")
    row = 2
    total = 0
    for m, df in resultats_par_marche.items():
        nb = len(df) if df is not None and not df.empty else 0
        total += nb
        ws1.cell(row, 1, f"🌍 {m.replace('_',' ')}").font = Font(name="Arial", size=10)
        ws1.cell(row, 2, nb).font = Font(name="Arial", size=10)
        row += 1
    ws1.cell(row, 1, "TOTAL").font = Font(name="Arial", bold=True, size=11)
    ws1.cell(row, 2, total).font = Font(name="Arial", bold=True, size=11)

    # Onglets par marché
    for marche, df in resultats_par_marche.items():
        if df is None or df.empty:
            print(f"   ⚠ '{marche}' vide — ignoré")
            continue
        ecrire_onglet(wb, df, marche.replace("_", " "), COULEURS.get(marche, "F2F2F2"))
        print(f"   ✅ '{marche.replace('_',' ')}' — {len(df)} agences")

    wb.save(fichier)
    print(f"\n✅ Fichier : {fichier}  |  Total : {total} agences")

# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

async def main():
    print("\n" + "="*65)
    print("  SCRAPER AGENCES — Vrai Chrome CDP")
    print("="*65)

    if not wait_for_chrome():
        print("\n❌ Chrome non disponible sur le port 9222.")
        print("   Lance d'abord :")
        print('   chrome.exe --remote-debugging-port=9222 --user-data-dir=C:\\ChromeBot')
        sys.exit(1)

    marchés_list = list(MARCHES.keys())
    total_req    = sum(len(v) for v in MARCHES.values())
    print(f"\n   Marchés  : {marchés_list}")
    print(f"   Requêtes : {total_req} au total")
    print(f"   Pages/req: {MAX_PAGES}")
    print(f"   Pause    : {PAUSE_MIN}–{PAUSE_MAX}s\n")

    rep = input("▶️  Lancer le scraping ? (o/n) : ").strip().lower()
    if rep != "o":
        print("Annulé.")
        return

    t0 = time.time()
    resultats_par_marche = {}

    async with async_playwright() as pw:
        browser = await pw.chromium.connect_over_cdp(CDP_URL)
        context = (browser.contexts[0] if browser.contexts
                   else await browser.new_context(viewport={"width": 1366, "height": 900}))
        page = await context.new_page()
        print(f"✅ Connecté à Chrome {browser.version}\n")

        for idx_m, (marche, requetes) in enumerate(MARCHES.items(), 1):
            print(f"\n[Marché {idx_m}/{len(MARCHES)}]")
            df = await scraper_marche_async(page, marche, requetes)
            if not df.empty:
                df = analyser_completude(df)
            resultats_par_marche[marche] = df

            # Sauvegarde intermédiaire tous les SAVE_EVERY marchés
            if idx_m % max(1, SAVE_EVERY // 5) == 0:
                temp = OUTPUT_FILE.replace(".xlsx", "_temp.xlsx")
                export_excel(resultats_par_marche, temp)
                print(f"   💾 Sauvegarde intermédiaire → {temp}")

        await page.close()

    export_excel(resultats_par_marche, OUTPUT_FILE)

    elapsed = int(time.time() - t0)
    total   = sum(len(df) for df in resultats_par_marche.values()
                  if df is not None and not df.empty)

    print("\n" + "="*65)
    print(f"  ✅ TERMINÉ en {elapsed//60}m{elapsed%60:02d}s")
    print(f"  📊 {total} agences uniques")
    print(f"  💾 {OUTPUT_FILE}")
    print("="*65)


if __name__ == "__main__":
    asyncio.run(main())