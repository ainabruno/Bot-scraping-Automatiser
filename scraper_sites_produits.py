import re
import json
import random
import asyncio
import argparse
import hashlib
from pathlib import Path
from datetime import datetime

import pandas as pd
from bs4 import BeautifulSoup
from playwright.async_api import async_playwright, TimeoutError as PWTimeout
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── CONFIG ───────────────────────────────────────────────────────────────────
CHECKPOINT_FILE = Path(".checkpoint_sites.json")
USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:123.0) Gecko/20100101 Firefox/123.0",
]

# ─── DICTIONNAIRE DE DÉTECTION PRODUIT ────────────────────────────────────────
# Format : mot_clé → (catégorie_produit, score_points)
# Plus le score est élevé, plus le mot est spécifique/fiable
KEYWORDS = {
    # ── Café ──────────────────────────────────────────────────────────────────
    "torréfacteur":     ("Café torréfié artisanal", 10),
    "torréfaction":     ("Café torréfié artisanal", 10),
    "café de spécialité":("Café torréfié artisanal", 10),
    "specialty coffee": ("Café torréfié artisanal", 10),
    "grain de café":    ("Café torréfié artisanal", 8),
    "café moulu":       ("Café torréfié artisanal", 8),
    "café en grain":    ("Café torréfié artisanal", 8),
    "arabica":          ("Café torréfié artisanal", 6),
    "robusta":          ("Café torréfié artisanal", 6),
    "espresso":         ("Café torréfié artisanal", 5),

    # ── Thé / Infusions ───────────────────────────────────────────────────────
    "thé vert":         ("Thé / Infusion", 10),
    "thé noir":         ("Thé / Infusion", 10),
    "thé blanc":        ("Thé / Infusion", 10),
    "thé rouge":        ("Thé / Infusion", 10),
    "infusion":         ("Thé / Infusion", 8),
    "tisane":           ("Thé / Infusion", 8),
    "rooibos":          ("Thé / Infusion", 8),
    "maté":             ("Thé / Infusion", 7),
    "herbal tea":       ("Thé / Infusion", 7),
    "plantes médicinales":("Thé / Infusion", 6),
    "herboristerie":    ("Thé / Infusion", 6),

    # ── Épices / Mélanges ──────────────────────────────────────────────────────
    "épices":           ("Épices / Mélanges", 8),
    "mélange d'épices": ("Épices / Mélanges", 10),
    "ras el hanout":    ("Épices / Mélanges", 10),
    "curry":            ("Épices / Mélanges", 8),
    "curcuma":          ("Épices / Mélanges", 7),
    "cumin":            ("Épices / Mélanges", 7),
    "paprika":          ("Épices / Mélanges", 7),
    "poivre":           ("Épices / Mélanges", 6),
    "cannelle":         ("Épices / Mélanges", 6),
    "gingembre":        ("Épices / Mélanges", 6),
    "herbes de provence":("Épices / Mélanges", 8),
    "fines herbes":     ("Épices / Mélanges", 7),
    "moutarde artisanale":("Épices / Mélanges", 9),
    "condiment":        ("Épices / Mélanges", 5),

    # ── Granola / Céréales ─────────────────────────────────────────────────────
    "granola":          ("Granola / Céréales", 10),
    "muesli":           ("Granola / Céréales", 10),
    "flocons d'avoine": ("Granola / Céréales", 9),
    "céréales artisanales":("Granola / Céréales", 9),
    "barre de céréales":("Granola / Céréales", 8),
    "porridge":         ("Granola / Céréales", 8),
    "biscuit bio":      ("Granola / Céréales", 7),
    "biscuit artisanal":("Granola / Céréales", 7),
    "farine bio":       ("Granola / Céréales", 6),

    # ── Fruits secs ───────────────────────────────────────────────────────────
    "fruits secs":      ("Fruits secs", 10),
    "noix de cajou":    ("Fruits secs", 10),
    "amande":           ("Fruits secs", 8),
    "noisette":         ("Fruits secs", 8),
    "noix":             ("Fruits secs", 7),
    "pistache":         ("Fruits secs", 8),
    "raisin sec":       ("Fruits secs", 8),
    "abricot sec":      ("Fruits secs", 8),
    "cranberry":        ("Fruits secs", 8),
    "goji":             ("Fruits secs", 9),
    "baies séchées":    ("Fruits secs", 9),
    "fruits déshydratés":("Fruits secs", 10),
    "lyophilisé":       ("Fruits secs", 8),

    # ── Compléments alimentaires / Poudres / Superfoods ───────────────────────
    "complément alimentaire":("Compléments alimentaires", 10),
    "superfood":        ("Compléments alimentaires", 10),
    "super aliment":    ("Compléments alimentaires", 10),
    "protéine en poudre":("Compléments alimentaires", 10),
    "whey":             ("Compléments alimentaires", 10),
    "spiruline":        ("Compléments alimentaires", 10),
    "chlorelle":        ("Compléments alimentaires", 10),
    "maca":             ("Compléments alimentaires", 9),
    "ashwagandha":      ("Compléments alimentaires", 9),
    "collagène":        ("Compléments alimentaires", 9),
    "probiotique":      ("Compléments alimentaires", 9),
    "prébiotique":      ("Compléments alimentaires", 9),
    "naturopathie":     ("Compléments alimentaires", 7),
    "phytothérapie":    ("Compléments alimentaires", 7),
    "poudre nutritionnelle":("Compléments alimentaires", 10),
    "nutrition sportive":("Compléments alimentaires", 8),

    # ── Cacao / Chocolat en poudre ────────────────────────────────────────────
    "cacao en poudre":  ("Cacao / Chocolat en poudre", 10),
    "chocolat en poudre":("Cacao / Chocolat en poudre", 10),
    "fève de cacao":    ("Cacao / Chocolat en poudre", 9),
    "cacao cru":        ("Cacao / Chocolat en poudre", 9),
    "chocolat artisanal":("Cacao / Chocolat en poudre", 7),
    "tablette de chocolat":("Cacao / Chocolat en poudre", 6),
    "praline":          ("Cacao / Chocolat en poudre", 6),

    # ── Produits bio conditionnés ──────────────────────────────────────────────
    "agriculture biologique":("Produits bio conditionnés", 7),
    "certifié bio":     ("Produits bio conditionnés", 7),
    "label bio":        ("Produits bio conditionnés", 7),
    "ab certifié":      ("Produits bio conditionnés", 8),
    "vrac bio":         ("Produits bio conditionnés", 9),
    "épicerie fine":    ("Produits bio conditionnés", 6),
    "produit naturel":  ("Produits bio conditionnés", 5),

    # ── Confitures / Conserves artisanales ────────────────────────────────────
    "confiture artisanale":("Confitures / Conserves artisanales", 10),
    "confiture maison": ("Confitures / Conserves artisanales", 10),
    "gelée de fruits":  ("Confitures / Conserves artisanales", 9),
    "conserve artisanale":("Confitures / Conserves artisanales", 9),
    "bocal artisanal":  ("Confitures / Conserves artisanales", 7),
}

# Mots qui signalent HORS CIBLE même si d'autres mots correspondent
NEGATIVE_KEYWORDS = [
    "vins", "champagne", "bière", "alcool", "distillerie",
    "boucherie", "charcuterie", "abattoir", "viande",
    "fromagerie", "fromage",
    "boulangerie", "pâtisserie",
    "restaurant", "traiteur", "livraison de repas",
    "plomberie", "électronique", "btp", "escalier", "transport",
]

# Pages produits à chercher en priorité
PRODUCT_PAGES = [
    "/produits", "/products", "/boutique", "/shop", "/store",
    "/nos-produits", "/notre-gamme", "/gamme", "/catalogue",
    "/cafe", "/café", "/the", "/the-infusion", "/epices", "/epicerie",
    "/granola", "/fruits-secs", "/complements", "/superfoods",
]


# ═══════════════════════════════════════════════════════════════════════════════
#  CHECKPOINT
# ═══════════════════════════════════════════════════════════════════════════════

class Checkpoint:
    def __init__(self):
        self.data: dict = {}
        if CHECKPOINT_FILE.exists():
            try:
                self.data = json.loads(CHECKPOINT_FILE.read_text(encoding="utf-8"))
                print(f"📂 Checkpoint : {len(self.data)} sites déjà analysés")
            except Exception:
                pass

    def save(self, url: str, result: dict):
        self.data[url] = result
        CHECKPOINT_FILE.write_text(
            json.dumps(self.data, ensure_ascii=False, indent=2), encoding="utf-8"
        )

    def get(self, url: str) -> dict | None:
        return self.data.get(url)

    def clear(self):
        self.data = {}
        if CHECKPOINT_FILE.exists():
            CHECKPOINT_FILE.unlink()


# ═══════════════════════════════════════════════════════════════════════════════
#  DÉTECTION PRODUIT À PARTIR DU TEXTE
# ═══════════════════════════════════════════════════════════════════════════════

def detect_product(text: str, url: str = "") -> dict:
    """
    Analyse le texte d'une page web et retourne :
    - produit     : catégorie détectée
    - score       : niveau de confiance (0-100)
    - mots_cles   : mots trouvés
    - hors_cible  : True si produit non-doypack
    """
    text_low = text.lower()

    # Check mots négatifs
    neg_found = [n for n in NEGATIVE_KEYWORDS if n in text_low]

    # Comptage des mots-clés positifs par catégorie
    scores: dict[str, int] = {}
    found_kws: dict[str, list] = {}

    for kw, (cat, pts) in KEYWORDS.items():
        if kw in text_low:
            scores[cat]    = scores.get(cat, 0) + pts
            found_kws[cat] = found_kws.get(cat, []) + [kw]

    if not scores:
        return {
            "produit":   "Indéterminé",
            "score":     0,
            "mots_cles": "",
            "hors_cible": len(neg_found) > 0,
            "neg_trouves": ", ".join(neg_found[:3]),
        }

    # Catégorie gagnante
    best_cat = max(scores, key=lambda c: scores[c])
    best_score = scores[best_cat]
    # Score normalisé sur 100
    score_norm = min(100, int(best_score / max(scores.values()) * 100)) if scores else 0
    kws = found_kws.get(best_cat, [])[:5]

    return {
        "produit":    best_cat,
        "score":      best_score,
        "mots_cles":  ", ".join(kws),
        "hors_cible": len(neg_found) > 2 and best_score < 10,
        "neg_trouves": ", ".join(neg_found[:3]),
        "all_scores": {k: v for k, v in sorted(scores.items(), key=lambda x: -x[1])},
    }


def extract_text(html: str, max_chars: int = 8000) -> str:
    """Extrait le texte propre d'une page HTML."""
    try:
        soup = BeautifulSoup(html, "html.parser")
        # Supprimer scripts, styles, nav, footer
        for tag in soup(["script", "style", "nav", "footer", "header", "meta"]):
            tag.decompose()
        text = soup.get_text(separator=" ", strip=True)
        # Nettoyer espaces multiples
        text = re.sub(r"\s+", " ", text)
        return text[:max_chars]
    except Exception:
        return html[:max_chars]


# ═══════════════════════════════════════════════════════════════════════════════
#  SCRAPING D'UN SITE
# ═══════════════════════════════════════════════════════════════════════════════

async def scrape_site(page, site: str, timeout_ms: int = 15000) -> dict:
    """
    Visite un site web :
    1. Homepage
    2. Cherche une page produits si homepage insuffisante
    Retourne le résultat de détection.
    """
    # Normaliser l'URL
    if not site.startswith("http"):
        site = "https://" + site
    site = site.rstrip("/")

    all_text = ""
    pages_visited = []

    # ── Étape 1 : Homepage ────────────────────────────────────────────────────
    try:
        await page.goto(site, wait_until="domcontentloaded", timeout=timeout_ms)
        await page.wait_for_timeout(random.randint(800, 1500))
        html = await page.content()
        text = extract_text(html)
        all_text += text
        pages_visited.append(site)
    except Exception as e:
        # Essayer HTTP si HTTPS échoue
        if site.startswith("https://"):
            try:
                await page.goto(site.replace("https://", "http://"),
                                wait_until="domcontentloaded", timeout=timeout_ms)
                html = await page.content()
                all_text += extract_text(html)
                pages_visited.append(site.replace("https://", "http://"))
            except Exception:
                return {"produit": "Site inaccessible", "score": 0,
                        "mots_cles": "", "hors_cible": False,
                        "pages_visitees": "", "erreur": str(e)[:100]}
        else:
            return {"produit": "Site inaccessible", "score": 0,
                    "mots_cles": "", "hors_cible": False,
                    "pages_visitees": "", "erreur": str(e)[:100]}

    # Détecter sur homepage
    result = detect_product(all_text, site)

    # ── Étape 2 : Page produits si score faible (<8) ──────────────────────────
    if result["score"] < 8:
        # Chercher d'abord via les liens de la page
        try:
            links = await page.eval_on_selector_all(
                "a[href]",
                "els => els.map(e => e.getAttribute('href'))"
            )
            product_links = []
            for href in links:
                if not href:
                    continue
                href_low = href.lower()
                # Chercher lien produit dans les hrefs de la page
                if any(p.strip("/") in href_low for p in PRODUCT_PAGES):
                    full = href if href.startswith("http") else site + href
                    product_links.append(full)

            # Visiter jusqu'à 2 pages produits
            for prod_url in list(dict.fromkeys(product_links))[:2]:
                try:
                    await page.goto(prod_url, wait_until="domcontentloaded",
                                    timeout=timeout_ms)
                    await page.wait_for_timeout(random.randint(500, 1000))
                    html2 = await page.content()
                    all_text += " " + extract_text(html2)
                    pages_visited.append(prod_url)
                    break
                except Exception:
                    pass

        except Exception:
            pass

        # Re-détecter avec texte enrichi
        if len(pages_visited) > 1:
            result = detect_product(all_text, site)

    result["pages_visitees"] = " | ".join(pages_visited[:3])
    result.pop("all_scores", None)
    return result


# ═══════════════════════════════════════════════════════════════════════════════
#  WORKER PARALLÈLE
# ═══════════════════════════════════════════════════════════════════════════════

STEALTH_JS = """
Object.defineProperty(navigator, 'webdriver', { get: () => undefined });
Object.defineProperty(navigator, 'languages', { get: () => ['fr-FR','fr','en-US'] });
window.chrome = { runtime: {} };
"""

async def worker(
    worker_id: int,
    queue: asyncio.Queue,
    results: list,
    checkpoint: Checkpoint,
    browser,
    timeout_ms: int,
):
    ctx  = await browser.new_context(
        viewport={"width": random.randint(1280, 1400), "height": 800},
        locale="fr-FR",
        user_agent=random.choice(USER_AGENTS),
        extra_http_headers={"Accept-Language": "fr-FR,fr;q=0.9"},
    )
    await ctx.add_init_script(STEALTH_JS)
    page = await ctx.new_page()

    # Bloquer images/fonts pour aller plus vite
    await page.route(
        "**/*.{png,jpg,jpeg,gif,svg,woff,woff2,ttf,css}",
        lambda r: r.abort()
    )

    print(f"🚀 Worker {worker_id} prêt")

    try:
        while True:
            try:
                i, total, idx, site, row_data = queue.get_nowait()
            except asyncio.QueueEmpty:
                break

            print(f"[W{worker_id}] [{i+1}/{total}] {str(row_data.get('Entreprise',''))[:40]:40s} → {site}")

            # Checkpoint
            cached = checkpoint.get(site)
            if cached:
                site_result = cached
                print(f"   ⏭️  Cache — {site_result.get('produit','?')} (score:{site_result.get('score',0)})")
            else:
                site_result = await scrape_site(page, site, timeout_ms)
                checkpoint.save(site, site_result)

            produit  = site_result.get("produit", "Indéterminé")
            score    = site_result.get("score", 0)
            mots     = site_result.get("mots_cles", "")
            icon = "✅" if score >= 8 else ("⚠️ " if score > 0 else "❌")
            print(f"   {icon} {produit} (score:{score}) | mots: {mots}")

            results.append({
                "idx": idx,
                "produit_site":   produit,
                "score_detection": score,
                "mots_cles_trouves": mots,
                "pages_visitees": site_result.get("pages_visitees", ""),
                "hors_cible_site": site_result.get("hors_cible", False),
                "erreur_site":    site_result.get("erreur", ""),
            })
            queue.task_done()

    finally:
        await ctx.close()


# ═══════════════════════════════════════════════════════════════════════════════
#  SAUVEGARDE EXCEL
# ═══════════════════════════════════════════════════════════════════════════════

def save_excel(df_original: pd.DataFrame, site_results: list, output_file: str):
    # Fusionner résultats dans le df original
    df = df_original.copy()
    df['Produit_Detecte']      = ""
    df['Score_Detection']      = ""
    df['Mots_Cles_Trouves']    = ""
    df['Pages_Visitees']       = ""
    df['Hors_Cible_Site']      = ""

    for r in site_results:
        idx = r["idx"]
        df.at[idx, 'Produit_Detecte']   = r.get("produit_site", "")
        df.at[idx, 'Score_Detection']   = r.get("score_detection", 0)
        df.at[idx, 'Mots_Cles_Trouves'] = r.get("mots_cles_trouves", "")
        df.at[idx, 'Pages_Visitees']    = r.get("pages_visitees", "")
        df.at[idx, 'Hors_Cible_Site']   = "OUI" if r.get("hors_cible_site") else ""

    # Colonnes finales
    COLS = [
        'Entreprise', 'Ville',
        'Produit_Detecte', 'Score_Detection', 'Mots_Cles_Trouves',
        'Nom_Dirigeant', 'Fonction',
        'Tel_Mobile_Direct', 'Tel_Entreprise', 'Site_Internet',
        'Source', 'Adresse', 'Note_Google', 'Nb_Avis',
        'Pages_Visitees', 'Hors_Cible_Site',
    ]
    available = [c for c in COLS if c in df.columns]
    extra = [c for c in df.columns if c not in available]
    df = df[available + extra]

    # Couleurs
    C_HEADER = "1F4E79"
    thin = Side(style="thin", color="CCCCCC")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook()

    # ── Onglet 1 : Tous les leads ──────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Tous les leads"

    COL_WIDTHS = {
        'Entreprise': 30, 'Ville': 18, 'Produit_Detecte': 28,
        'Score_Detection': 8, 'Mots_Cles_Trouves': 30,
        'Nom_Dirigeant': 20, 'Fonction': 25,
        'Tel_Mobile_Direct': 22, 'Tel_Entreprise': 16,
        'Site_Internet': 28, 'Source': 20,
        'Adresse': 32, 'Note_Google': 8, 'Nb_Avis': 8,
        'Pages_Visitees': 35, 'Hors_Cible_Site': 10,
    }

    for ci, col in enumerate(df.columns, 1):
        c = ws1.cell(1, ci, col)
        c.font      = Font(bold=True, name="Arial", size=9, color="FFFFFF")
        c.fill      = PatternFill("solid", fgColor=C_HEADER)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = bdr
        ws1.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(col, 15)
    ws1.row_dimensions[1].height = 30

    for ri, (_, row) in enumerate(df.iterrows(), 2):
        score = row.get("Score_Detection", 0)
        try:
            score_int = int(score) if str(score).strip() not in ("", "nan") else 0
        except Exception:
            score_int = 0

        if score_int >= 12:
            row_bg = "E8F5E9"   # vert — très bon match
        elif score_int >= 6:
            row_bg = "FFF9C4"   # jaune — match partiel
        elif score_int > 0:
            row_bg = "FFF3E0"   # orange — faible
        else:
            row_bg = "FAFAFA"   # gris — pas de site ou inaccessible

        for ci, col in enumerate(df.columns, 1):
            val = row.get(col, "")
            if pd.isna(val): val = ""
            c = ws1.cell(ri, ci, str(val) if val not in (None, "") else "")
            c.border = bdr
            c.font   = Font(name="Arial", size=9)

            if col == 'Score_Detection':
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.fill = PatternFill("solid", fgColor=row_bg)
            elif col == 'Hors_Cible_Site' and str(val) == "OUI":
                c.fill = PatternFill("solid", fgColor="FFCDD2")
                c.alignment = Alignment(horizontal="center")
            elif col == 'Tel_Mobile_Direct':
                c.fill = PatternFill("solid", fgColor="FFF9C4")
                c.alignment = Alignment(horizontal="center")
            else:
                c.fill = PatternFill("solid", fgColor=row_bg)
                c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        ws1.row_dimensions[ri].height = 16

    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}1"

    # ── Onglet 2 : Leads doypack-compatibles (score ≥ 6) ──────────────────
    df_good = df[df['Score_Detection'].apply(
        lambda x: int(x) >= 6 if str(x).strip() not in ("", "nan") else False
    )].copy()
    ws2 = wb.create_sheet(f"✅ Compatibles ({len(df_good)})")
    for ci, col in enumerate(df_good.columns, 1):
        c = ws2.cell(1, ci, col)
        c.font  = Font(bold=True, name="Arial", size=9, color="FFFFFF")
        c.fill  = PatternFill("solid", fgColor="2E7D32")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = bdr
        ws2.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(col, 15)
    ws2.row_dimensions[1].height = 28

    for ri, (_, row) in enumerate(df_good.iterrows(), 2):
        for ci, col in enumerate(df_good.columns, 1):
            val = row.get(col, "")
            if pd.isna(val): val = ""
            c = ws2.cell(ri, ci, str(val) if val not in (None, "") else "")
            c.border = bdr
            c.font   = Font(name="Arial", size=9)
            c.fill   = PatternFill("solid", fgColor="E8F5E9" if ri % 2 == 0 else "F1F8E9")
            c.alignment = Alignment(
                horizontal="center" if col in ("Score_Detection", "Note_Google", "Nb_Avis") else "left",
                vertical="center", wrap_text=True
            )
        ws2.row_dimensions[ri].height = 16
    ws2.freeze_panes = "A2"

    # ── Onglet 3 : Stats ──────────────────────────────────────────────────
    ws3 = wb.create_sheet("📊 Stats")
    ws3.column_dimensions['A'].width = 38
    ws3.column_dimensions['B'].width = 12

    ws3.merge_cells("A1:B1")
    t = ws3.cell(1, 1, "Analyse sites web — Détection produits")
    t.font      = Font(bold=True, name="Arial", size=13, color="FFFFFF")
    t.fill      = PatternFill("solid", fgColor=C_HEADER)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 28

    total      = len(df)
    with_site  = df['Site_Internet'].astype(str).str.strip().ne("").sum()
    detected   = df[df['Score_Detection'].apply(
        lambda x: int(x) > 0 if str(x).strip() not in ("", "nan") else False)].shape[0]
    good       = len(df_good)

    stats_rows = [
        ("RÉSULTATS ANALYSE", ""),
        ("Total leads", total),
        ("Avec site web", with_site),
        ("Produit détecté (score > 0)", detected),
        ("✅ Compatibles doypack (score ≥ 6)", good),
        ("", ""),
        ("PRODUITS DÉTECTÉS", ""),
    ]
    prod_counts = df_good['Produit_Detecte'].value_counts()
    for prod, cnt in prod_counts.items():
        stats_rows.append((f"   {prod}", cnt))

    stats_rows += [
        ("", ""),
        ("LÉGENDE COULEURS", ""),
        ("🟢 Vert — score ≥ 12 : très bon match", ""),
        ("🟡 Jaune — score 6–11 : match correct", ""),
        ("🟠 Orange — score 1–5 : faible indice", ""),
        ("⚪ Gris — site inaccessible ou score 0", ""),
    ]

    sec_fill = PatternFill("solid", fgColor="D6E4F0")
    for ri, (lbl, val) in enumerate(stats_rows, 3):
        is_sec = (val == "" and lbl and not lbl.startswith("🟢") and not lbl.startswith("🟡")
                  and not lbl.startswith("🟠") and not lbl.startswith("⚪"))
        c1 = ws3.cell(ri, 1, lbl)
        c2 = ws3.cell(ri, 2, val if val != "" else "")
        c1.font = Font(name="Arial", size=10, bold=is_sec)
        c2.font = Font(name="Arial", size=10, bold=True, color="1F4E79")
        if is_sec:
            c1.fill = sec_fill
            c2.fill = sec_fill
        ws3.row_dimensions[ri].height = 16

    wb.save(output_file)
    print(f"\n💾 Sauvegardé : {output_file}")
    print(f"   Total       : {total}")
    print(f"   Avec site   : {with_site}")
    print(f"   Détectés    : {detected}")
    print(f"   ✅ Compat.  : {good}")


# ═══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════════════════

async def main(args):
    # Lecture Excel
    try:
        df = pd.read_excel(args.input, encoding="utf-8-sig") if args.input.endswith(".csv") \
             else pd.read_excel(args.input)
    except Exception:
        df = pd.read_csv(args.input, encoding="latin-1")

    if args.limit:
        df = df.head(args.limit)

    print(f"\n📊 {len(df)} leads chargés")

    checkpoint = Checkpoint()
    if args.reset_checkpoint:
        checkpoint.clear()
        print("🗑️  Checkpoint effacé.")

    # Construire queue des sites à visiter
    queue   = asyncio.Queue()
    results = []
    count   = 0
    for idx, row in df.iterrows():
        site = str(row.get("Site_Internet", "")).strip()
        if site and site.lower() not in ("nan", "", "none"):
            queue.put_nowait((count, df['Site_Internet'].notna().sum(), idx, site, row.to_dict()))
            count += 1

    print(f"🌐 {queue.qsize()} sites à analyser avec {args.workers} workers\n")

    timeout_ms = args.timeout * 1000

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(
            headless=True,
            args=["--no-sandbox", "--disable-blink-features=AutomationControlled",
                  "--disable-dev-shm-usage", "--blink-settings=imagesEnabled=false"],
        )

        workers = [
            worker(w, queue, results, checkpoint, browser, timeout_ms)
            for w in range(min(args.workers, queue.qsize() or 1))
        ]

        # Auto-save toutes les 60s
        async def auto_save():
            while True:
                await asyncio.sleep(60)
                if results:
                    save_excel(df, results, args.output)
                    print(f"   💾 Auto-save ({len(results)} sites traités)")

        saver = asyncio.create_task(auto_save())
        await asyncio.gather(*workers)
        saver.cancel()
        await browser.close()

    save_excel(df, results, args.output)


def parse_args():
    p = argparse.ArgumentParser(description="Scrape les sites web pour détecter les produits")
    p.add_argument("--input",   required=True,  help="Fichier Excel source (.xlsx)")
    p.add_argument("--output",  required=True,  help="Fichier Excel résultat (.xlsx)")
    p.add_argument("--workers", type=int, default=4,  help="Pages parallèles (défaut: 4)")
    p.add_argument("--timeout", type=int, default=15, help="Timeout par site en secondes (défaut: 15)")
    p.add_argument("--limit",   type=int,             help="Tester sur N premières lignes")
    p.add_argument("--reset-checkpoint", action="store_true",
                   help="Efface le cache et re-analyse tout")
    return p.parse_args()


if __name__ == "__main__":
    asyncio.run(main(parse_args()))
