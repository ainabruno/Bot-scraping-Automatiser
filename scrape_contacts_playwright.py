# """
# scrape_contacts_playwright.py
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Script de scraping Playwright pour enrichir les contacts Airbnb Paris.
# Pour chaque fournisseur, il cherche sur Google :
#   • Téléphone
#   • Email
#   • Site web
#   • Adresse complète (si manquante)

# UTILISATION :
#   pip install playwright openpyxl pandas
#   playwright install chromium
#   python scrape_contacts_playwright.py

# Le script reprend là où il s'est arrêté (checkpoint JSON).
# Résultat final : airbnb_paris_contacts_ENRICHI.xlsx
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# """

# import asyncio
# import json
# import logging
# import random
# import re
# import time
# from pathlib import Path

# import pandas as pd
# from playwright.async_api import async_playwright

# # ── Config ─────────────────────────────────────────────────
# INPUT_FILE  = "airbnb_paris_contacts_sans_dup.xlsx"
# OUTPUT_FILE = "airbnb_paris_contacts_ENRICHI_1.xlsx"
# CHECKPOINT  = "checkpoint.json"

# HEADLESS          = True          # False → voir le navigateur
# CONCURRENCY       = 3             # requêtes simultanées (max 5 pour éviter ban)
# DELAY_MIN         = 1.5           # délai min entre requêtes (secondes)
# DELAY_MAX         = 3.5           # délai max
# REQUEST_TIMEOUT   = 15_000        # ms

# logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
# log = logging.getLogger(__name__)

# # ── Regex patterns ──────────────────────────────────────────
# PHONE_RE = re.compile(
#     r"(?:\+33[\s.\-]?|0)"
#     r"(?:[1-9](?:[\s.\-]?\d{2}){4})"
# )
# EMAIL_RE = re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}")
# URL_RE   = re.compile(
#     r"https?://(?:www\.)?"
#     r"(?!(?:airbnb|google|facebook|tripadvisor|booking|maps)\.)[\w\-]+\.[a-z]{2,}"
#     r"(?:/[\w\-._~:/?#\[\]@!$&'()*+,;=%]*)?"
# )

# # ── Helpers ─────────────────────────────────────────────────
# def clean_phone(raw: str) -> str:
#     digits = re.sub(r"[^\d+]", "", raw)
#     if digits.startswith("33"):
#         digits = "+33" + digits[2:]
#     elif digits.startswith("0") and len(digits) == 10:
#         digits = "+33" + digits[1:]
#     return digits if len(digits) >= 10 else ""


# def extract_from_text(text: str) -> dict:
#     result = {"telephone": "", "email": "", "site_web": ""}

#     phones = PHONE_RE.findall(text)
#     if phones:
#         cleaned = clean_phone(phones[0])
#         if cleaned:
#             result["telephone"] = cleaned

#     emails = EMAIL_RE.findall(text)
#     if emails:
#         # ignorer emails génériques / noreply
#         for e in emails:
#             if not any(x in e.lower() for x in ["noreply", "no-reply", "example", "test@"]):
#                 result["email"] = e.lower()
#                 break

#     urls = URL_RE.findall(text)
#     if urls:
#         result["site_web"] = urls[0].rstrip("/")

#     return result


# def build_query(row: pd.Series) -> str:
#     """Construit la requête Google optimale pour ce fournisseur."""
#     fournisseur = str(row.get("Fournisseur", "")).strip()
#     adresse     = str(row.get("Adresse complète", "")).strip().split(",")[0]

#     # Si le fournisseur ressemble à un prénom seul → ajouter contexte
#     if len(fournisseur.split()) <= 1 or fournisseur.lower() in ("dawit","loik","matthieu","adrian"):
#         experience = str(row.get("Nom de l'expérience", "")).strip()
#         return f'"{fournisseur}" Paris contact site téléphone "{experience[:40]}"'
    
#     return f'"{fournisseur}" Paris contact téléphone email site officiel'


# # ── Checkpoint ──────────────────────────────────────────────
# def load_checkpoint() -> dict:
#     if Path(CHECKPOINT).exists():
#         with open(CHECKPOINT) as f:
#             return json.load(f)
#     return {}


# def save_checkpoint(data: dict) -> None:
#     with open(CHECKPOINT, "w") as f:
#         json.dump(data, f, ensure_ascii=False, indent=2)


# # ── Playwright scraping ─────────────────────────────────────
# async def google_search(page, query: str) -> str:
#     """Effectue une recherche Google et retourne le texte brut de la page."""
#     try:
#         url = f"https://www.google.com/search?q={query}&hl=fr&gl=fr&num=5"
#         await page.goto(url, timeout=REQUEST_TIMEOUT, wait_until="domcontentloaded")

#         # Accepter cookies si présent
#         try:
#             btn = page.locator('button:has-text("Tout accepter"), button:has-text("Accept all")')
#             if await btn.count():
#                 await btn.first.click()
#                 await page.wait_for_timeout(500)
#         except Exception:
#             pass

#         # Attendre les résultats
#         await page.wait_for_selector("#search, #rso", timeout=10_000)
#         text = await page.inner_text("#search")
#         return text

#     except Exception as e:
#         log.warning(f"Search error for '{query[:50]}': {e}")
#         return ""


# async def get_site_details(context, site_url: str) -> dict:
#     """Visite le site officiel pour extraire téléphone et email."""
#     result = {"telephone": "", "email": ""}
#     if not site_url:
#         return result

#     # Pages contact courantes
#     paths = ["", "/contact", "/nous-contacter", "/about", "/a-propos"]

#     page = await context.new_page()
#     try:
#         for path in paths:
#             try:
#                 url = site_url.rstrip("/") + path
#                 await page.goto(url, timeout=REQUEST_TIMEOUT, wait_until="domcontentloaded")
#                 text = await page.inner_text("body")
#                 extracted = extract_from_text(text)
#                 if extracted["telephone"]:
#                     result["telephone"] = extracted["telephone"]
#                 if extracted["email"]:
#                     result["email"] = extracted["email"]
#                 if result["telephone"] and result["email"]:
#                     break
#                 await asyncio.sleep(0.5)
#             except Exception:
#                 continue
#     finally:
#         await page.close()

#     return result


# async def process_row(semaphore, context, index: int, row: pd.Series, checkpoint: dict) -> dict:
#     """Traite une ligne : cherche sur Google, puis visite le site."""
#     key = str(index)
#     if key in checkpoint:
#         log.info(f"[{index+1}] Déjà traité (checkpoint) — {row['Fournisseur']}")
#         return checkpoint[key]

#     result = {
#         "index": index,
#         "telephone": "",
#         "email": "",
#         "site_web": "",
#         "adresse_trouvee": "",
#     }

#     query = build_query(row)
#     log.info(f"[{index+1}] Recherche : {row['Fournisseur']} | Query: {query[:60]}")

#     async with semaphore:
#         page = await context.new_page()
#         try:
#             # 1) Recherche Google
#             text = await google_search(page, query)
#             extracted = extract_from_text(text)

#             result["telephone"] = extracted["telephone"]
#             result["email"]     = extracted["email"]
#             result["site_web"]  = extracted["site_web"]

#             # 2) Si on a un site, on visite pour plus de détails
#             if result["site_web"] and (not result["telephone"] or not result["email"]):
#                 details = await get_site_details(context, result["site_web"])
#                 result["telephone"] = result["telephone"] or details["telephone"]
#                 result["email"]     = result["email"]     or details["email"]

#             # 3) Délai aléatoire anti-ban
#             await asyncio.sleep(random.uniform(DELAY_MIN, DELAY_MAX))

#         except Exception as e:
#             log.error(f"[{index+1}] Erreur : {e}")
#         finally:
#             await page.close()

#     checkpoint[key] = result
#     save_checkpoint(checkpoint)

#     log.info(
#         f"[{index+1}] ✅ {row['Fournisseur']} → "
#         f"tel={result['telephone'] or '—'} | "
#         f"email={result['email'] or '—'} | "
#         f"site={result['site_web'] or '—'}"
#     )
#     return result


# # ── Main ────────────────────────────────────────────────────
# async def main():
#     df = pd.read_excel(INPUT_FILE)
#     checkpoint = load_checkpoint()

#     log.info(f"📋 {len(df)} lignes à traiter | Checkpoint: {len(checkpoint)} déjà faits")

#     semaphore = asyncio.Semaphore(CONCURRENCY)

#     async with async_playwright() as pw:
#         browser = await pw.chromium.launch(
#             headless=HEADLESS,
#             args=[
#                 "--no-sandbox",
#                 "--disable-blink-features=AutomationControlled",
#                 "--disable-dev-shm-usage",
#             ],
#         )

#         # Contexte avec user-agent réaliste
#         context = await browser.new_context(
#             user_agent=(
#                 "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
#                 "AppleWebKit/537.36 (KHTML, like Gecko) "
#                 "Chrome/124.0.0.0 Safari/537.36"
#             ),
#             locale="fr-FR",
#             timezone_id="Europe/Paris",
#             viewport={"width": 1280, "height": 800},
#         )

#         # Masquer les traces d'automatisation
#         await context.add_init_script(
#             "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
#         )

#         tasks = [
#             process_row(semaphore, context, i, row, checkpoint)
#             for i, row in df.iterrows()
#         ]
#         results = await asyncio.gather(*tasks)

#         await browser.close()

#     # ── Mise à jour du DataFrame ────────────────────────────
#     results_map = {r["index"]: r for r in results}

#     for i, row in df.iterrows():
#         r = results_map.get(i, {})
#         if r.get("telephone"):
#             df.at[i, "Telephone"] = r["telephone"]
#         if r.get("email"):
#             df.at[i, "email"] = r["email"]
#         if r.get("site_web"):
#             df.at[i, "Site web"] = r["site_web"]

#     # ── Export Excel ────────────────────────────────────────
#     from openpyxl import load_workbook
#     from openpyxl.styles import Font, PatternFill, Alignment, PatternFill
#     from openpyxl.utils import get_column_letter

#     df.to_excel(OUTPUT_FILE, index=False, engine="openpyxl")

#     # Mise en forme
#     wb = load_workbook(OUTPUT_FILE)
#     ws = wb.active

#     # En-têtes
#     header_fill = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
#     header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
#     for cell in ws[1]:
#         cell.fill = header_fill
#         cell.font = header_font
#         cell.alignment = Alignment(horizontal="center", vertical="center")

#     # Mise en évidence des cellules remplies (téléphone / email / site web)
#     cols_highlight = {"Telephone": "D9EAD3", "email": "D9EAD3", "Site web": "D9EAD3"}
#     col_idx = {cell.value: cell.column for cell in ws[1]}

#     for col_name, color in cols_highlight.items():
#         if col_name in col_idx:
#             c = col_idx[col_name]
#             for row in range(2, ws.max_row + 1):
#                 cell = ws.cell(row=row, column=c)
#                 if cell.value:
#                     cell.fill = PatternFill("solid", start_color=color, end_color=color)
#                 cell.font = Font(name="Arial", size=9)

#     # Largeurs auto
#     col_widths = {
#         1: 5, 2: 40, 3: 30, 4: 35, 5: 20, 6: 50, 7: 45, 8: 10, 9: 18, 10: 30, 11: 35
#     }
#     for col, width in col_widths.items():
#         ws.column_dimensions[get_column_letter(col)].width = width

#     ws.freeze_panes = "A2"
#     wb.save(OUTPUT_FILE)

#     # ── Stats finales ───────────────────────────────────────
#     df2 = pd.read_excel(OUTPUT_FILE)
#     tel   = df2["Telephone"].notna().sum()
#     email = df2["email"].notna().sum()
#     site  = df2["Site web"].notna().sum()
#     total = len(df2)

#     print("\n" + "━"*50)
#     print(f"✅ Fichier enrichi : {OUTPUT_FILE}")
#     print(f"   Téléphone trouvé : {tel}/{total} ({tel/total*100:.1f}%)")
#     print(f"   Email trouvé     : {email}/{total} ({email/total*100:.1f}%)")
#     print(f"   Site web trouvé  : {site}/{total} ({site/total*100:.1f}%)")
#     print("━"*50)


# if __name__ == "__main__":
#     asyncio.run(main())

"""
scrape_contacts_playwright.py  v3
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Stratégie multi-sources par type de fournisseur :

  ÉTAPE 1 — Scrape page Airbnb (URL déjà dans le fichier)
            → nom complet hôte, site web, Instagram, description
  ÉTAPE 2 — Google avec le NOM COMPLET trouvé sur Airbnb
            (bien meilleur que juste "Gladys Paris")
  ÉTAPE 3 — Pages Jaunes (pour entreprises)
  ÉTAPE 4 — DuckDuckGo fallback
  ÉTAPE 5 — Visite du site officiel trouvé

UTILISATION :
  pip install playwright openpyxl pandas
  playwright install chromium
  python scrape_contacts_playwright.py
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
"""

import asyncio, json, logging, random, re
from pathlib import Path
from urllib.parse import quote_plus

import pandas as pd
from playwright.async_api import async_playwright, Browser

# ── Config ───────────────────────────────────────────────────
INPUT_FILE  = "airbnb_paris_contacts_sans_dup.xlsx"
OUTPUT_FILE = "airbnb_paris_contacts_ENRICHI.xlsx"
CHECKPOINT  = "checkpoint.json"

HEADLESS     = True
CONCURRENCY  = 2      # workers (2 = safe, 3 = plus rapide mais risque CAPTCHA)
DELAY_MIN    = 1.8
DELAY_MAX    = 3.5
PAGE_TIMEOUT = 14_000

IGNORE_DOMAINS = re.compile(
    r"(airbnb|google|facebook|instagram|twitter|youtube|"
    r"tripadvisor|booking|yelp|maps\.google|wikipedia|linkedin|"
    r"pages-jaunes|pagesjaunes|lafourchette|thefork|"
    r"viator|getyourguide|expedia|klook)\.",
    re.I,
)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# ── Regex ────────────────────────────────────────────────────
PHONE_RE = re.compile(
    r"(?<!\d)(\+33[\s.\-]?[1-9](?:[\s.\-]?\d{2}){4}"
    r"|0[1-9](?:[\s.\-]?\d{2}){4})(?!\d)"
)
EMAIL_RE = re.compile(r"[a-zA-Z0-9._%+\-]{2,}@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,6}")
URL_RE   = re.compile(r"https?://(?:www\.)?([a-zA-Z0-9\-]+\.[a-z]{2,6})(?:/[^\s\"<>\"]*)?")
INSTA_RE = re.compile(r"instagram\.com/([a-zA-Z0-9_.]+)")
BAD_EMAILS = ("noreply","no-reply","example","test@","support@google",
              "privacy@","legal@","schema.org","sentry","wixpress")

# ── Helpers ──────────────────────────────────────────────────
def normalize_phone(raw: str) -> str:
    d = re.sub(r"[^\d+]", "", raw)
    if d.startswith("+33"):
        pass
    elif d.startswith("33") and len(d) == 11:
        d = "+" + d
    elif d.startswith("0") and len(d) == 10:
        d = "+33" + d[1:]
    else:
        return ""
    return d if len(d) >= 11 else ""


def extract(text: str) -> dict:
    out = {"telephone": "", "email": "", "site_web": "", "instagram": ""}
    for m in PHONE_RE.findall(text):
        p = normalize_phone(m)
        if p:
            out["telephone"] = p
            break
    for e in EMAIL_RE.findall(text):
        if not any(b in e.lower() for b in BAD_EMAILS) and len(e) < 80:
            out["email"] = e.lower()
            break
    for m in URL_RE.finditer(text):
        domain = m.group(1)
        if not IGNORE_DOMAINS.search(domain):
            url = m.group(0).rstrip(".,);\"'")
            out["site_web"] = "/".join(url.split("/")[:3])
            break
    ig = INSTA_RE.search(text)
    if ig and ig.group(1) not in ("p", "explore", "reel", "stories", "accounts"):
        out["instagram"] = f"https://instagram.com/{ig.group(1)}"
    return out


def is_company(name: str) -> bool:
    keywords = ["SAS","SARL","SRL","LLC","LTD","TOURS","CLUB","EXPERIENCE",
                "EXPERIENCES","PARIS","FOOD","RIVER","BOAT","COMPANY",
                "GROUP","AGENCY","STUDIO","SCHOOL","ATELIER","ASSOCIATION"]
    n = name.upper()
    return any(k in n for k in keywords)


def build_queries(row: pd.Series, airbnb_data: dict) -> list[str]:
    """
    Construit les requêtes selon le contexte.
    Utilise le nom complet trouvé sur Airbnb si disponible.
    """
    name     = str(row.get("Fournisseur", "")).strip()
    exp      = str(row.get("Nom de l'expérience", "")).strip()
    full_name = airbnb_data.get("host_name", "").strip()
    site      = airbnb_data.get("site_web", "")

    queries = []

    # Si on a déjà un site Airbnb → chercher avec ce nom complet
    if full_name and full_name.lower() != name.lower():
        queries.append(f'"{full_name}" Paris contact téléphone email')
        queries.append(f'"{full_name}" site officiel Paris')

    # Entreprise
    if is_company(name):
        queries.append(f'"{name}" Paris téléphone email site officiel')
        queries.append(f'"{name}" pages jaunes Paris contact')
    # Prénom seul ou multi-personnes → utiliser expérience
    elif len(name.split()) <= 2 or " ou " in name.lower():
        queries.append(f'"{name}" "{exp[:45]}" Paris contact')
        queries.append(f'{name} Paris "{exp[:35]}" site instagram')
    else:
        queries.append(f'"{name}" Paris téléphone email site officiel contact')

    return queries[:3]  # max 3 requêtes par fournisseur


# ── Checkpoint ───────────────────────────────────────────────
def load_cp() -> dict:
    if Path(CHECKPOINT).exists():
        with open(CHECKPOINT, encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_cp(data: dict) -> None:
    with open(CHECKPOINT, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


# ── Page navigation ───────────────────────────────────────────
async def get_text(page, url: str, wait_ms: int = 800) -> str:
    try:
        await page.goto(url, timeout=PAGE_TIMEOUT, wait_until="domcontentloaded")
        await page.wait_for_timeout(wait_ms)
        return await page.evaluate("document.body.innerText || ''")
    except Exception as e:
        log.debug(f"get_text error [{url[:55]}]: {e}")
        return ""


async def get_html(page, url: str) -> str:
    try:
        await page.goto(url, timeout=PAGE_TIMEOUT, wait_until="domcontentloaded")
        await page.wait_for_timeout(1000)
        return await page.content()
    except Exception as e:
        log.debug(f"get_html error [{url[:55]}]: {e}")
        return ""


async def accept_cookies(page):
    for label in ["Tout accepter","Accept all","Accepter tout","J'accepte","Agree","Fermer"]:
        try:
            btn = page.locator(f'button:has-text("{label}")')
            if await btn.count():
                await btn.first.click(timeout=2000)
                await page.wait_for_timeout(400)
                return
        except Exception:
            pass


# ── Source 1 : Page Airbnb ────────────────────────────────────
async def scrape_airbnb(page, airbnb_url: str) -> dict:
    """
    Scrape la page d'expérience Airbnb pour extraire :
    - Nom complet de l'hôte
    - Site web (si mentionné dans description)
    - Instagram (si mentionné)
    - Email (rare mais parfois dans description)
    """
    result = {"host_name": "", "site_web": "", "instagram": "", "email": "", "telephone": ""}
    if not airbnb_url or "airbnb" not in airbnb_url:
        return result

    html = await get_html(page, airbnb_url)
    if not html:
        return result

    # Nom de l'hôte : chercher dans les balises JSON-LD ou meta
    host_patterns = [
        r'"hostName"\s*:\s*"([^"]+)"',
        r'"name"\s*:\s*"([^"]{3,50})".*?"@type"\s*:\s*"Person"',
        r'Hébergé par\s+([A-ZÀÂÄÉÈÊËÎÏÔÖÙÛÜ][a-zàâäéèêëîïôöùûü\s\-]{2,30})',
        r'Hosted by\s+([A-Z][a-zA-Z\s\-]{2,30})',
        r'"givenName"\s*:\s*"([^"]+)"',
    ]
    for pat in host_patterns:
        m = re.search(pat, html)
        if m:
            candidate = m.group(1).strip()
            if 2 < len(candidate) < 50 and candidate.lower() not in ("airbnb","paris"):
                result["host_name"] = candidate
                break

    # Extraire tout le texte pour chercher contacts
    text = re.sub(r'<[^>]+>', ' ', html)
    text = re.sub(r'\s+', ' ', text)
    ex = extract(text)
    result["site_web"]   = result["site_web"]   or ex["site_web"]
    result["instagram"]  = result["instagram"]  or ex["instagram"]
    result["email"]      = result["email"]      or ex["email"]
    result["telephone"]  = result["telephone"]  or ex["telephone"]

    log.debug(f"  Airbnb → host='{result['host_name']}' site='{result['site_web']}' ig='{result['instagram']}'")
    return result


# ── Source 2 : Google ─────────────────────────────────────────
async def google(page, query: str) -> str:
    url  = f"https://www.google.com/search?q={quote_plus(query)}&hl=fr&gl=fr&num=6"
    text = await get_text(page, url, wait_ms=700)
    if not text or "Avant de continuer" in text or "consent.google" in page.url:
        await accept_cookies(page)
        await page.wait_for_timeout(500)
        text = await page.evaluate("document.body.innerText || ''")
    return text


# ── Source 3 : DuckDuckGo ─────────────────────────────────────
async def duckduckgo(page, query: str) -> str:
    url = f"https://html.duckduckgo.com/html/?q={quote_plus(query)}&kl=fr-fr"
    return await get_text(page, url, wait_ms=600)


# ── Source 4 : Pages Jaunes ───────────────────────────────────
async def pages_jaunes(page, name: str, ville: str = "Paris") -> dict:
    result = {"telephone": "", "site_web": "", "email": ""}
    try:
        url = f"https://www.pagesjaunes.fr/recherche/{quote_plus(name)}/{quote_plus(ville)}"
        text = await get_text(page, url, wait_ms=1000)
        ex   = extract(text)
        result["telephone"] = ex["telephone"]
        result["site_web"]  = ex["site_web"]
        result["email"]     = ex["email"]
    except Exception:
        pass
    return result


# ── Source 5 : Visit site ─────────────────────────────────────
async def visit_site(page, site_url: str) -> dict:
    result = {"telephone": "", "email": ""}
    if not site_url:
        return result
    for path in ["", "/contact", "/nous-contacter", "/contact-us", "/about"]:
        text = await get_text(page, site_url.rstrip("/") + path, wait_ms=600)
        ex   = extract(text)
        result["telephone"] = result["telephone"] or ex["telephone"]
        result["email"]     = result["email"]     or ex["email"]
        if result["telephone"] and result["email"]:
            break
        await asyncio.sleep(0.3)
    return result


# ── Source 6 : Instagram (extraire email/site du bio) ─────────
async def scrape_instagram(page, ig_url: str) -> dict:
    result = {"telephone": "", "email": "", "site_web": ""}
    if not ig_url:
        return result
    try:
        text = await get_text(page, ig_url, wait_ms=1500)
        ex   = extract(text)
        result.update({k: v for k, v in ex.items() if v})
    except Exception:
        pass
    return result


# ── Worker ────────────────────────────────────────────────────
async def worker(wid: int, queue: asyncio.Queue, browser: Browser,
                 df: pd.DataFrame, checkpoint: dict, results: dict):
    ctx = await browser.new_context(
        user_agent=(
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            f"Chrome/12{random.randint(0,4)}.0.0.0 Safari/537.36"
        ),
        locale="fr-FR",
        timezone_id="Europe/Paris",
        viewport={"width": 1280, "height": 900},
    )
    await ctx.add_init_script(
        "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
    )
    page = await ctx.new_page()

    try:
        while True:
            try:
                idx = queue.get_nowait()
            except asyncio.QueueEmpty:
                break

            key = str(idx)
            row = df.iloc[idx]
            name      = str(row["Fournisseur"]).strip()
            airbnb_url = str(row.get("URL", "")).strip()

            if key in checkpoint:
                log.info(f"[W{wid}] #{idx+1} SKIP — {name}")
                results[idx] = checkpoint[key]
                queue.task_done()
                continue

            r = {"index": idx, "telephone": "", "email": "", "site_web": "", "instagram": "", "host_name": ""}
            log.info(f"[W{wid}] #{idx+1} {name}")

            # ── ÉTAPE 1 : Scraper la page Airbnb ──────────────
            airbnb_data = await scrape_airbnb(page, airbnb_url)
            r["host_name"] = airbnb_data.get("host_name", "")
            for k in ("telephone", "email", "site_web", "instagram"):
                if airbnb_data.get(k): r[k] = airbnb_data[k]

            # ── ÉTAPE 2 : Google (avec nom complet si dispo) ──
            if not (r["telephone"] and r["email"]):
                queries = build_queries(row, airbnb_data)
                for q in queries:
                    text = await google(page, q)
                    ex   = extract(text)
                    for k, v in ex.items():
                        if v and not r.get(k): r[k] = v
                    if r["telephone"] and r["email"] and r["site_web"]:
                        break
                    await asyncio.sleep(random.uniform(0.8, 1.5))

            # ── ÉTAPE 3 : DuckDuckGo fallback ─────────────────
            if not r["telephone"] and not r["site_web"]:
                q2   = f'"{r["host_name"] or name}" Paris contact téléphone'
                text = await duckduckgo(page, q2)
                ex   = extract(text)
                for k, v in ex.items():
                    if v and not r.get(k): r[k] = v

            # ── ÉTAPE 4 : Pages Jaunes (entreprises) ──────────
            if is_company(name) and not r["telephone"]:
                pj = await pages_jaunes(page, name)
                for k, v in pj.items():
                    if v and not r.get(k): r[k] = v

            # ── ÉTAPE 5 : Visiter le site officiel ────────────
            if r["site_web"] and (not r["telephone"] or not r["email"]):
                details = await visit_site(page, r["site_web"])
                r["telephone"] = r["telephone"] or details["telephone"]
                r["email"]     = r["email"]     or details["email"]

            # ── ÉTAPE 6 : Instagram → extraire email/site ─────
            if r["instagram"] and (not r["email"] or not r["site_web"]):
                ig_data = await scrape_instagram(page, r["instagram"])
                r["site_web"] = r["site_web"] or ig_data.get("site_web","")
                r["email"]    = r["email"]    or ig_data.get("email","")

            log.info(
                f"[W{wid}] #{idx+1} ✅ {name}"
                + (f" ({r['host_name']})" if r['host_name'] and r['host_name'].lower() != name.lower() else "")
                + f"  ☎ {r['telephone'] or '—'}"
                + f"  ✉ {r['email'] or '—'}"
                + f"  🌐 {r['site_web'] or '—'}"
                + (f"  📸 {r['instagram']}" if r['instagram'] else "")
            )

            results[idx]    = r
            checkpoint[key] = r
            save_cp(checkpoint)
            queue.task_done()

            await asyncio.sleep(random.uniform(DELAY_MIN, DELAY_MAX))

    finally:
        await page.close()
        await ctx.close()


# ── Export Excel ──────────────────────────────────────────────
def export_excel(df: pd.DataFrame, results: dict):
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    # Ajouter colonne Instagram et Nom hôte si pas déjà là
    if "Instagram" not in df.columns:
        df.insert(df.columns.get_loc("Site web") + 1, "Instagram", "")
    if "Nom hôte Airbnb" not in df.columns:
        df.insert(df.columns.get_loc("Fournisseur") + 1, "Nom hôte Airbnb", "")

    for idx, r in results.items():
        if r.get("telephone"):   df.at[idx, "Telephone"]       = r["telephone"]
        if r.get("email"):       df.at[idx, "email"]           = r["email"]
        if r.get("site_web"):    df.at[idx, "Site web"]        = r["site_web"]
        if r.get("instagram"):   df.at[idx, "Instagram"]       = r["instagram"]
        if r.get("host_name"):   df.at[idx, "Nom hôte Airbnb"] = r["host_name"]

    df.to_excel(OUTPUT_FILE, index=False, engine="openpyxl")

    wb = load_workbook(OUTPUT_FILE)
    ws = wb.active

    hf   = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
    hfnt = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    for cell in ws[1]:
        cell.fill = hf
        cell.font = hfnt
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30

    col_map = {c.value: c.column for c in ws[1]}
    green  = PatternFill("solid", start_color="D9EAD3", end_color="D9EAD3")
    yellow = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")
    blue   = PatternFill("solid", start_color="D0E4F7", end_color="D0E4F7")

    contact_cols  = ["Telephone", "email", "Site web", "Instagram"]
    enriched_cols = ["Nom hôte Airbnb"]

    for col_name in contact_cols:
        c = col_map.get(col_name)
        if not c: continue
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = Font(name="Arial", size=9)
            cell.fill = green if cell.value else yellow

    for col_name in enriched_cols:
        c = col_map.get(col_name)
        if not c: continue
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = Font(name="Arial", size=9, italic=True)
            if cell.value:
                cell.fill = blue

    # Largeurs
    widths = {"#":5, "Nom de l'expérience":40, "Fournisseur":26, "Nom hôte Airbnb":22,
              "Adresse (rue)":30, "Ville / CP":15, "Adresse complète":40,
              "URL":40, "Statut":9, "Telephone":18, "email":28, "Site web":32, "Instagram":30}
    for i, col in enumerate(df.columns, 1):
        w = widths.get(col, 18)
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(OUTPUT_FILE)

    df2   = pd.read_excel(OUTPUT_FILE)
    total = len(df2)
    tel   = df2["Telephone"].notna().sum()
    mail  = df2["email"].notna().sum()
    site  = df2["Site web"].notna().sum()
    insta = df2["Instagram"].notna().sum() if "Instagram" in df2.columns else 0
    host  = df2["Nom hôte Airbnb"].notna().sum() if "Nom hôte Airbnb" in df2.columns else 0

    print("\n" + "━"*58)
    print(f"  ✅  {OUTPUT_FILE}")
    print(f"  📞  Téléphone      : {tel}/{total}  ({tel/total*100:.1f}%)")
    print(f"  ✉️   Email          : {mail}/{total}  ({mail/total*100:.1f}%)")
    print(f"  🌐  Site web       : {site}/{total}  ({site/total*100:.1f}%)")
    print(f"  📸  Instagram      : {insta}/{total}  ({insta/total*100:.1f}%)")
    print(f"  👤  Nom hôte trouvé: {host}/{total}  ({host/total*100:.1f}%)")
    print("━"*58)


# ── Main ──────────────────────────────────────────────────────
async def main():
    df         = pd.read_excel(INPUT_FILE)
    checkpoint = load_cp()
    results: dict = {}

    todo = [i for i in range(len(df)) if str(i) not in checkpoint]
    log.info(f"📋 Total:{len(df)}  Checkpoint:{len(checkpoint)}  À faire:{len(todo)}")

    for key, val in checkpoint.items():
        results[int(key)] = val

    queue: asyncio.Queue = asyncio.Queue()
    for i in todo:
        await queue.put(i)

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(
            headless=HEADLESS,
            args=[
                "--no-sandbox",
                "--disable-blink-features=AutomationControlled",
                "--disable-dev-shm-usage",
                "--disable-gpu",
            ],
        )
        await asyncio.gather(*[
            worker(wid, queue, browser, df, checkpoint, results)
            for wid in range(1, CONCURRENCY + 1)
        ])
        await browser.close()

    export_excel(df, results)


if __name__ == "__main__":
    asyncio.run(main())
