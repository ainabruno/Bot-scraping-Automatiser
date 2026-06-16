"""
serres_maps_scraper.py
======================
Scraper Google Maps orienté serres agricoles — CEE AGRI-EQ-108
Étape 1 uniquement : trouver ET pré-qualifier les entreprises avec serres probables.
Sales Navigator / Apollo = étapes suivantes, jamais avant.

Fonctionnement :
  1. Recherche Google Maps via SerpAPI avec des requêtes ciblées serres
  2. Score de confiance automatique (nom, catégorie, avis, mots-clés)
  3. Génère un lien satellite Google Maps pour vérification visuelle humaine
  4. Export Excel 3 onglets : A_Serres_Confirmes / B_A_Verifier / C_Rejetes
"""

import time
import re
import requests
import pandas as pd
from urllib.parse import quote

# ── CONFIGURATION ──────────────────────────────────────────────────────────────
SERP_API_KEY = "b4b185517e867b6f276ee8d9ff7b8986a3e90fda5854d9f614200ee5d8084823"
PAUSE_ENTRE_REQUETES = 1.5   # secondes (respecte les limites API)
MAX_RESULTS_PAR_REQUETE = 20  # Google Maps retourne 20 max par page

# ── REQUÊTES CIBLÉES SERRES ────────────────────────────────────────────────────
# Regroupées par type pour éviter les doublons et maximiser la couverture
REQUETES_SERRES = [
    # Maraîchage sous serre
    "serre maraîchère",
    "producteur tomates serre",
    "producteur concombres serre",
    "producteur salades serre",
    "maraîcher serre chauffée",
    "jeunes plants maraîchage",
    # Horticulture sous serre
    "horticulteur serre",
    "pépinière serre",
    "producteur fleurs serre",
    "production florale serre",
    "plantes ornementales serre",
    # Serristes / installateurs (prescripteurs)
    "constructeur serre agricole",
    "installateur serre horticole",
    "serre de production",
    # Coopératives / groupes
    "coopérative maraîchage serre",
    "coopérative horticole",
]

# ── MOTS-CLÉS POUR SCORING ─────────────────────────────────────────────────────

# Dans le nom ou la catégorie → fort signal de serre réelle
KEYWORDS_FORT = [
    "serre", "serres", "maraîch", "horticult", "horticulture",
    "pépinière", "pepiniere", "serriste", "sous serre",
    "tomates", "concombres", "fleurs coupe", "jeunes plants",
    "production florale", "plantes en pot", "plants maraîch",
    "cultures sous", "fraises serre",
]

# Dans le nom ou la catégorie → signal modéré, à vérifier
KEYWORDS_MOYEN = [
    "agricol", "exploitation", "earl", "gaec", "scea", "sas agricole",
    "producteur", "production", "fruits légumes", "légumes",
    "coopérative", "cooperative", "fleurs", "plantes",
    "jardinage", "culture",
]

# Dans le nom ou la catégorie → signal négatif = probablement PAS une serre de production
KEYWORDS_REJET = [
    "jardinerie", "garden center", "bricomarché", "leroy merlin",
    "point vert", "truffaut", "botanic", "gamm vert",
    "paysagiste", "paysage", "entretien espaces verts", "espaces verts",
    "supermarché", "hypermarché", "épicerie",
    "formation", "école", "lycée", "collège",
    "engrais", "fertilisant", "nutrition animale",
    "boulangerie", "boucherie", "restaurant",
    "conseil agricole", "chambre agriculture",
]

# Catégories Google Maps qui indiquent une vraie exploitation
CATEGORIES_SERRES = [
    "farm", "greenhouse", "flower grower", "vegetable farm",
    "plant nursery", "garden center production", "horticulture",
    "ferme", "exploitation agricole", "pépinière de production",
    "producteur de légumes", "producteur de fleurs",
    "serriste", "maraîcher",
]

# ── FONCTIONS UTILITAIRES ──────────────────────────────────────────────────────

def normalise(texte: str) -> str:
    """Lowercase sans accents pour comparaison."""
    if not texte:
        return ""
    t = texte.lower()
    for src, dst in [("é","e"),("è","e"),("ê","e"),("ë","e"),
                     ("à","a"),("â","a"),("ô","o"),("î","i"),
                     ("ù","u"),("û","u"),("ç","c")]:
        t = t.replace(src, dst)
    return t


def lien_satellite(lat: float, lon: float, zoom: int = 18) -> str:
    """Génère l'URL Google Maps satellite pour vérification visuelle."""
    if not lat or not lon:
        return ""
    return f"https://www.google.com/maps/@{lat},{lon},{zoom}z/data=!3m1!1e3"


def lien_maps_fiche(place_id: str) -> str:
    """Lien direct vers la fiche Google Maps."""
    if not place_id:
        return ""
    return f"https://www.google.com/maps/place/?q=place_id:{place_id}"


def scorer_lead(nom: str, categorie: str, nb_avis: int, note: float) -> tuple[int, str]:
    """
    Calcule un score de 0 à 100 et une catégorie de priorité.

    Retourne (score, "A" | "B" | "C" | "REJETE")
    """
    nom_n    = normalise(nom or "")
    cat_n    = normalise(categorie or "")
    texte_n  = nom_n + " " + cat_n

    # ── Vérification rejet immédiat ──
    for kw in KEYWORDS_REJET:
        if normalise(kw) in texte_n:
            return 0, "REJETE", f"rejet: {kw}"

    score = 0
    raisons = []

    # ── Mots-clés forts dans le nom (40 pts max) ──
    for kw in KEYWORDS_FORT:
        if normalise(kw) in texte_n:
            score += 20
            raisons.append(f"mot-cle fort: {kw}")
            if score >= 40:
                break

    # ── Mots-clés moyens (15 pts max, seulement si pas déjà fort) ──
    if score < 40:
        for kw in KEYWORDS_MOYEN:
            if normalise(kw) in texte_n:
                score += 8
                raisons.append(f"mot-cle moyen: {kw}")
                if score >= 15:
                    break

    # ── Catégorie Google Maps pertinente (20 pts) ──
    for cat in CATEGORIES_SERRES:
        if normalise(cat) in cat_n:
            score += 20
            raisons.append(f"categorie: {cat}")
            break

    # ── Taille estimée via nombre d'avis (20 pts) ──
    try:
        n = int(nb_avis) if nb_avis else 0
        if n >= 50:
            score += 20
            raisons.append(f"{n} avis (grande structure probable)")
        elif n >= 15:
            score += 10
            raisons.append(f"{n} avis (structure moyenne)")
        elif n >= 5:
            score += 5
    except (ValueError, TypeError):
        pass

    # ── Bonus forme juridique dans le nom (10 pts) ──
    for forme in ["earl", "gaec", "scea", "sas", "sarl", "sa "]:
        if forme in nom_n:
            score += 10
            raisons.append(f"forme juridique: {forme.upper()}")
            break

    # ── Classification ──
    score = min(score, 100)
    if score >= 60:
        priorite = "A"
    elif score >= 30:
        priorite = "B"
    elif score > 0:
        priorite = "C"
    else:
        priorite = "REJETE"

    raison_str = " | ".join(raisons) if raisons else "aucun signal fort"
    return score, priorite, raison_str


def scrape_une_requete(query: str, ville: str) -> list[dict]:
    """
    Interroge SerpAPI Google Maps pour une requête + ville.
    Retourne une liste de dicts bruts.
    """
    params = {
        "engine":   "google_maps",
        "q":        f"{query} {ville}",
        "api_key":  SERP_API_KEY,
        "hl":       "fr",
        "gl":       "fr",
        "type":     "search",
    }

    try:
        response = requests.get(
            "https://serpapi.com/search",
            params=params,
            timeout=30
        )
        response.raise_for_status()
        data = response.json()

        if "error" in data:
            print(f"      ⚠ Erreur API SerpAPI : {data['error']}")
            return []

        resultats = []
        for place in data.get("local_results", []):
            gps   = place.get("gps_coordinates", {})
            lat   = gps.get("latitude")
            lon   = gps.get("longitude")
            pid   = place.get("place_id", "")
            nb    = place.get("reviews", 0)
            note  = place.get("rating", 0)

            resultats.append({
                "nom":              place.get("title", ""),
                "categorie":        place.get("type", ""),
                "adresse":          place.get("address", ""),
                "telephone":        place.get("phone", ""),
                "site_web":         place.get("website", ""),
                "note":             note,
                "nb_avis":          nb,
                "horaires":         str(place.get("hours", "")),
                "latitude":         lat,
                "longitude":        lon,
                "place_id":         pid,
                "lien_maps_fiche":  lien_maps_fiche(pid),
                "lien_satellite":   lien_satellite(lat, lon),
                "requete_source":   query,
                "ville_recherche":  ville,
            })

        return resultats

    except requests.exceptions.RequestException as e:
        print(f"      ⚠ Erreur réseau : {e}")
        return []


def scraper_serres_ville(ville: str, requetes: list[str] = None) -> pd.DataFrame:
    """
    Lance toutes les requêtes pour une ville et consolide les résultats.
    Déduplique par nom + adresse.
    """
    if requetes is None:
        requetes = REQUETES_SERRES

    tous_resultats = []
    vus = set()  # pour dédupliquation

    for i, requete in enumerate(requetes, 1):
        print(f"  [{i:02d}/{len(requetes)}] '{requete} {ville}'", end=" ... ")
        resultats = scrape_une_requete(requete, ville)

        nouveaux = 0
        for r in resultats:
            cle = (normalise(r["nom"]), normalise(r["adresse"][:20] if r["adresse"] else ""))
            if cle not in vus and r["nom"]:
                vus.add(cle)
                tous_resultats.append(r)
                nouveaux += 1

        print(f"{len(resultats)} trouvés, {nouveaux} nouveaux")
        time.sleep(PAUSE_ENTRE_REQUETES)

    print(f"\n  → Total unique pour {ville} : {len(tous_resultats)}")
    return pd.DataFrame(tous_resultats) if tous_resultats else pd.DataFrame()


def qualifier_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """
    Ajoute les colonnes de scoring et de vérification.
    """
    if df.empty:
        return df

    scores, priorites, raisons = [], [], []

    for _, row in df.iterrows():
        resultat = scorer_lead(
            row.get("nom", ""),
            row.get("categorie", ""),
            row.get("nb_avis", 0),
            row.get("note", 0),
        )
        # scorer_lead retourne 3 valeurs
        score, priorite, raison = resultat
        scores.append(score)
        priorites.append(priorite)
        raisons.append(raison)

    df = df.copy()
    df["score_confiance"]       = scores
    df["priorite"]              = priorites
    df["raison_score"]          = raisons
    df["surface_estimee"]       = "A verifier par appel"
    df["serre_confirmee"]       = "NON - verifier satellite"
    df["statut_verification"]   = "EN ATTENTE"
    df["commentaire"]           = ""

    # Trier par priorité puis score décroissant
    ordre = {"A": 0, "B": 1, "C": 2, "REJETE": 3}
    df["_ordre"] = df["priorite"].map(ordre)
    df = df.sort_values(["_ordre", "score_confiance"], ascending=[True, False])
    df = df.drop(columns=["_ordre"])

    return df.reset_index(drop=True)


def export_excel(df: pd.DataFrame, fichier: str, ville: str):
    """
    Exporte en Excel avec 3 onglets colorés et liens cliquables.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils import get_column_letter

    # Séparer les 3 groupes
    df_a      = df[df["priorite"] == "A"].copy()
    df_b      = df[df["priorite"] == "B"].copy()
    df_c      = df[df["priorite"] == "C"].copy()
    df_rejete = df[df["priorite"] == "REJETE"].copy()

    # Colonnes à afficher (dans l'ordre)
    cols = [
        "priorite", "score_confiance", "serre_confirmee",
        "nom", "categorie", "adresse",
        "telephone", "site_web",
        "note", "nb_avis",
        "lien_satellite", "lien_maps_fiche",
        "surface_estimee", "statut_verification",
        "raison_score", "requete_source", "commentaire",
    ]
    cols_exist = [c for c in cols if c in df.columns]

    # Styles
    fills = {
        "header":  PatternFill("solid", start_color="185FA5", end_color="185FA5"),
        "A":       PatternFill("solid", start_color="EAF3DE", end_color="EAF3DE"),
        "B":       PatternFill("solid", start_color="FAEEDA", end_color="FAEEDA"),
        "C":       PatternFill("solid", start_color="F1EFE8", end_color="F1EFE8"),
        "REJETE":  PatternFill("solid", start_color="FCEBEB", end_color="FCEBEB"),
    }
    wh_font  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    bd_font  = Font(name="Arial", bold=True, size=10)
    nm_font  = Font(name="Arial", size=10)
    lk_font  = Font(name="Arial", size=10, color="185FA5", underline="single")
    thin     = Side(style="thin", color="D3D1C7")
    border   = Border(bottom=thin)

    wb = Workbook()
    wb.remove(wb.active)  # supprimer la feuille vide par défaut

    def ecrire_onglet(wb, df_onglet, titre, fill_data):
        ws = wb.create_sheet(titre)

        if df_onglet.empty:
            ws.cell(1, 1, f"Aucun lead {titre}").font = nm_font
            return

        df_out = df_onglet[cols_exist].copy()

        # En-têtes
        headers = {
            "priorite": "Priorité",
            "score_confiance": "Score /100",
            "serre_confirmee": "Serre confirmée",
            "nom": "Nom entreprise",
            "categorie": "Catégorie Google",
            "adresse": "Adresse",
            "telephone": "Téléphone",
            "site_web": "Site web",
            "note": "Note /5",
            "nb_avis": "Nb avis",
            "lien_satellite": "🛰 Satellite (cliquer)",
            "lien_maps_fiche": "📍 Fiche Maps",
            "surface_estimee": "Surface estimée",
            "statut_verification": "Statut vérif.",
            "raison_score": "Raison du score",
            "requete_source": "Requête source",
            "commentaire": "Commentaire",
        }

        # Ligne de titre
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols_exist))
        titre_cell = ws.cell(1, 1, f"LEADS SERRES — {titre} — {ville.upper()}")
        titre_cell.font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
        titre_cell.fill = fills["header"]
        titre_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 24

        # En-têtes colonnes
        for col_idx, col_name in enumerate(cols_exist, 1):
            cell = ws.cell(2, col_idx, headers.get(col_name, col_name))
            cell.font = wh_font
            cell.fill = fills["header"]
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[2].height = 36

        # Données
        for row_idx, (_, data_row) in enumerate(df_out.iterrows(), 3):
            prio = str(data_row.get("priorite", "C"))
            row_fill = fill_data or fills.get(prio, fills["C"])

            for col_idx, col_name in enumerate(cols_exist, 1):
                val = data_row[col_name]
                cell = ws.cell(row_idx, col_idx)

                # Liens cliquables
                if col_name in ("lien_satellite", "lien_maps_fiche", "site_web") and val:
                    cell.value = str(val)
                    cell.font = lk_font
                    cell.hyperlink = str(val)
                else:
                    cell.value = val if pd.notna(val) else ""
                    cell.font = nm_font if col_name != "nom" else bd_font

                cell.fill = row_fill
                cell.alignment = Alignment(vertical="top", wrap_text=(col_name in ("adresse", "raison_score", "commentaire")))
                cell.border = border

            ws.row_dimensions[row_idx].height = 28

        # Largeurs colonnes
        largeurs = {
            "priorite": 8, "score_confiance": 10, "serre_confirmee": 18,
            "nom": 30, "categorie": 20, "adresse": 35,
            "telephone": 16, "site_web": 28,
            "note": 8, "nb_avis": 8,
            "lien_satellite": 28, "lien_maps_fiche": 28,
            "surface_estimee": 20, "statut_verification": 16,
            "raison_score": 30, "requete_source": 22, "commentaire": 30,
        }
        for col_idx, col_name in enumerate(cols_exist, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = largeurs.get(col_name, 18)

        # Figer les 2 premières lignes
        ws.freeze_panes = "A3"

    # Créer les 4 onglets
    ecrire_onglet(wb, df_a,      "A_Serres_Confirmes",     fills["A"])
    ecrire_onglet(wb, df_b,      "B_A_Verifier",           fills["B"])
    ecrire_onglet(wb, df_c,      "C_Faible_Priorite",      fills["C"])
    ecrire_onglet(wb, df_rejete, "Z_Rejetes",              fills["REJETE"])

    # Onglet instructions
    ws_inst = wb.create_sheet("MODE EMPLOI", 0)
    ws_inst.column_dimensions["A"].width = 12
    ws_inst.column_dimensions["B"].width = 55

    instruct_fill = PatternFill("solid", start_color="185FA5", end_color="185FA5")
    ws_inst.merge_cells("A1:B1")
    c = ws_inst.cell(1, 1, "MODE EMPLOI — VERIFICATION SERRES AVANT SALES NAVIGATOR")
    c.font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    c.fill = instruct_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws_inst.row_dimensions[1].height = 26

    instructions = [
        ("ETAPE 1", "Aller dans l'onglet A_Serres_Confirmes — ce sont les meilleurs candidats"),
        ("ETAPE 2", "Cliquer sur le lien SATELLITE de chaque ligne — s'ouvre dans Google Maps"),
        ("ETAPE 3", "Verifier visuellement : voit-on des serres (grandes structures translucides) ?"),
        ("SI OUI",  "Mettre 'OUI' dans la colonne 'Serre confirmee' + noter la surface estimee"),
        ("SI NON",  "Mettre 'NON' + passer au suivant — ne pas continuer avec SalesNav / Apollo"),
        ("ETAPE 4", "Sur les serres confirmees seulement : chercher l'entreprise sur Pappers"),
        ("ETAPE 5", "Sur les serres confirmees seulement : chercher le dirigeant sur Sales Navigator"),
        ("ETAPE 6", "En dernier : recuperer mobile/email sur Apollo (jamais avant etape 3 validee)"),
        ("REGLE",   "Pas de serre visible = pas de lead. Mieux vaut 10 leads confirmes que 50 douteux."),
    ]
    ok_f  = PatternFill("solid", start_color="EAF3DE", end_color="EAF3DE")
    warn_f= PatternFill("solid", start_color="FAEEDA", end_color="FAEEDA")
    for i, (etape, texte) in enumerate(instructions, 2):
        f = ok_f if i % 2 == 0 else warn_f
        c1 = ws_inst.cell(i, 1, etape)
        c1.font = Font(name="Arial", bold=True, size=10)
        c1.fill = f
        c1.alignment = Alignment(horizontal="center", vertical="top")
        c2 = ws_inst.cell(i, 2, texte)
        c2.font = Font(name="Arial", size=10)
        c2.fill = f
        c2.alignment = Alignment(vertical="top", wrap_text=True)
        ws_inst.row_dimensions[i].height = 22

    wb.save(fichier)
    print(f"\n✅ Fichier sauvegardé : {fichier}")
    print(f"   Onglet A (serres probables) : {len(df_a)} leads")
    print(f"   Onglet B (à vérifier)       : {len(df_b)} leads")
    print(f"   Onglet C (faible signal)    : {len(df_c)} leads")
    print(f"   Rejetés                     : {len(df_rejete)} entrées")


# ── SCRIPT PRINCIPAL ───────────────────────────────────────────────────────────

ZONES_BRETAGNE = {
    "Saint-Pol-de-Leon": [
        "serre maraîchère Saint-Pol-de-Léon",
        "horticulteur serre Finistère nord",
        "pépinière production Saint-Pol-de-Léon",
        "producteur légumes serre Ceinture dorée",
    ],
    "Roscoff": [
        "serre maraîchère Roscoff",
        "producteur artichauts serre Roscoff",
        "coopérative maraîchage Roscoff",
    ],
    "Morlaix": [
        "serre maraîchère Morlaix",
        "horticulteur serre Morlaix",
        "producteur tomates serre Finistère",
    ],
    "Plougastel": [
        "producteur fraises serre Plougastel",
        "serre chauffée Plougastel",
        "maraîcher serre Brest",
    ],
    "Quimper": [
        "serre maraîchère Quimper",
        "pépinière serre Quimper",
        "horticulteur serre Quimper",
    ],
    "Rennes": [
        "serre maraîchère Rennes",
        "producteur tomates serre Rennes",
        "coopérative maraîchage Bretagne",
        "serre chauffée Ille-et-Vilaine",
    ],
}


def run_bretagne():
    """Lance le scraping complet de la Bretagne."""
    print("=" * 60)
    print("SCRAPING SERRES AGRICOLES — BRETAGNE")
    print("Etape 1 : Google Maps uniquement")
    print("Sales Navigator / Apollo = apres validation visuelle")
    print("=" * 60)

    tous_les_leads = []

    for ville, requetes_specifiques in ZONES_BRETAGNE.items():
        print(f"\n📍 {ville}")
        df_ville = scraper_serres_ville(ville, requetes=requetes_specifiques)

        if not df_ville.empty:
            df_qualifie = qualifier_dataframe(df_ville)
            tous_les_leads.append(df_qualifie)

    if not tous_les_leads:
        print("\n❌ Aucun résultat.")
        return

    df_final = pd.concat(tous_les_leads, ignore_index=True)

    # Dédupliquer sur le total (même nom + même ville)
    df_final["_cle"] = df_final["nom"].apply(normalise) + "|" + df_final["ville_recherche"].apply(normalise)
    df_final = df_final.drop_duplicates(subset=["_cle"]).drop(columns=["_cle"])
    df_final = df_final.sort_values(
        ["priorite", "score_confiance"],
        key=lambda x: x.map({"A": 0, "B": 1, "C": 2, "REJETE": 3}) if x.name == "priorite" else x,
        ascending=[True, False]
    ).reset_index(drop=True)

    # Export
    fichier = "serres_CEE_AGRI_Bretagne_ETAPE1_Maps.xlsx"
    export_excel(df_final, fichier, "Bretagne")

    # Résumé console
    print("\n" + "=" * 60)
    print("RÉSUMÉ")
    print("=" * 60)
    for prio in ["A", "B", "C", "REJETE"]:
        n = len(df_final[df_final["priorite"] == prio])
        emoji = {"A": "🟢", "B": "🟡", "C": "⚪", "REJETE": "🔴"}[prio]
        print(f"  {emoji} Priorité {prio} : {n}")
    print(f"\n  Total : {len(df_final)}")
    print("\n⚠  PROCHAINE ETAPE : ouvrir le fichier Excel,")
    print("   cliquer sur chaque lien SATELLITE et confirmer visuellement")
    print("   que des serres sont visibles AVANT d'utiliser Sales Navigator.")


def run_custom():
    """Mode interactif — n'importe quelle ville / région."""
    print("\n" + "=" * 60)
    print("SCRAPER SERRES — MODE PERSONNALISE")
    print("=" * 60)

    villes_input = input("Villes séparées par virgule (ex: Nantes,Angers,Saumur) : ").strip()
    villes = [v.strip() for v in villes_input.split(",") if v.strip()]

    if not villes:
        print("Aucune ville saisie.")
        return

    custom_requetes_input = input(
        "Requêtes personnalisées ? (laisser vide = requêtes standard serres) : "
    ).strip()
    requetes = (
        [r.strip() for r in custom_requetes_input.split(",") if r.strip()]
        or REQUETES_SERRES
    )

    tous = []
    for ville in villes:
        print(f"\n📍 {ville}")
        df_v = scraper_serres_ville(ville, requetes=requetes)
        if not df_v.empty:
            tous.append(qualifier_dataframe(df_v))

    if not tous:
        print("Aucun résultat.")
        return

    df_final = pd.concat(tous, ignore_index=True)
    region = "_".join(villes[:3])
    fichier = f"serres_CEE_{region}_ETAPE1_Maps.xlsx"
    export_excel(df_final, fichier, region)


# ── LANCEMENT ──────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("\n" + "=" * 60)
    print("  SCRAPER SERRES AGRICOLES — CEE AGRI-EQ-108")
    print("  Etape 1 : Maps + scoring + liens satellite")
    print("=" * 60)
    print("\n1. Bretagne complète (planning Jour 1)")
    print("2. Villes personnalisées")
    choix = input("\nChoix (1 ou 2) : ").strip()

    if choix == "2":
        run_custom()
    else:
        run_bretagne()