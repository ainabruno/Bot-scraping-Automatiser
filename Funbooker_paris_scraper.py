"""
Funbooker — Activités Paris Scraper v3
========================================
STRATÉGIE : Grands scrolls JS rapides
  1. Scroll instantané vers le bas (JS) → attend les appels API
  2. Répète jusqu'à saturation (plus rien de nouveau)
  → Beaucoup plus rapide que les petits scrolls de 800px
"""

import asyncio
import re
import time
import urllib.request
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from playwright.async_api import async_playwright, TimeoutError as PWTimeout

# ─── Configuration ─────────────────────────────────────────────────────────────

BASE_URL   = "https://www.funbooker.com"
URL_SEARCH = (
    "https://www.funbooker.com/fr/search/entertainment"
    "?mode=geo_only&hitsPerPage=27&requireAvailabilities=1"
    "&isBirthdayUrl=0&isTeamBuildingUrl=0&where=Paris"
    "&haveRoom=0&haveFood=0&isInstant=0&isTopSales=0"
    "&isPromos=0&isFastResponse=0&hasCateringService=0"
    "&hasMeetingRoom=0&withoutRadius=0"
    "&lat=48.85349500000001&lng=2.34839199999999"
    "&rootCategory=1&categories=1"
)
OUTPUT         = "funbooker_paris_activites_all1.xlsx"
CDP_URL        = "http://127.0.0.1:9222"
TOTAL_EXPECTED = 10745

# Pause APRÈS chaque grand scroll — attend que l'API réponde
WAIT_AFTER_SCROLL = 4.0   # secondes
# Pause supplémentaire si on n'a pas reçu de nouvelles données
WAIT_EXTRA        = 3.0
# Nombre de tentatives sans nouveaux résultats avant abandon
MAX_NO_CHANGE     = 5

# ─── Helpers ───────────────────────────────────────────────────────────────────

def clean(text: str) -> str:
    return " ".join(str(text).split()).strip()

def parse_price(raw) -> str:
    if isinstance(raw, (int, float)):
        return f"{raw} €"
    raw = str(raw).replace("\u00a0", " ").replace("&nbsp;", " ").strip()
    m = re.search(r"[\d\s,\.]+", raw)
    val = m.group(0).strip() if m else raw
    return f"{val} €" if val and "€" not in raw else raw.strip()

# ─── Chrome ────────────────────────────────────────────────────────────────────

def wait_for_chrome(max_attempts: int = 15) -> bool:
    print("🔍  Vérification du port 9222 (Chrome)...")
    for attempt in range(1, max_attempts + 1):
        try:
            urllib.request.urlopen(f"{CDP_URL}/json/version", timeout=2)
            print(f"✅  Chrome détecté sur {CDP_URL}")
            return True
        except Exception:
            print(f"    ⏳ Tentative {attempt}/{max_attempts}...")
            time.sleep(2)
    return False

async def dismiss_overlays(page) -> None:
    for sel in [
        "#onetrust-accept-btn-handler",
        "button:has-text('Accepter')",
        "button:has-text('Accept All')",
        "button:has-text('Tout accepter')",
        "button:has-text('Agree')",
        ".cookie-accept",
        "[aria-label='Close']",
    ]:
        try:
            btn = page.locator(sel).first
            if await btn.is_visible(timeout=1_500):
                await btn.click(timeout=3_000)
                await asyncio.sleep(0.4)
        except Exception:
            pass

# ─── Extraction API ────────────────────────────────────────────────────────────

def extract_activities_from_api(data) -> list[dict]:
    results = []

    def walk(obj, depth=0):
        if depth > 6:
            return
        if isinstance(obj, list):
            for item in obj:
                walk(item, depth + 1)
        elif isinstance(obj, dict):
            if "hits" in obj and isinstance(obj["hits"], list):
                for hit in obj["hits"]:
                    act = build_activity(hit)
                    if act:
                        results.append(act)
                return
            if "results" in obj and isinstance(obj["results"], list):
                walk(obj["results"], depth + 1)
                return
            for key in ("data", "items", "listings", "content", "payload", "activities"):
                if key in obj and isinstance(obj[key], (list, dict)):
                    walk(obj[key], depth + 1)
                    return
            act = build_activity(obj)
            if act:
                results.append(act)

    walk(data)
    return results


def build_activity(obj: dict) -> dict | None:
    if not isinstance(obj, dict):
        return None
    title = clean(
        obj.get("title") or obj.get("name") or
        obj.get("listing_title") or obj.get("label") or ""
    )
    if not title:
        return None

    slug = obj.get("slug") or obj.get("url") or obj.get("permalink") or ""
    if slug and slug.startswith("http"):
        url = slug
    elif slug:
        slug = slug.lstrip("/")
        url = f"{BASE_URL}/{slug}" if "/" in slug else f"{BASE_URL}/fr/annonce/{slug}/voir"
    else:
        oid = obj.get("id") or obj.get("objectID") or ""
        url = f"{BASE_URL}/fr/annonce/{oid}/voir" if oid else ""

    cat = obj.get("category") or obj.get("category_name") or obj.get("activity_type") or ""
    if isinstance(cat, dict):
        cat = cat.get("name") or cat.get("title") or ""
    cat = clean(cat)

    price_raw = (
        obj.get("price") or obj.get("min_price") or obj.get("price_from") or
        obj.get("price_per_person") or obj.get("starting_price") or ""
    )
    if isinstance(price_raw, dict):
        price_raw = price_raw.get("amount") or price_raw.get("value") or ""
    prix = parse_price(price_raw) if price_raw != "" else ""

    rating = (
        obj.get("rating") or obj.get("score") or obj.get("average_rating") or
        obj.get("note") or ""
    )
    if isinstance(rating, dict):
        rating = rating.get("average") or rating.get("score") or ""
    note = str(rating).replace(",", ".").strip() if rating else ""

    nb_avis = (
        obj.get("reviews_count") or obj.get("review_count") or
        obj.get("nb_reviews") or obj.get("ratings_count") or
        obj.get("number_of_reviews") or ""
    )
    if isinstance(nb_avis, dict):
        nb_avis = nb_avis.get("count") or nb_avis.get("total") or ""
    nb_avis = str(nb_avis).strip() if nb_avis else ""

    img = ""
    images = obj.get("images") or obj.get("photos") or []
    if isinstance(images, list) and images:
        first = images[0]
        img = (first.get("url") or first.get("src") or str(first)) if isinstance(first, dict) else str(first)
    if not img:
        img = str(
            obj.get("image") or obj.get("image_url") or
            obj.get("cover_image") or obj.get("thumbnail") or ""
        )

    return {
        "url": url, "categorie": cat, "titre": title,
        "prix": prix, "note": note, "nb_avis": nb_avis, "image_url": img,
    }

# ─── Scroll rapide JS ──────────────────────────────────────────────────────────

async def fast_scroll_collect(page, api_activities: list, api_seen: set) -> None:
    """
    Stratégie en 2 phases :
    
    PHASE 1 — Grands sauts JS (10 000px à la fois)
      → Scroll instantané, attend WAIT_AFTER_SCROLL secondes
      → Les appels API arrivent pendant l'attente
      → Continue jusqu'en bas de page
    
    PHASE 2 — Boucle "scroll to bottom" répétée
      → La page grandit à chaque fois (lazy load DOM)
      → On scrolle tout en bas, on attend, on re-scrolle
      → S'arrête quand plus rien de nouveau après MAX_NO_CHANGE tentatives
    """

    # ── Phase 1 : Grands sauts ─────────────────────────────────────────────
    print("\n⚡  PHASE 1 — Grands sauts JS (10 000px / saut)")
    print(f"    Attente de {WAIT_AFTER_SCROLL}s après chaque saut pour laisser l'API répondre\n")

    pos = 0
    jump = 10_000

    while True:
        pos += jump
        await page.evaluate(f"window.scrollTo(0, {pos})")
        await asyncio.sleep(WAIT_AFTER_SCROLL)

        count = len(api_activities)
        page_height = await page.evaluate("document.body.scrollHeight")
        pct = count / TOTAL_EXPECTED * 100
        print(f"    📦  {count:>4}/{TOTAL_EXPECTED}  ({pct:.0f}%)  | pos={pos:>7}px  page_h={page_height}px")

        if count >= TOTAL_EXPECTED:
            print(f"\n🎉  Objectif atteint en Phase 1 ! {count} activités.")
            return

        if pos >= page_height:
            print(f"\n    → Bas de page atteint ({page_height}px). Passage Phase 2.")
            break

    # ── Phase 2 : Boucle scroll-to-bottom ────────────────────────────────
    print(f"\n⚡  PHASE 2 — Boucle scroll-to-bottom (attend {WAIT_AFTER_SCROLL + WAIT_EXTRA}s)")
    print(f"    Arrêt après {MAX_NO_CHANGE} tentatives sans nouveaux résultats\n")

    no_change  = 0
    last_count = len(api_activities)

    while no_change < MAX_NO_CHANGE:
        await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
        await asyncio.sleep(WAIT_AFTER_SCROLL + WAIT_EXTRA)

        count = len(api_activities)
        pct   = count / TOTAL_EXPECTED * 100
        ph    = await page.evaluate("document.body.scrollHeight")

        if count > last_count:
            print(f"    📦  {count:>4}/{TOTAL_EXPECTED}  ({pct:.0f}%)  | page_h={ph}px  [+{count-last_count}]")
            no_change  = 0
            last_count = count
        else:
            no_change += 1
            print(f"    ⏳  {count:>4}/{TOTAL_EXPECTED}  ({pct:.0f}%)  | sans changement ({no_change}/{MAX_NO_CHANGE})")

        if count >= TOTAL_EXPECTED:
            print(f"\n🎉  Objectif atteint ! {count} activités.")
            return

    print(f"\n⚠️  Arrêt — {len(api_activities)}/{TOTAL_EXPECTED} activités récupérées.")

# ─── Scraping principal ────────────────────────────────────────────────────────

async def scrape() -> list[dict]:
    print("\n" + "="*65)
    print("  INSTRUCTIONS :")
    print("  1. Lance Chrome en mode debug :")
    print('     chrome.exe --remote-debugging-port=9222 --user-data-dir="C:\\ChromeDebug"')
    print("  2. Navigue vers :")
    print(f"     {URL_SEARCH[:80]}...")
    print("  3. Attends le chargement complet")
    print("  4. Reviens ici et appuie sur Entrée")
    print("="*65 + "\n")
    input("  → Prêt ? Appuie sur Entrée : ")

    if not wait_for_chrome():
        print("\n❌ Chrome ne répond pas sur le port 9222.")
        return []

    api_activities: list[dict] = []
    api_seen:       set        = set()

    async with async_playwright() as p:
        try:
            browser = await p.chromium.connect_over_cdp(CDP_URL)
            print(f"✅  Connecté à Chrome ({browser.version})\n")
        except Exception as e:
            print(f"❌ Connexion CDP échouée : {e}")
            return []

        context = (
            browser.contexts[0] if browser.contexts
            else await browser.new_context(viewport={"width": 1440, "height": 900})
        )
        page = await context.new_page()

        # ── Intercepteur API ──────────────────────────────────────────────
        async def on_response(response):
            if "/api/user/v1/search" not in response.url:
                return
            try:
                ct = response.headers.get("content-type", "")
                if "json" not in ct:
                    return
                data = await response.json()
                found = extract_activities_from_api(data)
                for act in found:
                    key = act["url"] or act["titre"]
                    if key and key not in api_seen:
                        api_seen.add(key)
                        api_activities.append(act)
            except Exception:
                pass

        page.on("response", on_response)

        # ── Navigation ────────────────────────────────────────────────────
        print("📄  Chargement de la page de recherche...")
        await page.goto(URL_SEARCH, wait_until="domcontentloaded", timeout=60_000)
        await asyncio.sleep(3.0)
        await dismiss_overlays(page)
        await asyncio.sleep(2.0)
        print(f"    Premier batch API : {len(api_activities)} activités reçues")

        # ── Scroll rapide ─────────────────────────────────────────────────
        await fast_scroll_collect(page, api_activities, api_seen)

        # Pause finale pour les derniers appels en transit
        await asyncio.sleep(3.0)
        await page.close()

    print(f"\n{'='*65}")
    print(f"  🎯 TOTAL FINAL : {len(api_activities)} activités")
    print(f"{'='*65}\n")
    return api_activities

# ─── Export Excel ──────────────────────────────────────────────────────────────

def save_excel(activities: list[dict], filename: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Funbooker Paris"

    HDR_BG = PatternFill("solid", fgColor="00B4D8")
    ODD    = PatternFill("solid", fgColor="F0FAFF")
    EVEN   = PatternFill("solid", fgColor="FFFFFF")
    LINK_C = "0077B6"
    thin   = Side(style="thin", color="90E0EF")
    brd    = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value     = "Funbooker — Activités à Paris"
    c.font      = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    c.fill      = HDR_BG
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:G2")
    c = ws["A2"]
    c.value = (
        f"Extrait le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
        f"  |  {len(activities)} activités"
    )
    c.font      = Font(name="Arial", italic=True, size=10, color="555555")
    c.fill      = PatternFill("solid", fgColor="E0F7FA")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    headers = [
        ("#", 5), ("URL", 52), ("Catégorie", 22),
        ("Titre", 55), ("Prix", 12), ("Note", 8), ("Nb Avis", 10),
    ]
    for col, (h, w) in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font      = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        c.fill      = HDR_BG
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = brd
        ws.column_dimensions[chr(64 + col)].width = w
    ws.row_dimensions[3].height = 22

    for i, act in enumerate(activities, 1):
        row  = i + 3
        fill = ODD if i % 2 else EVEN

        def cell(col, value, hyperlink=None, link=False):
            c = ws.cell(row=row, column=col, value=value)
            c.font = Font(
                name="Arial", size=10,
                color=LINK_C if link else "000000",
                underline="single" if link else None,
            )
            c.alignment = Alignment(vertical="center", wrap_text=(col == 4))
            c.fill   = fill
            c.border = brd
            if hyperlink:
                c.hyperlink = hyperlink

        cell(1, i)
        cell(2, act.get("url",""), hyperlink=act.get("url") or None, link=bool(act.get("url")))
        cell(3, act.get("categorie",""))
        cell(4, act.get("titre",""))
        cell(5, act.get("prix",""))
        cell(6, act.get("note",""))
        cell(7, act.get("nb_avis",""))
        ws.row_dimensions[row].height = 20

    ws.freeze_panes = "A4"
    wb.save(filename)
    print(f"💾  Excel sauvegardé : {filename}  ({len(activities)} activités)")

# ─── Main ──────────────────────────────────────────────────────────────────────

async def main():
    print("=" * 65)
    print("  Funbooker Paris — Scraper v3 (grands scrolls JS rapides)")
    print("=" * 65)

    activities = await scrape()
    if not activities:
        print("\n⚠️  Aucune activité collectée.")
        return

    save_excel(activities, OUTPUT)

    print("\n📋  Aperçu des 10 premières :")
    print(f"  {'#':>3}  {'Catégorie':<20}  {'Titre':<42}  {'Prix':<10}  Note")
    print("  " + "-"*95)
    for i, a in enumerate(activities[:10], 1):
        print(
            f"  {i:>3}  {a.get('categorie','')[:18]:<20}  "
            f"{a.get('titre','')[:40]:<42}  "
            f"{a.get('prix',''):<10}  {a.get('note','')}"
        )
    if len(activities) > 10:
        print(f"  … et {len(activities)-10} autres dans {OUTPUT}")

if __name__ == "__main__":
    asyncio.run(main())